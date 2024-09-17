import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import numpy as np
import concurrent.futures
import os
from typing import List, Tuple
from PIL import Image
import io

# 일반적인 배경색을 제외하는 함수
def is_common_background_color(avg_color, color_tolerance):
    white = np.array([1, 1, 1])
    gray = np.array([0.5, 0.5, 0.5])
    black = np.array([0, 0, 0])

    # 평균 색상과 기준 색상의 차이를 계산
    diff_white = np.linalg.norm(avg_color - white)
    diff_gray = np.linalg.norm(avg_color - gray)
    diff_black = np.linalg.norm(avg_color - black)

    # 색상 차이가 허용 오차 이내이면 일반적인 배경색으로 간주
    if diff_white <= color_tolerance or diff_gray <= color_tolerance or diff_black <= color_tolerance:
        return True
    return False

# 셀의 배경색을 얻는 함수
def get_cell_background_color(page, rect):
    # 셀 영역을 이미지로 클립
    pix = page.get_pixmap(clip=rect, colorspace=fitz.csRGB)
    img_data = pix.tobytes(output="png")
    # 이미지를 PIL 이미지로 로드
    image = Image.open(io.BytesIO(img_data))
    # 이미지를 numpy 배열로 변환하고 RGB 값을 [0,1] 범위로 정규화
    image_np = np.array(image) / 255.0
    # 평균 색상 계산
    avg_color = image_np.mean(axis=(0, 1))
    return avg_color

# 한 페이지에서 표와 셀의 배경색을 추출하는 함수
def extract_tables_from_page(page, color_tolerance):
    tables = page.find_tables()
    page_tables = []

    for table in tables:
        data = []
        for row in table.rows:
            row_data = []
            for cell in row.cells:
                cell_rect = cell.rect
                cell_text = cell.get_text().strip()
                avg_color = get_cell_background_color(page, cell_rect)
                if not is_common_background_color(avg_color, color_tolerance):
                    status = "추가"
                else:
                    status = "유지"
                row_data.append((cell_text, status))
            data.append(row_data)
        page_tables.append(data)
    return page_tables

# 병렬 처리를 위한 함수
def process_page(args):
    page_number, doc, color_tolerance = args
    page = doc.load_page(page_number)
    tables = extract_tables_from_page(page, color_tolerance)
    return tables

# 엑셀 파일로 저장하는 함수
def save_tables_to_excel(all_tables, output_excel_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extracted Tables"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    row_num = 1
    for tables in all_tables:
        for table in tables:
            for row in table:
                for col_num, (text, status) in enumerate(row, start=1):
                    cell = sheet.cell(row=row_num, column=col_num, value=text)
                    cell.alignment = Alignment(wrap_text=True)
                    if status == "추가":
                        cell.fill = yellow_fill
                row_num += 1
            row_num += 2  # 테이블 간 간격
    workbook.save(output_excel_path)
    print(f"Data saved to {output_excel_path}")

# 메인 함수
def main(pdf_path, output_excel_path, color_tolerance=0.3, max_workers=4):
    doc = fitz.open(pdf_path)
    num_pages = len(doc)
    all_tables = []

    # 병렬 처리 설정
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        # 각 페이지 번호와 필요한 인자를 튜플로 전달
        args = [(page_num, pdf_path, color_tolerance) for page_num in range(num_pages)]
        # 결과를 받아옴
        for result in executor.map(process_page_pool, args):
            all_tables.append(result)

    # 결과를 엑셀로 저장
    save_tables_to_excel(all_tables, output_excel_path)

# ProcessPoolExecutor에서 PDF 문서를 다시 열어야 함
def process_page_pool(args):
    page_number, pdf_path, color_tolerance = args
    doc = fitz.open(pdf_path)
    page = doc.load_page(page_number)
    tables = extract_tables_from_page(page, color_tolerance)
    return tables

if __name__ == "__main__":
    pdf_path = "/path/to/your.pdf"  # 처리할 PDF 파일 경로로 변경하세요
    output_excel_path = "/path/to/output.xlsx"  # 저장할 엑셀 파일 경로로 변경하세요
    color_tolerance = 0.3  # 색상 판별 기준 (0에 가까울수록 엄격)
    max_workers = 4  # 병렬 처리 시 사용할 프로세스 수

    main(pdf_path, output_excel_path, color_tolerance, max_workers)
