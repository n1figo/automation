# 1종, 2종, 3종이 있는 위치는 정확히 찾음
# 테이블 끝까지 파싱필요

import PyPDF2
import camelot
import pandas as pd
import fitz
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_text_from_pdf(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text = ""
        page_numbers = []
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text()
            text += page_text + "\n"
            page_numbers.extend([i+1] * len(page_text.split()))
    return text, page_numbers

def find_pages_with_keywords(text, keywords, page_numbers):
    results = {}
    for keyword in keywords:
        pages = []
        for i, (word, page) in enumerate(zip(text.split(), page_numbers)):
            if keyword in word:
                if page not in pages:
                    pages.append(page)
        results[keyword] = pages
    return results

def extract_tables_with_camelot(pdf_path, page_numbers):
    all_tables = []
    for page in page_numbers:
        print(f"Extracting tables from page {page} using Camelot...")
        tables = camelot.read_pdf(pdf_path, pages=str(page), flavor='lattice')
        all_tables.extend(tables
        
        
        
        )
    print(f"Found {len(all_tables)} tables in total")
    return all_tables

def process_tables(tables):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)
    return pd.DataFrame(processed_data)

def save_to_excel(df, output_path, title=None):
    wb = Workbook()
    ws = wb.active

    if title:
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=20, bold=True)
        ws.row_dimensions[1].height = 30

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Data saved to '{output_path}'")

def main():
    uploads_folder = "/workspaces/automation/uploads"
    output_folder = "/workspaces/automation/output"
    
    os.makedirs(output_folder, exist_ok=True)

    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_analysis.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    # Step 1: Find pages with keywords
    text, page_numbers = extract_text_from_pdf(pdf_path)
    keywords = ["[1종]", "[2종]", "[3종]", "선택특약"]
    keyword_results = find_pages_with_keywords(text, keywords, page_numbers)

    for keyword, pages in keyword_results.items():
        print(f"{keyword}이(가) 포함된 페이지: {pages}")

    # Step 2: Extract and process tables for "선택특약"
    special_clause_pages = keyword_results["선택특약"]
    tables = extract_tables_with_camelot(pdf_path, special_clause_pages)
    processed_df = process_tables(tables)

    # Extract title from the first page
    doc = fitz.open(pdf_path)
    first_page = doc[0]
    page_text = first_page.get_text("text")
    title = page_text.strip().split('\n')[0]
    print(f"Extracted title: {title}")

    # Save results to Excel
    save_to_excel(processed_df, output_excel_path, title=f"{title} - 분석 결과")

    print(f"All processed data has been saved to {output_excel_path}")

if __name__ == "__main__":
    main()