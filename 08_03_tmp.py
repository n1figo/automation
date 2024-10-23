import PyPDF2
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import os
import pandas as pd
import camelot
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class InsuranceDocumentAnalyzer:
    def __init__(self):
        self.section_patterns = {
            "종류": r'\[(\d)종\]',
            "특약유형": r'(상해관련|질병관련)\s*특약'
        }
        self.section_pages = {"[1종]": None, "[2종]": None, "[3종]": None}
        self.section_ranges = {}

    def find_section_pages(self, pdf_path: str) -> Dict[str, int]:
        """PDF에서 1종, 2종, 3종의 시작 페이지 찾기"""
        try:
            doc = fitz.open(pdf_path)
            for page_num in range(len(doc)):
                text = doc[page_num].get_text()
                
                # 종 패턴 찾기
                matches = re.finditer(self.section_patterns["종류"], text)
                for match in matches:
                    종_type = f"[{match.group(1)}종]"
                    if self.section_pages[종_type] is None:
                        self.section_pages[종_type] = page_num
                        logger.info(f"{종_type} 시작 페이지: {page_num + 1}")

            # 섹션 범위 설정
            sorted_pages = sorted([(k, v) for k, v in self.section_pages.items() if v is not None], 
                                key=lambda x: x[1])
            
            for i, (종_type, start_page) in enumerate(sorted_pages):
                if i + 1 < len(sorted_pages):
                    end_page = sorted_pages[i + 1][1]
                else:
                    end_page = len(doc)
                self.section_ranges[종_type] = (start_page, end_page)
                
            doc.close()
            return self.section_pages
            
        except Exception as e:
            logger.error(f"Error finding section pages: {e}")
            return {}

class TableExtractor:
    def __init__(self):
        self.font_size_threshold = 10
        self.title_max_length = 50

    def extract_tables_from_section(self, pdf_path: str, start_page: int, end_page: int) -> List[Tuple[str, pd.DataFrame]]:
        """섹션 범위 내의 표 추출"""
        try:
            results = []
            
            for page_num in range(start_page, end_page):
                # 페이지 텍스트 분석
                doc = fitz.open(pdf_path)
                page = doc[page_num]
                text = page.get_text()
                
                # 상해관련 또는 질병관련 특약 확인
                if re.search(r'(상해관련|질병관련)\s*특약', text):
                    # 표 위의 제목 찾기
                    title = self.extract_table_title(page)
                    
                    # 표 추출
                    tables = self.extract_with_camelot(pdf_path, page_num + 1)
                    
                    for table in tables:
                        df = self.clean_table(table.df)
                        if not df.empty:
                            results.append((title, df, page_num + 1))
                            
                doc.close()
                
            return results
            
        except Exception as e:
            logger.error(f"Error extracting tables from section: {e}")
            return []

    def extract_table_title(self, page) -> str:
        """표 위의 제목 추출"""
        try:
            blocks = page.get_text("dict")["blocks"]
            table_block = None
            title_block = None
            
            # 표 블록 찾기
            for block in blocks:
                if "lines" in block:
                    text = " ".join(span["text"] for line in block["lines"] 
                                  for span in line["spans"])
                    if "특약" in text and len(block["lines"]) > 1:
                        table_block = block
                        break
            
            if table_block:
                table_top = table_block["bbox"][1]
                # 표 위의 가장 가까운 텍스트 블록 찾기
                potential_titles = [
                    b for b in blocks if "lines" in b 
                    and b["bbox"][3] < table_top
                    and len(" ".join(span["text"] for line in b["lines"] 
                                   for span in line["spans"])) < self.title_max_length
                ]
                
                if potential_titles:
                    title_block = max(potential_titles, 
                                    key=lambda x: x["bbox"][3])
                    return " ".join(span["text"] for line in title_block["lines"] 
                                  for span in line["spans"])
            
            return "Untitled Table"
            
        except Exception as e:
            logger.error(f"Error extracting table title: {e}")
            return "Untitled Table"

    def extract_with_camelot(self, pdf_path: str, page_num: int) -> List:
        """Camelot을 사용한 표 추출"""
        try:
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num),
                flavor='lattice'
            )
            if not tables:
                tables = camelot.read_pdf(
                    pdf_path,
                    pages=str(page_num),
                    flavor='stream'
                )
            return tables
        except Exception as e:
            logger.error(f"Camelot extraction failed: {str(e)}")
            return []

    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """표 데이터 정제"""
        try:
            df = df.dropna(how='all')
            df = df[~df.iloc[:, 0].str.contains("※|주)", regex=False, na=False)]
            return df
        except Exception as e:
            logger.error(f"Error cleaning table: {e}")
            return pd.DataFrame()

class ExcelWriter:
    @staticmethod
    def save_to_excel(sections_data: Dict[str, List[Tuple[str, pd.DataFrame, int]]], output_path: str):
        """섹션별 데이터를 Excel 파일로 저장"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for section, tables in sections_data.items():
                    if not tables:
                        continue
                        
                    sheet_name = section.replace("[", "").replace("]", "")
                    current_row = 0
                    
                    for title, df, page_num in tables:
                        # 제목 쓰기
                        title_df = pd.DataFrame([[f"{title} (페이지: {page_num})"]], columns=[''])
                        title_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row,
                            index=False,
                            header=False
                        )
                        
                        # 표 데이터 쓰기
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row + 2,
                            index=False
                        )
                        
                        # 스타일 적용
                        worksheet = writer.sheets[sheet_name]
                        
                        # 제목 스타일링
                        title_cell = worksheet.cell(row=current_row + 1, column=1)
                        title_cell.font = Font(bold=True, size=12)
                        title_cell.fill = PatternFill(start_color='E6E6E6',
                                                    end_color='E6E6E6',
                                                    fill_type='solid')
                        
                        current_row += len(df) + 5

            logger.info(f"Successfully saved tables to {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")

def main():
    try:
        # 파일 경로 설정
        pdf_path = "/workspaces/automation/uploads/KB 9회주는 암보험Plus(무배당)(24.05)_요약서_10.1판매_v1.0_앞단.pdf"
        output_path = "보험특약표.xlsx"
        
        if not os.path.exists(pdf_path):
            logger.error("PDF file not found")
            return

        # 문서 분석기 초기화
        document_analyzer = InsuranceDocumentAnalyzer()
        section_pages = document_analyzer.find_section_pages(pdf_path)
        
        if not section_pages:
            logger.error("No sections found in the document")
            return

        # 표 추출기 초기화
        table_extractor = TableExtractor()
        sections_data = {}
        
        # 각 섹션별 표 추출
        for section, (start_page, end_page) in document_analyzer.section_ranges.items():
            logger.info(f"Processing {section} (pages {start_page + 1} to {end_page})")
            tables = table_extractor.extract_tables_from_section(pdf_path, start_page, end_page)
            sections_data[section] = tables

        # 결과 저장
        if any(sections_data.values()):
            ExcelWriter.save_to_excel(sections_data, output_path)
            logger.info("Processing completed successfully")
        else:
            logger.error("No tables extracted from any section")

    except Exception as e:
        logger.error(f"Processing error: {str(e)}")

if __name__ == "__main__":
    main()