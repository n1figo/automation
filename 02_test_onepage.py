import os
import logging
import fitz
import cv2
import numpy as np
from PIL import Image
import pandas as pd
from paddleocr import PPStructure
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import time
import gc

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
    datefmt='%Y/%m/%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

class PDFHighlightAnalyzer:
    def __init__(self):
        """초기화"""
        logger.info("PDFHighlightAnalyzer 초기화 시작")
        try:
            # PaddleOCR 초기화 - 최소 설정
            self.table_engine = PPStructure(
                show_log=True,
                table=True,
                ocr=False,  # OCR 비활성화
                layout=False,  # 레이아웃 분석 비활성화
                lang='en',
                use_angle_cls=False,
                recovery=False,
                cpu_threads=2,  # CPU 스레드 수 감소
                det_db_score_mode='fast',  # 빠른 모드
                det_limit_side_len=1024,  # 이미지 크기 제한
                max_batch_size=1  # 배치 크기 제한
            )
            logger.info("PaddleOCR 초기화 완료")

            # 색상 범위 정의 (HSV)
            self.color_ranges = {
                'yellow': [(20, 60, 60), (45, 255, 255)],
                'green': [(35, 60, 60), (85, 255, 255)],
                'blue': [(95, 60, 60), (145, 255, 255)]
            }
            logger.info("색상 범위 정의 완료")
            
        except Exception as e:
            logger.error(f"초기화 중 오류 발생: {str(e)}", exc_info=True)
            raise

    def preprocess_image(self, image):
        """이미지 전처리"""
        try:
            # 이미지 크기 제한
            height, width = image.shape[:2]
            max_dimension = 1024
            if width > max_dimension:
                scale = max_dimension / width
                width = max_dimension
                height = int(height * scale)
                image = cv2.resize(image, (width, height), interpolation=cv2.INTER_AREA)

            # 노이즈 제거 (매개변수 조정)
            denoised = cv2.fastNlMeansDenoisingColored(image, None, 7, 7, 5, 15)
            
            # 대비 향상
            lab = cv2.cvtColor(denoised, cv2.COLOR_BGR2LAB)
            l, a, b = cv2.split(lab)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            cl = clahe.apply(l)
            enhanced = cv2.merge((cl,a,b))
            
            result = cv2.cvtColor(enhanced, cv2.COLOR_LAB2BGR)
            
            # 메모리 정리
            del denoised, lab, l, a, b, cl, enhanced
            gc.collect()
            
            return result
            
        except Exception as e:
            logger.error(f"이미지 전처리 중 오류: {str(e)}", exc_info=True)
            raise

    def analyze_color(self, cell_img):
        """셀 이미지의 하이라이트 색상 분석"""
        try:
            # 이미지 크기 제한
            max_size = 256
            height, width = cell_img.shape[:2]
            if width > max_size or height > max_size:
                scale = min(max_size/width, max_size/height)
                cell_img = cv2.resize(cell_img, (int(width*scale), int(height*scale)))

            hsv = cv2.cvtColor(cell_img, cv2.COLOR_BGR2HSV)
            detected_colors = []
            
            for color_name, (lower, upper) in self.color_ranges.items():
                mask = cv2.inRange(hsv, np.array(lower), np.array(upper))
                pixel_count = np.sum(mask > 0)
                coverage = pixel_count / (mask.shape[0] * mask.shape[1])
                
                if coverage > 0.25:
                    detected_colors.append({
                        'color': color_name,
                        'coverage': coverage
                    })
                
                del mask
            
            del hsv
            gc.collect()
            
            if detected_colors:
                strongest_color = max(detected_colors, key=lambda x: x['coverage'])
                return [strongest_color['color']]
            return []
            
        except Exception as e:
            logger.error(f"색상 분석 중 오류: {str(e)}", exc_info=True)
            raise

    def extract_page_image(self, pdf_path, page_num):
        """PDF에서 특정 페이지 이미지 추출"""
        try:
            doc = fitz.open(pdf_path)
            page = doc[page_num - 1]
            
            # 낮은 해상도로 설정
            zoom = 1.0
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            doc.close()
            
            # 이미지 크기 제한
            width, height = img.size
            max_dimension = 1024
            if width > max_dimension:
                scale = max_dimension / width
                new_width = max_dimension
                new_height = int(height * scale)
                img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            
            result = np.array(img)
            
            # 메모리 정리
            del img, pix
            gc.collect()
            
            return result
            
        except Exception as e:
            logger.error(f"PDF 페이지 추출 중 오류: {str(e)}", exc_info=True)
            raise

    def analyze_highlighted_cells(self, table_data):
        """하이라이트된 셀들의 간단한 분석"""
        cells = table_data['highlighted_cells']
        if not cells:
            return None

        rows = [cell['row'] for cell in cells]
        cols = [cell['col'] for cell in cells]
        
        analysis = {
            'total_highlights': len(cells),
            'row_range': f"{min(rows)} - {max(rows)}",
            'col_range': f"{min(cols)} - {max(cols)}",
            'colors_used': list(set(sum([cell['colors'] for cell in cells], [])))
        }
        
        # 연속된 행 확인
        sorted_rows = sorted(set(rows))
        consecutive_rows = all(sorted_rows[i] + 1 == sorted_rows[i+1] 
                             for i in range(len(sorted_rows)-1))
        analysis['consecutive_rows'] = consecutive_rows
        
        return analysis

    def process_page(self, pdf_path, page_num):
        """특정 페이지 처리"""
        try:
            start_time = time.time()
            logger.info(f"페이지 {page_num} 처리 시작")
            
            # 메모리 정리
            gc.collect()
            
            # 이미지 추출 및 전처리
            image = self.extract_page_image(pdf_path, page_num)
            processed_image = self.preprocess_image(image)
            
            # 큰 이미지 메모리 해제
            del image
            gc.collect()
            
            # 표 분석
            result = self.table_engine(processed_image)
            tables_data = []
            
            # 처리된 이미지 메모리 해제
            del processed_image
            gc.collect()
            
            for idx, region in enumerate(result):
                if region['type'] == 'table':
                    logger.info(f"표 {idx+1} 분석 중")
                    table_img = region['img']
                    cells = region['cells']
                    
                    table_data = {
                        'table_index': idx,
                        'highlighted_cells': []
                    }
                    
                    for cell in cells:
                        bbox = cell['bbox']
                        cell_img = table_img[bbox[1]:bbox[3], bbox[0]:bbox[2]]
                        colors = self.analyze_color(cell_img)
                        
                        if colors:
                            cell_info = {
                                'row': cell['row_idx'],
                                'col': cell['col_idx'],
                                'text': cell['text'],
                                'colors': colors
                            }
                            table_data['highlighted_cells'].append(cell_info)
                            logger.debug(f"하이라이트 감지: {cell_info}")
                        
                        del cell_img
                    
                    if table_data['highlighted_cells']:
                        table_data['analysis'] = self.analyze_highlighted_cells(table_data)
                        tables_data.append(table_data)
                    
                    del table_img
                    gc.collect()
            
            process_time = time.time() - start_time
            logger.info(f"페이지 {page_num} 처리 완료 (소요시간: {process_time:.2f}초)")
            return tables_data
            
        except Exception as e:
            logger.error(f"페이지 처리 중 오류: {str(e)}", exc_info=True)
            raise

def save_to_excel(tables_data, output_path):
    """결과를 Excel 파일로 저장"""
    try:
        logger.info(f"Excel 저장 시작: {output_path}")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        for table_data in tables_data:
            if table_data['highlighted_cells']:
                sheet_name = f"Table_{table_data['table_index']}"
                ws = wb.create_sheet(sheet_name)
                
                # 제목 추가
                ws.cell(row=1, column=1, value="하이라이트 분석 결과").font = Font(bold=True, size=12)
                
                # 분석 결과 추가
                if table_data.get('analysis'):
                    row = 3
                    for key, value in table_data['analysis'].items():
                        ws.cell(row=row, column=1, value=key)
                        ws.cell(row=row, column=2, value=str(value))
                        row += 1
                
                # 데이터 헤더
                headers = ['행', '열', '내용', '하이라이트 색상']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row+2, column=col, value=header).font = Font(bold=True)
                
                # 데이터 추가
                yellow_fill = PatternFill(start_color='FFFF00', 
                                        end_color='FFFF00', 
                                        fill_type='solid')
                
                for i, cell in enumerate(table_data['highlighted_cells'], 1):
                    row_num = row + 2 + i
                    ws.cell(row=row_num, column=1, value=cell['row'])
                    ws.cell(row=row_num, column=2, value=cell['col'])
                    ws.cell(row=row_num, column=3, value=cell['text'])
                    ws.cell(row=row_num, column=4, value=', '.join(cell['colors']))
                    
                    # 하이라이트된 행 배경색 지정
                    for col in range(1, 5):
                        ws.cell(row=row_num, column=col).fill = yellow_fill
                
                # 열 너비 자동 조정
                for column_cells in ws.columns:
                    length = max(len(str(cell.value) or "") for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        wb.save(output_path)
        logger.info(f"Excel 파일 저장 완료: {output_path}")
        
    except Exception as e:
        logger.error(f"Excel 저장 중 오류: {str(e)}", exc_info=True)
        raise

def main():
    # 설정
    pdf_path = "/workspaces/automation/uploads/KB 9회주는 암보험Plus(무배당)(24.05)_요약서_10.1판매_v1.0_앞단.pdf"
    output_path = "page_59_analysis.xlsx"
    
    try:
        total_start_time = time.time()
        logger.info("프로그램 시작")
        
        # 메모리 정리
        gc.collect()
        
        # 분석기 초기화
        analyzer = PDFHighlightAnalyzer()
        
        # 59페이지 분석
        results = analyzer.process_page(pdf_path, page_num=59)
        
        # 결과 저장
        if results:
            save_to_excel(results, output_path)
            logger.info(f"분석 완료. 결과가 {output_path}에 저장되었습니다.")
        else:
            logger.warning("감지된 하이라이트가 없습니다.")
        
        total_time = time.time() - total_start_time
        logger.info(f"프로그램 종료 (총 소요시간: {total_time:.2f}초)")
        
    except Exception as e:
        logger.error(f"실행 중 오류 발생: {str(e)}", exc_info=True)
        raise

if __name__ == "__main__":
    main()