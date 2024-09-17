import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import os
from typing import List, Tuple, Any
from collections import defaultdict
import re

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 텍스트를 개별 보장 항목으로 분리
def split_coverage_items(text: str) -> List[str]:
    if not isinstance(text, str):
        return [str(text)]  # None 또는 다른 타입의 값을 문자열로 변환
    items = re.findall(r'[가-힣a-zA-Z0-9]+\([^)]+\)(?:\【[^】]+\】)?', text)
    return items if items else [text]

# 색상 체크 함수
def check_text_color(span) -> bool:
    common_colors = {(1, 1, 1), (0, 0, 0), (0.5, 0.5, 0.5)}  # RGB 값
    color = span.get('color', None)
    return color and color not in common_colors

# 텍스트 및 색상 추출 함수
def extract_text_and_colors(page: fitz.Page) -> List[Tuple[str, str]]:
    blocks = page.get_text("dict")["blocks"]
    extracted_text = []

    for block in blocks:
        if 'lines' in block:
            for line in block['lines']:
                for span in line['spans']:
                    text = span['text']
                    status = "추가" if check_text_color(span) else "유지"
                    extracted_text.append((text, status))
    
    return extracted_text

# PDF 테이블 추출 클래스
class PDFTableExtractor:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        
    def extract_tables_with_titles(self) -> List[Tuple[str, pd.DataFrame]]:
        all_tables = []
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            tables = self.extract_tables_from_page(page)
            titled_tables = self._assign_titles_to_tables(page, tables)
            all_tables.extend(titled_tables)
        return self._merge_tables_with_same_title(all_tables)
    
    def extract_tables_from_page(self, page: fitz.Page) -> List[Any]:
        tables = page.find_tables()
        return tables
    
    def _assign_titles_to_tables(self, page: fitz.Page, tables: List[Any]) -> List[Tuple[str, pd.DataFrame]]:
        titled_tables = []
        for table in tables:
            title = self._find_table_title(page, table)
            df = self._table_to_dataframe(table)
            titled_tables.append((title, df))
        return titled_tables
    
    def _find_table_title(self, page: fitz.Page, table: Any) -> str:
        blocks = page.get_text("dict")["blocks"]
        table_top = table.bbox[1]  # y0 좌표
        potential_titles = []
        for b in blocks:
            if 'lines' in b:
                for l in b['lines']:
                    for s in l['spans']:
                        if s['bbox'][3] < table_top and s['bbox'][3] > table_top - 50:
                            potential_titles.append(s['text'])
        
        if potential_titles:
            return " ".join(potential_titles).strip()
        return "Untitled Table"
    
    def _table_to_dataframe(self, table: Any) -> pd.DataFrame:
        data = table.extract()
        columns = data[0]
        rows = data[1:]
        
        processed_rows = []
        for row in rows:
            processed_row = []
            for cell in row:
                items = split_coverage_items(cell)
                processed_row.extend(items)
            processed_rows.append(processed_row)
        
        max_len = max(len(row) for row in processed_rows)
        columns = columns + [f'추가항목_{i}' for i in range(len(columns), max_len)]
        
        df = pd.DataFrame(processed_rows, columns=columns)
        return df
    
    def _merge_tables_with_same_title(self, tables: List[Tuple[str, pd.DataFrame]]) -> List[Tuple[str, pd.DataFrame]]:
        merged_tables = defaultdict(list)
        for title, df in tables:
            merged_tables[title].append(df)
        
        return [(title, pd.concat(dfs, ignore_index=True)) for title, dfs in merged_tables.items()]

# 엑셀 작성 및 변경사항 적용 클래스
class ExcelWriterWithChanges:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Extracted Tables"
        
    def write_tables_with_changes(self, tables: List[Tuple[str, pd.DataFrame]]):
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row = 1
        
        for title, df in tables:
            self.sheet.cell(row=row, column=1, value=title)
            row += 1
            
            for r in dataframe_to_rows(df, index=False, header=True):
                for col_num, cell_value in enumerate(r, start=1):
                    cell = self.sheet.cell(row=row, column=col_num)
                    cell.value = str(cell_value) if cell_value is not None else ''
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if "추가" in str(cell_value):
                        cell.fill = yellow_fill
                row += 1
            row += 2
        
        # 열 너비 자동 조정
        for column in self.sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.sheet.column_dimensions[column_letter].width = adjusted_width
        
        self.workbook.save(self.output_path)
        logger.info(f"Tables saved with changes to {self.output_path}")

# 메인 함수
def main(pdf_path: str, output_excel_path: str):
    try:
        extractor = PDFTableExtractor(pdf_path)
        
        # 테이블 추출 및 색상 감지
        tables = extractor.extract_tables_with_titles()
        
        # Excel 작성 및 변경사항 적용
        writer = ExcelWriterWithChanges(output_excel_path)
        writer.write_tables_with_changes(tables)
        
        logger.info("Table extraction, writing, and text extraction with changes completed successfully.")
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}", exc_info=True)

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables.xlsx"
    main(pdf_path, output_excel_path)