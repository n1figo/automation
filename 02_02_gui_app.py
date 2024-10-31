import PyPDF2
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import os
import pandas as pd
import camelot
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from sentence_transformers import SentenceTransformer
from scipy.spatial.distance import cosine
import cv2
from PIL import Image
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from datetime import datetime
import threading

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class HighlightDetector:
    def __init__(self):
        self.saturation_threshold = 30
        self.kernel_size = (5, 5)

    def pdf_to_image(self, page: fitz.Page) -> np.ndarray:
        """PDF 페이지를 이미지로 변환"""
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return np.array(img)

    def detect_highlights(self, image: np.ndarray) -> List[np.ndarray]:
        """하이라이트 영역 감지 및 컨투어 반환"""
        hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
        s = hsv[:, :, 1]
        v = hsv[:, :, 2]

        saturation_mask = s > self.saturation_threshold

        _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)

        kernel = np.ones(self.kernel_size, np.uint8)
        cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
        cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)

        contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        return contours

    def get_highlight_regions(self, contours: List[np.ndarray], image_height: int) -> List[Tuple[float, float]]:
        """컨투어를 하이라이트 영역으로 변환 (상단, 하단)"""
        regions = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            top = image_height - (y + h)
            bottom = image_height - y
            regions.append((top, bottom))
        return regions

    def check_highlight(self, row_range: Tuple[float, float], highlight_regions: List[Tuple[float, float]]) -> bool:
        """행이 하이라이트 영역과 겹치는지 확인"""
        row_top, row_bottom = row_range
        for region_top, region_bottom in highlight_regions:
            if (region_top <= row_top <= region_bottom) or \
               (region_top <= row_bottom <= region_bottom) or \
               (row_top <= region_top <= row_bottom) or \
               (row_top <= region_bottom <= row_bottom):
                return True
        return False

class TitleMatcher:
    def __init__(self):
        try:
            # 실행 파일 경로 또는 현재 스크립트 경로 기준으로 모델 경로 설정
            if getattr(sys, 'frozen', False):
                # PyInstaller로 생성된 실행 파일인 경우
                application_path = os.path.dirname(sys.executable)
            else:
                # 일반 Python 스크립트인 경우
                application_path = os.path.dirname(os.path.abspath(__file__))
                
            model_path = os.path.join(application_path, "models", "distiluse-base-multilingual-cased-v1")
            logger.info(f"모델 경로: {model_path}")
            
            self.model = SentenceTransformer(model_path)
            logger.info("로컬 모델 로드 성공")
        except Exception as e:
            logger.error(f"모델 로드 실패: {str(e)}")
            raise

    def get_embedding(self, text: str) -> np.ndarray:
        return self.model.encode(text)

    def calculate_similarity(self, text1: str, text2: str) -> float:
        emb1 = self.get_embedding(text1)
        emb2 = self.get_embedding(text2)
        return 1 - cosine(emb1, emb2)
    

class InsuranceDocumentAnalyzer:
    def __init__(self):
        self.section_patterns = {
            "종류": r'\[(\d)종\]',
        }
        self.section_pages = {"[1종]": None, "[2종]": None, "[3종]": None}
        self.section_ranges = {}
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]

    def find_section_pages(self, pdf_path: str) -> Dict[str, int]:
        """PDF에서 1종, 2종, 3종의 시작 페이지 찾기"""
        try:
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            for page_num in range(total_pages):
                text = doc[page_num].get_text()

                # 종 구분
                matches = re.finditer(self.section_patterns["종류"], text)
                for match in matches:
                    종_type = f"[{match.group(1)}종]"
                    if self.section_pages[종_type] is None:
                        self.section_pages[종_type] = page_num
                        logger.info(f"{종_type} 시작 페이지: {page_num + 1}")

            # 섹션 범위 설정
            self._set_section_ranges(total_pages)

            doc.close()
            return self.section_pages

        except Exception as e:
            logger.error(f"Error finding section pages: {e}")
            return {}

    def _set_section_ranges(self, total_pages):
        # 종 섹션 범위 설정
        found_sections = [v for v in self.section_pages.values() if v is not None]
        if not found_sections:
            logger.warning("1종, 2종, 3종 패턴을 찾지 못했습니다. 문서 전체를 대상으로 진행합니다.")
            self.section_ranges["전체"] = (0, total_pages)
        else:
            sorted_pages = sorted([(k, v) for k, v in self.section_pages.items() if v is not None],
                                  key=lambda x: x[1])

            for i, (종_type, start_page) in enumerate(sorted_pages):
                if i + 1 < len(sorted_pages):
                    end_page = sorted_pages[i + 1][1]
                else:
                    end_page = total_pages
                self.section_ranges[종_type] = (start_page, end_page)
                logger.info(f"{종_type} 범위: {start_page + 1} ~ {end_page}")

class TableExtractor:
    def __init__(self):
        self.matcher = TitleMatcher()
        self.highlight_detector = HighlightDetector()
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]
        self.max_distance = 50

    def get_titles_with_positions(self, page) -> List[Dict]:
        """페이지에서 제목과 위치 정보 추출"""
        titles = []
        blocks = page.get_text("dict")["blocks"]

        for block in blocks:
            if block.get("lines"):
                text = ""
                y_top = block["bbox"][1]
                y_bottom = block["bbox"][3]

                for line in block["lines"]:
                    for span in line["spans"]:
                        text += span["text"] + " "

                text = text.strip()
                if text and any(re.search(pattern, text) for pattern in self.title_patterns):
                    titles.append({
                        "text": text,
                        "y_top": y_top,
                        "y_bottom": y_bottom,
                        "bbox": block["bbox"],
                        "used": False
                    })

        return sorted(titles, key=lambda x: x["y_top"])

    def get_table_positions(self, table, page_height):
        """Camelot 표의 위치를 PyMuPDF 좌표계로 변환"""
        x0, y0, x1, y1 = table._bbox
        converted_y0 = page_height - y1  # 상단
        converted_y1 = page_height - y0  # 하단

        return {
            "y_top": converted_y0,
            "y_bottom": converted_y1,
            "bbox": (x0, converted_y0, x1, converted_y1)
        }

    def match_title_to_table(self, titles, table_position):
        """표와 가장 적절한 제목 매칭"""
        best_title = None
        min_distance = float('inf')

        for title in titles:
            if title["used"]:
                continue

            distance = table_position["y_top"] - title["y_bottom"]
            if 0 < distance < self.max_distance and distance < min_distance:
                best_title = title
                min_distance = distance

        if best_title:
            best_title["used"] = True
            return best_title["text"], min_distance
        return None, None

    def process_table_with_highlights(self, table, highlight_regions: List[Tuple[float, float]], page_height: int, table_index: int, page_num: int) -> pd.DataFrame:
        """테이블 처리 및 하이라이트 감지"""
        df = table.df.copy()
        x1, y1, x2, y2 = table._bbox

        table_height = y2 - y1
        num_rows = len(df)
        if num_rows == 0:
            return df

        row_height = table_height / num_rows
        processed_data = []

        for row_index in range(num_rows):
            row_data = df.iloc[row_index].copy()

            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height

            row_highlighted = self.highlight_detector.check_highlight((row_top, row_bottom), highlight_regions)
            row_data['변경사항'] = '추가' if row_highlighted else ''
            row_data['Table_Number'] = table_index + 1
            row_data['페이지'] = page_num
            processed_data.append(row_data)

        return pd.DataFrame(processed_data)
    

def extract_with_camelot(self, pdf_path: str, page_num: int) -> List:
        """Camelot을 사용한 표 추출"""
        try:
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num),
                flavor='lattice'
            )
            if not tables:
                tables = camelot.read_pdf(
                    pdf_path,
                    pages=str(page_num),
                    flavor='stream'
                )
            return tables
        except Exception as e:
            logger.error(f"Camelot extraction failed: {str(e)}")
            return []

    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """표 데이터 정제"""
        try:
            df = df.dropna(how='all')
            df = df[~df.iloc[:, 0].str.contains("※|주)", regex=False, na=False)]
            return df
        except Exception as e:
            logger.error(f"Error cleaning table: {e}")
            return pd.DataFrame()

    def extract_tables_from_section(self, pdf_path: str, start_page: int, end_page: int) -> List[Tuple[str, pd.DataFrame, int]]:
        try:
            results = []

            for page_num in range(start_page, end_page):
                logger.info(f"Processing page {page_num + 1}")
                doc = fitz.open(pdf_path)
                page = doc[page_num]

                # 페이지를 이미지로 변환하고 하이라이트 감지
                image = self.highlight_detector.pdf_to_image(page)
                contours = self.highlight_detector.detect_highlights(image)
                highlight_regions = self.highlight_detector.get_highlight_regions(contours, image.shape[0])

                # 제목 추출
                titles = self.get_titles_with_positions(page)

                # 표 추출
                tables = self.extract_with_camelot(pdf_path, page_num + 1)

                if tables:
                    page_height = page.rect.height

                    for table_idx, table in enumerate(tables):
                        table_position = self.get_table_positions(table, page_height)
                        title, distance = self.match_title_to_table(titles, table_position)

                        # 테이블 처리 및 하이라이트 적용
                        df = self.process_table_with_highlights(
                            table, highlight_regions, image.shape[0], table_idx, page_num + 1
                        )
                        df = self.clean_table(df)
                        if not df.empty:
                            if title:
                                logger.info(f"Found table {table_idx + 1} with title: {title} (distance: {distance:.1f})")
                            else:
                                title = f"Table_{table_idx + 1}"
                                logger.warning(f"No matching title found for table {table_idx + 1} on page {page_num + 1}")

                            results.append((title, df, page_num + 1))

                doc.close()

            return results

        except Exception as e:
            logger.error(f"Error extracting tables from section: {e}")
            return []

class ExcelWriter:
    @staticmethod
    def save_to_excel(sections_data: Dict[str, List[Tuple[str, pd.DataFrame, int]]], output_path: str):
        """엑셀로 저장하고 하이라이트 적용"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for section, tables in sections_data.items():
                    if not tables:
                        continue

                    # 시트 이름을 '1종', '2종', '3종'으로 제한
                    if section not in ["[1종]", "[2종]", "[3종]"]:
                        continue

                    sheet_name = section.replace("[", "").replace("]", "")
                    current_row = 0

                    for title, df, page_num in tables:
                        # 제목 쓰기 (페이지 번호 포함)
                        title_df = pd.DataFrame([[f"{title} (페이지: {page_num})"]], columns=[''])
                        title_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row,
                            index=False,
                            header=False
                        )

                        # 표 데이터 쓰기
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row + 2,
                            index=False
                        )

                        # 스타일 적용
                        workbook = writer.book
                        worksheet = writer.sheets[sheet_name]

                        # 제목 스타일링
                        title_cell = worksheet.cell(row=current_row + 1, column=1)
                        title_cell.font = Font(bold=True, size=12)
                        title_cell.fill = PatternFill(start_color='E6E6E6',
                                                      end_color='E6E6E6',
                                                      fill_type='solid')

                        # 하이라이트 스타일링
                        yellow_fill = PatternFill(start_color='FFFF00',
                                                  end_color='FFFF00',
                                                  fill_type='solid')

                        # '변경사항' 컬럼 인덱스 찾기
                        try:
                            change_col_index = df.columns.get_loc('변경사항') + 1  # OpenPyXL는 1부터 시작
                        except KeyError:
                            logger.warning("'변경사항' 컬럼을 찾을 수 없습니다.")
                            change_col_index = None

                        # 데이터 행 시작 위치 (제목 + 헤더)
                        data_start_row = current_row + 3

                        # '변경사항'이 '추가'인 행에 하이라이트 적용
                        if change_col_index is not None:
                            for idx in range(len(df)):
                                cell_value = worksheet.cell(row=data_start_row + idx, column=change_col_index).value
                                if cell_value == '추가':
                                    for col in range(1, len(df.columns) + 1):
                                        worksheet.cell(row=data_start_row + idx, column=col).fill = yellow_fill

                        # 자동 열 너비 조정
                        for column_cells in worksheet.columns:
                            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                            worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

                        current_row += len(df) + 5

            logger.info(f"Successfully saved tables to {output_path}")

        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")



class PDFAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("KB손해보험 상품개정 자동화 서비스")
        self.root.geometry("800x600")
        self.setup_gui()
        self.setup_logging()

    def setup_logging(self):
        # 로그 파일 설정
        log_filename = f'pdf_analyzer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
        file_handler = logging.FileHandler(log_filename, encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)

    def setup_gui(self):
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목
        title_label = ttk.Label(
            main_frame,
            text="KB손해보험 상품개정 자동화 서비스",
            font=("Helvetica", 16, "bold")
        )
        title_label.pack(pady=10)
        
        # 파일 선택 프레임
        file_frame = ttk.LabelFrame(main_frame, text="PDF 파일 선택", padding="10")
        file_frame.pack(fill=tk.X, pady=10)
        
        self.file_path_var = tk.StringVar()
        self.file_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=60)
        self.file_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        browse_button = ttk.Button(
            file_frame,
            text="찾아보기",
            command=self.browse_file
        )
        browse_button.pack(side=tk.LEFT, padx=5)
        
        # 진행 상태 표시
        self.progress_var = tk.StringVar(value="대기 중...")
        progress_label = ttk.Label(main_frame, textvariable=self.progress_var)
        progress_label.pack(pady=5)
        
        self.progress_bar = ttk.Progressbar(main_frame, mode='determinate', length=300)
        self.progress_bar.pack(pady=5)
        
        # 처리 시작 버튼
        self.process_button = ttk.Button(
            main_frame,
            text="분석 시작",
            command=self.process_pdf,
            state=tk.DISABLED
        )
        self.process_button.pack(pady=10)
        
        # 로그 영역
        log_frame = ttk.LabelFrame(main_frame, text="처리 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.log_text = tk.Text(log_frame, height=15)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def log_message(self, message, level="INFO"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        
        # 로거에도 기록
        if level == "INFO":
            logger.info(message)
        elif level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="PDF 파일 선택",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.process_button['state'] = tk.NORMAL
            self.log_message(f"파일이 선택되었습니다: {file_path}")

    def process_pdf_thread(self, pdf_path):
        try:
            # 출력 폴더 설정
            output_folder = os.path.join(os.path.dirname(pdf_path), "output")
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, f"보험특약표_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            
            # 처리 시작
            self.log_message("문서 분석 시작...")
            self.progress_bar['value'] = 10
            
            document_analyzer = InsuranceDocumentAnalyzer()
            document_analyzer.find_section_pages(pdf_path)
            
            if not document_analyzer.section_ranges:
                self.log_message("섹션을 찾지 못했습니다. 문서 전체를 대상으로 진행합니다.", "WARNING")
                doc = fitz.open(pdf_path)
                total_pages = len(doc)
                document_analyzer.section_ranges["전체"] = (0, total_pages)
                doc.close()
            
            self.log_message("표 추출 시작...")
            self.progress_bar['value'] = 30
            
            table_extractor = TableExtractor()
            sections_data = {}
            
            total_sections = len(document_analyzer.section_ranges)
            for idx, (section, (start_page, end_page)) in enumerate(document_analyzer.section_ranges.items(), 1):
                if section not in ["[1종]", "[2종]", "[3종]"]:
                    continue
                    
                progress = 30 + (idx / total_sections) * 50
                self.progress_bar['value'] = progress
                self.progress_var.set(f"{section} 처리 중...")
                
                self.log_message(f"처리 중: {section} (페이지 {start_page + 1} ~ {end_page})")
                tables = table_extractor.extract_tables_from_section(pdf_path, start_page, end_page)
                sections_data[section] = tables
                
            if any(sections_data.values()):
                self.progress_bar['value'] = 90
                self.progress_var.set("결과 저장 중...")
                ExcelWriter.save_to_excel(sections_data, output_path)
                self.progress_bar['value'] = 100
                
                success_message = f"처리가 완료되었습니다.\n저장 위치: {output_path}"
                self.log_message(success_message)
                messagebox.showinfo("완료", success_message)
            else:
                self.log_message("추출된 표가 없습니다.", "WARNING")
                messagebox.showwarning("경고", "추출된 표가 없습니다.")
                
        except Exception as e:
            error_message = f"처리 중 오류가 발생했습니다: {str(e)}"
            self.log_message(error_message, "ERROR")
            messagebox.showerror("오류", error_message)
            
        finally:
            self.progress_var.set("대기 중...")
            self.process_button['state'] = tk.NORMAL
            self.progress_bar['value'] = 0

    def process_pdf(self):
        pdf_path = self.file_path_var.get()
        if not pdf_path:
            messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
            return
            
        self.process_button['state'] = tk.DISABLED
        self.progress_bar['value'] = 0
        
        # 별도 스레드에서 처리
        thread = threading.Thread(target=self.process_pdf_thread, args=(pdf_path,))
        thread.daemon = True
        thread.start()

    def run(self):
        self.root.mainloop()

def main():
    try:
        app = PDFAnalyzerGUI()
        app.run()
    except Exception as e:
        logger.error(f"프로그램 실행 중 오류 발생: {str(e)}")
        raise

if __name__ == "__main__":
    main()