import sys
import os
from pathlib import Path
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import pandas as pd
import camelot
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import cv2
from PIL import Image
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from datetime import datetime
import threading

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'pdf_analyzer_{datetime.now().strftime("%Y%m%d")}.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class PDFDocument:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        self.section_patterns = {
            "상해": r"(?:◆|◇|■|□|▶)?\s*상해.*(?:특약|특별약관)",
            "질병": r"(?:◆|◇|■|□|▶)?\s*질병.*(?:특약|특별약관)",
            "상해및질병": r"(?:◆|◇|■|□|▶)?\s*(?:상해\s*및\s*질병|상해와\s*질병).*(?:특약|특별약관)"
        }
        self.sections = {
            "상해": {"start": None, "end": None},
            "질병": {"start": None, "end": None},
            "상해및질병": {"start": None, "end": None}
        }

    def find_section_ranges(self) -> Dict[str, Dict[str, int]]:
        """PDF에서 섹션별 페이지 범위 찾기"""
        try:
            print("\n=== PDF 섹션 분석 결과 ===")
            
            # 섹션 시작 페이지 찾기
            for page_num in range(len(self.doc)):
                text = self.doc[page_num].get_text()
                for section, pattern in self.section_patterns.items():
                    if not self.sections[section]["start"] and re.search(pattern, text):
                        self.sections[section]["start"] = page_num
                        logger.info(f"{section} 섹션 시작: {page_num + 1}페이지")
                        print(f"[{section}]")
                        print(f"시작 페이지: {page_num + 1}")

            # 섹션 끝 페이지 설정
            # 1. 상해섹션 끝 = 질병섹션 시작
            if self.sections["상해"]["start"] is not None and self.sections["질병"]["start"] is not None:
                self.sections["상해"]["end"] = self.sections["질병"]["start"]
            
            # 2. 질병섹션 끝 = 상해및질병섹션 시작
            if self.sections["질병"]["start"] is not None and self.sections["상해및질병"]["start"] is not None:
                self.sections["질병"]["end"] = self.sections["상해및질병"]["start"]
            
            # 3. 상해및질병 섹션은 한 페이지만
            if self.sections["상해및질병"]["start"] is not None:
                self.sections["상해및질병"]["end"] = self.sections["상해및질병"]["start"]

            # 상해섹션부터 상해및질병섹션까지 추출 범위 출력
            if self.sections["상해"]["start"] is not None and self.sections["상해및질병"]["start"] is not None:
                print("\n상해섹션부터 상해및질병섹션까지 추출을 시작합니다.")
                print("=" * 50)
                start_page = self.sections["상해"]["start"] + 1
                end_page = self.sections["상해및질병"]["start"] + 1
                print(f"{start_page}페이지부터 {end_page}페이지까지 추출을 시작합니다.")
                print("=" * 50 + "\n")

            # 섹션 범위 요약 출력
            print("\n=== 섹션 범위 요약 ===")
            for section, info in self.sections.items():
                if info["start"] is not None:
                    start = info["start"] + 1
                    if info["end"] is not None:
                        end = info["end"] + 1
                        print(f"{section}: {start}~{end}페이지")
                        logger.info(f"{section} 섹션 범위: {start} ~ {end} 페이지")
                    else:
                        print(f"{section}: {start}페이지부터 시작")
                        logger.info(f"{section} 섹션 범위: {start} ~ 마지막 페이지")

            # 테이블 추출 시작
            self.extract_tables()
            
            return self.sections

        except Exception as e:
            logger.error(f"페이지 범위 찾기 실패: {str(e)}")
            return self.sections

    def close(self):
        """문서 닫기"""
        if self.doc:
            self.doc.close()

    def get_page_count(self) -> int:
        """전체 페이지 수 반환"""
        return len(self.doc)

    def get_page_text(self, page_num: int) -> str:
        """특정 페이지의 텍스트 반환"""
        try:
            return self.doc[page_num].get_text()
        except Exception as e:
            logger.error(f"페이지 {page_num} 텍스트 추출 실패: {str(e)}")
            return ""

    def get_section_info(self, section_name: str) -> Optional[Dict[str, int]]:
        """특정 섹션의 정보 반환"""
        return self.sections.get(section_name)

    def is_valid_section(self, section_name: str) -> bool:
        """유효한 섹션인지 확인"""
        section = self.sections.get(section_name)
        if section is None:
            return False
        return section["start"] is not None and (
            section_name == "상해및질병" or 
            (section["end"] is not None and section["end"] >= section["start"])
        )

    def get_valid_sections(self) -> List[str]:
        """유효한 섹션 목록 반환"""
        return [
            section for section in self.sections.keys() 
            if self.is_valid_section(section)
        ]

class HighlightDetector:
    def __init__(self):
        self.yellow_range = {
            'lower': np.array([20, 100, 100]),
            'upper': np.array([30, 255, 255])
        }
        self.kernel_size = (5, 5)

    def pdf_to_image(self, page: fitz.Page) -> np.ndarray:
        """PDF 페이지를 이미지로 변환"""
        try:
            pix = page.get_pixmap(alpha=False)
            img_array = np.frombuffer(pix.samples, dtype=np.uint8)
            img_array = img_array.reshape(pix.height, pix.width, 3)
            return cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
        except Exception as e:
            logger.error(f"PDF 페이지 이미지 변환 실패: {str(e)}")
            return None

    def detect_highlights(self, image: np.ndarray) -> List[np.ndarray]:
        """하이라이트 영역 감지"""
        try:
            if image is None:
                return []

            # BGR -> HSV 변환
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            
            # 노란색 마스크 생성
            mask = cv2.inRange(hsv, self.yellow_range['lower'], self.yellow_range['upper'])
            
            # 노이즈 제거
            kernel = np.ones(self.kernel_size, np.uint8)
            mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
            mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
            
            # 윤곽선 찾기
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            return [cnt for cnt in contours if cv2.contourArea(cnt) > 100]
            
        except Exception as e:
            logger.error(f"하이라이트 감지 실패: {str(e)}")
            return []

    def get_highlight_regions(self, contours: List[np.ndarray], image_height: int) -> List[Tuple[int, int]]:
        """하이라이트된 영역의 위치 정보 추출"""
        try:
            regions = []
            for contour in contours:
                x, y, w, h = cv2.boundingRect(contour)
                top = image_height - (y + h)
                bottom = image_height - y
                regions.append((top, bottom))
            return regions
        except Exception as e:
            logger.error(f"하이라이트 영역 추출 실패: {str(e)}")
            return []

    def check_highlight(self, row_range: Tuple[float, float], highlight_regions: List[Tuple[float, float]]) -> bool:
        """특정 행이 하이라이트된 영역과 겹치는지 확인"""
        try:
            row_top, row_bottom = row_range
            for region_top, region_bottom in highlight_regions:
                overlap = min(row_bottom, region_bottom) - max(row_top, region_top)
                if overlap > 0:
                    return True
            return False
        except Exception as e:
            logger.error(f"하이라이트 체크 실패: {str(e)}")
            return False
        
class TableProcessor:
    def __init__(self):
        # 표준 컬럼 정의
        self.standard_columns = ['담보명', '지급사유', '지급금액', '변경사항', '페이지']
        
        # 섹션 제목
        self.section_titles = {
            'injury': '상해관련 특별약관',
            'disease': '질병관련 특별약관',
            'both': '상해 및 질병관련 특별약관'
        }
        
        # 무시할 패턴
        self.ignore_patterns = [
            r'이륜자동차',
            r'Page',
            r'^\s*$',
            r'특별약관$',
            r'약관$',
            r'^표$',
            r'^그림$'
        ]
        
        # 컬럼 매핑 정의
        self.column_mapping = {
            '보장명': '담보명',
            '급부명': '담보명',
            '보험금지급사유': '지급사유',
            '지급조건': '지급사유',
            '보험금액': '지급금액',
            '지급액': '지급금액',
            '비고': '변경사항'
        }

    def standardize_columns(self, df: pd.DataFrame, page_num: int) -> pd.DataFrame:
        """컬럼 표준화 및 페이지 정보 추가"""
        try:
            logger.info(f"컬럼 표준화 시작 (페이지 {page_num})")
            logger.info(f"원본 컬럼: {df.columns.tolist()}")
            
            # 데이터프레임 복사
            df = df.copy()
            
            # 컬럼이 모두 숫자인 경우 처리
            if all(str(col).isdigit() for col in df.columns):
                logger.info("숫자로 된 컬럼명 발견, 기본 컬럼명 적용")
                base_columns = ['담보명', '지급사유', '지급금액']
                if len(df.columns) >= len(base_columns):
                    df.columns = base_columns + list(df.columns[len(base_columns):])
                    logger.info(f"변경된 컬럼: {df.columns.tolist()}")

            # 컬럼명 매핑 적용
            mapped_cols = set()
            for old_col, new_col in self.column_mapping.items():
                if old_col in df.columns:
                    df = df.rename(columns={old_col: new_col})
                    mapped_cols.add(new_col)
                    logger.info(f"컬럼 매핑: {old_col} -> {new_col}")

            # 필수 컬럼 추가
            for col in self.standard_columns:
                if col not in df.columns:
                    df[col] = ''
                    logger.info(f"필수 컬럼 추가: {col}")

            # 페이지 번호 설정
            df['페이지'] = page_num
            logger.info(f"페이지 번호 설정: {page_num}")

            # 컬럼 순서 조정
            df = df[self.standard_columns]
            logger.info(f"최종 컬럼: {df.columns.tolist()}")
            
            if not df.empty:
                logger.info(f"데이터 샘플:\n{df.head(1)}")
            
            return df

        except Exception as e:
            logger.error(f"컬럼 표준화 실패: {str(e)}")
            return pd.DataFrame()

    def clean_table_content(self, df: pd.DataFrame) -> pd.DataFrame:
        """테이블 내용 정제 개선"""
        try:
            if df.empty:
                logger.warning("빈 데이터프레임")
                return df

            # 처리 전 상태 기록
            initial_rows = len(df)
            logger.info(f"정제 전 행 수: {initial_rows}")
            if not df.empty:
                logger.info(f"정제 전 첫 행:\n{df.iloc[0]}")

            # 컬럼별 데이터 정제
            for col in df.columns:
                if col != '페이지':
                    # 줄바꿈 문자를 공백으로 변경
                    df[col] = df[col].astype(str).apply(lambda x: ' '.join(x.split('\n')))
                    # 연속된 공백 제거
                    df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
                    # 앞뒤 공백 제거
                    df[col] = df[col].str.strip()
                    # None, nan 등 제거
                    df[col] = df[col].replace({'None': '', 'nan': '', 'NaN': ''})

            # 모든 컬럼이 빈 문자열인 행 제거
            df = df[~df.apply(lambda x: all(str(v).strip() == '' for v in x), axis=1)]
            
            # 불필요한 행 제거
            df = df[~df['담보명'].str.contains('|'.join(self.ignore_patterns), regex=True, na=False)]
            
            # 담보명이 빈 문자열이거나 숫자로만 구성된 행 제거
            df = df[~df['담보명'].str.match(r'^\s*\d*\s*$')]

            # 중복 제거
            df = df.drop_duplicates()

            # 처리 결과 로깅
            final_rows = len(df)
            logger.info(f"행 수 변화: {initial_rows} -> {final_rows}")
            if not df.empty:
                logger.info(f"정제 후 첫 행:\n{df.iloc[0]}")
            
            return df

        except Exception as e:
            logger.error(f"테이블 내용 정제 실패: {str(e)}")
            return df

    def standardize_columns(self, df: pd.DataFrame, page_num: int) -> pd.DataFrame:
        """컬럼 표준화 개선"""
        try:
            logger.info(f"컬럼 표준화 시작 (페이지 {page_num})")
            logger.info(f"원본 컬럼: {df.columns.tolist()}")
            
            # 데이터프레임 복사
            df = df.copy()
            
            # 컬럼이 모두 숫자인 경우 처리
            if all(str(col).isdigit() for col in df.columns):
                logger.info("숫자로 된 컬럼명 발견, 기본 컬럼명 적용")
                # 컬럼 수에 따라 다르게 처리
                if len(df.columns) >= 3:
                    new_cols = ['담보명', '지급사유', '지급금액']
                    df = df.iloc[:, :3]  # 처음 3개 컬럼만 사용
                    df.columns = new_cols
                else:
                    # 컬럼이 3개 미만인 경우 빈 데이터프레임 반환
                    logger.warning("유효하지 않은 컬럼 수")
                    return pd.DataFrame(columns=self.standard_columns)
            
            # 컬럼명 매핑 적용
            for old_col, new_col in self.column_mapping.items():
                if old_col in df.columns:
                    df = df.rename(columns={old_col: new_col})
            
            # 필수 컬럼 추가
            for col in self.standard_columns:
                if col not in df.columns:
                    df[col] = ''
            
            # 페이지 번호 설정
            df['페이지'] = page_num
            
            # 컬럼 순서 조정
            df = df[self.standard_columns]
            
            if not df.empty:
                logger.info(f"데이터 샘플:\n{df.head(1)}")
            
            return df

        except Exception as e:
            logger.error(f"컬럼 표준화 실패: {str(e)}")
            return pd.DataFrame(columns=self.standard_columns)

    def merge_tables(self, tables: List[pd.DataFrame], section_type: str) -> pd.DataFrame:
        """동일 섹션의 테이블 병합"""
        try:
            logger.info(f"{section_type} 섹션 테이블 병합 시작")
            
            if not tables:
                logger.warning("병합할 테이블 없음")
                return pd.DataFrame(columns=self.standard_columns)

            # 테이블 개수 및 크기 로깅
            logger.info(f"병합할 테이블 수: {len(tables)}")
            for i, df in enumerate(tables):
                logger.info(f"테이블 {i+1} 크기: {df.shape}")

            # 테이블 병합
            merged_df = pd.concat(tables, ignore_index=True)
            logger.info(f"병합 후 크기: {merged_df.shape}")

            # 중복 제거
            before_dedup = len(merged_df)
            merged_df = merged_df.drop_duplicates()
            after_dedup = len(merged_df)
            if before_dedup != after_dedup:
                logger.info(f"중복 제거: {before_dedup} -> {after_dedup}행")

            # 정렬
            merged_df = merged_df.sort_values(['페이지', '담보명'])
            logger.info("페이지와 담보명 기준으로 정렬 완료")

            # 결과 요약
            if not merged_df.empty:
                logger.info(f"병합 결과 샘플:\n{merged_df.head(1)}")
            
            return merged_df

        except Exception as e:
            logger.error(f"테이블 병합 실패: {str(e)}")
            return pd.DataFrame(columns=self.standard_columns)

    def validate_table(self, df: pd.DataFrame) -> bool:
        """테이블 유효성 검사"""
        try:
            # 빈 데이터프레임 체크
            if df.empty:
                logger.warning("빈 데이터프레임")
                return False

            # 필수 컬럼 존재 여부 체크
            missing_cols = [col for col in self.standard_columns if col not in df.columns]
            if missing_cols:
                logger.warning(f"필수 컬럼 없음: {missing_cols}")
                return False

            # 데이터 존재 여부 체크
            if df['담보명'].isna().all() or df['담보명'].str.strip().eq('').all():
                logger.warning("담보명 데이터 없음")
                return False

            logger.info("테이블 유효성 검사 통과")
            return True

        except Exception as e:
            logger.error(f"테이블 유효성 검사 실패: {str(e)}")
            return False

class TableExtractor:
    def __init__(self):
        self.table_processor = TableProcessor()
        self.highlight_detector = HighlightDetector()
        self.extraction_settings = {
            'lattice': {
                'line_scale': 40,
                'process_background': True,
                'copy_text': ['v']
            },
            'stream': {
                'edge_tol': 500,
                'row_tol': 10,
                'split_text': True
            }
        }
        self.section_mapping = {
            "상해": "injury",
            "질병": "disease",
            "상해및질병": "both"
        }

    def process_table(self, table, page_image: np.ndarray, page_num: int) -> pd.DataFrame:
        """테이블 처리"""
        try:
            logger.info(f"테이블 처리 시작 (페이지 {page_num + 1})")
            df = table.df.copy()
            
            # 기본 정제
            df = self.table_processor.standardize_columns(df, page_num + 1)
            if df.empty:
                logger.warning(f"표준화 후 빈 데이터프레임")
                return df
                
            df = self.table_processor.clean_table_content(df)
            if df.empty:
                logger.warning(f"정제 후 빈 데이터프레임")
                return df
            
            # 데이터 검증
            if df['담보명'].isna().all() or df['담보명'].str.strip().eq('').all():
                logger.warning("담보명 데이터 없음")
                return pd.DataFrame()
                
            # 하이라이트 감지
            contours = self.highlight_detector.detect_highlights(page_image)
            if contours:
                logger.info(f"페이지 {page_num + 1}에서 하이라이트 영역 감지됨")
                highlight_regions = self.highlight_detector.get_highlight_regions(
                    contours, page_image.shape[0]
                )
                
                # 변경사항 표시
                if highlight_regions:
                    bbox = table._bbox
                    if bbox:
                        x1, y1, x2, y2 = bbox
                        row_height = (y2 - y1) / len(df)
                        
                        for idx in range(len(df)):
                            row_top = y2 - (idx + 1) * row_height
                            row_bottom = y2 - idx * row_height
                            
                            if self.highlight_detector.check_highlight(
                                (row_top, row_bottom), highlight_regions
                            ):
                                df.loc[idx, '변경사항'] = '추가'
                                logger.info(f"페이지 {page_num + 1}의 {idx+1}번째 행에서 변경사항 감지")
            
            logger.info(f"페이지 {page_num + 1}의 테이블 처리 완료 ({len(df)} 행)")
            return df
                
        except Exception as e:
            logger.error(f"테이블 처리 실패 (페이지 {page_num + 1}): {str(e)}")
            return pd.DataFrame()

    def extract_tables(self, pdf_path: str, page_num: int) -> List[pd.DataFrame]:
        """페이지에서 테이블 추출"""
        try:
            # Lattice 방식으로 시도
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num + 1),
                flavor='lattice',
                **self.extraction_settings['lattice']
            )
            
            # Lattice 방식 실패 시 Stream 방식 시도
            if len(tables) == 0:
                tables = camelot.read_pdf(
                    pdf_path,
                    pages=str(page_num + 1),
                    flavor='stream',
                    **self.extraction_settings['stream']
                )
            
            # 정확도가 80% 이상인 테이블만 반환
            valid_tables = [table for table in tables if table.parsing_report['accuracy'] > 80]
            logger.info(f"페이지 {page_num + 1}에서 {len(valid_tables)}개의 유효한 테이블 추출")
            return valid_tables
            
        except Exception as e:
            logger.error(f"페이지 {page_num + 1} 테이블 추출 실패: {str(e)}")
            return []

    def extract_section_tables(self, pdf_doc: fitz.Document, start_page: int, 
                             end_page: int, section_type: str) -> Dict[str, pd.DataFrame]:
        """섹션의 모든 테이블 추출 및 처리"""
        tables_data = []
        
        try:
            # 페이지 범위 유효성 검사
            if not isinstance(start_page, int) or not isinstance(end_page, int):
                logger.error(f"페이지 번호가 정수가 아님: start_page={start_page}, end_page={end_page}")
                return {self.section_mapping[section_type]: pd.DataFrame()}

            if start_page > end_page:
                logger.error(f"시작 페이지가 끝 페이지보다 큼: {start_page} > {end_page}")
                return {self.section_mapping[section_type]: pd.DataFrame()}

            logger.info(f"{section_type} 섹션 테이블 추출 시작: {start_page + 1}페이지 ~ {end_page + 1}페이지")
            
            for page_num in range(start_page, end_page + 1):
                logger.info(f"페이지 {page_num + 1} 처리 중...")
                
                # 페이지 이미지 변환
                page = pdf_doc[page_num]
                image = self.highlight_detector.pdf_to_image(page)
                
                if image is None:
                    logger.warning(f"페이지 {page_num + 1} 이미지 변환 실패")
                    continue
                
                # 테이블 추출
                tables = self.extract_tables(pdf_doc.name, page_num)
                
                # 테이블 처리
                for table_idx, table in enumerate(tables):
                    try:
                        df = self.process_table(table, image, page_num)
                        if not df.empty:
                            tables_data.append(df)
                            logger.info(f"페이지 {page_num + 1}에서 테이블 {table_idx + 1} 추출 성공")
                    except Exception as e:
                        logger.error(f"페이지 {page_num + 1}의 테이블 {table_idx + 1} 처리 실패: {str(e)}")
                        continue
            
            # 섹션별 테이블 병합
            if tables_data:
                merged_df = self.table_processor.merge_tables(tables_data, section_type)
                logger.info(f"{section_type} 섹션 테이블 {len(tables_data)}개 병합 완료")
                # 매핑된 섹션 타입으로 반환
                mapped_section = self.section_mapping[section_type]
                return {mapped_section: merged_df}
            
            logger.warning(f"{section_type} 섹션에서 추출된 테이블 없음")
            return {self.section_mapping[section_type]: pd.DataFrame()}
            
        except Exception as e:
            logger.error(f"섹션 테이블 추출 실패: {str(e)}")
            return {self.section_mapping[section_type]: pd.DataFrame()}
        

class ExcelWriter:
    def __init__(self):
        self.styles = {
            'header': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'cell': {
                'alignment': Alignment(vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'highlight': {
                'fill': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            }
        }

    def apply_cell_style(self, cell, style_name: str):
        """스타일 적용"""
        for attr, value in self.styles[style_name].items():
            setattr(cell, attr, value)

    def write_table(self, worksheet, df: pd.DataFrame, start_row: int, 
                    title: str) -> int:
        """테이블 작성"""
        try:
            # 제목 작성
            title_cell = worksheet.cell(row=start_row, column=1, value=title)
            self.apply_cell_style(title_cell, 'header')
            worksheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=start_row,
                end_column=len(df.columns)
            )
            
            # 열 헤더 작성
            for col_idx, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=start_row + 2, column=col_idx, value=column)
                self.apply_cell_style(cell, 'header')
            
            # 데이터 작성
            for row_idx, row in enumerate(df.itertuples(index=False), start_row + 3):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    self.apply_cell_style(cell, 'cell')
                    
                    # 변경사항이 '추가'인 행 하이라이트
                    if col_idx == df.columns.get_loc('변경사항') + 1 and value == '추가':
                        for highlight_col in range(1, len(df.columns) + 1):
                            self.apply_cell_style(
                                worksheet.cell(row=row_idx, column=highlight_col),
                                'highlight'
                            )
            
            return start_row + len(df) + 4

        except Exception as e:
            logger.error(f"테이블 작성 실패: {str(e)}")
            return start_row + 1

    def adjust_column_widths(self, worksheet):
        """열 너비 자동 조정"""
        try:
            for column in worksheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except:
                        pass
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = \
                    min(max_length + 2, 50)
        except Exception as e:
            logger.error(f"열 너비 조정 실패: {str(e)}")

    def save_to_excel(self, tables: Dict[str, pd.DataFrame], output_path: str):
        """엑셀 파일로 저장"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "보장내용"
            current_row = 1
            
            # 섹션별 테이블 작성
            for section_type, title in [
                ('injury', '상해관련 특별약관'),
                ('disease', '질병관련 특별약관')
            ]:
                if section_type in tables and not tables[section_type].empty:
                    current_row = self.write_table(
                        ws, tables[section_type], current_row, title
                    )
            
            # 열 너비 조정
            self.adjust_column_widths(ws)
            
            # 파일 저장
            wb.save(output_path)
            logger.info(f"Excel 파일 저장 완료: {output_path}")
            
        except Exception as e:
            logger.error(f"Excel 파일 저장 실패: {str(e)}")


class ExcelWriter:
    def __init__(self):
        self.styles = {
            'header': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'cell': {
                'alignment': Alignment(vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'highlight': {
                'fill': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            }
        }

    def apply_cell_style(self, cell, style_name: str):
        """스타일 적용"""
        for attr, value in self.styles[style_name].items():
            setattr(cell, attr, value)

    def write_table(self, worksheet, df: pd.DataFrame, start_row: int, 
                    title: str) -> int:
        """테이블 작성"""
        try:
            # 제목 작성
            title_cell = worksheet.cell(row=start_row, column=1, value=title)
            self.apply_cell_style(title_cell, 'header')
            worksheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=start_row,
                end_column=len(df.columns)
            )
            
            # 열 헤더 작성
            for col_idx, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=start_row + 2, column=col_idx, value=column)
                self.apply_cell_style(cell, 'header')
            
            # 데이터 작성
            for row_idx, row in enumerate(df.itertuples(index=False), start_row + 3):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    self.apply_cell_style(cell, 'cell')
                    
                    # 변경사항이 '추가'인 행 하이라이트
                    if col_idx == df.columns.get_loc('변경사항') + 1 and value == '추가':
                        for highlight_col in range(1, len(df.columns) + 1):
                            self.apply_cell_style(
                                worksheet.cell(row=row_idx, column=highlight_col),
                                'highlight'
                            )
            
            return start_row + len(df) + 4

        except Exception as e:
            logger.error(f"테이블 작성 실패: {str(e)}")
            return start_row + 1

    def adjust_column_widths(self, worksheet):
        """열 너비 자동 조정"""
        try:
            for column in worksheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except:
                        pass
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = \
                    min(max_length + 2, 50)
        except Exception as e:
            logger.error(f"열 너비 조정 실패: {str(e)}")

    def save_to_excel(self, tables: Dict[str, pd.DataFrame], output_path: str):
        """엑셀 파일로 저장"""
        try:
            logger.info("Excel 파일 생성 시작")
            wb = Workbook()
            ws = wb.active
            ws.title = "보장내용"
            current_row = 1

            # 섹션별 테이블 작성
            section_titles = {
                'injury': '상해관련 특별약관',
                'disease': '질병관련 특별약관',
                'both': '상해 및 질병관련 특별약관'
            }
            
            for section_type, df in tables.items():
                if not df.empty:
                    logger.info(f"{section_type} 섹션 데이터 작성 중...")
                    title = section_titles.get(section_type, section_type)
                    current_row = self.write_table(ws, df, current_row, title)
                    logger.info(f"{section_type} 섹션 데이터 작성 완료: {len(df)}행")
                else:
                    logger.warning(f"{section_type} 섹션 데이터 없음")

            # 열 너비 조정
            self.adjust_column_widths(ws)
            
            # 파일 저장
            wb.save(output_path)
            logger.info(f"Excel 파일 저장 완료: {output_path}")
            
            # 저장된 데이터 확인
            logger.info(f"저장된 섹션: {list(tables.keys())}")
            for section_type, df in tables.items():
                logger.info(f"{section_type} 섹션: {len(df)}행")
            
        except Exception as e:
            logger.error(f"Excel 파일 저장 실패: {str(e)}")
            raise


class PDFAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("KB손해보험 상품 분석기")
        self.root.geometry("800x600")
        self.file_path_var = tk.StringVar()
        self.progress_var = tk.StringVar(value="대기 중...")
        self.setup_gui()

    def setup_gui(self):
        """GUI 구성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 제목
        title_label = ttk.Label(
            main_frame,
            text="보험약관 테이블 분석기",
            font=("Helvetica", 16, "bold")
        )
        title_label.pack(pady=10)

        # 파일 선택 프레임
        file_frame = ttk.LabelFrame(main_frame, text="PDF 파일 선택", padding="10")
        file_frame.pack(fill=tk.X, pady=10)

        file_entry = ttk.Entry(
            file_frame,
            textvariable=self.file_path_var,
            width=60
        )
        file_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        browse_button = ttk.Button(
            file_frame,
            text="찾아보기",
            command=self.browse_file
        )
        browse_button.pack(side=tk.LEFT, padx=5)

        # 진행 상태 표시
        progress_label = ttk.Label(
            main_frame,
            textvariable=self.progress_var
        )
        progress_label.pack(pady=5)

        self.progress_bar = ttk.Progressbar(
            main_frame,
            mode='determinate',
            length=300
        )
        self.progress_bar.pack(pady=5)

        # 분석 버튼
        self.analyze_button = ttk.Button(
            main_frame,
            text="분석 시작",
            command=self.start_analysis,
            state=tk.DISABLED
        )
        self.analyze_button.pack(pady=10)

        # 로그 영역
        log_frame = ttk.LabelFrame(main_frame, text="처리 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = tk.Text(log_frame, height=15, width=70)
        scrollbar = ttk.Scrollbar(
            log_frame,
            orient=tk.VERTICAL,
            command=self.log_text.yview
        )
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def browse_file(self):
        """파일 선택"""
        file_path = filedialog.askopenfilename(
            title="PDF 파일 선택",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.analyze_button['state'] = tk.NORMAL
            self.log_message(f"파일 선택됨: {file_path}")

    def log_message(self, message: str, level: str = "INFO"):
        """로그 메시지 추가"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {level}: {message}\n"
        
        self.log_text.insert(tk.END, log_entry)
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_progress(self, value: int, message: str = None):
        """진행 상태 업데이트"""
        self.progress_bar['value'] = value
        if message:
            self.progress_var.set(message)
        self.root.update_idletasks()

    def analyze_pdf(self, pdf_path: str):
        """PDF 분석 실행"""
        try:
            # 출력 폴더 생성
            output_dir = os.path.join(os.path.dirname(pdf_path), "output")
            os.makedirs(output_dir, exist_ok=True)

            self.log_message("PDF 분석 시작")
            self.update_progress(10, "PDF 문서 로드 중...")

            # PDF 문서 로드 및 섹션 분석
            pdf_doc = PDFDocument(pdf_path)
            sections = pdf_doc.find_section_ranges()

            if not any(sections[section]["start"] is not None for section in sections):
                self.log_message("섹션을 찾을 수 없습니다.", "WARNING")
                return

            # 테이블 추출기 초기화
            extractor = TableExtractor()
            all_tables = {}

            # 섹션별 테이블 추출
            section_count = len([s for s in sections if sections[s]["start"] is not None])
            current_section = 0

            for section_type, ranges in sections.items():
                if ranges["start"] is not None:
                    current_section += 1
                    progress = 10 + (current_section / section_count) * 60
                    self.update_progress(
                        progress,
                        f"{section_type} 섹션 테이블 추출 중..."
                    )
                    
                    tables = extractor.extract_section_tables(
                        pdf_doc.doc,
                        ranges["start"],
                        ranges["end"],
                        section_type
                    )
                    
                    all_tables.update(tables)

            # PDF 문서 닫기
            pdf_doc.close()

            if any(not df.empty for df in all_tables.values()):
                self.update_progress(70, "Excel 파일 생성 중...")
                
                # 결과 저장
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = os.path.join(
                    output_dir,
                    f"analyzed_tables_{timestamp}.xlsx"
                )
                
                excel_writer = ExcelWriter()
                excel_writer.save_to_excel(all_tables, output_path)
                
                self.update_progress(100, "완료")
                self.log_message(f"분석 완료. 결과 파일: {output_path}")
                messagebox.showinfo(
                    "완료",
                    f"분석이 완료되었습니다.\n저장 위치: {output_path}"
                )
            else:
                self.log_message("추출된 테이블이 없습니다.", "WARNING")
                messagebox.showwarning("경고", "추출된 테이블이 없습니다.")

        except Exception as e:
            self.log_message(f"오류 발생: {str(e)}", "ERROR")
            messagebox.showerror(
                "오류",
                f"처리 중 오류가 발생했습니다:\n{str(e)}"
            )

    def start_analysis(self):
        """분석 시작"""
        pdf_path = self.file_path_var.get()
        if not pdf_path:
            messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
            return

        self.analyze_button['state'] = tk.DISABLED
        self.progress_bar['value'] = 0
        self.progress_var.set("처리 중...")

        # 별도 스레드에서 처리
        threading.Thread(
            target=self.analyze_pdf,
            args=(pdf_path,),
            daemon=True
        ).start()

    def run(self):
        """애플리케이션 실행"""
        self.root.mainloop()

def main():
    try:
        app = PDFAnalyzerGUI()
        app.run()
    except Exception as e:
        logger.error(f"프로그램 실행 중 오류 발생: {str(e)}")
        raise

if __name__ == "__main__":
    main()
