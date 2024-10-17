import os
import re
import numpy as np  # numpy 임포트 추가
import PyPDF2
import camelot
import fitz  # PyMuPDF
import pandas as pd
from sentence_transformers import SentenceTransformer
import faiss
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_text_with_positions(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text_chunks = []
        for page_num, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                lines = text.split('\n')
                for line_num, line in enumerate(lines):
                    text_chunks.append({
                        'text': line,
                        'page': page_num + 1,
                        'line': line_num + 1
                    })
    return text_chunks

def preprocess_text(text):
    # 특수문자 및 공백 제거, 소문자로 변환
    text = re.sub(r'[\s\W_]+', '', text)
    text = text.lower()
    return text

def detect_table_boundaries(text_chunks):
    start_patterns = [
        r'표\s*\d+',
        r'선택\s*특약',
        r'상해관련\s*특약',
        r'질병관련\s*특약',
        r'\[1\s*종\]',
        r'\[2\s*종\]',
        r'\[3\s*종\]'
    ]
    end_patterns = [
        r'합\s*계',
        r'총\s*계',
        r'주\s*\)',
        r'※',
        r'결\s*론'
    ]

    tables = []
    table_start = None

    for i, chunk in enumerate(text_chunks):
        text = chunk['text']

        if table_start is None and any(re.search(pattern, text) for pattern in start_patterns):
            table_start = i
            continue

        if table_start is not None and any(re.search(pattern, text) for pattern in end_patterns):
            table_end = i

            start_page = text_chunks[table_start]['page']
            end_page = text_chunks[table_end]['page']
            pages = list(range(start_page, end_page + 1))

            context_start = max(0, table_start - 5)
            context_end = min(len(text_chunks), table_end + 5)
            context = text_chunks[context_start:context_end]

            tables.append({
                'start': text_chunks[table_start],
                'end': text_chunks[table_end],
                'pages': pages,
                'context': context
            })
            table_start = None

    return tables

def extract_text_above_bbox(page, bbox):
    x0, y0, x1, y1 = bbox  # bbox: (x0, y0, x1, y1)
    text_blocks = page.get_text("blocks")
    # 테이블 bbox의 y0보다 위에 있는 텍스트 블록 중 가장 아래에 있는 것 선택
    texts_above = []
    for block in text_blocks:
        if len(block) >= 5:
            bx0, by0, bx1, by1, text = block[:5]
            if by1 <= y0:  # 블록의 아래쪽 y좌표가 테이블의 위쪽 y좌표보다 작거나 같으면
                texts_above.append((by1, text))
    if texts_above:
        # y 좌표가 가장 큰 (테이블 바로 위에 있는) 텍스트 선택
        texts_above.sort(reverse=True)
        return texts_above[0][1].strip()
    else:
        return "제목 없음"

def extract_tables_with_camelot(pdf_path, tables_info):
    all_tables = []
    doc = fitz.open(pdf_path)
    for table_info in tables_info:
        pages = table_info['pages']
        pages_str = ','.join(map(str, pages))
        print(f"Extracting table from pages {pages_str} using Camelot...")
        try:
            tables = camelot.read_pdf(pdf_path, pages=pages_str, flavor='lattice')
        except Exception as e:
            print(f"Error extracting tables from pages {pages_str}: {e}")
            continue

        if not tables:
            continue

        for table in tables:
            df = table.df

            # 표의 바운딩 박스 가져오기
            bbox = table._bbox  # (x1, y1, x2, y2)
            # fitz는 (x0, y0, x1, y1)
            x0, y0, x1, y1 = bbox

            # 표 위의 텍스트 추출
            page_num = table.page - 1  # fitz는 0부터 시작
            page = doc.load_page(page_num)
            text_above_table = extract_text_above_bbox(page, bbox)

            # 추출한 텍스트를 전처리하여 제목으로 지정
            title = text_above_table.strip()

            all_tables.append({
                'dataframe': df,
                'title': title,
                'page': table.page
            })
    print(f"Found {len(all_tables)} tables in total")
    return all_tables

def process_tables(all_tables):
    processed_data = []
    for i, table_info in enumerate(all_tables):
        df = table_info['dataframe']
        title = table_info['title']
        page = table_info['page']
        df['Table_Number'] = i + 1
        df['Table_Title'] = title  # 여기에서 제목을 지정
        df['Page'] = page
        processed_data.append(df)
    if processed_data:
        return pd.concat(processed_data, ignore_index=True)
    else:
        return pd.DataFrame()

def save_tables_to_excel(tables_dict, output_path, document_title=None):
    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)
    
    for sheet_name, tables in tables_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        current_row = 1

        if document_title:
            ws.cell(row=current_row, column=1, value=document_title)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=1).font = Font(size=16, bold=True)
            current_row += 2  # 빈 줄 추가

        for idx, table_info in enumerate(tables):
            # table_info는 표 정보 dict
            if 'dataframe' not in table_info:
                print(f"Skipping table in sheet '{sheet_name}' as 'dataframe' key is missing.")
                continue  # 'dataframe' 키가 없으면 스킵

            df = table_info['dataframe']
            title = table_info['title']
            page = table_info['page']

            # 표 제목 추가
            ws.cell(row=current_row, column=1, value=f"표 {idx+1}: {title} (페이지 {page})")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=df.shape[1])
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            # 데이터프레임을 엑셀 시트에 추가
            for r in dataframe_to_rows(df, index=False, header=True):
                for c_idx, value in enumerate(r, start=1):
                    ws.cell(row=current_row, column=c_idx, value=value)
                    ws.cell(row=current_row, column=c_idx).alignment = Alignment(wrap_text=True)
                current_row += 1
            current_row += 1  # 표 사이에 빈 줄 추가

        # 열 너비 자동 조정
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    wb.save(output_path)
    print(f"Tables have been saved to {output_path}")

def split_text_into_chunks(text, chunk_size=200, overlap=50):
    words = text.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        chunk = " ".join(words[i:i + chunk_size])
        chunks.append(chunk)
    return chunks

def create_index(chunks):
    model = SentenceTransformer('distiluse-base-multilingual-cased')
    embeddings = model.encode(chunks)
    
    # 임베딩 정규화 (코사인 유사도 사용을 위함)
    embeddings = embeddings / np.linalg.norm(embeddings, axis=1, keepdims=True)
    
    dimension = embeddings.shape[1]
    index = faiss.IndexFlatIP(dimension)  # Inner Product를 사용하여 코사인 유사도 계산
    index.add(embeddings.astype('float32'))
    
    return index, model

def search_rag(query, index, model, chunks, page_numbers=None, k=5, threshold=0.5):
    query_vector = model.encode([query])
    query_vector = query_vector / np.linalg.norm(query_vector, axis=1, keepdims=True)
    
    distances, indices = index.search(query_vector.astype('float32'), k)
    
    results = []
    for idx, score in zip(indices[0], distances[0]):
        if score >= threshold:  # 임계값 이상인 경우만 결과에 추가
            result = {'content': chunks[idx], 'score': score}
            if page_numbers:
                result['page'] = page_numbers[idx]
            results.append(result)
    
    return results

def results_to_dataframe(results, query_type):
    df = pd.DataFrame(results)
    df['type'] = query_type
    return df

def find_pages_with_keyword_in_page(texts_by_page, keyword):
    pages = []
    preprocessed_keyword = preprocess_text(keyword)

    for page_num, text in texts_by_page.items():
        preprocessed_text = preprocess_text(text)
        if preprocessed_keyword in preprocessed_text:
            pages.append(page_num)
    return pages

def extract_text_from_page(page):
    return page.get_text("text")

def main():
    uploads_folder = "uploads"
    output_folder = "output"

    os.makedirs(output_folder, exist_ok=True)

    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_tables.xlsx")
    search_output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_search_results.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    # 전체 텍스트와 페이지 번호 추출
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        full_text = ""
        page_numbers = []
        texts_by_page = {}
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"
                page_numbers.extend([i+1] * len(page_text.split()))
                texts_by_page[i+1] = page_text  # 페이지별로 텍스트 저장

    # 텍스트 청크 생성 및 인덱스 구축 (RAG용)
    chunks = split_text_into_chunks(full_text)
    index, model = create_index(chunks)

    # RAG를 사용하여 선택특약 검색
    select_query = "선택특약"
    select_results = search_rag(select_query, index, model, chunks, page_numbers, k=10, threshold=0.5)
    select_df = results_to_dataframe(select_results, "선택특약")

    # 선택특약이 포함된 페이지 찾기
    select_pages = find_pages_with_keyword_in_page(texts_by_page, "선택특약")
    print("선택특약이 포함된 페이지:", select_pages)

    if not select_pages:
        print("선택특약이 포함된 페이지를 찾지 못했습니다.")
        return

    # 선택특약이 있는 페이지의 텍스트를 txt 파일로 저장
    doc = fitz.open(pdf_path)
    for page_num in select_pages:
        page = doc.load_page(page_num - 1)
        text = extract_text_from_page(page)
        txt_output_path = os.path.join(output_folder, f"page_{page_num}_text.txt")
        with open(txt_output_path, 'w', encoding='utf-8') as f:
            f.write(text)
        print(f"Page {page_num}의 텍스트가 {txt_output_path}에 저장되었습니다.")

    # 선택특약 페이지 내에서 1종, 2종, 3종 찾기 및 출력
    # 별도의 데이터프레임 생성
    종_sheets = {
        "1종": [],
        "2종": [],
        "3종": []
    }

    for page_num in select_pages:
        page_text = texts_by_page.get(page_num, "")
        종_detected = False
        if "1종" in page_text:
            print(f"Page {page_num}: 1종이 검출되었습니다.")
            종_detected = True
            종_sheets["1종"].append({
                'page': page_num,
                'text': page_text
            })
        if "2종" in page_text:
            print(f"Page {page_num}: 2종이 검출되었습니다.")
            종_detected = True
            종_sheets["2종"].append({
                'page': page_num,
                'text': page_text
            })
        if "3종" in page_text:
            print(f"Page {page_num}: 3종이 검출되었습니다.")
            종_detected = True
            종_sheets["3종"].append({
                'page': page_num,
                'text': page_text
            })
        if not 종_detected:
            print(f"Page {page_num}: 종이 검출되지 않았습니다.")

    # RAG를 사용하여 상해관련특약 검색
    injury_query = "상해관련특약"
    injury_results = search_rag(injury_query, index, model, chunks, page_numbers, k=10, threshold=0.5)
    injury_df = results_to_dataframe(injury_results, "상해관련특약")

    # RAG를 사용하여 질병관련특약 검색
    disease_query = "질병관련특약"
    disease_results = search_rag(disease_query, index, model, chunks, page_numbers, k=10, threshold=0.5)
    disease_df = results_to_dataframe(disease_results, "질병관련특약")

    # 결과 합치기
    final_search_df = pd.concat([select_df, injury_df, disease_df], ignore_index=True)

    # 엑셀 파일로 저장 (검색 결과)
    final_search_df.to_excel(search_output_excel_path, index=False, engine='openpyxl')
    print(f"검색 결과가 {search_output_excel_path}에 저장되었습니다.")

    # '질병관련 특별약관'이 나오면 표 추출 종료
    stop_keyword = "질병관련 특별약관"
    stop_pages = find_pages_with_keyword_in_page(texts_by_page, stop_keyword)
    if stop_pages:
        print(f"'질병관련 특별약관'이 포함된 페이지 {stop_pages}에서 표 추출을 종료합니다.")
        # 표 추출 시 질병관련 특별약관이 있는 페이지 이후는 제외
        min_stop_page = min(stop_pages)
        tables_info = detect_table_boundaries(extract_text_with_positions(pdf_path))
        filtered_tables_info = [table for table in tables_info if max(table['pages']) < min_stop_page]
    else:
        tables_info = detect_table_boundaries(extract_text_with_positions(pdf_path))
        filtered_tables_info = tables_info

    # Camelot과 PyMuPDF를 함께 사용하여 표와 제목 추출
    tables = extract_tables_with_camelot(pdf_path, filtered_tables_info)

    # 1종, 2종, 3종에 따라 시트 분류 (표만 분류)
    # 기존의 종_sheets에 추가하지 않음. 새로운 sheets_dict을 생성
    tables_sheets = {
        "1종": [],
        "2종": [],
        "3종": []
    }

    for table in tables:
        title = table['title']
        page = table['page']
        if "1종" in title:
            tables_sheets["1종"].append(table)
        elif "2종" in title:
            tables_sheets["2종"].append(table)
        elif "3종" in title:
            tables_sheets["3종"].append(table)

    # 엑셀로 저장
    # 별도의 시트로 저장
    if any(tables_sheets.values()):
        save_tables_to_excel(tables_sheets, output_excel_path, document_title=None)
    else:
        print("1종, 2종, 3종에 해당하는 표가 없습니다.")

    print(f"All processed tables have been saved to {output_excel_path}")

if __name__ == "__main__":
    main()
