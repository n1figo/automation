import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import fitz  # PyMuPDF
import os
from PIL import Image
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
from io import StringIO
import re

async def get_full_html(url, output_dir):
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page()
        await page.goto(url, wait_until="networkidle")
        
        await page.evaluate("""
            () => {
                const expandElements = (elements) => {
                    for (let elem of elements) {
                        if (window.getComputedStyle(elem).display === 'none') {
                            elem.style.display = 'block';
                        }
                        expandElements(elem.children);
                    }
                };
                expandElements(document.body.children);
            }
        """)
        
        html_content = await page.content()
        await browser.close()
    
    html_file_path = os.path.join(output_dir, "source.txt")
    with open(html_file_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"Full HTML source has been saved to {html_file_path}")
    return html_file_path

def extract_tables_from_html(html_file_path):
    try:
        with open(html_file_path, 'r', encoding='utf-8') as f:
            html_content = f.read()

        soup = BeautifulSoup(html_content, 'html.parser')
        tables = soup.find_all('table')
        
        if not tables:
            print("No tables found in the HTML content.")
            print(f"HTML content preview: {html_content[:1000]}...")
            return []
        
        dfs = [pd.read_html(StringIO(str(table)))[0] for table in tables]
        print(f"Extracted {len(dfs)} tables.")
        return dfs
    except Exception as e:
        print(f"Error extracting tables: {str(e)}")
        print("Detailed error information:")
        import traceback
        print(traceback.format_exc())
        return []

def extract_specific_table(html_file_path):
    try:
        with open(html_file_path, 'r', encoding='utf-8') as f:
            html_content = f.read()

        soup = BeautifulSoup(html_content, 'html.parser')
        
        # 정규표현식을 사용하여 유사한 텍스트를 찾습니다
        pattern = re.compile(r'상해.*특별약관.*')
        section = soup.find(string=pattern)
        
        if section:
            # 섹션을 찾았다면, 가장 가까운 테이블을 찾습니다
            table = section.find_next('table')
            if table:
                df = pd.read_html(StringIO(str(table)))[0]
                # 열 이름 설정
                if len(df.columns) >= 3:
                    df.columns = ['보장명', '지급사유', '지급금액']
                df = df.dropna(how='all').reset_index(drop=True)
                print("Extracted specific table successfully.")
                return df
            else:
                print("Specific table not found near the section.")
                return None
        else:
            print("Specific section not found.")
            return None
    except Exception as e:
        print(f"Error extracting specific table: {str(e)}")
        print("Detailed error information:")
        import traceback
        print(traceback.format_exc())
        return None

def save_original_tables_to_excel(dfs, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Original Tables"

    row = 1
    for i, df in enumerate(dfs, start=1):
        ws.cell(row=row, column=1, value=f'Table_{i}')
        row += 1

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [' '.join(col).strip() for col in df.columns.values]

        for col, header in enumerate(df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1

        for _, data_row in df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

        row += 2

    wb.save(output_excel_path)
    print(f"Original tables have been saved to {output_excel_path}")

def save_specific_table_to_excel(df, output_excel_path):
    if df is not None and not df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "상해관련 특별약관"

        for col, header in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col, value=header)

        for row, data in df.iterrows():
            for col, value in enumerate(data, start=1):
                ws.cell(row=row+2, column=col, value=value)

        wb.save(output_excel_path)
        print(f"Specific table has been saved to {output_excel_path}")
    else:
        print("No data to save.")

def process_tables(dfs):
    all_data = []
    for i, df in enumerate(dfs):
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [' '.join(col).strip() for col in df.columns.values]
        df['table_info'] = f'Table_{i+1}'
        all_data.append(df)
    
    if not all_data:
        print("No data extracted.")
        return pd.DataFrame()
    
    result = pd.concat(all_data, axis=0, ignore_index=True)
    
    print("Final DataFrame:")
    print(f"Columns: {result.columns.tolist()}")
    print(f"Shape: {result.shape}")
    
    return result

def is_color_highlighted(color):
    r, g, b = color
    if r == g == b:
        return False
    return max(r, g, b) > 200 and (max(r, g, b) - min(r, g, b)) > 30

def detect_highlights(image):
    width, height = image.size
    img_array = np.array(image)
    
    highlighted_rows = set()
    for y in range(height):
        for x in range(width):
            if is_color_highlighted(img_array[y, x]):
                highlighted_rows.add(y)
    
    if highlighted_rows:
        start_row = max(0, min(highlighted_rows) - 10 * height // 100)
        end_row = min(height, max(highlighted_rows) + 10 * height // 100)
        return [(0, start_row, width, end_row)]
    
    return []

def extract_highlighted_text_with_context(pdf_path, max_pages=20):
    print("Starting extraction of highlighted text from PDF...")
    doc = fitz.open(pdf_path)
    total_pages = min(len(doc), max_pages)
    highlighted_texts_with_context = []
    
    output_image_dir = os.path.join("output", "images")
    os.makedirs(output_image_dir, exist_ok=True)
    
    for page_num in range(total_pages):
        print(f"Processing page {page_num + 1}/{total_pages}")
        page = doc[page_num]
        
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        highlighted_sections = detect_highlights(img)
        
        if highlighted_sections:
            section = highlighted_sections[0]
            x0, y0, x1, y1 = section
            
            highlight_img = img.crop(section)
            
            image_filename = f"page_{page_num + 1}_highlight.png"
            image_path = os.path.join(output_image_dir, image_filename)
            highlight_img.save(image_path)
            
            text = page.get_text("text", clip=section)
            if text.strip():
                context = page.get_text("text", clip=section)
                highlighted_texts_with_context.append((context, text, page_num + 1, image_path))

    print(f"Finished extracting highlighted text from PDF (total {total_pages} pages)")
    return highlighted_texts_with_context

def compare_dataframes(df_before, highlighted_texts_with_context):
    print("Starting comparison of dataframes...")
    matching_rows = []

    for context, highlighted_text, page_num, image_path in highlighted_texts_with_context:
        context_lines = context.split('\n')
        for i in range(len(df_before)):
            match = True
            for j, line in enumerate(context_lines):
                if i+j >= len(df_before) or not any(str(cell).strip() in line for cell in df_before.iloc[i+j]):
                    match = False
                    break
            if match:
                matching_rows.extend(range(i, i+len(context_lines)))
                break

    matching_rows = sorted(set(matching_rows))
    df_matching = df_before.loc[matching_rows].copy()
    
    df_matching['일치'] = '일치'
    df_matching['하단 표 삽입요망'] = '하단 표 삽입요망'
    df_matching['PDF_페이지'] = ''
    df_matching['이미지_경로'] = ''
    
    for i, (_, _, page, path) in enumerate(highlighted_texts_with_context):
        if i < len(df_matching):
            df_matching.iloc[i, df_matching.columns.get_loc('PDF_페이지')] = page
            df_matching.iloc[i, df_matching.columns.get_loc('이미지_경로')] = path
    
    print(f"Finished comparison. Found {len(matching_rows)} matching rows")
    return df_matching

def save_to_excel(df, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    row = 1
    for table_info in df['table_info'].unique():
        table_df = df[df['table_info'] == table_info]
        
        ws.cell(row=row, column=1, value=table_info)
        row += 1
        
        for col, header in enumerate(table_df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1
        
        for _, data_row in table_df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1
        
        row += 2

    wb.save(output_excel_path)
    print(f"Results have been saved to {output_excel_path}")

async def main():
    print("Program start")
    try:
        url = "https://www.kbinsure.co.kr/CG302290001.ec#"
        pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
        output_dir = "/workspaces/automation/output"
        os.makedirs(output_dir, exist_ok=True)
        output_excel_path = os.path.join(output_dir, "comparison_results.xlsx")
        original_excel_path = os.path.join(output_dir, "변경전.xlsx")
        specific_table_excel_path = os.path.join(output_dir, "상해관련_특별약관.xlsx")

        html_file_path = await get_full_html(url, output_dir)
        
        # Extract and save all tables
        dfs = extract_tables_from_html(html_file_path)
        if dfs:
            save_original_tables_to_excel(dfs, original_excel_path)
            df_before = process_tables(dfs)
        else:
            print("Failed to extract tables. Please check the HTML file.")
            return

        # Extract and save specific table
        specific_df = extract_specific_table(html_file_path)
        if specific_df is not None:
            save_specific_table_to_excel(specific_df, specific_table_excel_path)
        else:
            print("Failed to extract the specific table.")

        # Compare with PDF
        highlighted_texts_with_context = extract_highlighted_text_with_context(pdf_path, max_pages=20)

        if not df_before.empty and highlighted_texts_with_context:
            df_matching = compare_dataframes(df_before, highlighted_texts_with_context)
            save_to_excel(df_matching, output_excel_path)
        else:
            print("Failed to extract tables or highlighted text. Please check the URL and PDF.")

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print("Detailed error information:")
        import traceback
        print(traceback.format_exc())
    
    print("Program end")

if __name__ == "__main__":
    asyncio.run(main())