import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import os
from PIL import Image
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
from io import StringIO
import pdfplumber
import re

async def get_full_html_and_tables(url, output_dir):
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.goto(url, wait_until="networkidle")
            
            # '보장내용' 탭 클릭
            await page.click('text="보장내용"')
            await page.wait_for_load_state('networkidle')
            
            # JavaScript 실행 후 페이지 내용 가져오기
            html_content = await page.content()
            
            # 테이블 데이터 추출
            tables = await page.evaluate('''
                () => {
                    const tables = document.querySelectorAll('.tab_cont[data-list="보장내용"] table');
                    return Array.from(tables).map(table => {
                        const rows = table.querySelectorAll('tr');
                        return Array.from(rows).map(row => {
                            const cells = row.querySelectorAll('th, td');
                            return Array.from(cells).map(cell => cell.innerText);
                        });
                    });
                }
            ''')
            
            # 전체 페이지 스크린샷 캡처
            await page.screenshot(path=os.path.join(output_dir, "full_page.png"), full_page=True)
            
            # 테이블 영역 스크린샷 캡처
            table_elements = await page.query_selector_all('.tab_cont[data-list="보장내용"] table')
            for i, table_element in enumerate(table_elements):
                try:
                    await table_element.scroll_into_view_if_needed()
                    await page.wait_for_timeout(1000)
                    await table_element.screenshot(path=os.path.join(output_dir, f"table_{i+1}.png"))
                except Exception as e:
                    print(f"Failed to capture screenshot for table {i+1}: {str(e)}")
            
            await browser.close()
        
        html_file_path = os.path.join(output_dir, "source.txt")
        with open(html_file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"Full HTML source and table screenshots have been saved to {output_dir}")
        return html_file_path, tables
    except Exception as e:
        print(f"Error in get_full_html_and_tables: {str(e)}")
        return None, []

def process_tables(tables):
    all_data = []
    for i, table in enumerate(tables):
        df = pd.DataFrame(table[1:], columns=table[0])
        df['table_info'] = f'Table_{i+1}'
        all_data.append(df)
    
    if not all_data:
        print("No data extracted.")
        return pd.DataFrame()
    
    result = pd.concat(all_data, axis=0, ignore_index=True)
    
    print("Final DataFrame:")
    print(f"Columns: {result.columns.tolist()}")
    print(f"Shape: {result.shape}")
    
    return result

def save_original_tables_to_excel(dfs, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Original Tables"

    row = 1
    for i, df in enumerate(dfs, start=1):
        ws.cell(row=row, column=1, value=f'Table_{i}')
        row += 1

        for col, header in enumerate(df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1

        for _, data_row in df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

        row += 2

    wb.save(output_excel_path)
    print(f"Original tables have been saved to {output_excel_path}")

def is_color_highlighted(color):
    r, g, b = color
    if r == g == b:
        return False
    return max(r, g, b) > 200 and (max(r, g, b) - min(r, g, b)) > 30

def detect_highlights_and_colors(image):
    width, height = image.size
    img_array = np.array(image)
    
    highlighted_rows = set()
    colored_rows = set()
    for y in range(height):
        for x in range(width):
            color = img_array[y, x]
            if color[0] > 200 and color[1] > 200 and color[2] < 100:  # Yellow highlight
                highlighted_rows.add(y)
            elif is_color_highlighted(color):
                colored_rows.add(y)
    
    all_special_rows = sorted(highlighted_rows.union(colored_rows))
    if all_special_rows:
        return all_special_rows, highlighted_rows, colored_rows
    return [], set(), set()

def extract_highlighted_text_and_tables(pdf_path, output_dir):
    print("Starting extraction of text and tables from PDF...")
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        texts_with_context = []
        all_tables = []
        start_extraction = False
        end_extraction = False
        
        for page_num, page in enumerate(pdf.pages):
            print(f"Processing page {page_num + 1}/{total_pages}")
            
            text = page.extract_text()
            tables = page.extract_tables()
            
            if "□ 신태약관" in text:
                start_extraction = True
            
            if start_extraction and not end_extraction:
                if "다. 보험료산출기초 및 공시이율" in text:
                    end_extraction = True
                
                texts_with_context.append((text, page_num + 1))
                
                # 표와 그 제목을 함께 추출
                lines = text.split('\n')
                for i, table in enumerate(tables):
                    title = ""
                    for line in reversed(lines[:lines.index(table[0][0])]):
                        if line.strip() and not re.match(r'^\d+$', line.strip()):  # 페이지 번호가 아닌 경우
                            title = line.strip()
                            break
                    all_tables.append((table, page_num + 1, title))
            
            if end_extraction:
                break
        
        # 모든 표와 텍스트를 하나의 Excel 파일로 저장
        excel_path = os.path.join(output_dir, "extracted_content.xlsx")
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # 텍스트 저장
            text_df = pd.DataFrame(texts_with_context, columns=['Text', 'Page'])
            text_df.to_excel(writer, sheet_name='Extracted Text', index=False)
            
            # 표 저장
            for i, (table, page_num, title) in enumerate(all_tables):
                df = pd.DataFrame(table[1:], columns=table[0])
                df.insert(0, 'Title', title)
                df['Page'] = page_num
                df.to_excel(writer, sheet_name=f'Table_{i+1}', index=False)
        
        print(f"Saved all extracted content to {excel_path}")

    print(f"Finished extracting text and tables from PDF")
    return texts_with_context, all_tables

def compare_dataframes(df_before, texts_with_context, output_dir):
    print("Starting comparison of dataframes...")
    df_result = df_before.copy() if not df_before.empty else pd.DataFrame(columns=['table_info'])
    df_result['PDF_페이지'] = ''
    df_result['이미지_경로'] = ''
    df_result['상태'] = '유지'

    for text, page_num, image_path, status, special_rows in texts_with_context:
        text_lines = text.split('\n')
        for i in range(len(df_result)):
            if any(str(cell).strip() in text for cell in df_result.iloc[i]):
                table_start = i
                while table_start > 0 and df_result.loc[table_start-1, 'table_info'] == df_result.loc[i, 'table_info']:
                    table_start -= 1
                table_end = i
                while table_end < len(df_result)-1 and df_result.loc[table_end+1, 'table_info'] == df_result.loc[i, 'table_info']:
                    table_end += 1
                
                df_result.loc[table_start:table_end, 'PDF_페이지'] = page_num
                df_result.loc[table_start:table_end, '이미지_경로'] = image_path
                
                # 강조된 행에 대해서만 상태 업데이트
                for j in range(table_start, table_end + 1):
                    if any(line.strip() in str(cell) for cell in df_result.iloc[j] for line in text_lines if line.strip()):
                        df_result.loc[j, '상태'] = status
                
                break

    # 상태 컬럼을 맨 오른쪽으로 이동
    cols = df_result.columns.tolist()
    cols.append(cols.pop(cols.index('상태')))
    df_result = df_result[cols]

    print(f"Finished comparison. Updated {len(df_result[df_result['상태'] != '유지'])} rows")
    return df_result

def save_to_excel(df, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    row = 1
    if 'table_info' in df.columns:
        for table_info in df['table_info'].unique():
            table_df = df[df['table_info'] == table_info]
            
            ws.cell(row=row, column=1, value=table_info)
            row += 1
            
            for col, header in enumerate(table_df.columns, start=1):
                ws.cell(row=row, column=col, value=header)
            row += 1
            
            for _, data_row in table_df.iterrows():
                for col, value in enumerate(data_row, start=1):
                    ws.cell(row=row, column=col, value=value)
                row += 1
            
            row += 2
    else:
        # 'table_info' 열이 없는 경우 전체 DataFrame을 그대로 저장
        for col, header in enumerate(df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1
        
        for _, data_row in df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    wb.save(output_excel_path)
    print(f"Results have been saved to {output_excel_path}")

async def main():
    print("Program start")
    try:
        url = "https://www.kbinsure.co.kr/CG302120001.ec"
        pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
        output_dir = "/workspaces/automation/output"
        os.makedirs(output_dir, exist_ok=True)
        output_excel_path = os.path.join(output_dir, "comparison_results.xlsx")
        original_excel_path = os.path.join(output_dir, "변경전.xlsx")

        html_file_path, tables = await get_full_html_and_tables(url, output_dir)
        
        df_before = pd.DataFrame()
        if tables:
            df_before = process_tables(tables)
            save_original_tables_to_excel([df_before], original_excel_path)
            print("Combined DataFrame:")
            print(df_before.head())
            print(f"Shape of combined DataFrame: {df_before.shape}")
        else:
            print("No tables extracted from the website. Proceeding with an empty DataFrame.")

        texts_with_context, all_tables = extract_highlighted_text_and_tables(pdf_path, output_dir)

        if texts_with_context or all_tables:
            print("Successfully extracted content from PDF.")
            print(f"Number of extracted tables: {len(all_tables)}")
        else:
            print("Failed to extract content from PDF. Please check the PDF file.")

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print("Detailed error information:")
        import traceback
        print(traceback.format_exc())
    
    print("Program end")

if __name__ == "__main__":
    asyncio.run(main())