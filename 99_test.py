import os
import base64
import requests
import fitz  # PyMuPDF
from openpyxl import Workbook
from PIL import Image
import io
import json
from dotenv import load_dotenv

# .env 파일에서 환경 변수 로드
load_dotenv()

# 지정된 PDF 파일 경로
PDF_PATH = "/workspaces/automation/uploads/1722922992_5._KB_5.10.10_24.05__0801_v1.0.pdf"
OUTPUT_FOLDER = "highlight_images"

def get_api_key():
    # 환경 변수에서 API 키 읽기
    api_key = os.getenv("OPENAI_API_KEY")
    
    if not api_key:
        print("API 키가 환경 변수에 설정되어 있지 않습니다.")
        print("OpenAI API 키를 얻는 방법:")
        print("1. https://platform.openai.com에 로그인하세요.")
        print("2. 우측 상단의 계정 아이콘을 클릭하고 'View API keys'를 선택하세요.")
        print("3. 'Create new secret key'를 클릭하여 새 API 키를 생성하세요.")
        print("4. 생성된 키를 안전한 곳에 복사하여 저장하세요.")
        print("\n주의: API 키는 한 번만 표시되므로 반드시 저장해두세요.")
        
        api_key = input("\nOpenAI API 키를 입력하세요: ").strip()
        
        # API 키를 .env 파일에 저장할지 묻기
        save_key = input("이 API 키를 .env 파일에 저장하시겠습니까? (y/n): ").lower()
        if save_key == 'y':
            with open(".env", "a") as env_file:
                env_file.write(f"\nOPENAI_API_KEY={api_key}")
            print("API 키가 .env 파일에 저장되었습니다.")
    
    return api_key

def extract_images_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    images = []
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        images.append(img)
    doc.close()
    return images

def encode_image(image):
    buffered = io.BytesIO()
    image.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode('utf-8')

def detect_highlights(image, api_key):
    base64_image = encode_image(image)
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    payload = {
        "model": "gpt-4-vision-preview",
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "이 이미지에서 하이라이트된 텍스트를 찾아 정확히 추출해주세요. 또한 하이라이트된 영역의 좌표(x, y, width, height)도 함께 제공해주세요. JSON 형식으로 응답해주세요. 하이라이트된 텍스트가 없다면 빈 리스트를 반환해주세요."
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64_image}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 500
    }

    try:
        response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        
        content = response.json()['choices'][0]['message']['content']
        return json.loads(content)
    except requests.exceptions.RequestException as e:
        print(f"API request failed: {e}")
        return []
    except (KeyError, json.JSONDecodeError) as e:
        print(f"Failed to parse API response: {e}")
        return []

def capture_highlight(image, coordinates, output_path):
    x, y, width, height = coordinates
    highlight = image.crop((x, y, x + width, y + height))
    highlight.save(output_path)

def create_excel_with_highlights(highlights, excel_filename):
    wb = Workbook()
    ws = wb.active
    
    ws.cell(row=1, column=1, value="하이라이트된 텍스트")
    ws.cell(row=1, column=2, value="이미지 파일 경로")
    for i, highlight in enumerate(highlights, start=2):
        ws.cell(row=i, column=1, value=highlight['text'])
        ws.cell(row=i, column=2, value=highlight['image_path'])
    
    wb.save(excel_filename)
    return excel_filename

def process_pdf(pdf_path, api_key):
    try:
        print(f"Processing PDF: {pdf_path}")
        
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"The file {pdf_path} does not exist.")
        
        if not os.path.exists(OUTPUT_FOLDER):
            os.makedirs(OUTPUT_FOLDER)
        
        images = extract_images_from_pdf(pdf_path)
        print(f"Extracted {len(images)} pages from PDF")
        
        highlights = []
        for i, image in enumerate(images):
            print(f"Processing page {i+1}")
            detected_highlights = detect_highlights(image, api_key)
            for j, highlight in enumerate(detected_highlights):
                if 'text' in highlight and 'coordinates' in highlight:
                    output_image_path = os.path.join(OUTPUT_FOLDER, f"highlight_page{i+1}_{j+1}.png")
                    capture_highlight(image, highlight['coordinates'], output_image_path)
                    highlights.append({
                        'text': highlight['text'],
                        'image_path': output_image_path
                    })
                    print(f"Detected highlight on page {i+1}: {highlight['text']}")
        
        if highlights:
            excel_filename = "output_highlights.xlsx"
            excel_path = create_excel_with_highlights(highlights, excel_filename)
            print(f"Excel file created: {excel_path}")
            print(f"Total highlighted sections: {len(highlights)}")
        else:
            print("No highlights were detected in the PDF.")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    api_key = get_api_key()
    process_pdf(PDF_PATH, api_key)