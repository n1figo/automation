import PyPDF2
import camelot
import pandas as pd
import fitz  # PyMuPDF
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from sentence_transformers import SentenceTransformer
import faiss
import numpy as np

def extract_text_with_positions(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text_chunks = []
        for page_num, page in enumerate(reader.pages):
            text = page.extract_text()
            lines = text.split('\n')
            for line_num, line in enumerate(lines):
                text_chunks.append({
                    'text': line,
                    'page': page_num + 1,
                    'line': line_num + 1
                })
        return text_chunks

def preprocess_text(text):
    # 특수문자 및 숫자 제거, 공백 제거, 소문자로 변환
    text = re.sub(r'[\s\W\d_]+', '', text)
    text = text.lower()
    return text

def detect_table_boundaries(text_chunks):
    start_patterns = [r'표\s*\d+', r'선택\s*특약', r'상해관련\s*특약', r'질병관련\s*특약', r'\[1\s*종\]', r'\[2\s*종\]', r'\[3\s*종\]']
    end_patterns = [r'합\s*계', r'총\s*계', r'주\s*\)', r'※', r'결\s*론']

    tables = []
    table_start = None

    for i, chunk in enumerate(text_chunks):
        text = chunk['text']

        if table_start is None and any(re.search(pattern, text) for pattern in start_patterns):
            table_start = i
            continue

        if table_start is not None and any(re.search(pattern, text) for pattern in end_patterns):
            table_end = i

            start_page = text_chunks[table_start]['page']
            end_page = text_chunks[table_end]['page']
            pages = list(range(start_page, end_page + 1))

            context_start = max(0, table_start - 5)
            context_end = min(len(text_chunks), table_end + 5)
            context = text_chunks[context_start:context_end]

            tables.append({
                'start': text_chunks[table_start],
                'end': text_chunks[table_end],
                'pages': pages,
                'context': context
            })
            table_start = None

    return tables

def extract_text_from_page(page):
    return page.get_text("text")

def extract_text_above_bbox(page, bbox):
    x0, y0, x1, y1 = bbox  # bbox: (x0, y0, x1, y1)
    text_blocks = page.get_text("blocks")
    # 테이블 bbox의 y0보다 위에 있는 텍스트 블록 중 가장 아래에 있는 것 선택
    texts_above = []
    for block in text_blocks:
        if len(block) >= 5:
            bx0, by0, bx1, by1, text = block[:5]
            if by1 <= y0:  # 블록의 아래쪽 y좌표가 테이블의 위쪽 y좌표보다 작거나 같으면
                texts_above.append((by1, text))
    if texts_above:
        # y 좌표가 가장 큰 (테이블 바로 위에 있는) 텍스트 선택
        texts_above.sort(reverse=True)
        return texts_above[0][1].strip()
    else:
        return "제목 없음"

def extract_tables_with_camelot(pdf_path, tables_info):
    all_tables = []
    doc = fitz.open(pdf_path)
    for table_info in tables_info:
        pages = table_info['pages']
        pages_str = ','.join(map(str, pages))
        print(f"Extracting table from pages {pages_str} using Camelot...")
        tables = camelot.read_pdf(pdf_path, pages=pages_str, flavor='lattice')

        if not tables:
            continue

        # 여러 페이지에서 추출된 테이블을 하나로 병합
        combined_df = pd.DataFrame()
        for table in tables:
            combined_df = pd.concat([combined_df, table.df], ignore_index=True)

        # 표 바로 위의 텍스트 추출
        first_page_number = pages[0] - 1  # fitz 모듈은 0부터 시작
        page = doc.load_page(first_page_number)
        table_bbox = tables[0]._bbox  # 첫 번째 테이블의 bbox 사용
        text_above_table = extract_text_above_bbox(page, table_bbox)

        all_tables.append({
            'dataframe': combined_df,
            'title': text_above_table,
            'pages': pages
        })
    print(f"Found {len(all_tables)} tables in total")
    return all_tables

def process_tables(all_tables):
    processed_data = []
    for i, table_info in enumerate(all_tables):
        df = table_info['dataframe']
        title = table_info['title']
        pages = table_info['pages']
        df['Table_Number'] = i + 1
        df['Table_Title'] = title
        df['Pages'] = ', '.join(map(str, pages))
        processed_data.append(df)
    if processed_data:
        return pd.concat(processed_data, ignore_index=True)
    else:
        return pd.DataFrame()

def save_to_excel(df_dict, output_path, title=None):
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    for sheet_name, df in df_dict.items():
        ws = wb.create_sheet(title=sheet_name)

        if title:
            ws.cell(row=1, column=1, value=f"{title} - {sheet_name}")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = Font(size=20, bold=True)
            ws.row_dimensions[1].height = 30
            start_row = 2
        else:
            start_row = 1

        current_row = start_row
        # 'Table_Number'로 그룹핑하여 각 테이블을 구분
        if 'Table_Number' in df.columns:
            grouped = df.groupby('Table_Number')
            for table_number, group in grouped:
                table_title = group['Table_Title'].iloc[0]
                pages = group['Pages'].iloc[0]

                # 테이블 제목 추가
                ws.cell(row=current_row, column=1, value=table_title)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
                title_cell = ws.cell(row=current_row, column=1)
                title_cell.font = Font(size=14, bold=True)
                ws.row_dimensions[current_row].height = 20
                current_row += 1

                # 헤더 바로 위에 표 위의 텍스트 추가 (폰트 크게, 볼드체)
                header_title = f"{table_title}"
                ws.cell(row=current_row, column=1, value=header_title)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
                header_cell = ws.cell(row=current_row, column=1)
                header_cell.font = Font(size=12, bold=True)
                ws.row_dimensions[current_row].height = 18
                current_row += 1

                # 테이블 데이터 작성
                for r_idx, row in enumerate(dataframe_to_rows(group.drop(['Table_Number', 'Table_Title', 'Pages'], axis=1), index=False, header=True), start=current_row):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    # 페이지 번호를 오른편에 추가
                    ws.cell(row=r_idx, column=len(row)+1, value=pages)
                current_row = r_idx + 2  # 각 테이블 후에 공백 추가
        else:
            # 데이터프레임이 비어 있는 경우
            ws.cell(row=current_row, column=1, value="데이터 없음")
            current_row += 1

        # 열 너비 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Data saved to '{output_path}'")

def split_text_into_chunks(text, chunk_size=200, overlap=50):
    words = text.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        chunk = " ".join(words[i:i + chunk_size])
        chunks.append(chunk)
    return chunks

def create_index(chunks):
    model = SentenceTransformer('distiluse-base-multilingual-cased')
    embeddings = model.encode(chunks)

    # 임베딩 정규화
    embeddings = embeddings / np.linalg.norm(embeddings, axis=1, keepdims=True)

    dimension = embeddings.shape[1]
    index = faiss.IndexFlatIP(dimension)
    index.add(embeddings.astype('float32'))

    return index, model

def search(query, index, model, chunks, page_numbers=None, k=10, threshold=0.5):
    query_vector = model.encode([query])
    query_vector = query_vector / np.linalg.norm(query_vector, axis=1, keepdims=True)

    distances, indices = index.search(query_vector.astype('float32'), k)

    results = []
    for idx, score in zip(indices[0], distances[0]):
        if score >= threshold:  # 임계값 이상인 경우만 결과에 추가
            result = {'content': chunks[idx], 'score': score}
            if page_numbers:
                result['page'] = page_numbers[idx]
            results.append(result)

    return results

def results_to_dataframe(results, query_type):
    df = pd.DataFrame(results)
    df['type'] = query_type
    return df

def find_pages_with_keyword(text, keyword, page_numbers):
    pages = []
    words = text.split()
    preprocessed_keyword = preprocess_text(keyword)
    preprocessed_words = [preprocess_text(word) for word in words]

    for word, page in zip(preprocessed_words, page_numbers):
        if preprocessed_keyword in word:
            if page not in pages:
                pages.append(page)
    return pages

def main():
    uploads_folder = "/workspaces/automation/uploads"
    output_folder = "/workspaces/automation/output"

    os.makedirs(output_folder, exist_ok=True)

    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_analysis.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    text_chunks = extract_text_with_positions(pdf_path)
    tables = detect_table_boundaries(text_chunks)

    doc = fitz.open(pdf_path)

    # 전체 텍스트와 페이지 번호 추출
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        full_text = ""
        page_numbers = []
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text()
            full_text += page_text + "\n"
            page_numbers.extend([i+1] * len(page_text.split()))

    # 텍스트 청크 생성 및 인덱스 구축
    chunks = split_text_into_chunks(full_text)
    index, model = create_index(chunks)

    # 선택특약 검색
    select_query = "선택특약"
    select_results = search(select_query, index, model, chunks, page_numbers, k=10, threshold=0.5)
    select_df = results_to_dataframe(select_results, "선택특약")

    # 상해관련 특약 검색
    injury_query = "상해관련특약"
    injury_results = search(injury_query, index, model, chunks, page_numbers, k=10, threshold=0.5)
    injury_df = results_to_dataframe(injury_results, "상해관련특약")

    # 결과 합치기
    final_df = pd.concat([select_df, injury_df], ignore_index=True)

    # 엑셀 파일로 저장
    search_output_path = os.path.join(output_folder, "insurance_special_clauses_search_results.xlsx")
    final_df.to_excel(search_output_path, index=False, engine='openpyxl')

    print(f"검색 결과가 {search_output_path}에 저장되었습니다.")

    # 키워드가 포함된 페이지 찾기 및 출력
    select_pages = find_pages_with_keyword(full_text, "선택특약", page_numbers)
    injury_pages = find_pages_with_keyword(full_text, "상해관련 특약", page_numbers)
    injury_special_pages = find_pages_with_keyword(full_text, "상해관련 특별약관", page_numbers)

    print("선택특약이 포함된 페이지:", select_pages)
    print("상해관련 특약이 포함된 페이지:", injury_pages)
    print("상해관련 특별약관이 포함된 페이지:", injury_special_pages)

    # 선택특약이 있는 페이지의 텍스트를 txt 파일로 저장
    for page_num in select_pages:
        page = doc.load_page(page_num - 1)
        text = extract_text_from_page(page)
        txt_output_path = os.path.join(output_folder, f"page_{page_num}_text.txt")
        with open(txt_output_path, 'w', encoding='utf-8') as f:
            f.write(text)
        print(f"Page {page_num}의 텍스트가 {txt_output_path}에 저장되었습니다.")

    # '상해관련 특약'이 파싱되었는지 확인
    if injury_pages or injury_special_pages:
        print("'상해관련 특약'이 문서에서 발견되었습니다.")
    else:
        print("'상해관련 특약'이 문서에서 발견되지 않았습니다.")

    # 기존 테이블 처리 및 엑셀 저장
    df_dict = {}
    types = ["[1종]", "[2종]", "[3종]", "선택특약", "상해관련특약", "질병관련특약"]
    type_tables_info = {t: [] for t in types}

    for table in tables:
        table_text = ' '.join([chunk['text'] for chunk in table['context']])
        for t in types:
            if re.search(preprocess_text(t), preprocess_text(table_text)):
                type_tables_info[t].append(table)
                break

    for insurance_type in ["[1종]", "[2종]", "[3종]", "선택특약", "상해관련특약", "질병관련특약"]:
        if type_tables_info[insurance_type]:
            type_tables = type_tables_info[insurance_type]
            camelot_tables = extract_tables_with_camelot(pdf_path, type_tables)
            df = process_tables(camelot_tables)
            df_dict[insurance_type.strip('[]')] = df

    # 첫 번째 페이지에서 제목 추출
    first_page = doc.load_page(0)
    page_text = extract_text_from_page(first_page)
    title = page_text.strip().split('\n')[0]
    print(f"Extracted title: {title}")

    # 결과를 엑셀로 저장
    save_to_excel(df_dict, output_excel_path, title=title)

    print(f"All processed data has been saved to {output_excel_path}")

if __name__ == "__main__":
    main()
