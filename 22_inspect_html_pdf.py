import camelot
import pandas as pd
import numpy as np
import cv2
import os
import fitz  # PyMuPDF
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
from difflib import SequenceMatcher
import re
import sys
import logging
import argparse
from playwright.sync_api import sync_playwright

def ensure_inspection_upload_folder():
    folder_name = 'inspection upload'
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        logging.info(f"'{folder_name}' 폴더를 생성했습니다.")
    else:
        logging.info(f"'{folder_name}' 폴더가 이미 존재합니다.")
    return folder_name

def ensure_inspection_output_folder():
    folder_name = 'inspection output'
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        logging.info(f"'{folder_name}' 폴더를 생성했습니다.")
    else:
        logging.info(f"'{folder_name}' 폴더가 이미 존재합니다.")
    return folder_name

def find_files_in_folder(folder_path):
    pdf_files = {}
    html_files = {}
    for file_name in os.listdir(folder_path):
        lower_file_name = file_name.lower()
        file_path = os.path.join(folder_path, file_name)
        if lower_file_name.endswith('.pdf'):
            if '요약서' in lower_file_name:
                pdf_files['요약서'] = file_path
            elif '가입예시' in lower_file_name:
                pdf_files['가입예시'] = file_path
        elif (lower_file_name.endswith('.html') or
              lower_file_name.endswith('.htm') or
              lower_file_name.endswith('.mhtml')):
            if '보장내용' in lower_file_name:
                html_files['보장내용'] = file_path
            elif '가입예시' in lower_file_name:
                html_files['가입예시'] = file_path
    return pdf_files, html_files

def extract_html_content_with_playwright(html_path):
    try:
        logging.info(f"Playwright를 사용하여 '{html_path}' 파일을 처리합니다.")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            page = context.new_page()

            # mhtml 파일을 로드
            page.goto(f'file://{os.path.abspath(html_path)}')

            # 페이지가 로드될 때까지 대기
            page.wait_for_load_state('networkidle')

            html_content = page.content()
            browser.close()

            soup = BeautifulSoup(html_content, 'html.parser')
            logging.info(f"'{html_path}' 파일의 HTML 콘텐츠를 추출했습니다.")
            return soup
    except Exception as e:
        logging.error(f"Playwright를 사용하여 HTML 내용을 추출하는 데 실패했습니다: {e}")
        sys.exit(1)

def extract_html_content(html_path):
    try:
        if html_path.lower().endswith('.mhtml'):
            # Playwright를 사용하여 mhtml 파일 처리
            soup = extract_html_content_with_playwright(html_path)
            return soup
        else:
            # 일반 HTML 파일 처리
            with open(html_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
            soup = BeautifulSoup(html_content, 'html.parser')
            logging.info(f"'{html_path}' 파일의 HTML 콘텐츠를 추출했습니다.")
            return soup
    except Exception as e:
        logging.error(f"HTML 내용을 추출하는 데 실패했습니다: {e}")
        sys.exit(1)

def extract_relevant_tables(soup):
    try:
        logging.info("HTML에서 관련 테이블을 추출합니다.")
        target_text = '상해관련 특별약관'
        def find_start_tag(soup, text):
            for element in soup.find_all(string=re.compile(re.escape(text))):
                return element.parent
            return None

        start_tag = find_start_tag(soup, target_text)
        if not start_tag:
            logging.error(f"'{target_text}'에 해당하는 시작 태그를 찾을 수 없습니다.")
            sys.exit(1)

        # 시작 태그 이후의 모든 테이블을 추출
        tables = []
        for sibling in start_tag.find_all_next():
            if sibling.name == 'table':
                tables.append(sibling)
            elif sibling.name == 'h1' or sibling.name == 'h2' or sibling.name == 'h3':
                # 새로운 섹션이 시작되면 중단
                break

        if not tables:
            logging.error("시작 태그 이후에 테이블을 찾을 수 없습니다.")
            sys.exit(1)

        logging.info(f"관련 테이블 {len(tables)}개를 추출했습니다.")
        # 제목과 테이블을 튜플로 반환 (제목은 target_text로 설정)
        relevant_tables = [(target_text, table) for table in tables]
        return relevant_tables
    except Exception as e:
        logging.error(f"HTML 테이블을 추출하는 데 실패했습니다: {e}")
        sys.exit(1)

def pdf_to_image(page):
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return np.array(img)

def detect_highlights(image):
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    
    # 여러 색상 범위 정의 (HSV)
    color_ranges = [
        ((20, 100, 100), (40, 255, 255)),  # 노란색
        ((100, 100, 100), (140, 255, 255)),  # 파란색
        ((125, 100, 100), (155, 255, 255))  # 보라색
    ]
    
    masks = []
    for lower, upper in color_ranges:
        mask = cv2.inRange(hsv, np.array(lower), np.array(upper))
        masks.append(mask)
    
    # 모든 마스크 결합
    combined_mask = np.zeros_like(masks[0])
    for mask in masks:
        combined_mask = cv2.bitwise_or(combined_mask, mask)
    
    kernel = np.ones((5,5), np.uint8)
    cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)
    
    contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    return contours

def get_highlight_regions(contours, image_height):
    regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        top = image_height - (y + h)
        bottom = image_height - y
        regions.append((top, bottom))
    return regions

def extract_tables_with_camelot(pdf_path, page_number):
    logging.info(f"{page_number} 페이지에서 Camelot을 사용하여 테이블을 추출합니다.")
    tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
    logging.info(f"{page_number} 페이지에서 {len(tables)}개의 테이블을 찾았습니다.")
    return tables

def process_tables(tables, highlight_regions, page_height):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        x1, y1, x2, y2 = table._bbox

        table_height = y2 - y1
        row_height = table_height / len(df)

        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            
            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height
            
            row_highlighted = check_highlight((row_top, row_bottom), highlight_regions)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)

    return pd.DataFrame(processed_data)

def check_highlight(row_range, highlight_regions):
    row_top, row_bottom = row_range
    for region_top, region_bottom in highlight_regions:
        if (region_top <= row_top <= region_bottom) or (region_top <= row_bottom <= region_bottom) or \
           (row_top <= region_top <= row_bottom) or (row_top <= region_bottom <= row_bottom):
            return True
    return False

def preprocess_text(text):
    # 공백 및 특수문자 제거
    text = re.sub(r'[^\w]', '', text)  # 특수문자 제거
    text = re.sub(r'\s+', '', text)    # 공백 제거
    return text

def compare_tables_and_generate_report(html_tables, pdf_tables, similarity_threshold):
    total_mismatches = 0
    results = []  # 엑셀 출력을 위한 데이터 저장 리스트

    logging.info(f"총 {len(html_tables)}개의 HTML 테이블과 {len(pdf_tables)}개의 PDF 테이블을 비교합니다.")

    for idx, (title, html_table) in enumerate(html_tables, start=1):
        logging.info(f"테이블 {idx}/{len(html_tables)}: '{title}' 비교 시작")
        # HTML 테이블 데이터 추출
        html_rows = html_table.find_all('tr')
        html_data = []
        for tr in html_rows:
            row = [td.get_text(strip=True) for td in tr.find_all(['td', 'th'])]
            html_data.append(row)

        # PDF 테이블 가져오기
        if idx-1 < len(pdf_tables):
            pdf_table = pdf_tables[idx-1]
            pdf_data = pdf_table.df.values.tolist()
            # 컬럼명 유지
            pdf_columns = pdf_data[0]
            pdf_data = pdf_data[1:]
        else:
            logging.warning(f"PDF 테이블이 부족합니다. 인덱스: {idx-1}")
            pdf_data = []
            pdf_columns = []

        # 결과를 저장하기 전에 제목을 추가
        results.append([f"제목: {title}"])
        max_rows = max(len(html_data), len(pdf_data))

        # 헤더 작성 (PDF 데이터의 컬럼명을 사용)
        html_header = html_data[0] if html_data else []
        pdf_header = pdf_columns if pdf_columns else []
        header = html_header + pdf_header + ['검수과정']

        # 헤더를 결과에 추가
        results.append(header)

        # 데이터 비교 및 저장
        for i in range(1, max_rows):
            html_row = html_data[i] if i < len(html_data) else [''] * len(html_header)
            pdf_row = pdf_data[i-1] if i-1 < len(pdf_data) else [''] * len(pdf_header)

            # 각 행의 데이터를 병합
            combined_row = html_row + pdf_row

            # 행 단위로 비교
            html_line = ''.join(html_row)
            pdf_line = ''.join(pdf_row)
            similarity = SequenceMatcher(None, preprocess_text(html_line), preprocess_text(pdf_line)).ratio()

            검수과정 = ''
            if similarity < similarity_threshold:
                검수과정 = '불일치'
                total_mismatches += 1

            combined_row.append(검수과정)
            results.append(combined_row)

            if i % 10 == 0:
                logging.debug(f"'{title}' 테이블의 {i}/{max_rows}번째 행 비교 완료")

        logging.info(f"테이블 '{title}' 비교 완료")

    return results, total_mismatches

def write_results_to_excel(data, output_path):
    try:
        wb = Workbook()
        ws = wb.active

        for row_idx, row in enumerate(data, start=1):
            ws.append(row)
            if row_idx % 100 == 0:
                logging.debug(f"{row_idx}개의 행을 엑셀에 기록했습니다.")

        wb.save(output_path)
        logging.info(f"검수 과정을 '{output_path}' 파일에 저장했습니다.")
    except Exception as e:
        logging.error(f"검수 과정을 엑셀 파일로 저장하는 데 실패했습니다: {e}")

def save_to_excel_with_highlight(df, output_path, title=None):
    wb = Workbook()
    ws = wb.active

    # 제목을 추가
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title)
        max_col = len(df.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=20, bold=True)
        ws.row_dimensions[1].height = 30  # 제목 행 높이 조정
        start_row = 2  # 데이터는 다음 행부터 시작

    # DataFrame을 Excel로 저장
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    if '변경사항' in df.columns:
        change_col_index = df.columns.get_loc('변경사항') + 1
    else:
        raise ValueError("DataFrame에 '변경사항' 컬럼이 없습니다.")

    for row in range(start_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=change_col_index).value
        if cell_value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)
    logging.info(f"데이터를 '{output_path}'에 저장했습니다. (하이라이트 적용됨)")

def main(similarity_threshold=0.95, log_level='INFO'):
    numeric_level = getattr(logging, log_level.upper(), None)
    if not isinstance(numeric_level, int):
        print(f"유효하지 않은 로그 레벨입니다: {log_level}")
        sys.exit(1)
    logging.basicConfig(level=numeric_level, format='%(asctime)s - %(levelname)s - %(message)s')

    folder_path = ensure_inspection_upload_folder()
    output_folder_path = ensure_inspection_output_folder()

    # 폴더 내의 파일 리스트 출력
    files_in_folder = os.listdir(folder_path)
    if files_in_folder:
        logging.info("폴더 내의 파일 리스트:")
        for file_name in files_in_folder:
            logging.info(f"- {file_name}")
    else:
        logging.info("폴더가 비어 있습니다.")

    pdf_files, html_files = find_files_in_folder(folder_path)

    # '요약서' PDF와 '보장내용' HTML 파일 비교
    if '요약서' in pdf_files and '보장내용' in html_files:
        pdf_path = pdf_files['요약서']
        html_path = html_files['보장내용']

        logging.info(f"비교할 PDF 파일: {pdf_path}")
        logging.info(f"비교할 HTML 파일: {html_path}")

        # PDF에서 제목 추출
        doc = fitz.open(pdf_path)
        first_page = doc[0]
        page_text = first_page.get_text("text")
        pdf_title = page_text.strip().split('\n')[0]  # 첫 번째 줄을 제목으로 가정
        logging.info(f"PDF에서 제목을 추출했습니다: {pdf_title}")

        # HTML 콘텐츠 추출
        soup = extract_html_content(html_path)

        # HTML에서 필요한 테이블 추출
        html_tables = extract_relevant_tables(soup)
        if not html_tables:
            logging.error("HTML에서 필요한 테이블을 찾을 수 없습니다.")
            sys.exit(1)

        # PDF에서 '상해관련 특별약관' 이후의 테이블 추출
        target_text = '상해관련 특별약관'
        page_num = None

        # '상해관련 특별약관'이 있는 페이지를 찾음
        for num, page in enumerate(doc, start=1):
            text = page.get_text()
            logging.debug(f"페이지 {num}의 텍스트 내용:\n{text}\n")
            if target_text in text:
                page_num = num
                logging.info(f"'{target_text}'가 {page_num}번째 페이지에 있습니다.")
                break

        if page_num is None:
            logging.error(f"'{target_text}'를 PDF에서 찾을 수 없습니다.")
            # 추가 디버깅: 모든 페이지에서 텍스트 검색
            logging.debug("PDF의 모든 페이지에서 텍스트를 검색합니다.")
            for num, page in enumerate(doc, start=1):
                text = page.get_text()
                if target_text in text:
                    logging.info(f"'{target_text}'가 {num}번째 페이지에 있습니다.")
                    page_num = num
                    break
            if page_num is None:
                logging.error(f"'{target_text}'를 PDF에서 여전히 찾을 수 없습니다. 스크립트를 종료합니다.")
                sys.exit(1)

        all_processed_data = []
        pdf_tables = []

        for page_number in range(page_num - 1, len(doc)):
            logging.info(f"Processing page: {page_number + 1}/{len(doc)}")

            page = doc[page_number]
            image = pdf_to_image(page)

            contours = detect_highlights(image)
            highlight_regions = get_highlight_regions(contours, image.shape[0])

            logging.info(f"Page {page_number + 1}: Detected {len(highlight_regions)} highlighted regions")

            tables = extract_tables_with_camelot(pdf_path, page_number + 1)

            if not tables:
                logging.info(f"Page {page_number + 1}: No tables extracted")
                continue

            # 테이블 저장 (검수 과정에 사용하기 위해)
            pdf_tables.extend(tables)

            processed_df = process_tables(tables, highlight_regions, image.shape[0])
            processed_df['Page_Number'] = page_number + 1
            all_processed_data.append(processed_df)

        if not all_processed_data:
            logging.error("No processed data available.")
            sys.exit(1)

        final_df = pd.concat(all_processed_data, ignore_index=True)

        # 결과를 엑셀 파일로 저장 (하이라이트된 변경 사항 포함)
        output_excel_path = os.path.join(output_folder_path, '검수과정.xlsx')
        save_to_excel_with_highlight(final_df, output_excel_path, title=pdf_title)

        # 테이블 비교 및 결과 생성
        comparison_results, total_mismatches = compare_tables_and_generate_report(html_tables, pdf_tables, similarity_threshold)

        # 비교 결과를 추가 시트에 저장
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "하이라이트된 변경 사항"

        for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws1.cell(row=r_idx, column=c_idx, value=value)

        ws2 = wb.create_sheet(title="테이블 비교 결과")
        for row_idx, row in enumerate(comparison_results, start=1):
            ws2.append(row)

        wb.save(output_excel_path)
        logging.info(f"검수 과정을 '{output_excel_path}' 파일에 저장했습니다.")

        if total_mismatches > 0:
            logging.warning(f"{total_mismatches}개의 불일치가 발견되었습니다.")
        else:
            logging.info("모든 테이블이 일치합니다. PASS")
    else:
        logging.error("'요약서' PDF 또는 '보장내용' HTML 파일을 찾을 수 없습니다.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="PDF와 HTML 문서를 비교하여 검수 과정을 자동화합니다.")
    parser.add_argument('--threshold', type=float, default=0.95, help="유사도 임계값 (기본값: 0.95)")
    parser.add_argument('--loglevel', default='INFO', help="로그 레벨 설정 (예: DEBUG, INFO, WARNING, ERROR)")
    args = parser.parse_args()

    main(similarity_threshold=args.threshold, log_level=args.loglevel)
