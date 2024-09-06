import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import os
from PIL import Image
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
from io import StringIO
import pdfplumber
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

async def get_full_html_and_tables(url, output_dir):
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.goto(url, wait_until="networkidle")
            
            await page.click('text="보장내용"')
            await page.wait_for_load_state('networkidle')
            
            html_content = await page.content()
            
            tables = await page.evaluate('''
                () => {
                    const tables = document.querySelectorAll('.tab_cont[data-list="보장내용"] table');
                    return Array.from(tables).map(table => {
                        const rows = table.querySelectorAll('tr');
                        return Array.from(rows).map(row => {
                            const cells = row.querySelectorAll('th, td');
                            return Array.from(cells).map(cell => cell.innerText);
                        });
                    });
                }
            ''')
            
            await page.screenshot(path=os.path.join(output_dir, "full_page.png"), full_page=True)
            
            table_elements = await page.query_selector_all('.tab_cont[data-list="보장내용"] table')
            for i, table_element in enumerate(table_elements):
                try:
                    await table_element.scroll_into_view_if_needed()
                    await page.wait_for_timeout(1000)
                    await table_element.screenshot(path=os.path.join(output_dir, f"table_{i+1}.png"))
                except Exception as e:
                    logger.error(f"Failed to capture screenshot for table {i+1}: {str(e)}")
            
            await browser.close()
        
        html_file_path = os.path.join(output_dir, "source.txt")
        with open(html_file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"Full HTML source and table screenshots have been saved to {output_dir}")
        return html_file_path, tables
    except Exception as e:
        logger.error(f"Error in get_full_html_and_tables: {str(e)}")
        return None, []

def process_tables(tables):
    all_data = []
    for i, table in enumerate(tables):
        df = pd.DataFrame(table[1:], columns=table[0])
        df['table_info'] = f'Table_{i+1}'
        all_data.append(df)
    
    if not all_data:
        logger.warning("No data extracted.")
        return pd.DataFrame()
    
    result = pd.concat(all_data, axis=0, ignore_index=True)
    
    logger.info(f"Final DataFrame: Columns: {result.columns.tolist()}, Shape: {result.shape}")
    
    return result

def is_color_highlighted(color):
    r, g, b = color
    if r == g == b:
        return False
    return max(r, g, b) > 200 and (max(r, g, b) - min(r, g, b)) > 30

def is_red_text(color):
    r, g, b = color
    return r > 200 and g < 100 and b < 100

def detect_highlights_and_colors(image):
    width, height = image.size
    img_array = np.array(image)
    
    highlighted_rows = set()
    red_text_rows = set()
    for y in range(height):
        for x in range(width):
            color = img_array[y, x]
            if color[0] > 200 and color[1] > 200 and color[2] < 100:  # Yellow highlight
                highlighted_rows.add(y)
            elif is_red_text(color):
                red_text_rows.add(y)
    
    all_special_rows = sorted(highlighted_rows.union(red_text_rows))
    return all_special_rows, highlighted_rows, red_text_rows

def extract_highlighted_text_and_tables(pdf_path, output_dir):
    logger.info("Starting extraction of highlighted text and tables from PDF...")
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        texts_with_context = []
        highlighted_pages = set()
        
        output_image_dir = os.path.join(output_dir, "images")
        os.makedirs(output_image_dir, exist_ok=True)
        
        for page_num, page in enumerate(pdf.pages):
            logger.info(f"Processing page {page_num + 1}/{total_pages}")
            
            img = page.to_image()
            pil_image = img.original
            
            all_special_rows, highlighted_rows, red_text_rows = detect_highlights_and_colors(pil_image)
            
            if all_special_rows:
                highlighted_pages.add(page_num + 1)
                
                image_filename = f"page_{page_num + 1}_full.png"
                image_path = os.path.join(output_image_dir, image_filename)
                pil_image.save(image_path)
                
                text = page.extract_text()
                text_lines = text.split('\n')
                
                for i, line in enumerate(text_lines):
                    status = '추가' if i in highlighted_rows or i in red_text_rows else '유지'
                    texts_with_context.append((line, page_num + 1, image_path, status))
            else:
                text = page.extract_text()
                texts_with_context.extend((line, page_num + 1, None, '유지') for line in text.split('\n'))
        
        # Extract tables from highlighted pages
        table_data = {}
        for page_num in highlighted_pages:
            page = pdf.pages[page_num - 1]
            tables = page.extract_tables()
            if tables:
                for i, table in enumerate(tables):
                    df = pd.DataFrame(table[1:], columns=table[0])
                    sheet_name = f"Page_{page_num}_Table_{i+1}"
                    table_data[sheet_name] = df
        
        if table_data:
            tables_excel_path = os.path.join(output_dir, "highlighted_tables.xlsx")
            with pd.ExcelWriter(tables_excel_path, engine='openpyxl') as writer:
                for sheet_name, df in table_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info(f"Saved all highlighted tables to {tables_excel_path}")

    logger.info(f"Finished extracting text and tables from PDF (total {total_pages} pages)")
    return texts_with_context

def compare_dataframes(df_before, texts_with_context):
    logger.info("Starting comparison of dataframes...")
    df_result = df_before.copy() if not df_before.empty else pd.DataFrame(columns=['table_info'])
    df_result['PDF_페이지'] = ''
    df_result['이미지_경로'] = ''
    df_result['상태'] = '유지'

    for text, page_num, image_path, status in texts_with_context:
        for i in range(len(df_result)):
            if text.strip() in ' '.join(str(cell) for cell in df_result.iloc[i]):
                df_result.loc[i, 'PDF_페이지'] = page_num
                df_result.loc[i, '이미지_경로'] = image_path
                df_result.loc[i, '상태'] = status
                break

    # Move '상태' column to the rightmost position
    cols = df_result.columns.tolist()
    cols.append(cols.pop(cols.index('상태')))
    df_result = df_result[cols]

    logger.info(f"Finished comparison. Updated {len(df_result[df_result['상태'] != '유지'])} rows")
    return df_result

def save_to_excel(df, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    row = 1
    if 'table_info' in df.columns:
        for table_info in df['table_info'].unique():
            table_df = df[df['table_info'] == table_info]
            
            ws.cell(row=row, column=1, value=table_info)
            row += 1
            
            for col, header in enumerate(table_df.columns, start=1):
                ws.cell(row=row, column=col, value=header)
            row += 1
            
            for _, data_row in table_df.iterrows():
                for col, value in enumerate(data_row, start=1):
                    ws.cell(row=row, column=col, value=value)
                row += 1
            
            row += 2
    else:
        for col, header in enumerate(df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1
        
        for _, data_row in df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    wb.save(output_excel_path)
    logger.info(f"Results have been saved to {output_excel_path}")

async def main():
    logger.info("Program start")
    try:
        url = "https://www.kbinsure.co.kr/CG302290001.ec#"
        pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
        output_dir = "/workspaces/automation/output"
        os.makedirs(output_dir, exist_ok=True)
        output_excel_path = os.path.join(output_dir, "comparison_results.xlsx")
        original_excel_path = os.path.join(output_dir, "변경전.xlsx")

        html_file_path, tables = await get_full_html_and_tables(url, output_dir)
        
        df_before = pd.DataFrame()
        if tables:
            df_before = process_tables(tables)
            with pd.ExcelWriter(original_excel_path, engine='openpyxl') as writer:
                df_before.to_excel(writer, sheet_name='Original Tables', index=False)
            logger.info(f"Original tables saved to {original_excel_path}")
        else:
            logger.warning("No tables extracted from the website. Proceeding with an empty DataFrame.")

        texts_with_context = extract_highlighted_text_and_tables(pdf_path, output_dir)

        if texts_with_context:
            df_matching = compare_dataframes(df_before, texts_with_context)
            save_to_excel(df_matching, output_excel_path)
        else:
            logger.error("Failed to extract highlighted text from PDF. Please check the PDF file.")

    except Exception as e:
        logger.error(f"Error occurred: {str(e)}", exc_info=True)
    
    logger.info("Program end")

if __name__ == "__main__":
    asyncio.run(main())