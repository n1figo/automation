import os
import shutil
import tempfile
from typing import List

import camelot
import cv2
import fitz
import numpy as np
import pandas as pd
from fastapi import FastAPI, File, UploadFile, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image

app = FastAPI(title="PDF 처리 웹 서비스")

templates = Jinja2Templates(directory="templates")

def pdf_to_image(page):
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return np.array(img)

def detect_highlights(image):
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    s = hsv[:, :, 1]
    v = hsv[:, :, 2]

    saturation_threshold = 30
    saturation_mask = s > saturation_threshold

    _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)

    kernel = np.ones((5, 5), np.uint8)
    cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)

    contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    return contours

def get_highlight_regions(contours, image_height):
    regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        top = image_height - (y + h)
        bottom = image_height - y
        regions.append((top, bottom))
    return regions

def extract_tables_with_camelot(pdf_path, page_number):
    print(f"Extracting tables from page {page_number} using Camelot...")
    try:
        tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
    except Exception as e:
        print(f"Camelot 에러: {e}")
        tables = camelot.TableList()
    print(f"Found {len(tables)} tables on page {page_number}")
    return tables

def process_tables(tables, highlight_regions, page_height):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        x1, y1, x2, y2 = table._bbox

        table_height = y2 - y1
        row_height = table_height / len(df)

        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            
            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height
            
            row_highlighted = check_highlight((row_top, row_bottom), highlight_regions)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)

    return pd.DataFrame(processed_data)

def check_highlight(row_range, highlight_regions):
    row_top, row_bottom = row_range
    for region_top, region_bottom in highlight_regions:
        if (region_top <= row_top <= region_bottom) or (region_top <= row_bottom <= region_bottom) or \
           (row_top <= region_top <= row_bottom) or (row_top <= region_bottom <= row_bottom):
            return True
    return False

def save_to_excel_with_highlight(df, output_path):
    df.to_excel(output_path, index=False)
    
    wb = load_workbook(output_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    if '변경사항' in df.columns:
        change_col_index = df.columns.get_loc('변경사항') + 1
    else:
        raise ValueError("DataFrame에 '변경사항' 컬럼이 없습니다.")

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=change_col_index).value
        if cell_value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows")

def process_pdf(pdf_path, output_excel_path):
    print("PDF에서 개정된 부분을 추출합니다...")

    doc = fitz.open(pdf_path)
    all_processed_data = []

    for page_number in range(len(doc)):
        print(f"처리 중인 페이지: {page_number + 1}/{len(doc)}")

        page = doc[page_number]
        image = pdf_to_image(page)

        contours = detect_highlights(image)
        highlight_regions = get_highlight_regions(contours, image.shape[0])

        print(f"페이지 {page_number + 1}: 감지된 강조 영역 수: {len(highlight_regions)}")

        tables = extract_tables_with_camelot(pdf_path, page_number + 1)

        if not tables:
            print(f"페이지 {page_number + 1}: 추출된 표가 없습니다.")
            continue

        processed_df = process_tables(tables, highlight_regions, image.shape[0])
        processed_df['Page_Number'] = page_number + 1
        all_processed_data.append(processed_df)

    if not all_processed_data:
        print("처리된 데이터가 없습니다.")
        raise ValueError("처리된 데이터가 없습니다.")

    final_df = pd.concat(all_processed_data, ignore_index=True)

    save_to_excel_with_highlight(final_df, output_excel_path)

    print(f"모든 페이지의 처리된 데이터가 {output_excel_path}에 저장되었습니다.")

@app.get("/", response_class=HTMLResponse, summary="홈페이지")
async def read_root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request, "message": ""})

@app.post("/upload/", summary="PDF 파일 업로드 및 처리")
async def upload_pdf(request: Request, file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.pdf'):
        return templates.TemplateResponse("index.html", {"request": request, "message": "PDF 파일만 업로드할 수 있습니다."})

    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_path = os.path.join(tmpdir, file.filename)
        
        # PDF 파일 저장
        with open(pdf_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        output_excel_path = os.path.join(tmpdir, "extracted_tables.xlsx")
        
        try:
            # PDF 처리
            process_pdf(pdf_path, output_excel_path)
        except Exception as e:
            error_message = f"PDF 처리 중 오류 발생: {e}"
            return templates.TemplateResponse("index.html", {"request": request, "message": error_message})
        
        # 엑셀 파일 반환
        return FileResponse(
            path=output_excel_path,
            filename="extracted_tables.xlsx",
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

@app.get("/docs", include_in_schema=False)
def custom_swagger_ui_html():
    return HTMLResponse(content=templates.get_template("index.html").render(message=""), status_code=200)
