
import os
import re
import numpy as np
import PyPDF2
import camelot
import fitz  # PyMuPDF
import pandas as pd
from sentence_transformers import SentenceTransformer
import faiss
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class TableEndDetector:
    def __init__(self):
        """
        표 끝 부분 검출을 위한 패턴 및 설정 초기화
        """
        # 찾고자 하는 패턴 정의
        self.patterns = {
            '상해관련': [
                r'상해관련\s*특약.*?(?:자세한\s*사항은\s*반드시\s*약관을\s*참고하시기\s*바랍니다)',
                r'상해\s*및\s*질병\s*관련\s*특약.*?(?:자세한\s*사항은\s*반드시\s*약관을\s*참고하시기\s*바랍니다)'
            ],
            '질병관련': [
                r'질병관련\s*특약.*?(?:자세한\s*사항은\s*반드시\s*약관을\s*참고하시기\s*바랍니다)',
                r'질병\s*및\s*상해\s*관련\s*특약.*?(?:자세한\s*사항은\s*반드시\s*약관을\s*참고하시기\s*바랍니다)'
            ]
        }
        
        # 다음 섹션 시작을 나타내는 패턴
        self.next_section_patterns = [
            r'보험료\s*납입면제\s*관련\s*특약',
            r'간병\s*관련\s*특약',
            r'실손의료비\s*보장\s*특약',
            r'기타\s*특약',
            r'제\s*\d+\s*장',
            r'보장내용\s*요약서',
            r'주요\s*보장내용'
        ]

        # 표 구분을 위한 패턴
        self.table_patterns = {
            'start': [
                r'구분',
                r'보장명',
                r'급여명',
                r'보험금의\s*지급사유',
                r'보험료\s*납입기간',
                r'보험기간'
            ],
            'end': [
                r'※\s*위\s*내용은\s*약관의\s*일부만을\s*요약한\s*것',
                r'※\s*기타\s*자세한\s*사항은\s*약관을\s*참고',
                r'상품내용\s*요약서'
            ]
        }

    def find_table_ends(self, texts_by_page: Dict[int, str]) -> Dict[str, List[Tuple[int, str, str]]]:
        """
        각 특약 유형별로 표 끝을 찾아서 반환
        """
        results = {
            '상해관련': [],
            '질병관련': []
        }
        
        found_disease = False  # 질병관련 특약을 찾았는지 표시
        disease_page = None    # 질병관련 특약이 발견된 페이지
        
        sorted_pages = sorted(texts_by_page.keys())
        
        for page_num in sorted_pages:
            text = texts_by_page[page_num]
            logger.debug(f"페이지 {page_num} 분석 중...")
            
            # 질병관련 특약을 찾은 후 다음 섹션이 시작되는지 확인
            if found_disease and disease_page != page_num:
                if any(re.search(pattern, text, re.IGNORECASE | re.MULTILINE) 
                      for pattern in self.next_section_patterns):
                    logger.info(f"페이지 {page_num}에서 다음 섹션 시작 발견. 검색 종료")
                    break

            for category, patterns in self.patterns.items():
                for pattern in patterns:
                    matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
                    for match in matches:
                        found_text = match.group(0).strip()
                        # 표 컨텍스트 확인
                        if self.verify_table_context(text, match.start(), match.end()):
                            context = self.get_context(text, match.start(), match.end(), window=300)
                            results[category].append((page_num, found_text, context))
                            logger.info(f"{category} 표 끝 발견 - 페이지 {page_num}")
                            
                            if category == '질병관련':
                                found_disease = True
                                disease_page = page_num
        
        return results

    def verify_table_context(self, text: str, start: int, end: int, window: int = 500) -> bool:
        """
        발견된 패턴이 실제로 표의 일부인지 확인
        """
        context = self.get_context(text, start, end, window)
        
        # 표 시작 패턴 확인
        has_start = any(re.search(pattern, context, re.IGNORECASE) 
                       for pattern in self.table_patterns['start'])
        
        # 표 끝 패턴 확인
        has_end = any(re.search(pattern, context, re.IGNORECASE) 
                     for pattern in self.table_patterns['end'])
        
        return has_start or has_end

    def get_context(self, text: str, start: int, end: int, window: int = 300) -> str:
        """
        매칭된 텍스트의 전후 컨텍스트를 추출
        """
        context_start = max(0, start - window)
        context_end = min(len(text), end + window)
        return text[context_start:context_end]

class PDFProcessor:
    def __init__(self, pdf_path: str):
        """
        PDF 처리를 위한 클래스 초기화
        """
        self.pdf_path = pdf_path
        self.texts_by_page = {}
        self.doc = None
        self.load_pdf()

    def load_pdf(self):
        """
        PDF 파일 로드 및 텍스트 추출
        """
        try:
            with open(self.pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                self.texts_by_page = {
                    i+1: page.extract_text() 
                    for i, page in enumerate(reader.pages)
                }
            self.doc = fitz.open(self.pdf_path)
            logger.info(f"PDF 로드 완료: 총 {len(self.texts_by_page)} 페이지")
        except Exception as e:
            logger.error(f"PDF 로드 중 오류 발생: {str(e)}")
            raise

    def extract_tables(self, page_numbers: List[int]) -> List[Dict[str, Any]]:
        """
        지정된 페이지에서 표 추출
        """
        tables = []
        try:
            for page_num in page_numbers:
                logger.info(f"페이지 {page_num}에서 표 추출 중...")
                camelot_tables = camelot.read_pdf(
                    self.pdf_path,
                    pages=str(page_num),
                    flavor='lattice'
                )
                
                for table in camelot_tables:
                    df = table.df
                    table_info = {
                        'dataframe': df,
                        'page': page_num,
                        'accuracy': table.parsing_report['accuracy'],
                        'whitespace': table.parsing_report['whitespace']
                    }
                    tables.append(table_info)
                    
        except Exception as e:
            logger.error(f"표 추출 중 오류 발생: {str(e)}")
        
        return tables

class ExcelExporter:
    def __init__(self, output_path: str):
        """
        엑셀 내보내기를 위한 클래스 초기화
        """
        self.output_path = output_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def create_sheet(self, name: str) -> None:
        """
        새로운 시트 생성
        """
        if name not in self.wb.sheetnames:
            self.wb.create_sheet(name)

    def add_table_to_sheet(self, sheet_name: str, table_info: Dict[str, Any], 
                          start_row: int = 1) -> int:
        """
        시트에 표 추가
        """
        ws = self.wb[sheet_name]
        current_row = start_row

        # 표 제목 및 메타데이터 추가
        title = f"페이지 {table_info['page']} - "
        title += f"정확도: {table_info.get('accuracy', 'N/A')}%"
        
        ws.cell(row=current_row, column=1, value=title)
        ws.merge_cells(
            start_row=current_row, 
            start_column=1, 
            end_row=current_row, 
            end_column=5
        )
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        # 데이터프레임 추가
        df = table_info['dataframe']
        for r in dataframe_to_rows(df, index=False, header=True):
            for c_idx, value in enumerate(r, start=1):
                cell = ws.cell(row=current_row, column=c_idx, value=value)
                cell.alignment = Alignment(wrap_text=True)
            current_row += 1

        return current_row + 1

    def save(self) -> None:
        """
        엑셀 파일 저장
        """
        try:
            self.wb.save(self.output_path)
            logger.info(f"엑셀 파일 저장 완료: {self.output_path}")
        except Exception as e:
            logger.error(f"엑셀 파일 저장 중 오류 발생: {str(e)}")
            raise

def main():
    try:
        # 파일 경로 설정
        uploads_folder = "uploads"
        output_folder = "output"
        os.makedirs(output_folder, exist_ok=True)

        pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
        if not pdf_files:
            logger.error("PDF 파일을 찾을 수 없습니다.")
            return

        pdf_file = pdf_files[0]
        pdf_path = os.path.join(uploads_folder, pdf_file)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_excel_path = os.path.join(
            output_folder, 
            f"{os.path.splitext(pdf_file)[0]}_{timestamp}_analysis.xlsx"
        )

        # PDF 처리
        pdf_processor = PDFProcessor(pdf_path)
        
        # 표 끝 검출
        detector = TableEndDetector()
        results = detector.find_table_ends(pdf_processor.texts_by_page)

        # 결과 정리
        found_pages = {
            '상해관련': [],
            '질병관련': []
        }

        for category, findings in results.items():
            if findings:
                found_pages[category].extend([page for page, _, _ in findings])
                print(f"\n{category} 특약 표 끝 위치:")
                for page_num, text, context in findings:
                    print(f"\n페이지 {page_num}:")
                    print("발견된 텍스트:", text)
                    print("\n주변 컨텍스트:")
                    print(context)
                    print("-" * 80)

        # 표 추출 및 엑셀 저장
        excel_exporter = ExcelExporter(output_excel_path)
        
        for category in ['상해관련', '질병관련']:
            if found_pages[category]:
                excel_exporter.create_sheet(category)
                tables = pdf_processor.extract_tables(found_pages[category])
                current_row = 1
                
                for table_info in tables:
                    current_row = excel_exporter.add_table_to_sheet(
                        category, 
                        table_info, 
                        current_row
                    )

        excel_exporter.save()
        logger.info("처리 완료")

    except Exception as e:
        logger.error(f"처리 중 오류 발생: {str(e)}")
        raise

if __name__ == "__main__":
    main()
