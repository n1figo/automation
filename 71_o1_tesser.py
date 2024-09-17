import fitz  # PyMuPDF (여전히 사용 중, PDF 정보에 접근)
import pytesseract
import cv2
from pytesseract import Output
from pdf2image import convert_from_path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import os
from typing import List, Tuple, Any  # List, Tuple, Any를 import

# PDF를 이미지로 변환
def convert_pdf_to_images(pdf_path: str):
    pages = convert_from_path(pdf_path)
    image_paths = []
    for i, page in enumerate(pages):
        image_path = f'page_{i+1}.png'
        page.save(image_path, 'PNG')
        image_paths.append(image_path)
    return image_paths

# 텍스트 및 배경색을 감지
def extract_text_and_background_from_image(image_path: str):
    image = cv2.imread(image_path)
    d = pytesseract.image_to_data(image, output_type=Output.DICT)

    extracted_data = []
    n_boxes = len(d['text'])
    for i in range(n_boxes):
        if int(d['conf'][i]) > 60:  # 신뢰도 60 이상만 처리
            (x, y, w, h) = (d['left'][i], d['top'][i], d['width'][i], d['height'][i])
            text = d['text'][i]

            # 텍스트 영역의 배경 색상 확인
            roi = image[y:y+h, x:x+w]
            background_color = cv2.mean(roi)[:3]
            background_color = tuple(map(int, background_color))  # BGR 값을 정수로 변환
                
            # 배경색이 특정 조건을 만족하는 경우 "추가"로 판단
            if not is_common_background_color(background_color):
                status = "추가"
            else:
                status = "유지"
            extracted_data.append((text, status, background_color))

    return extracted_data

# 일반적인 배경색 (흰색, 회색, 검정색)을 제외하는 함수
def is_common_background_color(color) -> bool:
    white = (255, 255, 255)
    gray = (127, 127, 127)
    black = (0, 0, 0)
    color_tolerance = 30  # 허용 오차

    def is_similar(c1, c2):
        return all(abs(c1[i] - c2[i]) <= color_tolerance for i in range(3))

    # 흰색, 회색, 검정색 배경은 일반적이므로 배경색으로 제외
    if is_similar(color, white) or is_similar(color, gray) or is_similar(color, black):
        return True
    return False

# 엑셀 작성 및 변경사항 적용 클래스
class ExcelWriterWithChanges:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Extracted Data"
        
    def write_data_with_changes(self, extracted_data):
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 추출된 데이터를 DataFrame으로 변환
        df = pd.DataFrame(extracted_data, columns=['텍스트', '상태', '배경색'])
        
        # "변경사항" 컬럼 추가
        df['변경사항'] = df['상태']
        
        # DataFrame을 엑셀로 작성
        for r in dataframe_to_rows(df, index=False, header=True):
            self.sheet.append(r)
            
        # 서식 적용
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, min_col=1, max_col=self.sheet.max_column):
            status_cell = row[df.columns.get_loc('변경사항')]
            if status_cell.value == "추가":
                for cell in row:
                    cell.fill = yellow_fill
            # 줄바꿈 설정
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        
        self.workbook.save(self.output_path)
        print(f"Data saved with changes to {self.output_path}")

# 메인 함수
def main(pdf_path: str, output_excel_path: str):
    # PDF를 이미지로 변환
    image_paths = convert_pdf_to_images(pdf_path)
    
    extracted_data = []
    
    # 각 이미지에서 텍스트와 배경색을 추출
    for image_path in image_paths:
        extracted_data.extend(extract_text_and_background_from_image(image_path))
    
    # 엑셀 파일로 작성
    writer = ExcelWriterWithChanges(output_excel_path)
    writer.write_data_with_changes(extracted_data)

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables.xlsx"
    main(pdf_path, output_excel_path)
