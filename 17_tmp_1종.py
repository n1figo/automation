import os
import re
import PyPDF2
import camelot
import fitz  # PyMuPDF
import pandas as pd
from sentence_transformers import SentenceTransformer
import numpy as np
import faiss
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_text_from_pdf(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        return {i+1: page.extract_text() for i, page in enumerate(reader.pages)}

def find_종_pages(texts_by_page, start_page=1, end_page=100):
    종_pages = {"[1종]": [], "[2종]": [], "[3종]": []}
    pattern = re.compile(r'\[(\d)종\]')
    
    for page_num, text in texts_by_page.items():
        if start_page <= page_num <= end_page:
            matches = pattern.findall(text)
            for match in matches:
                종 = f"[{match}종]"
                if 종 in 종_pages:
                    종_pages[종].append(page_num)
                    print(f"{종} 패턴을 {page_num}페이지에서 발견했습니다.")
    
    for 종, pages in 종_pages.items():
        if pages:
            print(f"{종}을(를) 다음 페이지에서 찾았습니다: {pages}")
        else:
            print(f"{종}을(를) 찾지 못했습니다.")
    
    return 종_pages

def create_embeddings(texts):
    model = SentenceTransformer('distiluse-base-multilingual-cased-v1')
    return model.encode(texts)

def search_rag(query, index, model, chunks, page_numbers, k=5, threshold=0.5):
    query_vector = model.encode([query])
    query_vector = query_vector / np.linalg.norm(query_vector, axis=1, keepdims=True)
    distances, indices = index.search(query_vector.astype('float32'), k)
    results = []
    for idx, score in zip(indices[0], distances[0]):
        if score >= threshold:
            results.append({'content': chunks[idx], 'score': score, 'page': page_numbers[idx]})
    return results

def extract_tables_from_pages(pdf_path, pages):
    all_tables = []
    for page in pages:
        try:
            tables = camelot.read_pdf(pdf_path, pages=str(page), flavor='lattice')
            for table in tables:
                all_tables.append({
                    'dataframe': table.df,
                    'page': page
                })
        except Exception as e:
            print(f"Error extracting tables from page {page}: {e}")
    return all_tables

def save_tables_to_excel(tables_dict, output_path, document_title=None):
    wb = Workbook()
    wb.remove(wb.active)
    
    for sheet_name, tables in tables_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        current_row = 1

        if document_title:
            ws.cell(row=current_row, column=1, value=document_title)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=1).font = Font(size=16, bold=True)
            current_row += 2

        for idx, table_info in enumerate(tables):
            df = table_info['dataframe']
            page = table_info['page']

            ws.cell(row=current_row, column=1, value=f"표 {idx+1} (페이지 {page})")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=df.shape[1])
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for r in dataframe_to_rows(df, index=False, header=True):
                for c_idx, value in enumerate(r, start=1):
                    ws.cell(row=current_row, column=c_idx, value=value)
                    ws.cell(row=current_row, column=c_idx).alignment = Alignment(wrap_text=True)
                current_row += 1
            current_row += 1

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    wb.save(output_path)
    print(f"Tables have been saved to {output_path}")

def main():
    uploads_folder = "uploads"
    output_folder = "output"
    os.makedirs(output_folder, exist_ok=True)

    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_analysis.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    texts_by_page = extract_text_from_pdf(pdf_path)
    
    print("\n--- 종 페이지 검색 결과 ---")
    종_pages = find_종_pages(texts_by_page)

    # RAG 설정
    full_text = " ".join(texts_by_page.values())
    chunks = [full_text[i:i+200] for i in range(0, len(full_text), 150)]
    page_numbers = []
    for page, text in texts_by_page.items():
        page_numbers.extend([page] * (len(text) // 150 + 1))
    
    model = SentenceTransformer('distiluse-base-multilingual-cased-v1')
    embeddings = create_embeddings(chunks)
    index = faiss.IndexFlatIP(embeddings.shape[1])
    index.add(embeddings.astype('float32'))

    tables_sheets = {
        "1종": [],
        "2종": [],
        "3종": []
    }

    for 종, pages in 종_pages.items():
        if pages:
            print(f"\n{종} 관련 표 검색 중...")
            # 해당 종이 있는 페이지와 그 다음 10페이지까지 검색
            search_pages = set(pages + [p for page in pages for p in range(page, min(page+11, max(texts_by_page.keys())))])
            tables = extract_tables_from_pages(pdf_path, search_pages)
            tables_sheets[종.strip('[]')] = tables
            print(f"{종}에서 {len(tables)}개의 표를 추출했습니다.")

    if any(tables_sheets.values()):
        save_tables_to_excel(tables_sheets, output_excel_path)
    else:
        print("추출된 표가 없습니다.")

    print(f"모든 처리된 표가 {output_excel_path}에 저장되었습니다.")

if __name__ == "__main__":
    main()