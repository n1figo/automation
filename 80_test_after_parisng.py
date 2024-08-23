import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import fitz  # PyMuPDF
import os
from PIL import Image
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
from io import StringIO

async def get_full_html(url, output_dir):
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page()
        await page.goto(url, wait_until="networkidle")
        
        await page.evaluate("""
            () => {
                const expandElements = (elements) => {
                    for (let elem of elements) {
                        if (window.getComputedStyle(elem).display === 'none') {
                            elem.style.display = 'block';
                        }
                        expandElements(elem.children);
                    }
                };
                expandElements(document.body.children);
            }
        """)
        
        html_content = await page.content()
        await browser.close()
    
    html_file_path = os.path.join(output_dir, "source.txt")
    with open(html_file_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"Full HTML source has been saved to {html_file_path}")
    return html_file_path

def extract_tables_from_html(html_file_path):
    try:
        with open(html_file_path, 'r', encoding='utf-8') as f:
            html_content = f.read()

        soup = BeautifulSoup(html_content, 'html.parser')
        tables = soup.find_all('table')
        
        if not tables:
            print("No tables found in the HTML content.")
            print(f"HTML content preview: {html_content[:1000]}...")
            return []
        
        dfs = []
        for i, table in enumerate(tables):
            try:
                df = pd.read_html(StringIO(str(table)))[0]
                dfs.append(df)
            except Exception as e:
                print(f"Failed to parse table {i+1}: {str(e)}")
                print(f"Table content: {table.prettify()[:500]}...")  # 테이블 내용의 일부를 출력

        if not dfs:
            print("No valid tables could be extracted.")
            return []

        print(f"Successfully extracted {len(dfs)} tables.")
        return dfs
    except Exception as e:
        print(f"Error extracting tables: {str(e)}")
        print("Detailed error information:")
        import traceback
        print(traceback.format_exc())
        return []

def save_original_tables_to_excel(dfs, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Original Tables"

    row = 1
    for i, df in enumerate(dfs, start=1):
        ws.cell(row=row, column=1, value=f'Table_{i}')
        row += 1

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [' '.join(col).strip() for col in df.columns.values]

        for col, header in enumerate(df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1

        for _, data_row in df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

        row += 2

    wb.save(output_excel_path)
    print(f"Original tables have been saved to {output_excel_path}")

def process_tables(dfs):
    all_data = []
    for i, df in enumerate(dfs):
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [' '.join(col).strip() for col in df.columns.values]
        df['table_info'] = f'Table_{i+1}'
        all_data.append(df)
    
    if not all_data:
        print("No data extracted.")
        return pd.DataFrame()
    
    result = pd.concat(all_data, axis=0, ignore_index=True)
    
    print("Final DataFrame:")
    print(f"Columns: {result.columns.tolist()}")
    print(f"Shape: {result.shape}")
    
    return result

def is_color_highlighted(color):
    r, g, b = color
    if r == g == b:
        return False
    return max(r, g, b) > 200 and (max(r, g, b) - min(r, g, b)) > 30

def detect_highlights_and_colors(image):
    width, height = image.size
    img_array = np.array(image)
    
    highlighted_rows = set()
    colored_rows = set()
    for y in range(height):
        for x in range(width):
            color = img_array[y, x]
            if color[0] > 200 and color[1] > 200 and color[2] < 100:  # Yellow highlight
                highlighted_rows.add(y)
            elif is_color_highlighted(color):
                colored_rows.add(y)
    
    all_special_rows = sorted(highlighted_rows.union(colored_rows))
    if all_special_rows:
        return all_special_rows, highlighted_rows, colored_rows
    return [], set(), set()

def extract_highlighted_text_with_context(pdf_path, output_dir, max_pages=20):
    print("Starting extraction of highlighted and colored text from PDF...")
    doc = fitz.open(pdf_path)
    total_pages = min(len(doc), max_pages)
    texts_with_context = []
    
    output_image_dir = os.path.join(output_dir, "images")
    os.makedirs(output_image_dir, exist_ok=True)
    
    for page_num in range(total_pages):
        print(f"Processing page {page_num + 1}/{total_pages}")
        page = doc[page_num]
        
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        all_special_rows, highlighted_rows, colored_rows = detect_highlights_and_colors(img)
        
        if all_special_rows:
            # 전체 페이지 캡처
            image_filename = f"page_{page_num + 1}_full.png"
            image_path = os.path.join(output_image_dir, image_filename)
            img.save(image_path)
            
            text = page.get_text("text")
            status = '수정' if highlighted_rows and colored_rows else \
                     '삭제' if highlighted_rows else \
                     '추가' if colored_rows else '유지'
            texts_with_context.append((text, page_num + 1, image_path, status, all_special_rows))

    print(f"Finished extracting highlighted and colored text from PDF (total {total_pages} pages)")
    return texts_with_context

def compare_dataframes(df_before, texts_with_context, output_dir):
    print("Starting comparison of dataframes...")
    df_result = df_before.copy()
    df_result['PDF_페이지'] = ''
    df_result['이미지_경로'] = ''
    df_result['상태'] = '유지'

    for text, page_num, image_path, status, special_rows in texts_with_context:
        text_lines = text.split('\n')
        for i in range(len(df_result)):
            if any(str(cell).strip() in text for cell in df_result.iloc[i]):
                table_start = i
                while table_start > 0 and df_result.loc[table_start-1, 'table_info'] == df_result.loc[i, 'table_info']:
                    table_start -= 1
                table_end = i
                while table_end < len(df_result)-1 and df_result.loc[table_end+1, 'table_info'] == df_result.loc[i, 'table_info']:
                    table_end += 1
                
                df_result.loc[table_start:table_end, 'PDF_페이지'] = page_num
                df_result.loc[table_start:table_end, '이미지_경로'] = image_path
                
                # 강조된 행에 대해서만 상태 업데이트
                for j in range(table_start, table_end + 1):
                    if any(line.strip() in str(cell) for cell in df_result.iloc[j] for line in text_lines if line.strip()):
                        df_result.loc[j, '상태'] = status
                
                # 강조된 표를 별도의 xlsx 파일로 저장
                table_df = df_result.loc[table_start:table_end].copy()
                table_filename = f"table_page_{page_num}.xlsx"
                table_path = os.path.join(output_dir, table_filename)
                table_df.to_excel(table_path, index=False)
                print(f"Saved highlighted table to {table_path}")
                
                break

    # 상태 컬럼을 맨 오른쪽으로 이동
    cols = df_result.columns.tolist()
    cols.append(cols.pop(cols.index('상태')))
    df_result = df_result[cols]

    print(f"Finished comparison. Updated {len(df_result[df_result['상태'] != '유지'])} rows")
    return df_result

def save_to_excel(df, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    row = 1
    for table_info in df['table_info'].unique():
        table_df = df[df['table_info'] == table_info]
        
        ws.cell(row=row, column=1, value=table_info)
        row += 1
        
        for col, header in enumerate(table_df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1
        
        for _, data_row in table_df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1
        
        row += 2

    wb.save(output_excel_path)
    print(f"Results have been saved to {output_excel_path}")

async def main():
    print("Program start")
    try:
        url = "https://www.kbinsure.co.kr/CG302290001.ec#"
        pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
        output_dir = "/workspaces/automation/output"
        os.makedirs(output_dir, exist_ok=True)
        output_excel_path = os.path.join(output_dir, "comparison_results.xlsx")
        original_excel_path = os.path.join(output_dir, "변경전.xlsx")

        html_file_path = await get_full_html(url, output_dir)
        dfs = extract_tables_from_html(html_file_path)
        if not dfs:
            print("Failed to extract tables. Please check the HTML file.")
            return

        save_original_tables_to_excel(dfs, original_excel_path)

        df_before = process_tables(dfs)
        if df_before.empty:
            print("No data processed.")
            return

        print("Combined DataFrame:")
        print(df_before.head())
        print(f"Shape of combined DataFrame: {df_before.shape}")

        texts_with_context = extract_highlighted_text_with_context(pdf_path, output_dir, max_pages=20)

        if not df_before.empty and texts_with_context:
            df_matching = compare_dataframes(df_before, texts_with_context, output_dir)
            save_to_excel(df_matching, output_excel_path)
        else:
            print("Failed to extract tables or highlighted text. Please check the URL and PDF.")

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print("Detailed error information:")
        import traceback
        print(traceback.format_exc())
    
    print("Program end")

if __name__ == "__main__":
    asyncio.run(main())