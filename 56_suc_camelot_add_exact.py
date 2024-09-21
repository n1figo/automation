import camelot
import pandas as pd
import numpy as np
import cv2
import os
import fitz
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DEBUG_MODE = True
IMAGE_OUTPUT_DIR = "/workspaces/automation/output/images"
os.makedirs(IMAGE_OUTPUT_DIR, exist_ok=True)

def pdf_to_image(page):
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return np.array(img)

def detect_highlights(image, page_num):
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    s = hsv[:,:,1]
    v = hsv[:,:,2]

    saturation_threshold = 30
    saturation_mask = s > saturation_threshold

    _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)

    kernel = np.ones((5,5), np.uint8)
    cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)

    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_num}_mask.png'), cleaned_mask)

    contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    contour_image = image.copy()
    cv2.drawContours(contour_image, contours, -1, (0, 255, 0), 2)
    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_num}_contours.png'), cv2.cvtColor(contour_image, cv2.COLOR_RGB2BGR))

    return contours

def get_highlight_regions(contours, image_height):
    regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        # OpenCV 좌표계를 PDF 좌표계로 변환
        top = image_height - (y + h)
        bottom = image_height - y
        regions.append((top, bottom))
    return regions

def extract_tables_with_camelot(pdf_path, page_number):
    print(f"Extracting tables from page {page_number} using Camelot...")
    tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
    print(f"Found {len(tables)} tables on page {page_number}")
    return tables

def process_tables(tables, highlight_regions, page_height):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        x1, y1, x2, y2 = table._bbox  # 올바른 좌표 언패킹

        # PDF 좌표계에서 y1은 하단, y2는 상단
        table_height = y2 - y1
        row_height = table_height / len(df)

        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            
            # 행의 상단과 하단 y 좌표 계산 (PDF 좌표계 사용)
            # 상단부터 계산하기 위해 y2에서부터 감소
            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height
            
            row_highlighted = check_highlight((row_top, row_bottom), highlight_regions)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)

    return pd.DataFrame(processed_data)

def check_highlight(row_range, highlight_regions):
    row_top, row_bottom = row_range
    for region_top, region_bottom in highlight_regions:
        # 행과 강조 영역이 겹치는지 확인
        if (region_top <= row_top <= region_bottom) or (region_top <= row_bottom <= region_bottom) or \
           (row_top <= region_top <= row_bottom) or (row_top <= region_bottom <= row_bottom):
            return True
    return False

def save_to_excel_with_highlight(df, output_path):
    df.to_excel(output_path, index=False)
    
    wb = load_workbook(output_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    if '변경사항' in df.columns:
        change_col_index = df.columns.get_loc('변경사항') + 1
    else:
        raise ValueError("DataFrame에 '변경사항' 컬럼이 없습니다.")

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=change_col_index).value
        if cell_value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows")

def main(pdf_path, output_excel_path):
    print("PDF에서 개정된 부분을 추출합니다...")

    doc = fitz.open(pdf_path)
    page_number = 50  # 51페이지 (0-based index)

    page = doc[page_number]
    image = pdf_to_image(page)

    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number + 1}_original.png'), cv2.cvtColor(image, cv2.COLOR_RGB2BGR))

    contours = detect_highlights(image, page_number + 1)
    highlight_regions = get_highlight_regions(contours, image.shape[0])

    highlighted_image = image.copy()
    for top, bottom in highlight_regions:
        # PDF 좌표계를 OpenCV 좌표계로 변환하여 그리기
        cv2.rectangle(highlighted_image, (0, image.shape[0] - bottom), (image.shape[1], image.shape[0] - top), (0, 255, 0), 2)
    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number + 1}_highlighted.png'), cv2.cvtColor(highlighted_image, cv2.COLOR_RGB2BGR))

    print(f"감지된 강조 영역 수: {len(highlight_regions)}")
    print(f"강조 영역: {highlight_regions}")

    # Camelot을 사용하여 표 추출
    tables = extract_tables_with_camelot(pdf_path, page_number + 1)

    if not tables:
        print("추출된 표가 없습니다.")
        return

    # 추출된 표 처리
    processed_df = process_tables(tables, highlight_regions, image.shape[0])

    print(processed_df)

    save_to_excel_with_highlight(processed_df, output_excel_path)

    print(f"처리된 데이터가 {output_excel_path}에 저장되었습니다.")

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables_camelot.xlsx"
    main(pdf_path, output_excel_path)
