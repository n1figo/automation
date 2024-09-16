import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import StringIO
import fitz  # PyMuPDF

def is_black_or_white_or_gray(color):
    r, g, b = color
    is_black_or_dark = max(r, g, b) < 50  # 약간 어두운 색상도 제외
    is_white = r > 240 and g > 240 and b > 240  # 흰색에 가까운 색상
    is_gray = abs(r - g) < 10 and abs(g - b) < 10 and abs(r - b) < 10  # 회색 (RGB 값의 차이가 10 이하)
    return is_black_or_dark or is_white or is_gray

def extract_colored_texts(pdf_path):
    print("Extracting colored texts from PDF...")
    colored_texts = set()
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    for page_num, page in enumerate(doc):
        print(f"Processing page {page_num + 1}/{total_pages}")
        blocks = page.get_text("dict")["blocks"]
        for block in blocks:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        text = span["text"].strip()
                        if text:
                            color = span["color"]
                            # Convert color integer to RGB tuple
                            r = int((color >> 16) & 255)
                            g = int((color >> 8) & 255)
                            b = int(color & 255)
                            if not is_black_or_white_or_gray((r, g, b)):
                                colored_texts.add(text)
    print(f"Extracted {len(colored_texts)} colored text elements from PDF.")
    return colored_texts

def mark_changes_in_dataframe(df_before, colored_texts):
    print("Marking changes in DataFrame...")
    df_before = df_before.copy()
    df_before['상태'] = '유지'  # Default value is '유지'
    for index, row in df_before.iterrows():
        row_changed = False
        for col in df_before.columns:
            if col != 'table_info' and col != '상태':
                cell_value = str(row[col]).strip()
                if cell_value in colored_texts:
                    row_changed = True
                    break
        if row_changed:
            df_before.at[index, '상태'] = '추가'
    print(f"Marked {len(df_before[df_before['상태'] == '추가'])} rows as '추가'")
    return df_before

async def get_full_html_and_tables(url, output_dir):
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.goto(url, wait_until="networkidle")
            
            # '보장내용' 탭 클릭
            await page.click('text="보장내용"')
            await page.wait_for_load_state('networkidle')
            
            # JavaScript 실행 후 페이지 내용 가져오기
            html_content = await page.content()
            
            # '선택약관' 이후의 테이블 데이터 추출
            tables = await page.evaluate('''
                () => {
                    const tables = [];
                    let selectionElement = document.evaluate(
                        '//text()[contains(., "선택약관")]/ancestor::*[1]',
                        document,
                        null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE,
                        null
                    ).singleNodeValue;
                    if (selectionElement) {
                        let elem = selectionElement.nextElementSibling;
                        while (elem) {
                            if (elem.tagName.toLowerCase() === 'table') {
                                const rows = elem.querySelectorAll('tr');
                                const tableData = Array.from(rows).map(row => {
                                    const cells = row.querySelectorAll('th, td');
                                    return Array.from(cells).map(cell => cell.innerText.trim());
                                });
                                tables.push(tableData);
                            }
                            elem = elem.nextElementSibling;
                        }
                    }
                    return tables;
                }
            ''')
            
            await browser.close()
        
        html_file_path = os.path.join(output_dir, "source.txt")
        with open(html_file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"Full HTML source and table data have been saved to {output_dir}")
        return html_file_path, tables
    except Exception as e:
        print(f"Error in get_full_html_and_tables: {str(e)}")
        return None, []

def process_tables(tables):
    all_data = []
    for i, table in enumerate(tables):
        if len(table) > 1:
            df = pd.DataFrame(table[1:], columns=table[0])
        else:
            df = pd.DataFrame(table)
        df['table_info'] = f'Table_{i+1}'
        all_data.append(df)
    
    if not all_data:
        print("No data extracted.")
        return pd.DataFrame()
    
    result = pd.concat(all_data, axis=0, ignore_index=True)
    
    print("Final DataFrame:")
    print(f"Columns: {result.columns.tolist()}")
    print(f"Shape: {result.shape}")
    
    return result

def save_original_tables_to_excel(dfs, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Original Tables"

    row = 1
    for i, df in enumerate(dfs, start=1):
        ws.cell(row=row, column=1, value=f'Table_{i}')
        row += 1

        for col, header in enumerate(df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1

        for _, data_row in df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

        row += 2

    wb.save(output_excel_path)
    print(f"Original tables have been saved to {output_excel_path}")

def save_to_excel(df, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    row = 1
    if 'table_info' in df.columns:
        for table_info in df['table_info'].unique():
            table_df = df[df['table_info'] == table_info]

            ws.cell(row=row, column=1, value=table_info)
            row += 1

            for col, header in enumerate(table_df.columns, start=1):
                ws.cell(row=row, column=col, value=header)
            row += 1

            for _, data_row in table_df.iterrows():
                for col, value in enumerate(data_row, start=1):
                    ws.cell(row=row, column=col, value=value)
                row += 1

            row += 2
    else:
        # 'table_info' 열이 없는 경우 전체 DataFrame을 그대로 저장
        for col, header in enumerate(df.columns, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1

        for _, data_row in df.iterrows():
            for col, value in enumerate(data_row, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    wb.save(output_excel_path)
    print(f"Results have been saved to {output_excel_path}")

async def main():
    print("Program start")
    try:
        url = "https://www.kbinsure.co.kr/CG302290001.ec#"
        pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
        output_dir = "/workspaces/automation/output"
        os.makedirs(output_dir, exist_ok=True)
        output_excel_path = os.path.join(output_dir, "comparison_results.xlsx")
        original_excel_path = os.path.join(output_dir, "변경전.xlsx")

        html_file_path, tables = await get_full_html_and_tables(url, output_dir)
        
        df_before = pd.DataFrame()
        if tables:
            df_before = process_tables(tables)
            save_original_tables_to_excel([df_before], original_excel_path)
            print("Combined DataFrame:")
            print(df_before.head())
            print(f"Shape of combined DataFrame: {df_before.shape}")
        else:
            print("No tables extracted from the website. Proceeding with an empty DataFrame.")

        colored_texts = extract_colored_texts(pdf_path)

        if not df_before.empty and colored_texts:
            df_matching = mark_changes_in_dataframe(df_before, colored_texts)
            print("Columns in df_matching:", df_matching.columns)
            save_to_excel(df_matching, output_excel_path)
        else:
            print("Failed to extract data from website or colored text from PDF. Please check the inputs.")

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print("Detailed error information:")
        import traceback
        print(traceback.format_exc())
    
    print("Program end")

if __name__ == "__main__":
    asyncio.run(main())
