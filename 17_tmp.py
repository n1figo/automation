import os
import re
import PyPDF2
import camelot
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_text_with_positions(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        texts_by_page = {}
        for page_num, page in enumerate(reader.pages, start=1):
            text = page.extract_text()
            if text:
                texts_by_page[page_num] = text
    return texts_by_page

def find_종_pages(texts_by_page, start_page=1, end_page=100):
    종_pages = {"1종": [], "2종": [], "3종": []}
    pattern = re.compile(r'\[(\d)종\]')
    
    for page_num, text in texts_by_page.items():
        if start_page <= page_num <= end_page:
            matches = pattern.findall(text)
            for match in matches:
                종 = f"{match}종"
                if 종 in 종_pages:
                    종_pages[종].append(page_num)
    
    return 종_pages

def validate_종_info(texts_by_page, 종_pages):
    validated_pages = {"1종": [], "2종": [], "3종": []}
    for 종, pages in 종_pages.items():
        for page in pages:
            text = texts_by_page.get(page, "")
            if re.search(rf'\[{종}\].*가입기준', text, re.DOTALL):
                validated_pages[종].append(page)
    return validated_pages

def extract_종_info(texts_by_page, validated_pages):
    종_info = {}
    for 종, pages in validated_pages.items():
        for page in pages:
            text = texts_by_page.get(page, "")
            match = re.search(rf'\[{종}\](.*?)가입기준.*?:(.+)', text, re.DOTALL)
            if match:
                종_type = match.group(1).strip()
                가입기준 = match.group(2).strip()
                종_info[종] = {"type": 종_type, "가입기준": 가입기준, "page": page}
    return 종_info

def detect_table_boundaries(texts_by_page):
    tables = []
    for page_num, text in texts_by_page.items():
        if re.search(r'표\s*\d+|선택\s*특약|상해관련\s*특약|질병관련\s*특약|\[\d\s*종\]', text):
            tables.append({'pages': [page_num], 'text': text})
    return tables

def extract_text_above_bbox(page, bbox):
    x0, y0, x1, y1 = bbox
    text_blocks = page.get_text("blocks")
    texts_above = []
    for block in text_blocks:
        if len(block) >= 5:
            bx0, by0, bx1, by1, text = block[:5]
            if by1 <= y0:
                texts_above.append((by1, text))
    if texts_above:
        texts_above.sort(reverse=True)
        return texts_above[0][1].strip()
    else:
        return "제목 없음"

def extract_tables_with_camelot(pdf_path, tables_info):
    all_tables = []
    doc = fitz.open(pdf_path)
    for table_info in tables_info:
        pages = table_info['pages']
        pages_str = ','.join(map(str, pages))
        print(f"Extracting table from pages {pages_str} using Camelot...")
        try:
            tables = camelot.read_pdf(pdf_path, pages=pages_str, flavor='lattice')
        except Exception as e:
            print(f"Error extracting tables from pages {pages_str}: {e}")
            continue

        if not tables:
            continue

        for table in tables:
            df = table.df
            bbox = table._bbox
            page_num = table.page - 1
            page = doc.load_page(page_num)
            text_above_table = extract_text_above_bbox(page, bbox)
            title = text_above_table.strip()

            all_tables.append({
                'dataframe': df,
                'title': title,
                'page': table.page
            })
    print(f"Found {len(all_tables)} tables in total")
    return all_tables

def process_tables(all_tables):
    processed_data = []
    for i, table_info in enumerate(all_tables):
        df = table_info['dataframe']
        title = table_info['title']
        page = table_info['page']
        df['Table_Number'] = i + 1
        df['Table_Title'] = title
        df['Page'] = page
        processed_data.append(df)
    if processed_data:
        return pd.concat(processed_data, ignore_index=True)
    else:
        return pd.DataFrame()

def save_tables_to_excel(tables_dict, output_path, document_title=None):
    wb = Workbook()
    wb.remove(wb.active)
    
    for sheet_name, tables in tables_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        current_row = 1

        if document_title:
            ws.cell(row=current_row, column=1, value=document_title)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=1).font = Font(size=16, bold=True)
            current_row += 2

        for idx, table_info in enumerate(tables):
            if 'dataframe' not in table_info:
                print(f"Skipping table in sheet '{sheet_name}' as 'dataframe' key is missing.")
                continue

            df = table_info['dataframe']
            title = table_info['title']
            page = table_info['page']

            ws.cell(row=current_row, column=1, value=f"표 {idx+1}: {title} (페이지 {page})")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=df.shape[1])
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for r in dataframe_to_rows(df, index=False, header=True):
                for c_idx, value in enumerate(r, start=1):
                    ws.cell(row=current_row, column=c_idx, value=value)
                    ws.cell(row=current_row, column=c_idx).alignment = Alignment(wrap_text=True)
                current_row += 1
            current_row += 1

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    wb.save(output_path)
    print(f"Tables have been saved to {output_path}")

def main():
    uploads_folder = "uploads"
    output_folder = "output"
    os.makedirs(output_folder, exist_ok=True)

    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_analysis.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    texts_by_page = extract_text_with_positions(pdf_path)
    종_pages = find_종_pages(texts_by_page, start_page=1, end_page=100)
    validated_pages = validate_종_info(texts_by_page, 종_pages)
    종_info = extract_종_info(texts_by_page, validated_pages)

    for 종, info in 종_info.items():
        print(f"{종} 정보:")
        print(f"  페이지: {info['page']}")
        print(f"  유형: {info['type']}")
        print(f"  가입기준: {info['가입기준']}")
        print()

    tables_info = detect_table_boundaries(texts_by_page)
    tables = extract_tables_with_camelot(pdf_path, tables_info)

    tables_sheets = {
        "1종": [],
        "2종": [],
        "3종": [],
        "기타": []
    }

    for table in tables:
        title = table['title']
        if "1종" in title:
            tables_sheets["1종"].append(table)
        elif "2종" in title:
            tables_sheets["2종"].append(table)
        elif "3종" in title:
            tables_sheets["3종"].append(table)
        else:
            tables_sheets["기타"].append(table)

    # 첫 번째 페이지에서 문서 제목 추출
    doc = fitz.open(pdf_path)
    first_page = doc.load_page(0)
    first_page_text = first_page.get_text()
    document_title = first_page_text.strip().split('\n')[0]

    if any(tables_sheets.values()):
        save_tables_to_excel(tables_sheets, output_excel_path, document_title=document_title)
    else:
        print("추출된 표가 없습니다.")

    print(f"모든 처리된 표가 {output_excel_path}에 저장되었습니다.")

if __name__ == "__main__":
    main()