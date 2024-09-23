import camelot
import pandas as pd
import numpy as np
import cv2
import os
import fitz  # PyMuPDF
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz
import pytesseract  # Tesseract OCR 라이브러리
import re

# Tesseract 실행 파일 경로 설정 (필요 시)
# 예: pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# 디버그 모드 설정
DEBUG_MODE = True
IMAGE_OUTPUT_DIR = "/workspaces/automation/output/images"
TXT_OUTPUT_DIR = "/workspaces/automation/output/texts"
os.makedirs(IMAGE_OUTPUT_DIR, exist_ok=True)
os.makedirs(TXT_OUTPUT_DIR, exist_ok=True)

def resize_image(image, scale_factor=2):
    """
    이미지의 해상도를 증가시키는 함수.

    Args:
        image (numpy.ndarray): 원본 이미지.
        scale_factor (float): 이미지 크기 증가 배율.

    Returns:
        numpy.ndarray: 크기가 증가된 이미지.
    """
    width = int(image.shape[1] * scale_factor)
    height = int(image.shape[0] * scale_factor)
    dim = (width, height)
    resized = cv2.resize(image, dim, interpolation=cv2.INTER_LINEAR)
    return resized

def sharpen_image(image):
    """
    이미지 샤프닝을 적용하는 함수.

    Args:
        image (numpy.ndarray): 입력 이미지 (그레이스케일).

    Returns:
        numpy.ndarray: 샤프닝이 적용된 이미지.
    """
    # 샤프닝 커널 정의
    kernel = np.array([[0, -1, 0],
                       [-1, 5, -1],
                       [0, -1, 0]])
    sharpened = cv2.filter2D(image, -1, kernel)
    return sharpened

def post_process_ocr(text):
    """
    OCR 결과 후처리 함수. 자주 발생하는 오류를 수정합니다.
    """
    # 예시: 잘못 인식된 문자를 수정
    text = re.sub(r'\|', ' ', text)  # | 문자를 공백으로 대체
    text = re.sub(r'\[유사망', '유사망', text)  # 특정 패턴 수정
    # 추가적인 교정 로직을 여기에 추가
    return text

def pdf_to_image(page):
    """
    PDF 페이지를 이미지로 변환하는 함수.

    Args:
        page (fitz.Page): PDF 페이지 객체.

    Returns:
        numpy.ndarray: 변환된 이미지 배열.
    """
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return np.array(img)

def detect_highlights(image, page_num):
    """
    강조색(하이라이트) 영역을 감지하고 컨투어를 반환하는 함수
    """
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    s = hsv[:,:,1]
    v = hsv[:,:,2]

    saturation_threshold = 30
    saturation_mask = s > saturation_threshold

    _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)

    kernel = np.ones((5,5), np.uint8)
    cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)

    # 디버그 이미지 저장
    if DEBUG_MODE:
        cleaned_mask_path = os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_num}_cleaned_mask.png')
        cv2.imwrite(cleaned_mask_path, cleaned_mask)
        print(f"Debug: Saved cleaned_mask as '{cleaned_mask_path}'")

    contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    contour_image = image.copy()
    cv2.drawContours(contour_image, contours, -1, (0, 255, 0), 2)
    contours_image_path = os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_num}_contours.png')
    cv2.imwrite(contours_image_path, cv2.cvtColor(contour_image, cv2.COLOR_RGB2BGR))
    if DEBUG_MODE:
        print(f"Debug: Saved contours image as '{contours_image_path}'")

    return contours

def get_highlight_regions(contours, image_height):
    """
    컨투어를 기반으로 강조된 영역의 좌표를 반환하는 함수
    """
    regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        # OpenCV 좌표계를 PDF 좌표계로 변환
        top = image_height - (y + h)
        bottom = image_height - y
        regions.append((top, bottom, x, x + w))  # x 좌표도 포함
    return regions

def convert_highlights_to_pdf(image, highlight_regions, output_pdf_path):
    """
    강조색 영역을 PDF로 변환하는 함수
    """
    highlighted_image = image.copy()
    for top, bottom, left, right in highlight_regions:
        cv2.rectangle(highlighted_image, (left, top), (right, bottom), (0, 255, 0), 2)
    
    # PIL 이미지로 변환
    pil_image = Image.fromarray(cv2.cvtColor(highlighted_image, cv2.COLOR_BGR2RGB))
    
    # PDF로 저장
    pil_image.save(output_pdf_path, "PDF", resolution=100.0)
    print(f"강조 영역이 포함된 PDF를 저장했습니다: {output_pdf_path}")

def extract_tables_with_camelot(pdf_path, page_number, output_txt_path):
    """
    Camelot을 사용하여 PDF의 특정 페이지에서 테이블을 추출하고 텍스트를 저장하는 함수
    """
    print(f"Extracting tables from page {page_number} using Camelot...")
    try:
        tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
        print(f"Found {len(tables)} tables on page {page_number}")
        
        # 텍스트 추출 및 저장
        with open(output_txt_path, 'w', encoding='utf-8') as txt_file:
            for i, table in enumerate(tables):
                txt_file.write(f"Table {i+1}:\n")
                txt_file.write(table.df.to_string(index=False))
                txt_file.write('\n\n')
        
        print(f"텍스트가 {output_txt_path}에 저장되었습니다.")
        return tables
    except Exception as e:
        print(f"Camelot을 사용하여 테이블을 추출하는 중 오류 발생: {e}")
        return []

def ocr_image_tesseract_enhanced(image_array, debug_mode=False):
    """
    Enhanced Tesseract OCR with additional preprocessing steps.
    """
    try:
        # 1. 그레이스케일 변환
        gray = cv2.cvtColor(image_array, cv2.COLOR_BGR2GRAY)
        if debug_mode:
            grayscale_path = os.path.join(IMAGE_OUTPUT_DIR, 'debug_grayscale.png')
            cv2.imwrite(grayscale_path, gray)
            print(f"그레이스케일 변환 완료: {grayscale_path}")

        # 2. 이미지 해상도 향상 (스케일 팩터 2배)
        resized = resize_image(gray, scale_factor=2)
        if debug_mode:
            resized_path = os.path.join(IMAGE_OUTPUT_DIR, 'debug_resized.png')
            cv2.imwrite(resized_path, resized)
            print(f"이미지 해상도 향상 완료: {resized_path}")

        # 3. 노이즈 제거 (Median Blur)
        denoised = cv2.medianBlur(resized, 3)
        if debug_mode:
            denoised_path = os.path.join(IMAGE_OUTPUT_DIR, 'debug_denoised.png')
            cv2.imwrite(denoised_path, denoised)
            print(f"노이즈 제거 완료: {denoised_path}")

        # 4. 이미지 추가 확대 (스케일 팩터 1.5배) - 총 3배 확대
        further_resized = resize_image(denoised, scale_factor=1.5)
        if debug_mode:
            further_resized_path = os.path.join(IMAGE_OUTPUT_DIR, 'debug_further_resized.png')
            cv2.imwrite(further_resized_path, further_resized)
            print(f"이미지 추가 확대 완료: {further_resized_path}")

        # 5. 샤프닝 적용하여 텍스트 선명하게
        sharpened = sharpen_image(further_resized)
        if debug_mode:
            sharpened_path = os.path.join(IMAGE_OUTPUT_DIR, 'debug_sharpened.png')
            cv2.imwrite(sharpened_path, sharpened)
            print(f"샤프닝 적용 완료: {sharpened_path}")

        # 6. 대비 향상 (CLAHE 적용)
        clahe = cv2.createCLAHE(clipLimit=1.5, tileGridSize=(8,8))
        enhanced = clahe.apply(sharpened)
        if debug_mode:
            clahe_path = os.path.join(IMAGE_OUTPUT_DIR, 'debug_clahe.png')
            cv2.imwrite(clahe_path, enhanced)
            print(f"대비 향상 완료: {clahe_path}")

        # 7. 팽창(Dilation) 적용하여 텍스트 강조
        kernel = np.ones((1,1), np.uint8)
        dilated = cv2.dilate(enhanced, kernel, iterations=1)
        if debug_mode:
            dilate_path = os.path.join(IMAGE_OUTPUT_DIR, 'debug_dilate.png')
            cv2.imwrite(dilate_path, dilated)
            print(f"팽창(Dilation) 적용 완료: {dilate_path}")

        # PIL 이미지로 변환
        pil_image = Image.fromarray(dilated)

        # Tesseract OCR 수행 (한글 언어 설정: 'kor') 및 --oem 3, --psm 6 설정
        custom_config = r'--oem 3 --psm 6'
        ocr_text = pytesseract.image_to_string(pil_image, lang='kor', config=custom_config)

        return ocr_text.strip()

    except Exception as e:
        print(f"OCR 처리 중 오류 발생: {e}")
        return ""

def extract_highlighted_text_tesseract(pdf_path, page_number, highlight_regions, debug_mode=False):
    """
    강조된 영역에서 텍스트를 추출하는 함수 (Tesseract 사용)
    """
    try:
        doc = fitz.open(pdf_path)
        page = doc.load_page(page_number - 1)  # 0-based index
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        extracted_texts = []

        txt_file_path = os.path.join(TXT_OUTPUT_DIR, f'page_{page_number}_highlighted_texts.txt')
        with open(txt_file_path, 'w', encoding='utf-8') as txt_file:
            for idx, (top, bottom, left, right) in enumerate(highlight_regions):
                try:
                    # 이미지 좌표계로 변환
                    img_height, img_width, _ = np.array(img).shape
                    y0 = img_height - bottom
                    y1 = img_height - top
                    x0 = left
                    x1 = right

                    # 영역 크롭
                    cropped_img = img.crop((x0, y0, x1, y1))
                    cropped_img_path = os.path.join(IMAGE_OUTPUT_DIR, f'highlight_{page_number}_{idx}.png')
                    cropped_img.save(cropped_img_path)

                    if debug_mode:
                        print(f"Saved cropped image: {cropped_img_path}")

                    # OCR 수행 (Enhanced Tesseract 사용)
                    ocr_text = ocr_image_tesseract_enhanced(np.array(cropped_img), debug_mode=debug_mode)

                    if debug_mode:
                        print(f"OCR Text from region {idx}: {ocr_text}")

                    extracted_texts.append(ocr_text)

                    # txt 파일에 저장
                    txt_file.write(f"Region {idx}: {ocr_text}\n")
                except Exception as e:
                    print(f"Exception occurred while processing region {idx}: {e}")
                    extracted_texts.append('')
                    txt_file.write(f"Region {idx}: \n")
                    continue

        print(f"Extracted highlighted texts have been saved to '{txt_file_path}'")
        return extracted_texts

    except Exception as e:
        print(f"텍스트 추출 중 오류 발생: {e}")
        return []

def process_tables(tables, highlight_regions, page_height, debug_mode=False):
    """
    추출된 테이블을 처리하여 데이터프레임으로 반환하는 함수
    """
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        x1, y1, x2, y2 = table._bbox  # 올바른 좌표 언패킹

        # PDF 좌표계에서 y1은 하단, y2는 상단
        table_height = y2 - y1
        row_height = table_height / len(df)

        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()

            # 행의 상단과 하단 y 좌표 계산 (PDF 좌표계 사용)
            # 상단부터 계산하기 위해 y2에서부터 감소
            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height

            row_highlighted = check_highlight((row_top, row_bottom, 0, 0), highlight_regions)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)

    return pd.DataFrame(processed_data)

def check_highlight(row_range, highlight_regions):
    """
    특정 행이 강조된 영역과 겹치는지 확인하는 함수
    """
    row_top, row_bottom, _, _ = row_range
    for region_top, region_bottom, _, _ in highlight_regions:
        # 행과 강조 영역이 겹치는지 확인
        if (region_top <= row_top <= region_bottom) or (region_top <= row_bottom <= region_bottom) or \
           (row_top <= region_top <= row_bottom) or (row_top <= region_bottom <= row_bottom):
            return True
    return False

def match_highlighted_texts_with_table_fuzzy(highlighted_texts, table_df, threshold=80, debug_mode=False):
    """
    FuzzyWuzzy를 사용하여 OCR로 추출한 텍스트와 테이블 행을 유사도 기반으로 매칭하는 함수
    """
    for ocr_text in highlighted_texts:
        if not ocr_text:
            continue
        # 각 행을 순회하며 OCR 텍스트와의 유사도 계산
        for idx, row in table_df.iterrows():
            # 변경사항이 이미 '추가'인 경우 건너뜀
            if row['변경사항'] == '추가':
                continue
            row_text = ' '.join(row.drop(['변경사항', 'Table_Number']).astype(str))
            similarity = fuzz.partial_ratio(ocr_text, row_text)
            if similarity >= threshold:
                table_df.at[idx, '변경사항'] = '추가'
                if debug_mode:
                    print(f"Match found: OCR '{ocr_text}' matches Table row '{row_text}' with similarity {similarity}")
            elif similarity < threshold and table_df.at[idx, '변경사항'] == '추가':
                table_df.at[idx, '변경사항'] = ''
                if debug_mode:
                    print(f"Similarity below threshold for row {idx}: '{ocr_text}' not similar to '{row_text}'")
    return table_df

def load_existing_excel_data(excel_path, key_columns):
    """
    기존 엑셀 데이터를 로드하여 데이터프레임으로 반환하는 함수
    """
    if not os.path.exists(excel_path):
        print(f"Excel 파일이 존재하지 않습니다: {excel_path}")
        return pd.DataFrame()  # 빈 데이터프레임 반환
    try:
        df_existing = pd.read_excel(excel_path)
        df_existing = df_existing.drop_duplicates(subset=key_columns)
        print(f"Loaded existing Excel data with {len(df_existing)} rows.")
        return df_existing
    except Exception as e:
        print(f"기존 엑셀 데이터를 로드하는 중 오류 발생: {e}")
        return pd.DataFrame()

def update_excel_data(df_new, df_existing, key_columns, output_path, debug_mode=False):
    """
    새로운 데이터와 기존 데이터를 비교하여 추가 및 삭제를 수행하고, 엑셀 파일로 저장하는 함수
    """
    # 새로운 데이터에서 '변경사항'이 '추가'인 행만 선택
    df_to_add = df_new[df_new['변경사항'] == '추가']

    if not df_existing.empty:
        # 기존 데이터에서 새로운 데이터와 매칭되는 행을 찾아 제거
        df_to_remove = df_existing[~df_existing[key_columns].apply(tuple,1).isin(df_new[key_columns].apply(tuple,1))]
    else:
        df_to_remove = pd.DataFrame()  # 빈 데이터프레임

    if debug_mode:
        print(f"Rows to add: {len(df_to_add)}")
        print(f"Rows to remove: {len(df_to_remove)}")

    # 엑셀 업데이트
    df_updated = pd.concat([df_existing, df_to_add], ignore_index=True)
    if not df_to_remove.empty:
        df_updated = df_updated[~df_updated[key_columns].apply(tuple,1).isin(df_to_remove[key_columns].apply(tuple,1))]

    # 중복 제거
    df_updated = df_updated.drop_duplicates(subset=key_columns)

    # 엑셀 저장
    df_updated.to_excel(output_path, index=False)
    print(f"Updated Excel data has been saved to '{output_path}'.")

def save_to_excel_with_highlight(df, output_path, debug_mode=False):
    """
    데이터프레임을 엑셀 파일로 저장하고, '변경사항' 컬럼이 '추가'인 행을 노란색으로 강조하는 함수
    """
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    if '변경사항' in df.columns:
        change_col_index = df.columns.get_loc('변경사항') + 1
    else:
        raise ValueError("DataFrame에 '변경사항' 컬럼이 없습니다.")

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=change_col_index).value
        if cell_value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows.")

def compare_excel_and_extracted_text(excel_df, extracted_text_path):
    """
    엑셀 데이터와 추출된 텍스트를 비교하는 함수
    """
    with open(extracted_text_path, 'r', encoding='utf-8') as file:
        extracted_text = file.read()
    
    for idx, row in excel_df.iterrows():
        if row['변경사항'] == '추가':
            row_text = ' '.join(row.astype(str))
            if row_text not in extracted_text:
                excel_df.at[idx, '변경사항'] = ''
    
    return excel_df

def main(pdf_path, output_excel_path, existing_excel_path, key_columns):
    print("PDF에서 개정된 부분을 추출합니다...")

    doc = fitz.open(pdf_path)
    page_number = 51  # 51페이지

    page = doc.load_page(page_number - 1)
    image = pdf_to_image(page)

    original_image_path = os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number}_original.png')
    cv2.imwrite(original_image_path, cv2.cvtColor(image, cv2.COLOR_RGB2BGR))
    if DEBUG_MODE:
        print(f"Saved original image: {original_image_path}")

    contours = detect_highlights(image, page_number)
    highlight_regions = get_highlight_regions(contours, image.shape[0])

    if DEBUG_MODE:
        print(f"감지된 강조 영역 수: {len(highlight_regions)}")
        print(f"강조 영역: {highlight_regions}")

    # 강조 영역을 PDF로 변환
    highlight_pdf_path = os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number}_highlights.pdf')
    convert_highlights_to_pdf(image, highlight_regions, highlight_pdf_path)

    # Camelot을 사용하여 강조 영역 PDF에서 테이블 추출 및 텍스트 저장
    extracted_text_path = os.path.join(TXT_OUTPUT_DIR, f'page_{page_number}_extracted_text.txt')
    tables = extract_tables_with_camelot(highlight_pdf_path, 1, extracted_text_path)

    if not tables:
        print("추출된 표가 없습니다.")
        return

    # 추출된 표 처리
    processed_df = process_tables(tables, highlight_regions, image.shape[0], debug_mode=DEBUG_MODE)

    # 기존 엑셀 데이터 로드
    df_existing = load_existing_excel_data(existing_excel_path, key_columns)

    if df_existing.empty:
        print(f"기존 엑셀 파일이 없습니다: {existing_excel_path}")
        return

    # 엑셀 데이터와 추출된 텍스트 비교
    df_updated = compare_excel_and_extracted_text(df_existing, extracted_text_path)

    # 업데이트된 엑셀 파일 저장
    output_excel_path_v2 = output_excel_path.replace('.xlsx', '_v2.xlsx')
    df_updated.to_excel(output_excel_path_v2, index=False)
    print(f"업데이트된 엑셀 파일이 {output_excel_path_v2}에 저장되었습니다.")

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables_camelot_updated.xlsx"
    existing_excel_path = "/workspaces/automation/output/extracted_tables_camelot.xlsx"
    key_columns = ['Column1', 'Column2', 'Column3']  # 엑셀에서 고유 식별에 사용할 컬럼 이름으로 수정하세요
    main(pdf_path, output_excel_path, existing_excel_path, key_columns)