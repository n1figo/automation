import camelot
import pandas as pd
import numpy as np
import cv2
import os
import fitz
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def pdf_to_image(page):
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return np.array(img)

def detect_highlights(image):
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    
    # 여러 색상 범위 정의 (HSV)
    color_ranges = [
        ((20, 100, 100), (40, 255, 255)),  # 노란색
        ((100, 100, 100), (140, 255, 255)),  # 파란색
        ((125, 100, 100), (155, 255, 255))  # 보라색
    ]
    
    masks = []
    for lower, upper in color_ranges:
        mask = cv2.inRange(hsv, np.array(lower), np.array(upper))
        masks.append(mask)
    
    # 모든 마스크 결합
    combined_mask = np.zeros_like(masks[0])
    for mask in masks:
        combined_mask = cv2.bitwise_or(combined_mask, mask)
    
    kernel = np.ones((5,5), np.uint8)
    cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)
    
    contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    return contours

def get_highlight_regions(contours, image_height):
    regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        top = image_height - (y + h)
        bottom = image_height - y
        regions.append((top, bottom))
    return regions

def extract_tables_with_camelot(pdf_path, page_number):
    print(f"Extracting tables from page {page_number} using Camelot...")
    tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
    print(f"Found {len(tables)} tables on page {page_number}")
    return tables

def process_tables(tables, highlight_regions, page_height):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        x1, y1, x2, y2 = table._bbox

        table_height = y2 - y1
        row_height = table_height / len(df)

        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            
            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height
            
            row_highlighted = check_highlight((row_top, row_bottom), highlight_regions)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)

    return pd.DataFrame(processed_data)

def check_highlight(row_range, highlight_regions):
    row_top, row_bottom = row_range
    for region_top, region_bottom in highlight_regions:
        if (region_top <= row_top <= region_bottom) or (region_top <= row_bottom <= region_bottom) or \
           (row_top <= region_top <= row_bottom) or (row_top <= region_bottom <= row_bottom):
            return True
    return False

def save_to_excel_with_highlight(df, output_path, title=None):
    wb = Workbook()
    ws = wb.active

    # 제목을 추가
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title)
        max_col = len(df.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=20, bold=True)
        ws.row_dimensions[1].height = 30  # 제목 행 높이 조정
        start_row = 2  # 데이터는 다음 행부터 시작

    # DataFrame을 Excel로 저장
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    if '변경사항' in df.columns:
        change_col_index = df.columns.get_loc('변경사항') + 1
    else:
        raise ValueError("DataFrame에 '변경사항' 컬럼이 없습니다.")

    for row in range(start_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=change_col_index).value
        if cell_value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    # 셀의 텍스트 줄바꿈 설정
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 최대 열 너비 설정
    MAX_COLUMN_WIDTH = 50  # 필요한 경우 조정

    # 열 너비 조정
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = min((max_length + 2), MAX_COLUMN_WIDTH)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 행 높이 자동 조정
    for row in ws.iter_rows(min_row=start_row):
        max_line_count = 1
        for cell in row:
            if cell.value:
                line_count = str(cell.value).count('\n') + 1
                if line_count > max_line_count:
                    max_line_count = line_count
        ws.row_dimensions[row[0].row].height = max_line_count * 15  # 필요한 경우 조정

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows and auto-adjusted cell sizes")

def main():
    uploads_folder = "/workspaces/automation/uploads"
    output_folder = "/workspaces/automation/output"
    
    # 출력 폴더 생성
    os.makedirs(output_folder, exist_ok=True)

    # 업로드 폴더에서 첫 번째 PDF 파일 가져오기
    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_extracted_tables.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    doc = fitz.open(pdf_path)

    # 첫 페이지에서 제목 추출
    first_page = doc[0]
    page_text = first_page.get_text("text")
    title = page_text.strip().split('\n')[0]  # 첫 번째 줄을 제목으로 가정
    print(f"Extracted title: {title}")

    all_processed_data = []

    for page_number in range(len(doc)):
        print(f"Processing page: {page_number + 1}/{len(doc)}")

        page = doc[page_number]
        image = pdf_to_image(page)

        contours = detect_highlights(image)
        highlight_regions = get_highlight_regions(contours, image.shape[0])

        print(f"Page {page_number + 1}: Detected {len(highlight_regions)} highlighted regions")

        tables = extract_tables_with_camelot(pdf_path, page_number + 1)

        if not tables:
            print(f"Page {page_number + 1}: No tables extracted")
            continue

        processed_df = process_tables(tables, highlight_regions, image.shape[0])
        processed_df['Page_Number'] = page_number + 1
        all_processed_data.append(processed_df)

    if not all_processed_data:
        print("No processed data available.")
        return

    final_df = pd.concat(all_processed_data, ignore_index=True)

    # 제목을 포함하여 엑셀로 저장
    save_to_excel_with_highlight(final_df, output_excel_path, title=title)

    print(f"All processed data has been saved to {output_excel_path}")

if __name__ == "__main__":
    main()
