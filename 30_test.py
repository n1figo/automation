import camelot
import pandas as pd
import numpy as np
import cv2
import os
import fitz  # PyMuPDF
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz
import easyocr  # EasyOCR 라이브러리

# EasyOCR 리더 초기화 (한국어 지원)
reader = easyocr.Reader(['ko'])  # 첫 실행 시 모델 다운로드

# 디버그 모드 설정
DEBUG_MODE = True
IMAGE_OUTPUT_DIR = "/workspaces/automation/output/images"
TXT_OUTPUT_DIR = "/workspaces/automation/output/texts"
os.makedirs(IMAGE_OUTPUT_DIR, exist_ok=True)
os.makedirs(TXT_OUTPUT_DIR, exist_ok=True)

def pdf_to_image(page):
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return np.array(img)

def detect_highlights(image, page_num):
    """
    강조색(하이라이트) 영역을 감지하고 컨투어를 반환하는 함수
    """
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    
    # 노란색 하이라이트 감지를 위한 HSV 범위 설정
    lower_yellow = np.array([20, 100, 100])
    upper_yellow = np.array([30, 255, 255])
    mask = cv2.inRange(hsv, lower_yellow, upper_yellow)

    # Morphological operations to clean up the mask
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=2)

    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_num}_mask.png'), mask)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    contour_image = image.copy()
    cv2.drawContours(contour_image, contours, -1, (0, 255, 0), 2)
    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_num}_contours.png'), cv2.cvtColor(contour_image, cv2.COLOR_RGB2BGR))

    return contours

def get_highlight_regions(contours, image_height):
    """
    컨투어를 기반으로 강조된 영역의 좌표를 반환하는 함수
    """
    regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        # OpenCV 좌표계를 PDF 좌표계로 변환
        top = image_height - (y + h)
        bottom = image_height - y
        regions.append((top, bottom, x, x + w))  # x 좌표도 포함
    return regions

def extract_tables_with_camelot(pdf_path, page_number):
    """
    Camelot을 사용하여 PDF의 특정 페이지에서 테이블을 추출하는 함수
    """
    print(f"Extracting tables from page {page_number} using Camelot...")
    tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
    print(f"Found {len(tables)} tables on page {page_number}")
    return tables

def ocr_image_easyocr(image):
    """
    EasyOCR을 사용하여 이미지에서 한글 텍스트를 추출하는 함수
    """
    try:
        # 이미지 전처리: 그레이스케일 변환
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # 노이즈 제거 (Median Blur)
        gray = cv2.medianBlur(gray, 3)
        
        # 대비 향상
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        gray = clahe.apply(gray)
        
        # 이진화 (Adaptive Thresholding)
        thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY, 31, 2)
        
        # 기울기 보정 (Deskewing)
        coords = np.column_stack(np.where(thresh > 0))
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
        (h, w) = thresh.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        thresh = cv2.warpAffine(thresh, M, (w, h),
                                flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        
        # OCR 수행 (EasyOCR 사용)
        results = reader.readtext(thresh, detail=0, paragraph=True)
        ocr_text = ' '.join(results)
        
        return ocr_text.strip()
    except Exception as e:
        print(f"OCR 처리 중 오류 발생: {e}")
        return ""

def extract_highlighted_text_easyocr(pdf_path, page_number, highlight_regions):
    """
    강조된 영역에서 텍스트를 추출하는 함수 (EasyOCR 사용)
    """
    doc = fitz.open(pdf_path)
    page = doc.load_page(page_number - 1)  # 0-based index
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    extracted_texts = []

    txt_file_path = os.path.join(TXT_OUTPUT_DIR, f'page_{page_number}_highlighted_texts.txt')
    with open(txt_file_path, 'w', encoding='utf-8') as txt_file:
        for idx, (top, bottom, left, right) in enumerate(highlight_regions):
            try:
                # 이미지 좌표계로 변환
                img_height, img_width, _ = np.array(img).shape
                y0 = img_height - bottom
                y1 = img_height - top
                x0 = left
                x1 = right

                # 영역 크롭
                cropped_img = img.crop((x0, y0, x1, y1))
                cropped_img_path = os.path.join(IMAGE_OUTPUT_DIR, f'highlight_{page_number}_{idx}_cropped.png')
                cropped_img.save(cropped_img_path)

                # OCR 수행 (EasyOCR 사용)
                ocr_text = ocr_image_easyocr(np.array(cropped_img))

                if DEBUG_MODE:
                    print(f"OCR Text from region {idx}: {ocr_text}")
                extracted_texts.append(ocr_text)

                # txt 파일에 저장
                txt_file.write(f"Region {idx}: {ocr_text}\n")
            except Exception as e:
                print(f"Exception occurred while processing region {idx}: {e}")
                extracted_texts.append('')
                txt_file.write(f"Region {idx}: \n")
                continue

    print(f"Extracted highlighted texts have been saved to '{txt_file_path}'")
    return extracted_texts

def process_tables(tables, highlight_regions, page_height):
    """
    추출된 테이블을 처리하여 데이터프레임으로 반환하는 함수
    """
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        x1, y1, x2, y2 = table._bbox  # 올바른 좌표 언패킹

        # PDF 좌표계에서 y1은 하단, y2는 상단
        table_height = y2 - y1
        row_height = table_height / len(df)

        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            
            # 행의 상단과 하단 y 좌표 계산 (PDF 좌표계 사용)
            # 상단부터 계산하기 위해 y2에서부터 감소
            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height
            
            row_highlighted = check_highlight((row_top, row_bottom, 0, 0), highlight_regions)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)

    return pd.DataFrame(processed_data)

def check_highlight(row_range, highlight_regions):
    """
    특정 행이 강조된 영역과 겹치는지 확인하는 함수
    """
    row_top, row_bottom, _, _ = row_range
    for region_top, region_bottom, _, _ in highlight_regions:
        # 행과 강조 영역이 겹치는지 확인
        if (region_top <= row_top <= region_bottom) or (region_top <= row_bottom <= region_bottom) or \
           (row_top <= region_top <= row_bottom) or (row_top <= region_bottom <= row_bottom):
            return True
    return False

def match_highlighted_texts_with_table_fuzzy(highlighted_texts, table_df, threshold=80):
    """
    FuzzyWuzzy를 사용하여 OCR로 추출한 텍스트와 테이블 행을 유사도 기반으로 매칭하는 함수
    """
    for ocr_text in highlighted_texts:
        if not ocr_text:
            continue
        # 각 행을 순회하며 OCR 텍스트와의 유사도 계산
        for idx, row in table_df.iterrows():
            row_text = ' '.join(row.drop(['변경사항', 'Table_Number']).astype(str))
            similarity = fuzz.partial_ratio(ocr_text, row_text)
            if similarity >= threshold:
                table_df.at[idx, '변경사항'] = '추가'
            elif similarity < threshold and table_df.at[idx, '변경사항'] == '추가':
                table_df.at[idx, '변경사항'] = ''
                if DEBUG_MODE:
                    print(f"Similarity below threshold for row {idx}: '{ocr_text}' not similar to '{row_text}'")
    return table_df

def save_to_excel_with_highlight(df, output_path):
    """
    데이터프레임을 엑셀 파일로 저장하고, '변경사항' 컬럼이 '추가'인 행을 노란색으로 강조하는 함수
    """
    df.to_excel(output_path, index=False)
    
    wb = load_workbook(output_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    if '변경사항' in df.columns:
        change_col_index = df.columns.get_loc('변경사항') + 1
    else:
        raise ValueError("DataFrame에 '변경사항' 컬럼이 없습니다.")

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=change_col_index).value
        if cell_value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows")

def main(pdf_path, output_excel_path, use_easyocr=False):
    print("PDF에서 개정된 부분을 추출합니다...")

    doc = fitz.open(pdf_path)
    page_number = 51  # 51페이지

    page = doc.load_page(page_number - 1)
    image = pdf_to_image(page)

    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number}_original.png'), cv2.cvtColor(image, cv2.COLOR_RGB2BGR))

    contours = detect_highlights(image, page_number)
    highlight_regions = get_highlight_regions(contours, image.shape[0])

    highlighted_image = image.copy()
    for top, bottom, _, _ in highlight_regions:
        # PDF 좌표계를 OpenCV 좌표계로 변환하여 그리기
        cv2.rectangle(highlighted_image, (0, image.shape[0] - bottom), (image.shape[1], image.shape[0] - top), (0, 255, 0), 2)
    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number}_highlighted.png'), cv2.cvtColor(highlighted_image, cv2.COLOR_RGB2BGR))

    print(f"감지된 강조 영역 수: {len(highlight_regions)}")
    print(f"강조 영역: {highlight_regions}")

    # Camelot을 사용하여 표 추출
    tables = extract_tables_with_camelot(pdf_path, page_number)

    if not tables:
        print("추출된 표가 없습니다.")
        return

    # OCR 방식 선택
    if use_easyocr:
        # EasyOCR을 통해 강조된 텍스트 추출
        extracted_texts = extract_highlighted_text_easyocr(pdf_path, page_number, highlight_regions)
    else:
        # Tesseract를 통해 강조된 텍스트 추출
        extracted_texts = extract_highlighted_text_tesseract(pdf_path, page_number, highlight_regions)

    # 추출된 표 처리
    processed_df = process_tables(tables, highlight_regions, image.shape[0])

    # FuzzyWuzzy를 사용하여 OCR 텍스트와 테이블 행 매칭
    processed_df = match_highlighted_texts_with_table_fuzzy(extracted_texts, processed_df, threshold=80)

    # '변경사항' 컬럼을 기반으로 엑셀 저장
    save_to_excel_with_highlight(processed_df, output_excel_path)

    print(f"처리된 데이터가 {output_excel_path}에 저장되었습니다.")

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables_camelot.xlsx"
    # OCR 방식을 변경하려면 use_easyocr=True로 설정
    main(pdf_path, output_excel_path, use_easyocr=False)
