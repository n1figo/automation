import requests
from bs4 import BeautifulSoup
import fitz  # PyMuPDF
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import os
import re

def extract_text_from_url(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    # 필요한 텍스트 추출 로직 구현
    # 예: 특정 div 내의 텍스트만 추출
    text = soup.find('div', class_='content-area').get_text(separator='\n')
    return text

def is_color_highlighted(color):
    # 흰색과 검정색 제외
    return color not in [(1, 1, 1), (0, 0, 0)] and any(c != 1 for c in color)

def extract_highlighted_text_with_context_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    highlighted_texts_with_context = []
    for page in doc:
        blocks = page.get_text("dict")["blocks"]
        lines = page.get_text("text").split('\n')
        for block in blocks:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        if is_color_highlighted(span["color"]):
                            highlighted_text = span["text"]
                            line_index = lines.index(highlighted_text)
                            context = '\n'.join(lines[max(0, line_index-5):line_index])
                            highlighted_texts_with_context.append((context, highlighted_text))
    return highlighted_texts_with_context

def find_matches(html_text, highlighted_texts_with_context):
    matches = []
    for context, highlighted_text in highlighted_texts_with_context:
        escaped_context = re.escape(re.sub(r'[^\w\s]', '', context))
        pattern = re.compile(escaped_context, re.IGNORECASE)
        match = pattern.search(html_text)
        if match:
            end_index = match.end()
            matches.append((end_index, highlighted_text))
    return matches

def capture_surrounding_area(driver, element, filename):
    location = element.location
    size = element.size
    png = driver.get_screenshot_as_png()
    im = PILImage.open(io.BytesIO(png))
    left = location['x']
    top = max(0, location['y'] - 100)
    right = location['x'] + size['width']
    bottom = location['y'] + size['height'] + 100
    im = im.crop((left, top, right, bottom))
    im.save(filename)

def create_excel_report(matches, captures, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "변경 사항"

    for i, (match, capture) in enumerate(zip(matches, captures), start=1):
        ws.cell(row=i, column=1, value=match[1])  # 음영 처리된 텍스트
        ws.cell(row=i, column=2, value="")  # 디자이너가 입력할 열
        img = Image(capture)
        ws.add_image(img, f'C{i}')

    wb.save(output_path)

def main():
    url = "https://www.kbinsure.co.kr/CG302120001.ec"
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_dir = "/workspaces/automation/output"
    os.makedirs(output_dir, exist_ok=True)

    html_text = extract_text_from_url(url)
    highlighted_texts_with_context = extract_highlighted_text_with_context_from_pdf(pdf_path)
    matches = find_matches(html_text, highlighted_texts_with_context)

    driver = webdriver.Chrome()  # Selenium WebDriver 초기화
    driver.get(url)

    captures = []
    for match in matches:
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//*[contains(text(), '{match[1]}')]"))
            )
            filename = f"{output_dir}/capture_{match[0]}.png"
            capture_surrounding_area(driver, element, filename)
            captures.append(filename)
        except Exception as e:
            print(f"엘리먼트를 찾을 수 없습니다: {match[1]}. 오류: {e}")

    driver.quit()

    create_excel_report(matches, captures, f"{output_dir}/변경사항_보고서.xlsx")

if __name__ == "__main__":
    main()
