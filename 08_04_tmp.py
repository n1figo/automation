import PyPDF2
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import os
import pandas as pd
import camelot
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from sentence_transformers import SentenceTransformer

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class InsuranceDocumentAnalyzer:
    def __init__(self):
        self.section_patterns = {
            "종류": r'\[(\d)종\]',
            "특약유형": r'(상해관련|질병관련)\s*특약'
        }
        self.section_pages = {"[1종]": None, "[2종]": None, "[3종]": None}
        self.section_ranges = {}

    def find_section_pages(self, pdf_path: str) -> Dict[str, int]:
        """PDF에서 1종, 2종, 3종의 시작 페이지 찾기"""
        try:
            doc = fitz.open(pdf_path)
            for page_num in range(len(doc)):
                text = doc[page_num].get_text()
                
                # 종 패턴 찾기
                matches = re.finditer(self.section_patterns["종류"], text)
                for match in matches:
                    종_type = f"[{match.group(1)}종]"
                    if self.section_pages[종_type] is None:
                        self.section_pages[종_type] = page_num
                        logger.info(f"{종_type} 시작 페이지: {page_num + 1}")

            # 섹션 범위 설정
            sorted_pages = sorted([(k, v) for k, v in self.section_pages.items() if v is not None], 
                                key=lambda x: x[1])
            
            for i, (종_type, start_page) in enumerate(sorted_pages):
                if i + 1 < len(sorted_pages):
                    end_page = sorted_pages[i + 1][1]
                else:
                    end_page = len(doc)
                self.section_ranges[종_type] = (start_page, end_page)
                
            doc.close()
            return self.section_pages
            
        except Exception as e:
            logger.error(f"Error finding section pages: {e}")
            return {}


class TableExtractor:
    def __init__(self):
        self.font_size_threshold = 10
        self.title_max_length = 50
        self.model = SentenceTransformer('distiluse-base-multilingual-cased-v1')

    def extract_with_camelot(self, pdf_path: str, page_num: int) -> List:
        """Camelot을 사용한 표 추출"""
        try:
            # 먼저 lattice 모드로 시도
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num),
                flavor='lattice',
                line_scale=40,
                copy_text=['v']
            )
            
            # lattice 모드가 실패하면 stream 모드로 시도
            if not tables:
                tables = camelot.read_pdf(
                    pdf_path,
                    pages=str(page_num),
                    flavor='stream',
                    edge_tol=500,
                    row_tol=10
                )
                
            if tables:
                logger.info(f"Successfully extracted {len(tables)} tables from page {page_num}")
                # 테이블 품질 검사
                for i, table in enumerate(tables):
                    accuracy = table.parsing_report.get('accuracy', 0)
                    whitespace = table.parsing_report.get('whitespace', 100)
                    logger.info(f"Table {i+1} - Accuracy: {accuracy}%, Whitespace: {whitespace}%")
            
            return tables
            
        except Exception as e:
            logger.error(f"Camelot extraction failed on page {page_num}: {str(e)}")
            return []

    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """표 데이터 정제"""
        try:
            # 빈 행/열 제거
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            
            # 특수문자로 시작하는 행 제거
            df = df[~df.iloc[:, 0].str.contains("※|주)", regex=False, na=False)]
            
            # 빈 문자열 처리
            df = df.replace(r'^\s*$', np.nan, regex=True)
            
            # 불필요한 공백 제거
            df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
            
            return df
            
        except Exception as e:
            logger.error(f"Error cleaning table: {e}")
            return pd.DataFrame()

    def extract_tables_from_section(self, pdf_path: str, start_page: int, end_page: int) -> List[Tuple[str, pd.DataFrame, int]]:
        """섹션 범위 내의 표 추출"""
        try:
            results = []
            doc = fitz.open(pdf_path)
            
            for page_num in range(start_page, end_page):
                logger.info(f"Processing page {page_num + 1}")
                page = doc[page_num]
                text = page.get_text()
                
                # 상해관련 또는 질병관련 특약 확인
                if re.search(r'(상해관련|질병관련)\s*특약', text):
                    logger.info(f"Found 특약 section on page {page_num + 1}")
                    
                    # 표 추출
                    tables = self.extract_with_camelot(pdf_path, page_num + 1)
                    
                    for idx, table in enumerate(tables):
                        df = self.clean_table(table.df)
                        if not df.empty:
                            # 표의 위치 정보 얻기
                            table_bbox = table.cells[0][0].bbox  # 첫 번째 셀의 bbox
                            page_height = table.parsing_report['page_bbox'][3]
                            # PDF 좌표계를 Fitz 좌표계로 변환
                            table_top = page_height - table_bbox[3]
                            
                            # 제목 추출
                            title = self.extract_table_title(page)
                            logger.info(f"Extracted title for table {idx + 1}: {title}")
                            results.append((title, df, page_num + 1))

            doc.close()
            return results
            
        except Exception as e:
            logger.error(f"Error extracting tables from section: {e}")
            traceback.print_exc()  # 상세한 에러 정보 출력
            return []

    def extract_table_title(self, page) -> str:
        """RAG를 활용한 표 위의 제목 추출"""
        try:
            blocks = page.get_text("dict")["blocks"]
            table_block = None
            title_candidates = []
            
            # 표 블록 찾기
            for block in blocks:
                if "lines" in block:
                    text = " ".join(span["text"] for line in block["lines"] 
                                  for span in line["spans"])
                    if "특약" in text and len(block["lines"]) > 1:
                        table_block = block
                        break
            
            if table_block:
                table_top = table_block["bbox"][1]
                
                # 표 위의 텍스트 블록들 수집
                for block in blocks:
                    if "lines" in block and block["bbox"][3] < table_top:
                        text = " ".join(span["text"] for line in block["lines"] 
                                      for span in line["spans"])
                        
                        # 텍스트 길이 및 기본 필터링
                        if 5 < len(text) < self.title_max_length and not any(c in text for c in ['※', '►', '▶']):
                            title_candidates.append({
                                'text': text,
                                'distance': table_top - block["bbox"][3],
                                'bbox': block["bbox"]
                            })
                
                if title_candidates:
                    # RAG 검색을 위한 쿼리와 참조 텍스트 준비
                    queries = [
                        "보험금을 지급하는 사유",
                        "보장하는 내용",
                        "특약 보장 내용",
                        "보험금 지급사유"
                    ]
                    
                    # 후보 텍스트들의 임베딩 생성
                    candidate_embeddings = self.model.encode([c['text'] for c in title_candidates])
                    query_embeddings = self.model.encode(queries)
                    
                    # 각 후보에 대한 점수 계산
                    scores = []
                    for idx, candidate in enumerate(title_candidates):
                        # 의미적 유사도 점수
                        semantic_score = np.max([
                            np.dot(candidate_embeddings[idx], query_emb)
                            for query_emb in query_embeddings
                        ])
                        
                        # 거리 기반 점수 (가까울수록 높은 점수)
                        distance_score = 1 / (1 + candidate['distance'] / 100)
                        
                        # 텍스트 특성 점수
                        characteristic_score = 0
                        if any(keyword in candidate['text'] for keyword in ['보장', '특약', '진단', '수술', '입원']):
                            characteristic_score += 0.3
                        
                        # 최종 점수 계산
                        final_score = (
                            semantic_score * 0.5 +  # 의미적 유사도
                            distance_score * 0.3 +  # 물리적 거리
                            characteristic_score    # 텍스트 특성
                        )
                        
                        scores.append(final_score)
                    
                    # 가장 높은 점수를 받은 제목 선택
                    best_idx = np.argmax(scores)
                    return title_candidates[best_idx]['text']
            
            return "Untitled Table"
            
        except Exception as e:
            logger.error(f"Error extracting table title with RAG: {e}")
            return "Untitled Table"

class ExcelWriter:
    @staticmethod
    def save_to_excel(sections_data: Dict[str, List[Tuple[str, pd.DataFrame, int]]], output_path: str):
        """섹션별 데이터를 Excel 파일로 저장"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for section, tables in sections_data.items():
                    if not tables:
                        continue
                        
                    sheet_name = section.replace("[", "").replace("]", "")
                    current_row = 0
                    
                    for title, df, page_num in tables:
                        # 제목 쓰기
                        title_df = pd.DataFrame([[f"{title} (페이지: {page_num})"]], columns=[''])
                        title_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row,
                            index=False,
                            header=False
                        )
                        
                        # 표 데이터 쓰기
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row + 2,
                            index=False
                        )
                        
                        # 스타일 적용
                        worksheet = writer.sheets[sheet_name]
                        
                        # 제목 스타일링
                        title_cell = worksheet.cell(row=current_row + 1, column=1)
                        title_cell.font = Font(bold=True, size=12)
                        title_cell.fill = PatternFill(start_color='E6E6E6',
                                                    end_color='E6E6E6',
                                                    fill_type='solid')
                        
                        current_row += len(df) + 5

            logger.info(f"Successfully saved tables to {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")

def main():
    try:
        # 파일 경로 설정
        pdf_path = "/workspaces/automation/uploads/KB 9회주는 암보험Plus(무배당)(24.05)_요약서_10.1판매_v1.0_앞단.pdf"
        output_path = "보험특약표.xlsx"
        
        if not os.path.exists(pdf_path):
            logger.error("PDF file not found")
            return

        # 문서 분석기 초기화
        document_analyzer = InsuranceDocumentAnalyzer()
        section_pages = document_analyzer.find_section_pages(pdf_path)
        
        if not section_pages:
            logger.error("No sections found in the document")
            return

        # 표 추출기 초기화
        table_extractor = TableExtractor()
        sections_data = {}
        
        # 각 섹션별 표 추출
        for section, (start_page, end_page) in document_analyzer.section_ranges.items():
            logger.info(f"Processing {section} (pages {start_page + 1} to {end_page})")
            tables = table_extractor.extract_tables_from_section(pdf_path, start_page, end_page)
            sections_data[section] = tables

        # 결과 저장
        if any(sections_data.values()):
            ExcelWriter.save_to_excel(sections_data, output_path)
            logger.info("Processing completed successfully")
        else:
            logger.error("No tables extracted from any section")

    except Exception as e:
        logger.error(f"Processing error: {str(e)}")

if __name__ == "__main__":
    main()