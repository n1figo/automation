import PyPDF2
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import os
import pandas as pd
import camelot
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import tkinter as tk
from tkinter import filedialog, messagebox

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('insurance_analyzer.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class InsuranceDocumentAnalyzer:
    def __init__(self):
        self.section_patterns = {
            "종류": r'\[(\d)종\]',
            "특약유형": r'(상해관련|질병관련)\s*특약'
        }
        self.section_pages = {"[1종]": None, "[2종]": None, "[3종]": None}
        self.section_ranges = {}
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]

    def find_section_pages(self, pdf_path: str) -> Dict[str, int]:
        """PDF에서 1종, 2종, 3종의 시작 페이지 찾기"""
        try:
            doc = fitz.open(pdf_path)
            for page_num in range(len(doc)):
                text = doc[page_num].get_text()
                
                matches = re.finditer(self.section_patterns["종류"], text)
                for match in matches:
                    종_type = f"[{match.group(1)}종]"
                    if self.section_pages[종_type] is None:
                        self.section_pages[종_type] = page_num
                        logger.info(f"{종_type} 시작 페이지: {page_num + 1}")

            sorted_pages = sorted([(k, v) for k, v in self.section_pages.items() if v is not None], 
                                key=lambda x: x[1])
            
            for i, (종_type, start_page) in enumerate(sorted_pages):
                if i + 1 < len(sorted_pages):
                    end_page = sorted_pages[i + 1][1]
                else:
                    end_page = len(doc)
                self.section_ranges[종_type] = (start_page, end_page)
                
            doc.close()
            return self.section_pages
            
        except Exception as e:
            logger.error(f"Error finding section pages: {e}")
            return {}

class TableExtractor:
    def __init__(self):
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]
        self.max_distance = 50  # 표와 제목 사이의 최대 거리

    def get_titles_with_positions(self, page) -> List[Dict]:
        """페이지에서 제목과 위치 정보 추출"""
        titles = []
        blocks = page.get_text("dict")["blocks"]
        
        for block in blocks:
            if block.get("lines"):
                text = ""
                y_top = block["bbox"][1]
                y_bottom = block["bbox"][3]
                
                for line in block["lines"]:
                    for span in line["spans"]:
                        text += span["text"] + " "
                
                text = text.strip()
                if text and any(re.search(pattern, text) for pattern in self.title_patterns):
                    titles.append({
                        "text": text,
                        "y_top": y_top,
                        "y_bottom": y_bottom,
                        "bbox": block["bbox"],
                        "used": False
                    })
        
        return sorted(titles, key=lambda x: x["y_top"])

    def get_table_positions(self, table, page_height):
        """Camelot 표의 위치를 PyMuPDF 좌표계로 변환"""
        x0, y0, x1, y1 = table._bbox
        converted_y0 = page_height - y1  # 상단
        converted_y1 = page_height - y0  # 하단
        
        return {
            "y_top": converted_y0,
            "y_bottom": converted_y1,
            "bbox": (x0, converted_y0, x1, converted_y1)
        }

    def match_title_to_table(self, titles, table_position):
        """표와 가장 적절한 제목 매칭"""
        best_title = None
        min_distance = float('inf')
        
        for title in titles:
            if title["used"]:
                continue
                
            distance = table_position["y_top"] - title["y_bottom"]
            if 0 < distance < self.max_distance and distance < min_distance:
                best_title = title
                min_distance = distance
        
        if best_title:
            best_title["used"] = True
            return best_title["text"], min_distance
        return None, None

    def extract_tables_from_section(self, pdf_path: str, start_page: int, end_page: int) -> List[Tuple[str, pd.DataFrame]]:
        """섹션 범위 내의 표 추출"""
        try:
            results = []
            
            for page_num in range(start_page, end_page):
                doc = fitz.open(pdf_path)
                page = doc[page_num]
                
                # 제목 추출
                titles = self.get_titles_with_positions(page)
                
                # 표 추출
                tables = self.extract_with_camelot(pdf_path, page_num + 1)
                
                if tables:
                    page_height = page.rect.height
                    
                    for table in tables:
                        table_position = self.get_table_positions(table, page_height)
                        title, distance = self.match_title_to_table(titles, table_position)
                        
                        df = self.clean_table(table.df)
                        if not df.empty:
                            if title:
                                logger.info(f"Found table with title: {title} (distance: {distance:.1f})")
                            else:
                                title = "Untitled Table"
                                logger.warning(f"No matching title found for table on page {page_num + 1}")
                            
                            results.append((title, df, page_num + 1))
                            
                doc.close()
                
            return results
            
        except Exception as e:
            logger.error(f"Error extracting tables from section: {e}")
            return []

    def extract_with_camelot(self, pdf_path: str, page_num: int) -> List:
        """Camelot을 사용한 표 추출"""
        try:
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num),
                flavor='lattice'
            )
            if not tables:
                tables = camelot.read_pdf(
                    pdf_path,
                    pages=str(page_num),
                    flavor='stream'
                )
            return tables
        except Exception as e:
            logger.error(f"Camelot extraction failed: {str(e)}")
            return []

    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """표 데이터 정제"""
        try:
            df = df.dropna(how='all')
            df = df[~df.iloc[:, 0].str.contains("※|주)", regex=False, na=False)]
            return df
        except Exception as e:
            logger.error(f"Error cleaning table: {e}")
            return pd.DataFrame()

class ExcelWriter:
    @staticmethod
    def save_to_excel(sections_data: Dict[str, List[Tuple[str, pd.DataFrame, int]]], output_path: str):
        """섹션별 데이터를 Excel 파일로 저장"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for section, tables in sections_data.items():
                    if not tables:
                        continue
                        
                    sheet_name = section.replace("[", "").replace("]", "")
                    current_row = 0
                    
                    for title, df, page_num in tables:
                        # 제목 쓰기
                        title_df = pd.DataFrame([[f"{title} (페이지: {page_num})"]], columns=[''])
                        title_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row,
                            index=False,
                            header=False
                        )
                        
                        # 표 데이터 쓰기
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row + 2,
                            index=False
                        )
                        
                        # 스타일 적용
                        worksheet = writer.sheets[sheet_name]
                        
                        # 제목 스타일링
                        title_cell = worksheet.cell(row=current_row + 1, column=1)
                        title_cell.font = Font(bold=True, size=12)
                        title_cell.fill = PatternFill(start_color='E6E6E6',
                                                    end_color='E6E6E6',
                                                    fill_type='solid')
                        
                        current_row += len(df) + 5

            logger.info(f"Successfully saved tables to {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")

def main():
    try:
        # 파일 선택 다이얼로그
        root = tk.Tk()
        root.withdraw()  # 메인 창 숨기기
        
        pdf_path = filedialog.askopenfilename(
            title="PDF 파일 선택",
            filetypes=[("PDF files", "*.pdf")]
        )
        
        if not pdf_path:
            logger.error("No file selected")
            return

        # 출력 파일 경로 설정
        output_path = os.path.splitext(pdf_path)[0] + "_분석결과.xlsx"
        
        logger.info(f"Selected PDF file: {pdf_path}")
        logger.info(f"Output will be saved to: {output_path}")

        if not os.path.exists(pdf_path):
            logger.error("PDF file not found")
            messagebox.showerror("오류", "PDF 파일을 찾을 수 없습니다.")
            return

        # 문서 분석기 초기화
        document_analyzer = InsuranceDocumentAnalyzer()
        section_pages = document_analyzer.find_section_pages(pdf_path)
        
        if not section_pages:
            logger.error("No sections found in the document")
            messagebox.showwarning("경고", "문서에서 섹션을 찾을 수 없습니다.")
            return

        # 표 추출기 초기화
        table_extractor = TableExtractor()
        sections_data = {}
        
        # 각 섹션별 표 추출
        for section, (start_page, end_page) in document_analyzer.section_ranges.items():
            logger.info(f"Processing {section} (pages {start_page + 1} to {end_page})")
            tables = table_extractor.extract_tables_from_section(pdf_path, start_page, end_page)
            sections_data[section] = tables

        # 결과 저장
        if any(sections_data.values()):
            ExcelWriter.save_to_excel(sections_data, output_path)
            logger.info("Processing completed successfully")
            messagebox.showinfo("완료", f"처리가 완료되었습니다.\n결과 파일: {output_path}")
        else:
            logger.error("No tables extracted from any section")
            messagebox.showwarning("경고", "추출된 표가 없습니다.")

    except Exception as e:
        error_msg = f"처리 중 오류 발생: {str(e)}"
        logger.error(error_msg)
        messagebox.showerror("오류", error_msg)

if __name__ == "__main__":
    main()