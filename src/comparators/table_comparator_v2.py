import pandas as pd
import logging
from typing import List, Tuple
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import re
import os

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

class TableComparator:
    """보험 상품 테이블 비교 및 업데이트 클래스"""

    def __init__(self):
        # 기본 설정
        self.comparison_columns = ['보장명', '보험금액']

        # 스타일 정의
        self.changed_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 변경: 초록색
        self.new_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')     # 신규: 노란색
        self.deleted_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid') # 삭제: 분홍색

    def _normalize_money_amount(self, amount: str) -> str:
        """보험금액 정규화"""
        if pd.isna(amount):
            return ''

        amount = str(amount).strip()

        # 숫자와 단위 추출
        number_pattern = r'[\d,.]+'
        numbers = re.findall(number_pattern, amount)
        if not numbers:
            return amount

        # 숫자 정규화
        normalized_number = numbers[0].replace(',', '')
        try:
            value = float(normalized_number)
            # 천 단위 콤마 추가
            normalized_number = f"{value:,.0f}"
        except:
            pass

        # 단위 처리
        unit = ''
        if '만원' in amount or '만' in amount:
            unit = '만원'
        elif '원' in amount:
            unit = '원'

        return f"{normalized_number}{unit}"

    def prepare_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """데이터 전처리"""
        try:
            # 컬럼명 정규화
            df.columns = [col.strip() for col in df.columns]

            # 필요한 컬럼이 있는지 확인
            required_columns = {'보장명', '보험금액'}
            available_columns = set(df.columns)

            # 매핑이 필요한 컬럼명들
            column_mapping = {
                '담보명': '보장명',
                '보장내용': '보장명',
                '급부명': '보장명',
                '보험가입금액': '보험금액',
                '보장금액': '보험금액',
                '가입금액': '보험금액'
            }

            # 컬럼 매핑 적용
            df = df.rename(columns=column_mapping)

            # 필수 컬럼 체크
            missing_columns = required_columns - set(df.columns)
            if missing_columns:
                raise ValueError(f"필수 컬럼이 누락됨: {missing_columns}")

            # 필요한 컬럼만 선택
            df = df[self.comparison_columns].copy()

            # 문자열 정규화
            for col in df.columns:
                if df[col].dtype == "object":
                    df[col] = df[col].str.strip()
                    df[col] = df[col].str.replace(r'\s+', ' ', regex=True)

            # 보험금액 정규화
            df['보험금액'] = df['보험금액'].apply(self._normalize_money_amount)

            # 결측값 처리
            df = df.fillna('')

            # 중복 제거
            df = df.drop_duplicates()

            return df

        except Exception as e:
            logger.error(f"데이터 전처리 중 오류 발생: {str(e)}")
            raise

    def compare_and_update(self, web_df: pd.DataFrame, pdf_df: pd.DataFrame) -> Tuple[pd.DataFrame, List[int]]:
        """웹페이지 추출 결과를 PDF 추출 결과로 업데이트"""
        try:
            # 데이터 전처리
            web_df = self.prepare_data(web_df)
            pdf_df = self.prepare_data(pdf_df)

            logger.info(f"비교 시작 - 웹페이지 데이터: {len(web_df)}행, PDF 데이터: {len(pdf_df)}행")

            # 보장명 기준으로 병합
            merged_df = pd.merge(
                web_df,
                pdf_df,
                on='보장명',
                how='outer',
                suffixes=('_웹', '_PDF')
            )

            # 변경 상태 분석
            merged_df['변경상태'] = merged_df.apply(self._determine_change_status, axis=1)

            # 보험금액 업데이트 (PDF 데이터로 업데이트)
            merged_df['보험금액'] = merged_df['보험금액_PDF'].combine_first(merged_df['보험금액_웹'])

            # 하이라이트가 필요한 셀 위치 저장
            highlight_cells = []
            for idx, row in merged_df.iterrows():
                if row['변경상태'] == '업데이트':
                    highlight_cells.append(idx)

            # 결과 컬럼 정리
            result_columns = ['보장명', '보험금액', '변경상태']
            final_df = merged_df[result_columns]

            # 변경 통계 로깅
            changes = final_df['변경상태'].value_counts()
            logger.info("변경 사항 요약:")
            for status, count in changes.items():
                logger.info(f"- {status}: {count}건")

            return final_df, highlight_cells

        except Exception as e:
            logger.error(f"테이블 비교 및 업데이트 중 오류 발생: {str(e)}")
            raise

    def _determine_change_status(self, row: pd.Series) -> str:
        """행별 변경 상태 판단"""
        if pd.isna(row['보험금액_웹']) and pd.notna(row['보험금액_PDF']):
            return '신규추가'
        elif pd.notna(row['보험금액_웹']) and pd.isna(row['보험금액_PDF']):
            return '삭제'
        elif pd.notna(row['보험금액_웹']) and pd.notna(row['보험금액_PDF']):
            if row['보험금액_웹'] != row['보험금액_PDF']:
                return '업데이트'
            else:
                return '유지'
        else:
            return '알 수 없음'

    def save_to_excel(self, df: pd.DataFrame, highlight_cells: List[int], output_path: str):
        """결과를 엑셀 파일로 저장"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='업데이트결과')

                # 워크시트 가져오기
                ws = writer.sheets['업데이트결과']

                # 기본 스타일 설정
                header_font = Font(bold=True)
                center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # 헤더 스타일 적용
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = center_alignment

                # 데이터 정렬 및 하이라이트 적용
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    status_cell = row[2]  # 변경상태 열
                    amount_cell = row[1]  # 보험금액 열

                    # 모든 셀 가운데 정렬
                    for cell in row:
                        cell.alignment = center_alignment

                    # 상태별 색상 적용
                    if status_cell.value == '신규추가':
                        for cell in row:
                            cell.fill = self.new_fill
                    elif status_cell.value == '삭제':
                        for cell in row:
                            cell.fill = self.deleted_fill
                    elif status_cell.value == '업데이트':
                        for cell in row:
                            cell.fill = self.changed_fill

                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

            logger.info(f"업데이트 결과가 저장됨: {output_path}")

        except Exception as e:
            logger.error(f"엑셀 저장 중 오류 발생: {str(e)}")
            raise

def main():
    # 웹페이지 추출기와 PDF 추출기의 엑셀 파일 경로를 지정하세요.
    web_excel_path = '웹페이지추출결과.xlsx'  # [웹페이지추출기]의 결과 파일 경로
    pdf_excel_path = 'PDF추출결과.xlsx'     # [PDF추출기]의 결과 파일 경로
    output_excel_path = '업무정의서_업데이트.xlsx'  # 퍼블리셔에게 전달할 엑셀 파일 경로

    # 엑셀 파일 읽기
    try:
        logger.info("데이터 비교 시작...")
        web_df = pd.read_excel(web_excel_path)
        pdf_df = pd.read_excel(pdf_excel_path)

        comparator = TableComparator()
        # 여기서 compare_tables를 compare_and_update로 변경
        updated_df, highlight_cells = comparator.compare_and_update(web_df, pdf_df)
        comparator.save_to_excel(updated_df, highlight_cells, output_excel_path)
        print(f"업데이트된 엑셀 파일이 저장되었습니다: {output_excel_path}")
    except Exception as e:
        logger.error(f"프로세스 중 오류 발생: {str(e)}")
        print(f"오류가 발생했습니다: {str(e)}")
        print("자세한 내용은 로그 파일을 확인해주세요.")

if __name__ == "__main__":
    main()
