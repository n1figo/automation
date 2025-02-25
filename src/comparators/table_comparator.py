import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os
from pathlib import Path

# 기본 경로 설정
BASE_DIR = Path("D:/github/pdf_local")
DATA_DIR = BASE_DIR / "data"
INPUT_DIR = DATA_DIR / "input"
OUTPUT_DIR = DATA_DIR / "output"

# 파일 경로 설정
web_tables_path = OUTPUT_DIR / "web_tables.xlsx"
pdf_extracted_male_path = OUTPUT_DIR / "pdf_extracted_20241113_154304.xlsx"
pdf_extracted_female_path = OUTPUT_DIR / "pdf_extracted_female.xlsx"  # 여성 기준 파일
output_path = OUTPUT_DIR / "변경사항_업무정의서.xlsx"

# 디렉토리 생성
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 파일 존재 여부 확인
if not web_tables_path.exists():
    raise FileNotFoundError(f"웹 테이블 파일을 찾을 수 없습니다: {web_tables_path}")
if not pdf_extracted_male_path.exists():
    raise FileNotFoundError(f"PDF 추출 파일(남성)을 찾을 수 없습니다: {pdf_extracted_male_path}")
if not pdf_extracted_female_path.exists():
    print(f"PDF 추출 파일(여성)을 찾을 수 없습니다. 여성 데이터는 제외하고 처리합니다.")
    process_female = False
else:
    process_female = True

# Load data
web_tables_df = pd.read_excel(web_tables_path)
pdf_extracted_male_df = pd.read_excel(pdf_extracted_male_path)
if process_female:
    pdf_extracted_female_df = pd.read_excel(pdf_extracted_female_path)

# 컬럼 이름을 일치시키기
web_tables_df.columns = ['Coverage', 'Amount', 'Term', 'Male Premium', 'Female Premium']
pdf_extracted_male_df.columns = ['Coverage', 'Amount', 'Premium', 'Term']
if process_female:
    pdf_extracted_female_df.columns = ['Coverage', 'Amount', 'Premium', 'Term']

# Define styles for highlighting
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 신규추가
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # 변경
strike_font = Font(strike=True)  # 삭제

# Workbook setup
wb = Workbook()
ws = wb.active
ws.title = '변경사항 업무정의서'

# Add headers
headers = ['Coverage', 'Amount', 'Premium', 'Term', 'Change Status']
ws.append(headers)

# Helper function to process and highlight changes
def compare_and_update(pdf_df, gender):
    for idx, row in pdf_df.iterrows():
        coverage = row['Coverage']
        amount = row['Amount']
        premium = row['Premium']
        term = row['Term']
        change_status = []

        # Gender-specific column selection in web_tables_df
        if gender == 'Male':
            web_row = web_tables_df[(web_tables_df['Coverage'] == coverage)]
            web_premium_col = 'Male Premium'
        else:
            web_row = web_tables_df[(web_tables_df['Coverage'] == coverage)]
            web_premium_col = 'Female Premium'

        if not web_row.empty:
            web_row = web_row.iloc[0]  # Only need the first match

            # Check for amount change
            if amount != web_row['Amount']:
                amount = web_row['Amount']
                change_status.append('Amount updated')
                amount_cell_fill = green_fill
            else:
                amount_cell_fill = None

            # Check for premium change
            if premium != web_row[web_premium_col]:
                premium = web_row[web_premium_col]
                change_status.append('Premium updated')
                premium_cell_fill = green_fill
            else:
                premium_cell_fill = None

            # Check for term change
            if term != web_row['Term']:
                term = web_row['Term']
                change_status.append('Term updated')
                term_cell_fill = green_fill
            else:
                term_cell_fill = None

            status = ', '.join(change_status) if change_status else 'No changes'
            row_values = [coverage, amount, premium, term, status]
            ws.append(row_values)

            # Apply cell styles
            row_num = ws.max_row
            if amount_cell_fill:
                ws.cell(row=row_num, column=2).fill = amount_cell_fill
            if premium_cell_fill:
                ws.cell(row=row_num, column=3).fill = premium_cell_fill
            if term_cell_fill:
                ws.cell(row=row_num, column=4).fill = term_cell_fill
        else:
            # If not in web_tables_df, mark as deleted
            status = 'Deleted'
            row_values = [coverage, amount, premium, term, status]
            ws.append(row_values)
            row_num = ws.max_row
            for cell in ws[row_num]:
                cell.font = strike_font

# Compare and update for male and female
compare_and_update(pdf_extracted_male_df, 'Male')
if process_female:
    compare_and_update(pdf_extracted_female_df, 'Female')

# Add rows for new entries in web_tables_df not found in pdf_extracted
for idx, row in web_tables_df.iterrows():
    coverage = row['Coverage']
    if not pdf_extracted_male_df[pdf_extracted_male_df['Coverage'] == coverage].empty:
        continue  # Already processed in male data

    if process_female and not pdf_extracted_female_df[pdf_extracted_female_df['Coverage'] == coverage].empty:
        continue  # Already processed in female data

    # New entry
    amount = row['Amount']
    term = row['Term']
    male_premium = row['Male Premium']
    female_premium = row['Female Premium'] if process_female else "N/A"
    status = 'New Entry'
    row_values = [coverage, amount, f"{male_premium}/{female_premium}", term, status]
    ws.append(row_values)

    # Highlight as new entry
    row_num = ws.max_row
    for cell in ws[row_num]:
        cell.fill = yellow_fill

# Save the workbook
wb.save(output_path)
print(f"변경사항 업무정의서 파일이 생성되었습니다: {output_path}")