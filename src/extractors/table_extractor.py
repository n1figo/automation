import logging
from typing import List, Optional, Dict, Any, Tuple
import pandas as pd
from playwright.sync_api import Page
from datetime import datetime
import os
from io import StringIO
import numpy as np
from bs4 import BeautifulSoup, Comment
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)8s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

class ImprovedTableExtractor:
    """개선된 웹 테이블 추출기"""
    
    def __init__(self, page: Page):
        self.page = page
        self.extracted_tables: List[pd.DataFrame] = []
        self.stats: Dict[str, Any] = self._init_stats()
        
    def _init_stats(self) -> Dict[str, Any]:
        return {
            'total_tables': 0,
            'successful_extractions': 0,
            'failed_extractions': 0,
            'start_time': datetime.now()
        }

    async def _scroll_page(self):
        """페이지 전체를 동적으로 스크롤하며 콘텐츠 로드"""
        try:
            # 초기 높이 가져오기
            last_height = await self.page.evaluate('document.body.scrollHeight')
            
            while True:
                # 현재 뷰포트 높이 가져오기
                viewport_height = await self.page.evaluate('window.innerHeight')
                total_height = await self.page.evaluate('document.body.scrollHeight')
                
                logger.info(f"현재 스크롤 진행 중: {total_height}px / 이전 높이: {last_height}px")
                
                # 작은 단위로 부드럽게 스크롤
                current_scroll = 0
                step = min(100, viewport_height // 4)  # 한 번에 스크롤할 픽셀 (더 작은 단위로 조정)
                
                while current_scroll < total_height:
                    await self.page.evaluate(f'window.scrollTo(0, {current_scroll})')
                    current_scroll += step
                    await self.page.wait_for_timeout(50)  # 더 짧은 대기 시간
                    
                    # 동적 로딩 확인
                    try:
                        await self.page.wait_for_function(
                            """() => {
                                const elements = document.querySelectorAll('table, div, p');
                                return Array.from(elements).every(el => el.getBoundingClientRect().top !== 0);
                            }""",
                            timeout=1000
                        )
                    except:
                        pass  # 타임아웃은 무시하고 계속 진행
                
                # 페이지 맨 아래로 스크롤
                await self.page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                await self.page.wait_for_timeout(2000)  # 충분한 로딩 대기 시간
                
                # 새로운 높이 확인
                new_height = await self.page.evaluate('document.body.scrollHeight')
                
                # 더 이상 높이 변화가 없으면 종료
                if new_height == last_height:
                    logger.info("페이지 끝에 도달: 모든 콘텐츠 로드 완료")
                    break
                    
                last_height = new_height
            
            # 맨 위로 천천히 스크롤 백
            current_scroll = last_height
            while current_scroll > 0:
                current_scroll -= step
                await self.page.evaluate(f'window.scrollTo(0, {current_scroll})')
                await self.page.wait_for_timeout(50)
            
            await self.page.wait_for_timeout(1000)  # 최종 안정화 대기
            logger.info("스크롤 완료: 페이지 맨 위로 이동")
            return True
            
        except Exception as e:
            logger.error(f"스크롤 중 오류 발생: {str(e)}")
            return False

    async def _click_example_tab(self) -> bool:
        """가입예시 탭 클릭"""
        try:
            # 여러 선택자 시도
            selectors = [
                '#tabexmpl',
                'a:has-text("가입예시")',
                '[role="tab"]:has-text("가입예시")',
                'button:has-text("가입예시")'
            ]
            
            for selector in selectors:
                try:
                    if element := await self.page.wait_for_selector(selector, timeout=5000):
                        await element.click()
                        logger.info(f"가입예시 탭 클릭 성공 (selector: {selector})")
                        await self.page.wait_for_load_state('networkidle')
                        await self.page.wait_for_timeout(2000)  # 추가 대기
                        return True
                except Exception:
                    continue
            
            logger.warning("가입예시 탭을 찾을 수 없음")
            return False
            
        except Exception as e:
            logger.error(f"가입예시 탭 클릭 실패: {str(e)}")
            return False

    async def _ensure_content_loaded(self):
        """페이지의 모든 콘텐츠가 로드되었는지 확인"""
        try:
            # 네트워크 유휴 상태 대기
            await self.page.wait_for_load_state('networkidle', timeout=10000)
            
            # 주요 컨텐츠 요소 대기
            selectors = [
                'table',
                '.example-section',
                '#tabPanel4',
                '[class*="content"]',
                '[class*="table"]'
            ]
            
            for selector in selectors:
                try:
                    await self.page.wait_for_selector(selector, timeout=5000)
                    logger.debug(f"컨텐츠 요소 발견: {selector}")
                except:
                    continue
            
            # JavaScript 실행 완료 대기
            await self.page.wait_for_function("""
                () => {
                    return !document.querySelector('.loading') && 
                           !document.querySelector('[class*="spinner"]') &&
                           !document.querySelector('[class*="loading"]');
                }
            """, timeout=5000)
            
            return True
            
        except Exception as e:
            logger.warning(f"콘텐츠 로딩 확인 중 오류: {str(e)}")
            return False

    async def _find_end_comment(self) -> bool:
        """'가입예시 본문내용 끝' 주석 찾기"""
        try:
            found = await self.page.evaluate("""
                () => {
                    const walker = document.createTreeWalker(
                        document.body,
                        NodeFilter.SHOW_COMMENT,
                        null,
                        false
                    );
                    let comment;
                    while (comment = walker.nextNode()) {
                        if (comment.nodeValue.includes('가입예시 본문내용 끝')) {
                            return true;
                        }
                    }
                    return false;
                }
            """)
            
            if found:
                logger.info("'가입예시 본문내용 끝' 주석 발견")
                return True
            else:
                logger.warning("'가입예시 본문내용 끝' 주석을 찾을 수 없습니다")
                return False
                
        except Exception as e:
            logger.error(f"주석 확인 중 오류: {str(e)}")
            return False

    async def _get_example_content_section(self) -> Optional[str]:
        """가입예시 섹션 내용 추출"""
        try:
            # 스크롤 완료 후 전체 HTML 가져오기
            if not await self._scroll_page():
                return None
                
            # 주석 확인
            if not await self._find_end_comment():
                logger.warning("끝 주석을 찾을 수 없어 전체 내용을 검색합니다")
            
            full_html = await self.page.content()
            soup = BeautifulSoup(full_html, 'html.parser')
            
            # HTML 주석 찾기
            start_comment = None
            end_comment = None
            
            for comment in soup.find_all(string=lambda text: isinstance(text, Comment)):
                comment_text = str(comment).strip()
                if "가입예시 본문내용 시작" in comment_text:
                    start_comment = comment
                    logger.info("가입예시 시작 주석 발견")
                elif "가입예시 본문내용 끝" in comment_text:
                    end_comment = comment
                    logger.info("가입예시 끝 주석 발견")
            
            if start_comment:
                content = []
                current = start_comment.next_element
                
                while current and (not end_comment or current != end_comment):
                    if current.name == 'table':
                        content.append(str(current))
                    current = current.next_element
                    
                if content:
                    logger.info(f"{len(content)}개의 테이블 발견")
                    return '\n'.join(content)
            
            # 대체 방법으로 시도
            logger.warning("주석 기반 검색 실패, 대체 방법으로 시도")
            
            # tabPanel4 내부의 테이블 찾기
            tab_panel = soup.find('div', {'id': 'tabPanel4'})
            if tab_panel:
                tables = tab_panel.find_all('table')
                if tables:
                    logger.info(f"tabPanel4에서 {len(tables)}개의 테이블 발견")
                    return '\n'.join(str(table) for table in tables)
            
            return None

        except Exception as e:
            logger.error(f"가입예시 섹션 추출 실패: {str(e)}")
            return None

    def _parse_html_to_df(self, html: str) -> List[pd.DataFrame]:
        """HTML을 DataFrame으로 변환"""
        try:
            html_io = StringIO(f"<table>{html}</table>")
            return pd.read_html(html_io)
        except Exception as e:
            logger.debug(f"HTML 파싱 실패: {str(e)}")
            return []

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """DataFrame 정제"""
        try:
            # MultiIndex 처리
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [' - '.join(str(level) for level in col if pd.notna(level)).strip() 
                            for col in df.columns]
            
            # 컬럼명 정리
            df.columns = [str(col).strip() for col in df.columns]
            
            # NaN 처리
            df = df.replace([np.nan, 'nan', 'NaN', 'NULL', 'null', ''], None)
            
            # 빈 행/열 제거
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            # 데이터 타입 처리
            for col in df.columns:
                try:
                    df[col] = df[col].astype(str)
                except:
                    continue
            
            return df.reset_index(drop=True)
            
        except Exception as e:
            logger.error(f"DataFrame 정제 실패: {str(e)}")
            return df

    def _is_example_table(self, df: pd.DataFrame) -> bool:
        """가입예시 관련 표인지 확인 - 개선된 버전"""
        example_keywords = [
            # 기존 키워드
            '가입금액', '보험료', '보장내용', '가입나이',
            '보험기간', '납입기간', '갱신형', '비갱신형',
            '담보명', '보장금액', '보험가입금액',
            # 추가된 키워드
            '연령', '남자', '여자', '보장보험료', '예시',
            '나이', '세', '원', '기간', '납입'
        ]
        
        try:
            # 컬럼명 확인
            columns_text = ' '.join([str(col).lower() for col in df.columns])
            if any(keyword.lower() in columns_text for keyword in example_keywords):
                return True
            
            # 데이터 확인 (전체 데이터로 확장)
            all_data = []
            for row in df.values:
                all_data.extend([str(cell).lower() for cell in row])
            all_data_text = ' '.join(all_data)
            
            # 숫자와 '원' 또는 '세'가 함께 있는지 확인
            has_numbers = bool(re.search(r'\d+', all_data_text))
            has_units = '원' in all_data_text or '세' in all_data_text
            
            if has_numbers and has_units:
                return True
                
            if any(keyword.lower() in all_data_text for keyword in example_keywords):
                return True
            
            # 최소 크기 조건 완화
            if df.shape[0] >= 2 and df.shape[1] >= 2:
                # 숫자 데이터가 충분히 있는지 확인
                numeric_count = sum(1 for cell in all_data if re.search(r'\d+', str(cell)))
                if numeric_count / len(all_data) > 0.3:  # 30% 이상이 숫자인 경우
                    return True
            
            return False
            
        except Exception as e:
            logger.error(f"테이블 검증 중 오류: {str(e)}")
            return False

    def _get_parent_section_title(self, table) -> Optional[str]:
        """테이블이 속한 섹션의 제목 찾기"""
        try:
            current = table
            for _ in range(5):  # 최대 5단계 상위까지 확인
                if not current.parent:
                    break
                current = current.parent
                
                # 제목으로 보이는 요소 찾기
                for tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                    if title := current.find(tag):
                        return title.get_text().strip()
                
                # class나 id로 제목 찾기
                title_classes = ['title', 'header', 'section-title']
                for cls in title_classes:
                    if title := current.find(class_=cls):
                        return title.get_text().strip()
            
            return None
        except:
            return None
    
    async def extract_tables(self) -> List[pd.DataFrame]:
        """테이블 추출 메인 로직"""
        try:
            # 가입예시 탭 클릭
            if not await self._click_example_tab():
                logger.error("가입예시 탭 클릭 실패")
                return []
            
            # 콘텐츠 로딩 확인
            if not await self._ensure_content_loaded():
                logger.warning("페이지 로딩이 완료되지 않았을 수 있습니다")
            
            # 가입예시 섹션 내용 가져오기 (이미 스크롤 포함)
            content = await self._get_example_content_section()
            if not content:
                logger.warning("가입예시 섹션을 찾을 수 없습니다")
                return []

            tables = []
            soup = BeautifulSoup(content, 'html.parser')
            
            # 테이블 태그 추출
            for table in soup.find_all('table'):
                try:
                    # pandas로 테이블 파싱
                    html_io = StringIO(str(table))
                    dfs = pd.read_html(html_io)
                    
                    for df in dfs:
                        if not df.empty and self._is_example_table(df):
                            cleaned_df = self._clean_dataframe(df)
                            if not cleaned_df.empty:
                                # 테이블 위치 정보 추가
                                parent_section = self._get_parent_section_title(table)
                                if parent_section:
                                    cleaned_df.insert(0, '섹션', parent_section)
                                tables.append(cleaned_df)
                                
                except Exception as e:
                    logger.debug(f"테이블 파싱 실패: {str(e)}")
                    continue

            # 결과 저장
            self.extracted_tables = tables
            logger.info(f"총 {len(tables)}개의 테이블이 추출되었습니다")
            return tables

        except Exception as e:
            logger.error(f"테이블 추출 실패: {str(e)}")
            return []

    async def export_to_excel(self, output_filepath: str) -> Tuple[str, bool]:
        """추출된 테이블을 엑셀 파일로 저장 - 모든 데이터를 A열부터 정렬"""
        try:
            if not self.extracted_tables:
                logger.warning("저장할 테이블이 없습니다.")
                return "", False

            # 디렉토리 생성
            os.makedirs(os.path.dirname(output_filepath), exist_ok=True)

            # 엑셀 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = 'Combined_Tables'
            current_row = 1

            # 각 테이블 처리
            for i, df in enumerate(self.extracted_tables, 1):
                if df is not None and not df.empty:
                    # 테이블 구분선 및 정보 추가
                    ws.cell(row=current_row, column=1, value='=' * 100)
                    current_row += 1
                    ws.cell(row=current_row, column=1, value=f'Table {i} - Rows: {df.shape[0]}, Columns: {df.shape[1]}')
                    current_row += 1
                    ws.cell(row=current_row, column=1, value='-' * 100)
                    current_row += 1

                    # 컬럼 헤더 작성
                    for col_idx, column_name in enumerate(df.columns, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=str(column_name))
                        cell.font = Font(bold=True)
                    current_row += 1

                    # 데이터 작성
                    for _, row in df.iterrows():
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=str(value))
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                        current_row += 1

                    # 테이블 간 간격 추가
                    current_row += 2

            # 열 너비 자동 조정
            column_widths = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        current_width = len(str(cell.value))
                        column_letter = get_column_letter(cell.column)
                        current_max = column_widths.get(column_letter, 0)
                        column_widths[column_letter] = max(current_max, min(current_width + 2, 50))

            for column_letter, width in column_widths.items():
                ws.column_dimensions[column_letter].width = width

            # 모든 셀에 정렬 적용
            for row in ws.rows:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # 파일 저장
            output_filepath = os.path.abspath('data/output/web_tables.xlsx')
            wb.save(output_filepath)
            logger.info(f"Excel 파일 저장 완료: {output_filepath}")
            return output_filepath, True

        except Exception as e:
            logger.error(f"Excel 저장 실패: {str(e)}")
            return "", False


async def main():
    from playwright.async_api import async_playwright

    url = "https://www.kbinsure.co.kr/CG302230001.ec"

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        page = await context.new_page()

        try:
            await page.goto(url)
            await page.wait_for_load_state('networkidle')

            extractor = ImprovedTableExtractor(page)
            tables = await extractor.extract_tables()

            if tables:
                # 파일 경로를 직접 지정
                output_filepath = 'data/output/web_tables.xlsx'
                filepath, success = await extractor.export_to_excel(output_filepath=output_filepath)
                if success:
                    logger.info(f"테이블 추출 완료: {filepath}")
                else:
                    logger.error("테이블 저장 실패")
            else:
                logger.warning("추출된 테이블이 없습니다")

        except Exception as e:
            logger.error(f"실행 오류: {str(e)}")
        finally:
            await browser.close()


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
