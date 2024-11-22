# src/extractors/pdf_extractor.py

import camelot
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
from pathlib import Path
from datetime import datetime
import logging
import re
import fitz
import PyPDF2
from typing import List, Dict, Tuple

# 상수 정의
TARGET_PAGES = {
    'subscription': '상품제안서_가입담보(계약자용)',
    'subscription_table': '피보험자님의 가입내용',
    'summary': '상품제안서_요약해약환급금예시(계약자용)'
}

# 로깅 설정
BASE_DIR = Path("D:/github/pdf_local")
OUTPUT_DIR = BASE_DIR / "data/output"
log_file = OUTPUT_DIR / f'pdf_table_extraction_{datetime.now().strftime("%Y%m%d")}.log'

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def process_insurance_document(pdf_path: str) -> str:
    """PDF 문서 처리 및 결과 Excel 파일 생성"""
    try:
        output_dir = Path("data/output")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"PDF 파일 처리 시작: {pdf_path}")
        extractor = PDFTableExtractor(pdf_path)
        target_pages = extractor.find_target_pages()
        tables_data = {}
        
        # 가입담보 표 추출
        if 'subscription' in target_pages:
            page_num = target_pages['subscription']
            logger.info(f"가입담보 표 추출 시도 (페이지: {page_num})")
            tables = extractor.extract_tables(page_num, 'subscription')
            if tables:
                tables_data[page_num] = {
                    'tables': tables,
                    'type': 'subscription'
                }
                logger.info(f"가입담보 표 추출 성공: {len(tables)}개 표 발견")
        
        # 해약환급금 표 추출
        if 'summary' in target_pages:
            for page_num in target_pages['summary']:
                logger.info(f"해약환급금 표 추출 시도 (페이지: {page_num})")
                tables = extractor.extract_tables(page_num, 'refund')
                if tables:
                    tables_data[page_num] = {
                        'tables': tables,
                        'type': 'refund'
                    }
                    logger.info(f"해약환급금 표 추출 성공: {len(tables)}개 표 발견")

        if tables_data:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"pdf_extracted_{timestamp}.xlsx"
            extractor.save_to_excel(tables_data, str(output_path))
            logger.info(f"Excel 파일 생성 완료: {output_path}")
            return str(output_path)
        else:
            logger.error("추출된 표가 없습니다.")
            return ""

    except Exception as e:
        logger.error(f"PDF 처리 중 오류 발생: {str(e)}")
        return ""


class PDFTableExtractor:
    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        self.refund_columns = ['경과기간', '납입보험료', '예상해약환급금', '예상해약환급률']

    def find_target_pages(self) -> dict:
        """마커를 기준으로 목표 페이지 찾기"""
        target_pages = {}
        try:
            print("\n=== PDF 페이지 분석 중... ===")
            for page_num in range(len(self.doc)):
                text = self.doc[page_num].get_text()
                
                # 가입담보 페이지 찾기
                if PAGE_MARKERS['subscription'] in text and PAGE_MARKERS['subscription_table'] in text:
                    target_pages['subscription'] = page_num + 1
                    print(f"[가입담보] 페이지 발견: {page_num + 1}페이지")
                    logger.info(f"가입담보 페이지 발견: {page_num + 1}페이지")
                
                # 요약해약환급금 페이지 찾기
                if PAGE_MARKERS['summary'] in text:
                    target_pages['summary'] = [page_num + 1, page_num + 2]  # 현재 페이지와 다음 페이지
                    print(f"[요약해약환급금] 페이지 발견: {page_num + 1}, {page_num + 2}페이지")
                    logger.info(f"요약해약환급금 페이지 발견: {page_num + 1}, {page_num + 2}페이지")
            
            if not target_pages:
                print("\n[경고] 목표 페이지를 찾을 수 없습니다.")
                logger.warning("목표 페이지를 찾을 수 없습니다.")
            
            return target_pages
                
        except Exception as e:
            logger.error(f"페이지 검색 중 오류 발생: {str(e)}")
            return {}

    def process_refund_table(self, df):
        """해약환급금 표 처리"""
        try:
            if len(df.columns) >= len(self.refund_columns):
                df.columns = self.refund_columns

            # 헤더 행 제거
            df = df[~df.iloc[:, 0].str.contains('경과기간|해약환급금', na=False)]
            
            # 보험나이 행 처리
            df.loc[df.iloc[:, 0].str.contains('보험나이', na=False), '경과기간'] = '(보험나이)'

            # 경과기간 정리
            def clean_period(x):
                if pd.isna(x) or '보험나이' in str(x):
                    return x
                match = re.search(r'(\d+)[년개월]', str(x))
                if match:
                    num = match.group(1)
                    unit = '년' if '년' in str(x) else '개월'
                    return f'{num}{unit}'
                return x

            df['경과기간'] = df['경과기간'].apply(clean_period)

            # 금액 컬럼 처리
            for col in ['납입보험료', '예상해약환급금']:
                df[col] = df[col].apply(lambda x: 
                    re.sub(r'[^\d.]', '', str(x)) if pd.notnull(x) and str(x).strip() != '' else '')
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna('')

            # 환급률 처리
            df['예상해약환급률'] = df['예상해약환급률'].apply(lambda x: 
                str(x).replace('%', '').strip() if pd.notnull(x) and str(x).strip() != '' else '')
            
            # 빈 행 제거
            df = df[~df.apply(lambda x: all(str(val).strip() == '' for val in x), axis=1)]
            
            return df

        except Exception as e:
            logger.error(f"해약환급금 표 처리 중 오류 발생: {str(e)}")
            return df

    def clean_subscription_table(self, df):
        """가입담보 표 정제"""
        try:
            # 헤더 행 제거
            df = df[~df.apply(lambda x: x.astype(str).str.contains('피보험자님의 가입내용').any(), axis=1)]
            
            # 첫 번째 열의 번호 제거
            if not df.empty and len(df.columns) > 0:
                df.iloc[:, 0] = df.iloc[:, 0].apply(lambda x: re.sub(r'^\d+\.*\s*', '', str(x)))
            
            # 불필요한 공백 제거
            df = df.apply(lambda x: x.str.strip() if pd.notnull(x) else x)
            
            return df
        except Exception as e:
            logger.error(f"가입담보 표 정제 중 오류 발생: {str(e)}")
            return df

    def extract_tables(self, page_num, table_type='subscription'):
        """표 추출 함수"""
        try:
            # 먼저 자동 컬럼 감지로 시도
            tables = camelot.read_pdf(
                self.pdf_path,
                pages=str(page_num),
                flavor='stream',
                edge_tol=50,
                row_tol=10,
                split_text=True
            )
            
            # 표가 제대로 추출되지 않은 경우 lattice 모드로 재시도
            if not tables or (table_type == 'refund' and len(tables) < 2):
                tables_lattice = camelot.read_pdf(
                    self.pdf_path,
                    pages=str(page_num),
                    flavor='lattice'
                )
                if tables_lattice:
                    tables = tables_lattice
            
            if tables:
                logger.info(f"{page_num} 페이지에서 {len(tables)} 개의 표 추출 성공")
                processed_tables = []
                
                for idx, table in enumerate(tables):
                    df = table.df
                    
                    # 유효한 표인지 확인
                    if len(df.columns) >= 3 and len(df) > 1:  # 최소 조건 체크
                        if table_type == 'refund':
                            # 해약환급금 표 여부 확인
                            if any(df.apply(lambda x: x.astype(str).str.contains('환급|납입|경과').any(), axis=1)):
                                df = self.process_refund_table(df)
                                processed_tables.append(df)
                        else:
                            df = self.clean_subscription_table(df)
                            processed_tables.append(df)
                
                return processed_tables
            else:
                logger.warning(f"{page_num} 페이지에서 표를 찾을 수 없습니다.")
                return None
                
        except Exception as e:
            logger.error(f"{page_num} 페이지에서 표 추출 중 오류 발생: {str(e)}")
            return None

    def save_to_excel(self, tables_data, output_path):
        """결과를 Excel 파일로 저장"""
        wb = Workbook()
        ws = wb.active
        ws.title = '보험 계약 정보'
        current_row = 1

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for page_num, page_info in tables_data.items():
            tables = page_info['tables']
            table_type = page_info['type']
            
            # 섹션 제목 설정
            section_title = "피보험자 가입내용" if table_type == 'subscription' else "요약해약환급금예시"
            header_cell = ws.cell(row=current_row, column=1, value=section_title)
            header_cell.font = Font(bold=True, size=12)
            header_cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            current_row += 2

            for df in tables:
                # 컬럼 헤더
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                current_row += 1

                # 데이터 입력
                for _, row in df.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
                    current_row += 1

                current_row += 2

            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        logger.info(f"Excel 파일 저장 완료: {output_path}")

def process_insurance_documents():
    """보험 문서 처리 메인 함수"""
    try:
        # PDF 파일 찾기
        pdf_files = list(INPUT_DIR.glob("*.[Pp][Dd][Ff]"))
        if not pdf_files:
            logger.error("PDF 파일을 찾을 수 없습니다.")
            return [], []

        results = []
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        for pdf_file in pdf_files:
            file_name = pdf_file.name.lower()
            gender = "male" if '남자' in file_name else "female" if '여자' in file_name else "unknown"
            
            if gender != "unknown":
                logger.info(f"{gender.upper()} PDF 파일 처리 중: {pdf_file.name}")
                
                extractor = PDFTableExtractor(str(pdf_file))
                target_pages = extractor.find_target_pages()
                tables_data = {}
                
                # 가입담보 표 추출
                if 'subscription' in target_pages:
                    page_num = target_pages['subscription']
                    tables = extractor.extract_tables(page_num, 'subscription')
                    if tables:
                        tables_data[page_num] = {
                            'tables': tables,
                            'type': 'subscription'
                        }
                
                # 해약환급금 표 추출
                if 'summary' in target_pages:
                    for page_num in target_pages['summary']:
                        tables = extractor.extract_tables(page_num, 'refund')
                        if tables:
                            tables_data[page_num] = {
                                'tables': tables,
                                'type': 'refund'
                            }

                if tables_data:
                    output_path = OUTPUT_DIR / f"pdf_extracted_{gender}_{timestamp}.xlsx"
                    extractor.save_to_excel(tables_data, str(output_path))
                    results.append((gender, str(output_path)))
                    logger.info(f"{gender.upper()} 데이터 Excel 파일 생성 완료: {output_path}")

        return results

    except Exception as e:
        logger.error(f"문서 처리 중 오류 발생: {str(e)}")
        return []

def main():
    try:
        # 출력 디렉토리 생성
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        print("\n=== PDF 보험계약 정보 추출 시작 ===")
        logger.info("PDF 데이터 추출 시작")
        
        results = process_insurance_documents()
        
        if results:
            print("\n=== 처리 완료 ===")
            for gender, output_path in results:
                print(f"{gender.upper()} 데이터 Excel 파일 생성 완료: {output_path}")
        else:
            print("\n처리할 PDF 파일을 찾을 수 없거나 처리 중 오류가 발생했습니다.")
            print("로그 파일을 확인해주세요.")

    except Exception as e:
        logger.error(f"실행 중 오류 발생: {str(e)}")
        print(f"\n오류 발생: {str(e)}")
        print("자세한 내용은 로그 파일을 확인해주세요.")
        
def cleanup():
    """리소스 정리"""
    try:
        # 임시 파일 정리
        temp_files = list(OUTPUT_DIR.glob("*.tmp"))
        for temp_file in temp_files:
            try:
                temp_file.unlink()
            except Exception as e:
                logger.error(f"임시 파일 삭제 중 오류 발생: {str(e)}")
                
        # 로그 파일 관리 (30일 이상 된 로그 파일 삭제)
        log_files = list(OUTPUT_DIR.glob("pdf_table_extraction_*.log"))
        current_time = datetime.now()
        for log_file in log_files:
            try:
                file_time = datetime.fromtimestamp(log_file.stat().st_mtime)
                if (current_time - file_time).days > 30:
                    log_file.unlink()
            except Exception as e:
                logger.error(f"로그 파일 정리 중 오류 발생: {str(e)}")
                
    except Exception as e:
        logger.error(f"리소스 정리 중 오류 발생: {str(e)}")

if __name__ == "__main__":
    try:
        main()
    finally:
        cleanup()
