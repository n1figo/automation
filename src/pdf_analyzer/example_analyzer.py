import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict
import logging

# 상수 정의
TARGET_PAGES = {
    'subscription': '상품제안서_가입담보',
    'summary': '상품제안서_요약해약환급금예시'
}

logger = logging.getLogger(__name__)

class ExamplePDFAnalyzer:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.logger = logging.getLogger(__name__)
        self.pdf_document = None
        self.load_pdf()

    def load_pdf(self):
        """PDF 문서를 로드하는 메서드"""
        try:
            self.pdf_document = pdfplumber.open(self.pdf_path)
            self.logger.info(f"PDF 문서 로드 성공: {self.pdf_path}")
        except Exception as e:
            self.logger.error(f"PDF 문서 로드 중 오류 발생: {str(e)}")
            self.pdf_document = None

    def __del__(self):
        """소멸자: PDF 문서 닫기"""
        try:
            if self.pdf_document is not None:
                self.pdf_document.close()
        except Exception as e:
            self.logger.error(f"PDF 문서 닫기 중 오류 발생: {str(e)}")

    def get_total_pages(self) -> Optional[int]:
        """PDF 문서의 총 페이지 수를 반환"""
        try:
            if self.pdf_document is not None:
                return len(self.pdf_document.pages)
            else:
                self.logger.error("PDF 문서가 로드되지 않았습니다.")
                return None
        except Exception as e:
            self.logger.error(f"총 페이지 수를 가져오는 중 오류 발생: {str(e)}")
            return None

    def find_target_pages(self) -> Dict[str, int]:
        """특정 제목에 해당하는 페이지를 찾기"""
        target_pages = {}
        try:
            if self.pdf_document is None:
                raise ValueError("PDF 문서가 로드되지 않았습니다.")

            for page_num, page in enumerate(self.pdf_document.pages):
                text = page.extract_text()
                if text is None:
                    continue

                normalized_text = text.replace(" ", "")
                
                if "상품제안서_가입담보" in normalized_text:
                    target_pages['subscription'] = page_num + 1
                    self.logger.info(f"[상품제안서_가입담보] 페이지 발견: {page_num + 1}페이지")

                if "상품제안서_요약해약환급금예시" in normalized_text:
                    target_pages['summary'] = page_num + 1
                    self.logger.info(f"[상품제안서_요약해약환급금예시] 페이지 발견: {page_num + 1}페이지")

            return target_pages
        except Exception as e:
            self.logger.error(f"페이지 검색 중 오류 발생: {str(e)}")
            return {}

    def extract_tables(self, start_page: int, end_page: Optional[int] = None) -> List[pd.DataFrame]:
        """PDF에서 테이블을 추출"""
        tables = []
        try:
            if self.pdf_document is None:
                raise ValueError("PDF 문서가 로드되지 않았습니다.")

            if end_page is None:
                end_page = len(self.pdf_document.pages)

            for page_num in range(start_page - 1, end_page):  # 1-based to 0-based
                try:
                    page = self.pdf_document.pages[page_num]
                    tables_on_page = page.extract_tables()
                    
                    for table in tables_on_page:
                        if table and len(table) > 1:  # 헤더와 데이터가 있는 경우만
                            df = pd.DataFrame(table[1:], columns=table[0])
                            # 빈 행과 열 제거
                            df = df.dropna(how='all').dropna(axis=1, how='all')
                            if not df.empty:
                                tables.append(df)
                                
                except Exception as e:
                    self.logger.error(f"{page_num + 1}페이지 테이블 추출 중 오류: {str(e)}")
                    continue

            if not tables:
                self.logger.warning("추출된 테이블이 없습니다.")
            else:
                self.logger.info(f"{len(tables)}개의 표를 추출했습니다.")

            return tables

        except Exception as e:
            self.logger.error(f"테이블 추출 중 오류 발생: {str(e)}")
            return []

    def save_to_excel(self, tables_data: Dict[str, List[pd.DataFrame]]) -> str:
        """결과를 Excel 파일로 저장"""
        try:
            output_dir = Path("data/output")
            output_dir.mkdir(parents=True, exist_ok=True)
            
            wb = Workbook()
            ws = wb.active
            ws.title = '보험 계약 정보'
            current_row = 1

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for section, tables in tables_data.items():
                # 섹션 제목
                section_title = f"{section} 표"
                header_cell = ws.cell(row=current_row, column=1, value=section_title)
                header_cell.font = Font(bold=True, size=12)
                header_cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                current_row += 2

                for df in tables:
                    # 컬럼 헤더
                    for col_idx, col_name in enumerate(df.columns, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
                    current_row += 1

                    # 데이터
                    for _, row in df.iterrows():
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=value)
                            cell.border = border
                            cell.alignment = Alignment(wrap_text=True, vertical='center')
                        current_row += 1

                    current_row += 2

            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            output_path = output_dir / f"pdf_extracted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(output_path)
            self.logger.info(f"Excel 파일 저장 완료: {output_path}")
            return str(output_path)

        except Exception as e:
            self.logger.error(f"Excel 저장 중 오류 발생: {str(e)}")
            raise

    def analyze(self) -> str:
        """PDF 분석 실행"""
        try:
            target_pages = self.find_target_pages()
            if not target_pages:
                return "대상 페이지를 찾을 수 없습니다."

            tables_data = {}

            # 가입담보 페이지 처리
            if 'subscription' in target_pages:
                start_page = target_pages['subscription']
                tables = self.extract_tables(start_page, start_page + 1)
                if tables:
                    tables_data['상품제안서_가입담보'] = tables

            # 요약해약환급금예시 페이지 처리
            if 'summary' in target_pages:
                start_page = target_pages['summary']
                tables = self.extract_tables(start_page, start_page + 1)
                if tables:
                    tables_data['상품제안서_요약해약환급금예시'] = tables

            if tables_data:
                return self.save_to_excel(tables_data)
            else:
                return "No tables found for the given pages."

        except Exception as e:
            self.logger.error(f"분석 중 오류 발생: {str(e)}")
            return f"분석 중 오류 발생: {str(e)}"