from typing import Dict, List, Union, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

class ExcelProcessor:
    def __init__(self):
        self.header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def process_tables_to_excel(self, tables_data: Dict[str, Union[pd.DataFrame, dict]], output_path: str) -> bool:
        """
        테이블 데이터를 Excel 파일로 저장
        
        Args:
            tables_data: 테이블 데이터 (DataFrame 또는 dict)
            output_path: 저장할 Excel 파일 경로
            
        Returns:
            bool: 저장 성공 여부
        """
        try:
            wb = Workbook()
            wb.remove(wb.active)  # 기본 시트 제거
            
            for sheet_name, data in tables_data.items():
                ws = wb.create_sheet(sheet_name)
                
                if isinstance(data, pd.DataFrame):
                    self._write_dataframe(ws, data)
                elif isinstance(data, dict):
                    self._write_nested_data(ws, data)
                
                # 열 너비 자동 조정
                self._adjust_column_widths(ws)
            
            wb.save(output_path)
            logger.info(f"Excel 파일 저장 완료: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel 파일 저장 중 오류 발생: {str(e)}")
            return False
    
    def _write_dataframe(self, ws, df: pd.DataFrame, start_row: int = 1):
        """DataFrame을 워크시트에 작성"""
        # 헤더 작성
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 데이터 작성
        for r_idx, row in enumerate(df.values, start_row + 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    def _write_nested_data(self, ws, data: dict, start_row: int = 1):
        """중첩된 딕셔너리 데이터를 워크시트에 작성"""
        current_row = start_row
        
        for section, tables in data.items():
            # 섹션 제목
            cell = ws.cell(row=current_row, column=1, value=section)
            cell.font = Font(bold=True, size=12)
            cell.fill = self.header_fill
            current_row += 2
            
            if isinstance(tables, pd.DataFrame):
                self._write_dataframe(ws, tables, current_row)
                current_row += len(tables) + 3
            elif isinstance(tables, dict):
                for sub_section, df in tables.items():
                    if isinstance(df, pd.DataFrame):
                        # 하위 섹션 제목
                        cell = ws.cell(row=current_row, column=1, value=sub_section)
                        cell.font = Font(bold=True)
                        current_row += 1
                        
                        self._write_dataframe(ws, df, current_row)
                        current_row += len(df) + 3
    
    def _adjust_column_widths(self, ws):
        """열 너비 자동 조정"""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # 최대 50
            ws.column_dimensions[column].width = adjusted_width

    def create_refined_excel(self, input_path: str, output_path: str) -> str:
        """
        보장내용 엑셀 파일을 정제하고 변경사항 행에 노란색 음영 추가
        
        Args:
            input_path (str): 원본 엑셀 파일 경로
            output_path (str): 정제된 엑셀 파일 경로
        
        Returns:
            str: 생성된 엑셀 파일 경로
        """
        try:
            # 엑셀 파일 로드
            xls = pd.ExcelFile(input_path)
            
            # 결과를 저장할 워크북 생성
            wb = Workbook()
            wb.remove(wb.active)  # 기본 시트 제거
            
            # 각 시트별로 처리
            for sheet_name in xls.sheet_names:
                # 데이터프레임 로드
                df = pd.read_excel(input_path, sheet_name=sheet_name)
                
                # 데이터 정제
                df = self._refine_dataframe(df)
                
                # 워크시트 생성
                ws = wb.create_sheet(title=sheet_name)
                
                # 헤더 작성
                for col, header in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = self.header_fill
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 데이터 작성
                for r_idx, row in enumerate(df.values, 2):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = self.border
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
                        
                        # 변경사항 행에 노란색 음영 추가
                        if '변경사항' in df.columns:
                            change_col_idx = list(df.columns).index('변경사항') + 1
                            if c_idx == change_col_idx and str(value).strip() == '추가':
                                cell.fill = PatternFill(
                                    start_color='FFFF00',  # 노란색
                                    end_color='FFFF00', 
                                    fill_type='solid'
                                )
                
                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # 엑셀 파일 저장
            wb.save(output_path)
            
            logger.info(f"정제된 엑셀 파일 생성 완료: {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"엑셀 파일 정제 중 오류 발생: {str(e)}")
            raise

    def _refine_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """데이터프레임 정제 메서드"""
        try:
            # 빈 행 제거
            df = df.dropna(how='all')
            
            # 중복 행 제거
            df = df.drop_duplicates()
            
            # 문자열 컬럼 정제
            for col in df.columns:
                if df[col].dtype == object:
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].str.replace('\n', ' ')
                    df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
            
            return df
        
        except Exception as e:
            logger.error(f"데이터프레임 정제 중 오류 발생: {str(e)}")
            return df
