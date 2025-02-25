from typing import List, Dict, Optional, Tuple
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import re
import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from src.analyzers.pdf_analyzer import PDFAnalyzer
from src.analyzers.example_analyzer import ExamplePDFAnalyzer
from src.extractors.html_extractor import HTMLFileExtractor
import os


class ExcelWriter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        self.current_row = 1
        
        # 스타일 설정
        self.header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def write_table(self, df: pd.DataFrame, title: str, sheet_name: Optional[str] = None) -> None:
        """테이블 작성"""
        if sheet_name:
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
            else:
                ws = self.wb.create_sheet(title=sheet_name)
        else:
            ws = self.wb.active

        # 제목 작성
        cell = ws.cell(row=self.current_row, column=1, value=title)
        cell.font = Font(bold=True, size=12)
        cell.fill = self.header_fill
        self.current_row += 2

        # 컬럼 헤더
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=self.current_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(wrap_text=True)
        self.current_row += 1

        # 데이터
        for _, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=self.current_row, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True)
            self.current_row += 1
        
        self.current_row += 2

    def save(self) -> None:
        """Excel 파일 저장"""
        # 열 너비 자동 조정
        for ws in self.wb.worksheets:
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        self.wb.save(self.output_path)


class PDFAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("PDF 보험약관 분석기")
        self.root.geometry("800x600")
        self.setup_logging()
        self.setup_gui()

    def setup_logging(self):
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        
        log_filename = log_dir / f'pdf_analyzer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def setup_gui(self):
        """GUI 설정"""
        # 메인 프레임
        self.main_frame = ttk.Frame(self.root, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # 제목
        title_label = ttk.Label(
            self.main_frame,
            text="KB손해보험 상품개정 자동화 서비스",
            font=("Helvetica", 16, "bold")
        )
        title_label.pack(pady=10)

        # 1. 보장내용 PDF 파일 택 프레임
        coverage_frame = ttk.LabelFrame(self.main_frame, text="보장내용 PDF 선택", padding="10")
        coverage_frame.pack(fill=tk.X, pady=5)

        self.coverage_path_var = tk.StringVar()
        coverage_entry = ttk.Entry(coverage_frame, textvariable=self.coverage_path_var, width=60)
        coverage_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        coverage_button = ttk.Button(
            coverage_frame,
            text="찾아보기",
            command=lambda: self.browse_file(self.coverage_path_var, [("PDF files", "*.pdf")])
        )
        coverage_button.pack(side=tk.LEFT, padx=5)

        # 2. 가입예시 PDF 파일 선택 프레임 (파일 선택으로 유지)
        example_pdf_frame = ttk.LabelFrame(self.main_frame, text="가입예시 PDF 선택", padding="10")
        example_pdf_frame.pack(fill=tk.X, pady=5)

        self.example_pdf_path_var = tk.StringVar()
        example_pdf_entry = ttk.Entry(example_pdf_frame, textvariable=self.example_pdf_path_var, width=60)
        example_pdf_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        example_pdf_button = ttk.Button(
            example_pdf_frame,
            text="파일 선택",
            command=lambda: self.browse_file(self.example_pdf_path_var, [("PDF files", "*.pdf")])  # 파일 선택
        )
        example_pdf_button.pack(side=tk.LEFT, padx=5)

        # MHTML 파일 선택 프레임 추가
        mhtml_frame = ttk.LabelFrame(self.main_frame, text="MHTML 파일 선택", padding="10")
        mhtml_frame.pack(fill=tk.X, pady=5)

        self.mhtml_path_var = tk.StringVar()
        mhtml_entry = ttk.Entry(mhtml_frame, textvariable=self.mhtml_path_var, width=60)
        mhtml_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        mhtml_button = ttk.Button(
            mhtml_frame,
            text="MHTML 파일 선택",
            command=lambda: self.browse_file(self.mhtml_path_var, [("MHTML files", "*.mhtml")])  # MHTML 파일 선택
        )
        mhtml_button.pack(side=tk.LEFT, padx=5)

        # 처리 옵션 레임
        options_frame = ttk.LabelFrame(self.main_frame, text="처리 옵션", padding="10")
        options_frame.pack(fill=tk.X, pady=5)

        # 체크박스들을 가로로 배치할 프레임
        checkbox_frame = ttk.Frame(options_frame)
        checkbox_frame.pack(fill=tk.X, padx=5, pady=5)

        # 체크박스 변수들
        self.coverage_analysis_var = tk.BooleanVar(value=True)
        self.example_analysis_var = tk.BooleanVar(value=True)

        # 체크박스들
        ttk.Checkbutton(
            checkbox_frame, 
            text="보장내용 분석",
            variable=self.coverage_analysis_var
        ).pack(side=tk.LEFT, padx=20)

        ttk.Checkbutton(
            checkbox_frame, 
            text="가입예시 분석",
            variable=self.example_analysis_var
        ).pack(side=tk.LEFT, padx=20)

        # 진행 상태 표시
        self.progress_var = tk.StringVar(value="대기 중...")
        progress_label = ttk.Label(self.main_frame, textvariable=self.progress_var)
        progress_label.pack(pady=5)

        self.progress_bar = ttk.Progressbar(self.main_frame, mode='determinate', length=300)
        self.progress_bar.pack(pady=5)

        # 처리 시작 버튼
        self.process_button = ttk.Button(
            self.main_frame,
            text="분석 시작",
            command=self.process_start,
            state=tk.DISABLED
        )
        self.process_button.pack(pady=10)

        # 로그 영역
        log_frame = ttk.LabelFrame(self.main_frame, text="처리 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = tk.Text(log_frame, height=15)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def check_process_button_state(self):
        """분석 시작 버튼 활성화 상태 확인"""
        if (self.coverage_path_var.get() or 
            self.example_pdf_path_var.get()):
            self.process_button['state'] = tk.NORMAL
        else:
            self.process_button['state'] = tk.DISABLED

    def browse_file(self, string_var, filetypes):
        """파일 탐색기 실행"""
        file_path = filedialog.askopenfilename(title="파일 선택", filetypes=filetypes)  # 파일 선택으로 변경
        if file_path:
            string_var.set(file_path)
            self.check_process_button_state()

    def process_pdf(self):
        """PDF 파일 처리 수행"""
        try:
            # 버튼 비활성화 및 프로그레스바 초기화
            self.process_button['state'] = tk.DISABLED
            self.progress_bar['value'] = 0
            
            pdf_path = self.file_path_var.get()
            if not pdf_path:
                messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
                self.process_button['state'] = tk.NORMAL
                return

            # 분석 작업을 별도 스레드에서 실행
            thread = threading.Thread(target=self.analyze_pdf, args=(pdf_path,))
            thread.daemon = True  # 메인 프로그램 종료 시 스레드도 함께 종료
            thread.start()

        except Exception as e:
            self.log_message(f"처리 시작 중 오류 발생: {str(e)}", "ERROR")
            self.process_button['state'] = tk.NORMAL
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다:\n{str(e)}")

    def analyze_pdf(self, pdf_path: str):
        """PDF 분석 수행"""
        try:
            self.update_progress(10, "PDF 분석 시작...")
            analyzer = PDFAnalyzer(pdf_path, self.logger)
            
            self.update_progress(30, "섹션 분석 중...")
            output_path = analyzer.analyze()
            
            self.update_progress(100, "처리 완료")
            self.log_message(f"분석 완료. 결과 저장 위치: {output_path}")
            messagebox.showinfo("완료", f"분석이 료되습니다.\n저장 위치: {output_path}")

        except Exception as e:
            self.log_message(f"PDF 분석 중 오류 발생: {str(e)}", "ERROR")
            messagebox.showerror("오류", f"분석 중 오류가 발생했니다: {str(e)}")
        finally:
            self.process_button['state'] = tk.NORMAL
            self.progress_var.set("대기 중...")

    def update_progress(self, value, message):
        """진행 상태 업데이트"""
        self.progress_bar['value'] = value
        self.progress_var.set(message)
        self.root.update_idletasks()

    def log_message(self, message, level="INFO"):
        """로그 메시지 출력"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        
        if level == "INFO":
            self.logger.info(message)
        elif level == "ERROR":
            self.logger.error(message)
        elif level == "WARNING":
            self.logger.warning(message)

    def find_sections(self, doc):
        """PDF에서 섹션을 찾아 반환"""
        sections = {}
        
        # 1.  종별 검색
        type_pattern = r'\[(\d)종\]'
        type_pages = []
        
        for page_num in range(len(doc)):
            text = doc[page_num].get_text()
            matches = re.finditer(type_pattern, text)
            for match in matches:
                type_num = match.group(1)
                type_pages.append((page_num, f"[{type_num}종]"))
        
        # 종별 정렬
        type_pages = sorted(type_pages)
        
        if type_pages:
            self.log_message(f"종별 구분 발견: {[t[1] for t in type_pages]}")
            # TODO: 종별 섹션 처리 추가
            sections["종별처리필요"] = (0, len(doc) - 1)  # 임시 처리
        else:
            # 2. 종별 구분이 없는 경우 일반 섹션 ��색
            self.log_message("종별 구분이 없습니다. 일반 섹션 검색을 시작합니다.")
            current_section = None
            section_start = None
            
            for page_num in range(len(doc)):
                text = doc[page_num].get_text()
                
                for section_type, pattern in self.section_patterns.items():
                    if re.search(pattern, text):
                        if current_section:
                            sections[current_section] = (section_start, page_num - 1)
                        current_section = section_type
                        section_start = page_num
                        self.log_message(f"[시작] {section_type}: {page_num + 1}페이지")
                        break
            
            # 마지막 섹션 처리
            if current_section and section_start is not None:
                sections[current_section] = (section_start, len(doc) - 1)
        
        if not sections:
            self.log_message("섹션을 찾을 수 없습니다.", "WARNING")
        
        return sections
    
    def process_start(self):
        """분석 작업 시작"""
        try:
            self.process_button['state'] = tk.DISABLED
            self.progress_bar['value'] = 0
            
            coverage_path = self.coverage_path_var.get()
            example_pdf_path = self.example_pdf_path_var.get()
            
            # 가입예시 PDF만 선택된 경우 처리
            if self.example_analysis_var.get() and example_pdf_path and not coverage_path:
                thread = threading.Thread(
                    target=self.process_files, 
                    args=(coverage_path, self.mhtml_path_var.get(), example_pdf_path)  # MHTML 경로 전달
                )
                thread.start()
                return

            # 기존 직 지: 보장내용 PDF가 선택 우
            if not any([coverage_path, example_pdf_path]):
                messagebox.showerror("오류", "최소한 하나의 파일을 선택해주세요.")
                self.process_button['state'] = tk.NORMAL
                return

            thread = threading.Thread(
                target=self.process_files, 
                args=(coverage_path, None, example_pdf_path)  # HTML path는 None으로 설정
            )
            thread.start()

        except Exception as e:
            self.log_message(f"처리 시작 중 오류 발생: {str(e)}", "ERROR")
            self.process_button['state'] = tk.NORMAL


    # app.py의 HTML 처리 관련 메서드 수정

# app.py의 HTML 처리 관련 메서드 수정

    def process_files(self, coverage_path: Optional[str], mhtml_path: Optional[str], 
                     example_pdf_path: Optional[str]) -> None:
        """파일 처리 작업 수행"""
        try:
            self.update_progress(0, "처리 시작...")
            results = {'web': [], 'coverage': [], 'example': []}

            # MHTML 파일 처리
            if mhtml_path and os.path.exists(mhtml_path):  # mhtml_path가 None이 아닌지 확인
                self.log_message(f"MHTML 파일 경로: {mhtml_path}")  # MHTML 파일 경로 로그
                self.update_progress(20, "웹 페이지 분석 중...")
                try:
                    extractor = HTMLFileExtractor(mhtml_path)  # MHTML 파일 경로 사용
                    web_tables = extractor.extract_tables()
                    
                    # 웹 테이블 추출 결과 로그
                    if web_tables:
                        results['web'] = web_tables
                        self.log_message(f"{len(web_tables)}개의 웹 테이블 추출 완료")
                    else:
                        self.log_message("웹 테이블이 추출되지 않았습니다.", "WARNING")  # 테이블이 없을 경우 경고 로그
                except Exception as e:
                    self.log_message(f"웹 페이지 처리 중 오류: {str(e)}", "ERROR")
                    print(f"오류 발생: {str(e)}")  # 오류 메시지 출력
            else:
                self.log_message("MHTML 파일이 존재하지 않거나 경로가 잘못되었습니다.", "ERROR")  # 파일 경로 오류 로그

            # 보장내용 PDF 처리
            if coverage_path and self.coverage_analysis_var.get():
                self.update_progress(40, "보장내용 PDF 분석 중...")
                try:
                    analyzer = PDFAnalyzer(coverage_path, self.logger)
                    output_path = analyzer.analyze()
                    if output_path:
                        results['coverage'] = output_path
                        self.log_message(f"보장내용 PDF 분석 완료: {output_path}")
                except Exception as e:
                    self.log_message(f"보장내용 PDF 분석 중 오류: {str(e)}", "ERROR")

            # 가입예시 PDF 처리
            if example_pdf_path and self.example_analysis_var.get():
                self.update_progress(60, "가입예시 PDF 분석 중...")
                try:
                    analyzer = ExamplePDFAnalyzer(example_pdf_path)
                    output_path = analyzer.analyze()
                    if output_path:
                        results['example'] = output_path
                        self.log_message(f"가입예시 PDF 분석 완료: {output_path}")
                except Exception as e:
                    self.log_message(f"가입예시 PDF 분석 중 오류: {str(e)}", "ERROR")

            # 결과 저장
            self._save_results(results)

        except Exception as e:
            self.log_message(f"처리 중 오류 발생: {str(e)}", "ERROR")
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다:\n{str(e)}")
        finally:
            self.process_button['state'] = tk.NORMAL
            self.progress_var.set("대기 중...")

    def save_web_tables(self, tables: List[pd.DataFrame], output_path: str) -> None:
        """웹 테이블 저장"""
        try:
            excel_writer = ExcelWriter(output_path)
            
            for idx, df in enumerate(tables, 1):
                # 메타데이터 추가
                df = df.copy()
                df.insert(0, '테이블번호', idx)
                df.insert(0, '데이터출처', 'Web')
                
                # 테이블 저장
                title = f'Table_{idx}'
                excel_writer.write_table(df, title)
            
            excel_writer.save()
            self.log_message(f"웹 테이블을 {output_path}에 저장했습니다.")
            
        except Exception as e:
            self.log_message(f"웹 테이블 저장 중 오류 발생: {str(e)}", "ERROR")
            raise

    def _add_table_metadata(self, df: pd.DataFrame, table_num: int) -> pd.DataFrame:
        """테이블 메타데이터 추가"""
        df = df.copy()
        df.insert(0, '테이블번호', table_num)
        df.insert(0, '데이터출처', 'Web')
        return df


    def run(self):
        """애플리케이션 실행"""
        self.root.mainloop()

    def browse_folder(self, string_var):
        """폴더 탐색기 실행"""
        folder_path = filedialog.askdirectory(title="폴더 선택")
        if folder_path:
            string_var.set(folder_path)
            self.check_process_button_state()

    def browse_mhtml_files(self, folder_path):
        """선택한 폴더 내의 MHTML 파일 찾기"""
        mhtml_files = [f for f in os.listdir(folder_path) if f.endswith('.mhtml')]
        if mhtml_files:
            self.log_message(f"{len(mhtml_files)}개의 MHTML 파일을 찾았습니다.")
            for mhtml_file in mhtml_files:
                full_path = os.path.join(folder_path, mhtml_file)
                extractor = HTMLFileExtractor(full_path)
                tables = extractor.extract_tables()
                if tables:
                    self.log_message(f"{mhtml_file}에서 {len(tables)}개의 테이블을 성공적으로 추출했습니다.")
                else:
                    self.log_message(f"{mhtml_file}에서 테이블 추��에 실패했습니다.", "WARNING")
        else:
            self.log_message("선택한 폴더에 MHTML 파일이 없습니다.", "WARNING")

    def _save_results(self, results: Dict[str, any]) -> None:
        """결과 저장 및 완료 메시지 표시"""
        if not any(results.values()):
            self.log_message("추출된 데이터가 없습니다.", "WARNING")
            messagebox.showwarning("완료", "추출된 데이터가 없습니다.")
            return

        output_dir = Path("data/output")
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        saved_files = []

        # 웹 데이터 저장
        if results['web']:
            web_output = output_dir / f"web_tables_{timestamp}.xlsx"
            self.save_web_tables(results['web'], web_output)
            saved_files.append(str(web_output))

        # PDF 결과 파일 추가
        for key in ['coverage', 'example']:
            if results[key]:
                saved_files.append(results[key])

        self.update_progress(100, "처리 완료")
        result_message = "분석이 완료되었습니다.\n저장된 파일:\n" + "\n".join(saved_files)
        messagebox.showinfo("완료", result_message)
