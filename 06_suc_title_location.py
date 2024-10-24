from langchain_community.document_loaders import PyPDFLoader
import camelot
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import re
import fitz
import numpy as np

def get_titles_with_positions(pdf_path, page_num):
    """제목과 위치 정보 추출"""
    doc = fitz.open(pdf_path)
    page = doc[page_num - 1]
    blocks = page.get_text("dict")["blocks"]
    
    titles = []
    title_patterns = [
        r'기본계약',
        r'의무부가계약',
        r'[\w\s]+관련\s*특약',
        r'[\w\s]+계약',
        r'[\w\s]+보장'
    ]
    
    for block in blocks:
        if block.get("lines"):
            text = ""
            y_top = block["bbox"][1]
            y_bottom = block["bbox"][3]
            
            for line in block["lines"]:
                for span in line["spans"]:
                    text += span["text"] + " "
            
            text = text.strip()
            if text and any(re.search(pattern, text) for pattern in title_patterns):
                titles.append({
                    "text": text,
                    "y_top": y_top,
                    "y_bottom": y_bottom,
                    "bbox": block["bbox"],
                    "used": False
                })
    
    doc.close()
    return sorted(titles, key=lambda x: x["y_top"])

def get_table_positions(table, page_height):
    """Camelot 표의 위치를 PyMuPDF 좌표계로 변환"""
    return {
        "y_top": page_height - table._bbox[3],
        "y_bottom": page_height - table._bbox[1],
        "bbox": (
            table._bbox[0],
            page_height - table._bbox[3],
            table._bbox[2],
            page_height - table._bbox[1]
        )
    }

def match_titles_to_tables(titles, tables, max_distance=50):
    """가장 가까운 미사용 제목을 표와 매칭"""
    matches = []
    
    for table in tables:
        closest_title = None
        min_distance = float('inf')
        
        for title in titles:
            if title["used"]:
                continue
                
            # 표와 제목 사이의 거리 계산
            distance = table["y_top"] - title["y_bottom"]
            
            # 제목이 표보다 위에 있고, 허용 거리 내에 있으며,
            # 지금까지 찾은 것보다 더 가까운 경우
            if 0 < distance < max_distance and distance < min_distance:
                closest_title = title
                min_distance = distance
        
        if closest_title:
            closest_title["used"] = True
            matches.append({
                "title": closest_title["text"],
                "title_bbox": closest_title["bbox"],
                "table_bbox": table["bbox"],
                "distance": min_distance
            })
        else:
            matches.append({
                "title": None,
                "title_bbox": None,
                "table_bbox": table["bbox"],
                "distance": None
            })
    
    return matches

def process_page(pdf_path, page_num):
    """페이지의 표와 제목을 처리"""
    print(f"\n=== {page_num}페이지 처리 시작 ===")
    
    # 제목 추출
    titles = get_titles_with_positions(pdf_path, page_num)
    print("\n발견된 제목들:")
    for title in titles:
        print(f"- {title['text']} (y: {title['y_top']:.1f} - {title['y_bottom']:.1f})")
    
    # 표 추출
    tables = camelot.read_pdf(
        pdf_path,
        pages=str(page_num),
        flavor='lattice'
    )
    if not tables:
        tables = camelot.read_pdf(
            pdf_path,
            pages=str(page_num),
            flavor='stream'
        )
    print(f"\n추출된 표 수: {len(tables)}")
    
    # 페이지 크기 정보 가져오기
    doc = fitz.open(pdf_path)
    page = doc[page_num - 1]
    page_height = page.rect.height
    doc.close()
    
    # 표 위치 정보 추출
    table_positions = [get_table_positions(table, page_height) for table in tables]
    
    # 제목과 표 매칭
    matches = match_titles_to_tables(titles, table_positions)
    
    # 결과 생성
    results = []
    for i, (match, table) in enumerate(zip(matches, tables)):
        title = match["title"] if match["title"] else f"표 {i+1} (제목 없음)"
        
        # 거리 출력 수정
        distance_str = f"{match['distance']:.1f}" if match['distance'] is not None else "N/A"
        print(f"\n표 {i+1}:")
        print(f"- 제목: {title}")
        print(f"- 거리: {distance_str}")
        
        results.append({
            'title': title,
            'table': table.df,
            'page': page_num,
            'title_bbox': match["title_bbox"],
            'table_bbox': match["table_bbox"],
            'distance': match["distance"]
        })
    
    return results

def save_to_excel(results, output_path):
    """추출된 표와 제목을 Excel 파일로 저장"""
    wb = Workbook()
    ws = wb.active
    current_row = 1

    for i, item in enumerate(results, 1):
        # 제목과 위치 정보
        title_cell = ws.cell(row=current_row, column=1, 
                           value=f"{item['title']} (Page: {item['page']})")
        title_cell.font = Font(bold=True, size=12)
        title_cell.fill = PatternFill(start_color='E6E6E6', 
                                    end_color='E6E6E6', 
                                    fill_type='solid')
        
        # 위치 정보 추가 - 수정된 부분
        if item['distance'] is not None:
            distance_str = f"{item['distance']:.1f}"
            ws.cell(row=current_row, column=2,
                   value=f"Distance: {distance_str}")
        else:
            ws.cell(row=current_row, column=2,
                   value="Distance: N/A")
        
        current_row += 2

        # 표 데이터
        df = item['table']
        for r_idx, row in enumerate(df.values):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=current_row + r_idx, 
                             column=c_idx + 1, 
                             value=value)
                cell.alignment = Alignment(wrap_text=True)

        current_row += len(df) + 3

    wb.save(output_path)
    print(f"\n결과가 {output_path}에 저장되었습니다.")

def main():
    # 설정
    pdf_path = "/workspaces/automation/uploads/KB 9회주는 암보험Plus(무배당)(24.05)_요약서_10.1판매_v1.0_앞단.pdf"
    output_path = "test_page59_tables.xlsx"
    test_page = 59

    try:
        # 처리 실행
        results = process_page(pdf_path, test_page)
        
        # Excel 저장
        save_to_excel(results, output_path)
        
        print("\n처리 완료!")
        for i, result in enumerate(results, 1):
            print(f"표 {i}: {result['title']}")
            
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        raise

if __name__ == "__main__":
    main()