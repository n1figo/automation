import PyPDF2
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import os
import pandas as pd
import camelot
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from sentence_transformers import SentenceTransformer
from scipy.spatial.distance import cosine
import cv2
from PIL import Image

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class TitleMatcher:
    def __init__(self):
        self.model = SentenceTransformer('distiluse-base-multilingual-cased-v1')
        
    def get_embedding(self, text):
        return self.model.encode(text)
    
    def calculate_similarity(self, text1, text2):
        emb1 = self.get_embedding(text1)
        emb2 = self.get_embedding(text2)
        return 1 - cosine(emb1, emb2)

class HighlightDetector:
    @staticmethod
    def pdf_to_image(page):
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return np.array(img)

    @staticmethod
    def detect_highlights(image):
        hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
        s = hsv[:,:,1]
        v = hsv[:,:,2]

        saturation_threshold = 30
        saturation_mask = s > saturation_threshold

        _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)

        kernel = np.ones((5,5), np.uint8)
        cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
        cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)

        contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        return contours

    @staticmethod
    def get_highlight_regions(contours, image_height):
        regions = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            top = image_height - (y + h)
            bottom = image_height - y
            regions.append((top, bottom))
        return regions

    @staticmethod
    def check_highlight(row_range, highlight_regions):
        row_top, row_bottom = row_range
        for region_top, region_bottom in highlight_regions:
            if (region_top <= row_top <= region_bottom) or \
               (region_top <= row_bottom <= region_bottom) or \
               (row_top <= region_top <= row_bottom) or \
               (row_top <= region_bottom <= row_bottom):
                return True
        return False

class InsuranceDocumentAnalyzer:
    def __init__(self):
        self.section_patterns = {
            "종류": r'\[(\d)종\]',
            "특약유형": r'(상해관련|질병관련)\s*특약'
        }
        self.section_pages = {"[1종]": None, "[2종]": None, "[3종]": None}
        self.section_ranges = {}
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]

    def find_section_pages(self, pdf_path: str) -> Dict[str, int]:
        """PDF에서 1종, 2종, 3종의 시작 페이지 찾기"""
        try:
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            for page_num in range(total_pages):
                text = doc[page_num].get_text()
                
                matches = re.finditer(self.section_patterns["종류"], text)
                for match in matches:
                    종_type = f"[{match.group(1)}종]"
                    if self.section_pages[종_type] is None:
                        self.section_pages[종_type] = page_num
                        logger.info(f"{종_type} 시작 페이지: {page_num + 1}")

            # 섹션을 찾았는지 확인
            found_sections = [v for v in self.section_pages.values() if v is not None]
            if not found_sections:
                logger.warning("1종, 2종, 3종 패턴을 찾지 못했습니다. 문서 전체를 대상으로 진행합니다.")
                # 문서 전체를 하나의 섹션으로 간주
                self.section_ranges["전체"] = (0, total_pages)
            else:
                # 찾은 섹션들로 범위 설정
                sorted_pages = sorted([(k, v) for k, v in self.section_pages.items() if v is not None], 
                                    key=lambda x: x[1])
                
                for i, (종_type, start_page) in enumerate(sorted_pages):
                    if i + 1 < len(sorted_pages):
                        end_page = sorted_pages[i + 1][1]
                    else:
                        end_page = total_pages
                    self.section_ranges[종_type] = (start_page, end_page)
            doc.close()
            return self.section_pages
            
        except Exception as e:
            logger.error(f"Error finding section pages: {e}")
            return {}

class TableExtractor:
    def __init__(self):
        self.matcher = TitleMatcher()
        self.highlight_detector = HighlightDetector()
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]
        self.max_distance = 50

    def get_titles_with_positions(self, page) -> List[Dict]:
        """페이지에서 제목과 위치 정보 추출"""
        titles = []
        blocks = page.get_text("dict")["blocks"]
        
        for block in blocks:
            if block.get("lines"):
                text = ""
                y_top = block["bbox"][1]
                y_bottom = block["bbox"][3]
                
                for line in block["lines"]:
                    for span in line["spans"]:
                        text += span["text"] + " "
                
                text = text.strip()
                if text and any(re.search(pattern, text) for pattern in self.title_patterns):
                    titles.append({
                        "text": text,
                        "y_top": y_top,
                        "y_bottom": y_bottom,
                        "bbox": block["bbox"],
                        "used": False
                    })
        
        return sorted(titles, key=lambda x: x["y_top"])

    def get_table_positions(self, table, page_height):
        """Camelot 표의 위치를 PyMuPDF 좌표계로 변환"""
        x0, y0, x1, y1 = table._bbox
        converted_y0 = page_height - y1  # 상단
        converted_y1 = page_height - y0  # 하단
        
        return {
            "y_top": converted_y0,
            "y_bottom": converted_y1,
            "bbox": (x0, converted_y0, x1, converted_y1)
        }

    def match_title_to_table(self, titles, table_position):
        """표와 가장 적절한 제목 매칭"""
        best_title = None
        min_distance = float('inf')
        
        for title in titles:
            if title["used"]:
                continue
                
            distance = table_position["y_top"] - title["y_bottom"]
            if 0 < distance < self.max_distance and distance < min_distance:
                best_title = title
                min_distance = distance
        
        if best_title:
            best_title["used"] = True
            return best_title["text"], min_distance
        return None, None

    def process_table_with_highlights(self, table, highlight_regions, page_height):
        df = table.df.copy()
        x1, y1, x2, y2 = table._bbox
        table_height = y2 - y1
        row_height = table_height / len(df)

        # Add highlight column
        df['변경사항'] = ''

        for row_index in range(len(df)):
            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height
            
            if self.highlight_detector.check_highlight((row_top, row_bottom), highlight_regions):
                df.loc[row_index, '변경사항'] = '추가'

        return df

    def extract_tables_from_section(self, pdf_path: str, start_page: int, end_page: int) -> List[Tuple[str, pd.DataFrame, int]]:
        try:
            results = []
            
            for page_num in range(start_page, end_page):
                doc = fitz.open(pdf_path)
                page = doc[page_num]
                
                # 하이라이트 감지
                image = self.highlight_detector.pdf_to_image(page)
                contours = self.highlight_detector.detect_highlights(image)
                highlight_regions = self.highlight_detector.get_highlight_regions(contours, image.shape[0])
                
                # 제목 추출
                titles = self.get_titles_with_positions(page)
                
                # 표 추출
                tables = self.extract_with_camelot(pdf_path, page_num + 1)
                
                if tables:
                    page_height = page.rect.height
                    
                    for table in tables:
                        table_position = self.get_table_positions(table, page_height)
                        title, distance = self.match_title_to_table(titles, table_position)
                        
                        # 하이라이트 정보를 포함한 테이블 처리
                        df = self.process_table_with_highlights(table, highlight_regions, page_height)
                        df = self.clean_table(df)
                        
                        if not df.empty:
                            if title:
                                logger.info(f"Found table with title: {title} (distance: {distance:.1f})")
                            else:
                                title = "Untitled Table"
                                logger.warning(f"No matching title found for table on page {page_num + 1}")
                            
                            results.append((title, df, page_num + 1))
                            
                doc.close()
                
            return results
            
        except Exception as e:
            logger.error(f"Error extracting tables from section: {e}")
            return []

    def extract_with_camelot(self, pdf_path: str, page_num: int) -> List:
        """Camelot을 사용한 표 추출"""
        try:
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num),
                flavor='lattice'
            )
            if not tables:
                tables = camelot.read_pdf(
                    pdf_path,
                    pages=str(page_num),
                    flavor='stream'
                )
            return tables
        except Exception as e:
            logger.error(f"Camelot extraction failed: {str(e)}")
            return []

    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """표 데이터 정제"""
        try:
            df = df.dropna(how='all')
            df = df[~df.iloc[:, 0].str.contains("※|주)", regex=False, na=False)]
            return df
        except Exception as e:
            logger.error(f"Error cleaning table: {e}")
            return pd.DataFrame()

class ExcelWriter:
    @staticmethod
    def save_to_excel(sections_data: Dict[str, List[Tuple[str, pd.DataFrame, int]]], output_path: str):
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for section, tables in sections_data.items():
                    if not tables:
                        continue
                        
                    sheet_name = section.replace("[", "").replace("]", "")
                    current_row = 0
                    
                    for title, df, page_num in tables:
                        # 제목 쓰기
                        title_df = pd.DataFrame([[f"{title} (페이지: {page_num})"]], columns=[''])
                        title_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row,
                            index=False,
                            header=False
                        )
                        
                        # 표 데이터 쓰기
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row + 2,
                            index=False
                        )
                        
                        # 스타일 적용
                        worksheet = writer.sheets[sheet_name]
                        
                        # 제목 스타일링
                        title_cell = worksheet.cell(row=current_row + 1, column=1)
                        title_cell.font = Font(bold=True, size=12)
                        title_cell.fill = PatternFill(start_color='E6E6E6',
                                                    end_color='E6E6E6',
                                                    fill_type='solid')
                        
                        # 하이라이트 스타일링
                        yellow_fill = PatternFill(start_color='FFFF00',
                                                end_color='FFFF00',
                                                fill_type='solid')
                        
                        if '변경사항' in df.columns:
                            change_col_index = df.columns.get_loc('변경사항') + 1
                            for idx, row in enumerate(df.itertuples(), start=1):
                                if getattr(row, '변경사항') == '추가':
                                    for col in range(1, len(df.columns) + 1):
                                        cell = worksheet.cell(row=current_row + 2 + idx, column=col)
                                        cell.fill = yellow_fill
                        
                        current_row += len(df) + 5

            logger.info(f"Successfully saved tables to {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")

def main():
    try:
        # 파일 경로 설정
        pdf_path = "/workspaces/automation/uploads/KB 9회주는 암보험Plus(무배당)(24.05)_요약서_10.1판매_v1.0_앞단.pdf"
        output_path = "/workspaces/output/보험특약표.xlsx"
        
        if not os.path.exists(pdf_path):
            logger.error("PDF file not found")
            return

        # 문서 분석기 초기화
        logger.info("문서 분석 시작...")
        document_analyzer = InsuranceDocumentAnalyzer()
        section_pages = document_analyzer.find_section_pages(pdf_path)

        # 섹션 범위가 설정되었는지 확인
        if not document_analyzer.section_ranges:
            logger.warning("섹션을 찾지 못했습니다. 문서 전체를 대상으로 진행합니다.")
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            document_analyzer.section_ranges["전체"] = (0, total_pages)
            doc.close()

        # 표 추출기 초기화
        logger.info("표 추출 시작...")
        table_extractor = TableExtractor()
        sections_data = {}
        
        # 각 섹션별 표 추출
        for section, (start_page, end_page) in document_analyzer.section_ranges.items():
            logger.info(f"Processing {section} (pages {start_page + 1} to {end_page})")
            logger.info(f"하이라이트 및 표 분석 중... ({section})")
            tables = table_extractor.extract_tables_from_section(pdf_path, start_page, end_page)
            sections_data[section] = tables

        # 결과 저장
        if any(sections_data.values()):
            logger.info("엑셀 파일 생성 중...")
            ExcelWriter.save_to_excel(sections_data, output_path)
            logger.info(f"처리 완료. 결과가 {output_path}에 저장되었습니다.")
        else:
            logger.error("추출된 표가 없습니다.")

    except Exception as e:
        logger.error(f"Processing error: {str(e)}")
        raise

if __name__ == "__main__":
    try:
        main()
        logger.info("프로그램이 성공적으로 완료되었습니다.")
    except Exception as e:
        logger.error(f"프로그램 실행 중 오류 발생: {str(e)}")
