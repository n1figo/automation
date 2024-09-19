import camelot
import pandas as pd
import numpy as np
import cv2
import os
from PIL import Image
import fitz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 디버깅 모드 설정
DEBUG_MODE = True

# 이미지 저장 경로 설정
IMAGE_OUTPUT_DIR = "/workspaces/automation/output/images"
os.makedirs(IMAGE_OUTPUT_DIR, exist_ok=True)

def pdf_to_image(page):
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return np.array(img)

def detect_highlights(image, page_num):
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    s = hsv[:,:,1]
    v = hsv[:,:,2]

    saturation_threshold = 30
    saturation_mask = s > saturation_threshold

    _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)

    kernel = np.ones((5,5), np.uint8)
    cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)

    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_num}_mask.png'), cleaned_mask)

    contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    contour_image = image.copy()
    cv2.drawContours(contour_image, contours, -1, (0, 255, 0), 2)
    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_num}_contours.png'), cv2.cvtColor(contour_image, cv2.COLOR_RGB2BGR))

    return contours

def get_capture_regions(contours, image_height, image_width):
    if not contours:
        return []

    capture_height = image_height // 3
    sorted_contours = sorted(contours, key=lambda c: cv2.boundingRect(c)[1])

    regions = []
    current_region = None

    for contour in sorted_contours:
        x, y, w, h = cv2.boundingRect(contour)

        if current_region is None:
            current_region = [max(0, y - capture_height//2), min(image_height, y + h + capture_height//2)]
        elif y - current_region[1] < capture_height//2:
            current_region[1] = min(image_height, y + h + capture_height//2)
        else:
            regions.append(current_region)
            current_region = [max(0, y - capture_height//2), min(image_height, y + h + capture_height//2)]

    if current_region:
        regions.append(current_region)

    return regions

def extract_tables_with_camelot(pdf_path, page_number):
    print(f"Extracting tables from page {page_number} using Camelot...")
    try:
        tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
        print(f"Found {len(tables)} tables on page {page_number}")
        return tables
    except Exception as e:
        print(f"Error extracting tables: {e}")
        return []

def process_tables(tables, highlight_regions, page_height):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        # Camelot의 테이블 위치 정보 사용
        y1, x1, y2, x2 = table._bbox
        
        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            
            # 행의 y 좌표 계산 (PDF 좌표계에서 이미지 좌표계로 변환)
            row_y = page_height - (y1 + (row_index + 1) * (y2 - y1) / (len(df) + 1))
            
            row_highlighted = check_highlight(row_y, highlight_regions)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)

    return pd.DataFrame(processed_data)

def check_highlight(row_y, highlight_regions):
    for region in highlight_regions:
        if region[0] <= row_y <= region[1]:
            return True
    return False

def save_to_excel_with_highlight(df, output_path):
    df.to_excel(output_path, index=False)
    
    # 엑셀 파일 열기
    wb = load_workbook(output_path)
    ws = wb.active

    # 노란색 배경 스타일 정의
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # '변경사항' 열의 인덱스 찾기
    change_col_index = df.columns.get_loc('변경사항') + 1  # Excel은 1부터 시작하므로 1을 더함

    # 각 행을 순회하며 '추가' 항목에 노란색 배경 적용
    for row in range(2, ws.max_row + 1):  # 2부터 시작 (헤더 제외)
        if ws.cell(row=row, column=change_col_index).value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows")

def main(pdf_path, output_excel_path):
    print("Extracting tables and detecting highlights from PDF...")

    page_number = 50  # 51페이지 (0-based index)

    # PyMuPDF로 PDF 열기
    doc = fitz.open(pdf_path)
    page = doc[page_number]
    image = pdf_to_image(page)

    # 원본 이미지 저장
    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number + 1}_original.png'), cv2.cvtColor(image, cv2.COLOR_RGB2BGR))

    # 강조된 영역 탐지
    contours = detect_highlights(image, page_number + 1)
    highlight_regions = get_capture_regions(contours, image.shape[0], image.shape[1])

    # 강조 영역이 표시된 이미지 저장
    highlighted_image = image.copy()
    for region in highlight_regions:
        cv2.rectangle(highlighted_image, (0, region[0]), (image.shape[1], region[1]), (0, 255, 0), 2)
    cv2.imwrite(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number + 1}_highlighted.png'), cv2.cvtColor(highlighted_image, cv2.COLOR_RGB2BGR))

    print(f"감지된 강조 영역 수: {len(highlight_regions)}")
    print(f"강조 영역: {highlight_regions}")

    # Camelot을 사용하여 표 추출
    tables = extract_tables_with_camelot(pdf_path, page_number + 1)

    if not tables:
        print("No tables were extracted. Exiting.")
        return

    # 추출된 표 처리
    processed_df = process_tables(tables, highlight_regions, image.shape[0])

    if processed_df.empty:
        print("No data to process. Exiting.")
        return

    # 처리된 데이터 출력
    print(processed_df)

    # 엑셀로 저장 (하이라이트 포함)
    save_to_excel_with_highlight(processed_df, output_excel_path)

    print(f"Processed data saved to {output_excel_path}")

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables_camelot_highlighted.xlsx"
    main(pdf_path, output_excel_path)