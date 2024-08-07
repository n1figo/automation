import os
import re
import logging
import tempfile
from flask import Flask, render_template, request, send_file
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import cv2
import numpy as np
import fitz
from PIL import Image
import pytesseract

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def setup_browser():
    playwright = sync_playwright().start()
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()
    return playwright, browser, page

def capture_full_page(url):
    playwright, browser, page = setup_browser()
    try:
        page.goto(url)
        page.wait_for_timeout(5000)  # Wait for the page to load
        screenshot_path = "full_page.png"
        page.screenshot(path=screenshot_path, full_page=True)
        return screenshot_path
    finally:
        browser.close()
        playwright.stop()

def pdf_to_image(pdf_path, page_num):
    doc = fitz.open(pdf_path)
    page = doc.load_page(page_num)
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    doc.close()
    return np.array(img)

def detect_highlights(image):
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    s = hsv[:,:,1]
    v = hsv[:,:,2]
    
    saturation_threshold = 30
    saturation_mask = s > saturation_threshold
    
    _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)
    
    kernel = np.ones((5,5), np.uint8)
    cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)
    
    contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    return contours

def get_capture_regions(contours, image_height, image_width):
    if not contours:
        return []

    capture_height = image_height // 3
    
    sorted_contours = sorted(contours, key=lambda c: cv2.boundingRect(c)[1])
    
    regions = []
    current_region = None
    
    for contour in sorted_contours:
        x, y, w, h = cv2.boundingRect(contour)
        
        if current_region is None:
            current_region = [max(0, y - capture_height//2), min(image_height, y + h + capture_height//2)]
        elif y - current_region[1] < capture_height//2:
            current_region[1] = min(image_height, y + h + capture_height//2)
        else:
            regions.append(current_region)
            current_region = [max(0, y - capture_height//2), min(image_height, y + h + capture_height//2)]
    
    if current_region:
        regions.append(current_region)
    
    return regions

def clean_text_for_excel(text):
    # 제어 문자 제거
    text = ''.join(char for char in text if ord(char) >= 32)
    
    # 특수 문자 처리
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
    
    # 이모지 제거 (선택적)
    text = text.encode('ascii', 'ignore').decode('ascii')
    
    return text

def process_pdf(pdf_path, output_folder, homepage_image_path, wb):
    ws = wb.create_sheet("변경사항")

    # 변경 전 홈페이지 이미지 삽입
    img = XLImage(homepage_image_path)
    ws.add_image(img, 'A1')

    row = 1
    col = 'G'  # 하이라이트 이미지 시작 열

    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    doc.close()

    for page_num in range(total_pages):
        logger.info(f"Processing page {page_num + 1} of {total_pages}")
        
        image = pdf_to_image(pdf_path, page_num)
        contours = detect_highlights(image)
        
        if not contours:
            logger.info(f"No highlights detected on page {page_num + 1}")
            continue
        
        regions = get_capture_regions(contours, image.shape[0], image.shape[1])
        
        for i, (start_y, end_y) in enumerate(regions):
            highlighted_region = image[start_y:end_y, 0:image.shape[1]]
            
            output_path = os.path.join(output_folder, f"page_{page_num + 1}_highlights_{i + 1}.png")
            cv2.imwrite(output_path, cv2.cvtColor(highlighted_region, cv2.COLOR_RGB2BGR))

            # OCR 수행
            text = pytesseract.image_to_string(highlighted_region, lang='kor+eng')
            cleaned_text = clean_text_for_excel(text)

            # 엑셀에 정보 추가
            ws.cell(row=row, column=1, value=f"Page {page_num + 1}, Region {i + 1}")
            ws.cell(row=row, column=2, value=cleaned_text)

            # 하이라이트 이미지 삽입
            img = XLImage(output_path)
            ws.add_image(img, f'{col}{row}')

            # 빨간 네모 표시 (변경 전 이미지에)
            draw_red_rectangle(ws, 'A1', start_y, end_y, image.shape[1])

            row += 20  # 이미지 크기에 따라 조정 필요
            if row > 1000:  # 엑셀 행 제한에 근접하면 열 변경
                row = 1
                col = chr(ord(col) + 1)

    return wb

def draw_red_rectangle(ws, cell, start_y, end_y, width):
    img = ws._images[0]  # 첫 번째 이미지 (변경 전 홈페이지 이미지)
    img_width = img.width
    img_height = img.height

    border = Border(left=Side(style='thin', color='FF0000'),
                    right=Side(style='thin', color='FF0000'),
                    top=Side(style='thin', color='FF0000'),
                    bottom=Side(style='thin', color='FF0000'))

    start_col = ws[cell].column
    start_row = ws[cell].row

    # 이미지 크기에 맞게 비율 조정
    start_y_adj = int(start_y * 1.0 * img_height / img.height)
    end_y_adj = int(end_y * 1.0 * img_height / img.height)
    width_adj = int(width * 1.0 * img_width / img.width)

    for col in range(start_col, start_col + width_adj):
        for row in range(start_row + start_y_adj, start_row + end_y_adj):
            ws.cell(row=row, column=col).border = border

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        try:
            url = request.form['url']
            logger.info(f"Processing URL: {url}")
            
            # 웹페이지 캡처
            homepage_image_path = capture_full_page(url)
            logger.info(f"Webpage captured: {homepage_image_path}")
            
            # 엑셀 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "홈페이지 캡처"
            
            # 홈페이지 캡처 이미지를 첫 번째 시트에 추가
            img = XLImage(homepage_image_path)
            img.width = 800
            img.height = 600
            ws.add_image(img, 'A1')
            
            # PDF 파일 처리
            if 'file' in request.files:
                file = request.files['file']
                if file and file.filename.endswith('.pdf'):
                    logger.info(f"Processing PDF file: {file.filename}")
                    # 임시 파일로 PDF 저장
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                        file.save(temp_file.name)
                        pdf_path = temp_file.name

                    # PDF 처리 및 하이라이트 감지
                    output_folder = "highlight_images"
                    if not os.path.exists(output_folder):
                        os.makedirs(output_folder)
                    
                    try:
                        wb = process_pdf(pdf_path, output_folder, homepage_image_path, wb)
                        logger.info("PDF processing completed")
                    except Exception as e:
                        logger.error(f"Error processing PDF: {str(e)}", exc_info=True)
                        # 오류 발생 시에도 계속 진행
                    
                    # 임시 파일 삭제
                    os.unlink(pdf_path)
                else:
                    logger.warning("Invalid file type uploaded")
                    return "Please upload a PDF file", 400
            
            # 엑셀 파일 저장
            output_dir = 'output/excel'
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            excel_filename = "output.xlsx"
            excel_path = os.path.join(output_dir, excel_filename)
            wb.save(excel_path)
            logger.info(f"Excel file created: {excel_path}")
            
            return send_file(excel_path, as_attachment=True)
        except Exception as e:
            logger.error(f"An error occurred: {str(e)}", exc_info=True)
            return "An error occurred while processing your request", 500

    return render_template('index.html')

if __name__ == "__main__":
    app.run(debug=True)