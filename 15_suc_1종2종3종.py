import os
import re
import numpy as np
import PyPDF2
import camelot
import fitz  # PyMuPDF
import pandas as pd
from sentence_transformers import SentenceTransformer
import faiss
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def preprocess_text(text):
    text = re.sub(r'[\s\W_]+', '', text)
    text = text.lower()
    return text

def find_종_pages(texts_by_page, select_pages):
    종_pages = {"[1종]": [], "[2종]": [], "[3종]": []}
    pattern = re.compile(r'\[(\d)종\]')
    
    search_pages = set()
    for page in select_pages:
        search_pages.update([page-1, page, page+1])
    
    for page_num in search_pages:
        if page_num in texts_by_page:
            text = texts_by_page[page_num]
            matches = pattern.findall(text)
            for match in matches:
                종 = f"[{match}종]"
                if 종 in 종_pages:
                    종_pages[종].append(page_num)
                    print(f"{종} 패턴을 {page_num}페이지에서 발견했습니다.")
    
    for 종, pages in 종_pages.items():
        if pages:
            print(f"{종}이 발견된 페이지: {pages}")
        else:
            print(f"{종}을 찾지 못했습니다.")
    
    return 종_pages

def split_text_into_chunks(text, chunk_size=200, overlap=50):
    words = text.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        chunk = " ".join(words[i:i + chunk_size])
        chunks.append(chunk)
    return chunks

def create_index(chunks):
    model = SentenceTransformer('distiluse-base-multilingual-cased')
    embeddings = model.encode(chunks)
    embeddings = embeddings / np.linalg.norm(embeddings, axis=1, keepdims=True)
    dimension = embeddings.shape[1]
    index = faiss.IndexFlatIP(dimension)
    index.add(embeddings.astype('float32'))
    return index, model

def search_rag(query, index, model, chunks, page_numbers=None, k=5, threshold=0.5):
    query_vector = model.encode([query])
    query_vector = query_vector / np.linalg.norm(query_vector, axis=1, keepdims=True)
    distances, indices = index.search(query_vector.astype('float32'), k)
    results = []
    for idx, score in zip(indices[0], distances[0]):
        if score >= threshold:
            result = {'content': chunks[idx], 'score': score}
            if page_numbers:
                result['page'] = page_numbers[idx]
            results.append(result)
    return results

def find_pages_with_keyword(texts_by_page, keyword):
    pages = []
    preprocessed_keyword = preprocess_text(keyword)
    for page_num, text in texts_by_page.items():
        preprocessed_text = preprocess_text(text)
        if preprocessed_keyword in preprocessed_text:
            pages.append(page_num)
    return pages

def extract_tables_with_camelot(pdf_path, pages):
    all_tables = []
    for page in pages:
        print(f"Extracting table from page {page} using Camelot...")
        try:
            tables = camelot.read_pdf(pdf_path, pages=str(page), flavor='lattice')
            for table in tables:
                df = table.df
                all_tables.append({
                    'dataframe': df,
                    'page': page
                })
        except Exception as e:
            print(f"Error extracting tables from page {page}: {e}")
    
    print(f"Found {len(all_tables)} tables in total")
    return all_tables

def save_tables_to_excel(tables_dict, output_path, document_title=None):
    wb = Workbook()
    wb.remove(wb.active)
    
    for sheet_name, tables in tables_dict.items():
        ws = wb.create_sheet(title=sheet_name.replace("[", "").replace("]", ""))
        current_row = 1

        if document_title:
            ws.cell(row=current_row, column=1, value=document_title)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=1).font = Font(size=16, bold=True)
            current_row += 2

        for idx, table_info in enumerate(tables):
            if 'dataframe' not in table_info:
                print(f"Skipping table in sheet '{sheet_name}' as 'dataframe' key is missing.")
                continue

            df = table_info['dataframe']
            page = table_info['page']

            ws.cell(row=current_row, column=1, value=f"표 {idx+1} (페이지 {page})")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=df.shape[1])
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for r in dataframe_to_rows(df, index=False, header=True):
                for c_idx, value in enumerate(r, start=1):
                    ws.cell(row=current_row, column=c_idx, value=value)
                    ws.cell(row=current_row, column=c_idx).alignment = Alignment(wrap_text=True)
                current_row += 1
            current_row += 1

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    wb.save(output_path)
    print(f"Tables have been saved to {output_path}")

def find_table_end(texts_by_page, start_page, end_keywords):
    for page_num, text in texts_by_page.items():
        if page_num >= start_page:
            for keyword in end_keywords:
                if keyword in text:
                    print(f"표 끝 키워드 '{keyword}'를 {page_num}페이지에서 발견했습니다.")
                    return page_num - 1  # 키워드가 있는 페이지 직전까지가 표의 끝
    return None  # 끝을 찾지 못한 경우

def main():
    uploads_folder = "uploads"
    output_folder = "output"
    os.makedirs(output_folder, exist_ok=True)

    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_tables.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        full_text = ""
        page_numbers = []
        texts_by_page = {}
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"
                page_numbers.extend([i+1] * len(page_text.split()))
                texts_by_page[i+1] = page_text

    chunks = split_text_into_chunks(full_text)
    index, model = create_index(chunks)

    select_pages = find_pages_with_keyword(texts_by_page, "선택특약")
    print("선택특약이 포함된 페이지:", select_pages)

    종_pages = find_종_pages(texts_by_page, select_pages)

    tables_sheets = {
        "[1종]": [],
        "[2종]": [],
        "[3종]": [],
        "기타": []
    }

    end_keywords = ["상해관련 특약", "질병관련 특약", "기타 특약"]
    for page in select_pages:
        query = "선택특약 표"
        page_text = texts_by_page[page]
        page_chunks = split_text_into_chunks(page_text)
        page_index, _ = create_index(page_chunks)
        results = search_rag(query, page_index, model, page_chunks, k=1)
        if results:
            start_position = page_text.index(results[0]['content'])
            print(f"페이지 {page}에서 선택특약 표 시작 위치: {start_position}")

            end_page = find_table_end(texts_by_page, page, end_keywords)
            if end_page:
                print(f"선택특약 표가 끝나는 페이지: {end_page}")
            else:
                print(f"선택특약 표의 끝을 찾지 못했습니다.")

    tables = extract_tables_with_camelot(pdf_path, select_pages)

    for table in tables:
        page = table['page']
        for 종, pages in 종_pages.items():
            if page in pages:
                tables_sheets[종].append(table)
                break
        else:
            tables_sheets["기타"].append(table)

    doc = fitz.open(pdf_path)
    first_page = doc.load_page(0)
    first_page_text = first_page.get_text("text")
    document_title = first_page_text.strip().split('\n')[0]

    if any(tables_sheets.values()):
        save_tables_to_excel(tables_sheets, output_excel_path, document_title=document_title)
    else:
        print("추출된 표가 없습니다.")

    print(f"모든 처리된 표가 {output_excel_path}에 저장되었습니다.")

if __name__ == "__main__":
    main()