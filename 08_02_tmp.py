import PyPDF2
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import os
import pandas as pd
import camelot
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from sentence_transformers import SentenceTransformer
import faiss
import datetime

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class InsuranceDocumentAnalyzer:
    def __init__(self):
        self.section_patterns = {
            "종류": r'\[(\d)종\]',
            "특약유형": r'(상해관련|질병관련)\s*특약'
        }
        self.section_pages = {"[1종]": None, "[2종]": None, "[3종]": None}
        self.section_ranges = {}
        self.log_file = f"extraction_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        # RAG 설정
        self.model = SentenceTransformer('distiluse-base-multilingual-cased-v1')
        
    def log_extraction(self, message: str):
        """로그 메시지를 파일에 저장"""
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(f"{datetime.datetime.now()}: {message}\n")

    def create_embeddings(self, texts: List[str]) -> np.ndarray:
        """텍스트에 대한 임베딩 생성"""
        return self.model.encode(texts)

    def find_section_pages(self, pdf_path: str) -> Dict[str, int]:
        """PDF에서 1종, 2종, 3종의 시작 페이지 찾기"""
        try:
            doc = fitz.open(pdf_path)
            all_text_embeddings = []
            all_texts = []
            
            # 전체 문서의 텍스트 수집 및 임베딩
            for page_num in range(len(doc)):
                text = doc[page_num].get_text()
                all_texts.append(text)
                all_text_embeddings.append(self.create_embeddings([text])[0])
                
                # 종 패턴 찾기
                matches = re.finditer(self.section_patterns["종류"], text)
                for match in matches:
                    종_type = f"[{match.group(1)}종]"
                    if self.section_pages[종_type] is None:
                        self.section_pages[종_type] = page_num
                        log_msg = f"{종_type} 시작 페이지: {page_num + 1}"
                        logger.info(log_msg)
                        self.log_extraction(log_msg)

            # FAISS 인덱스 생성
            embeddings = np.array(all_text_embeddings)
            index = faiss.IndexFlatL2(embeddings.shape[1])
            index.add(embeddings.astype('float32'))

            # 섹션 범위 설정
            sorted_pages = sorted([(k, v) for k, v in self.section_pages.items() if v is not None], 
                                key=lambda x: x[1])
            
            for i, (종_type, start_page) in enumerate(sorted_pages):
                if i + 1 < len(sorted_pages):
                    end_page = sorted_pages[i + 1][1]
                else:
                    end_page = len(doc)
                self.section_ranges[종_type] = (start_page, end_page)
                
                # 관련성 분석 로깅
                section_text = " ".join(all_texts[start_page:end_page])
                self.log_extraction(f"\n{'='*50}")
                self.log_extraction(f"{종_type} 섹션 분석 결과:")
                self.log_extraction(f"페이지 범위: {start_page+1} - {end_page}")
                self.log_extraction(f"섹션 텍스트 샘플 (처음 200자):\n{section_text[:200]}...")
                
                # 특약 관련 키워드 검색
                for pattern_name, pattern in self.section_patterns.items():
                    if pattern_name != "종류":
                        matches = re.finditer(pattern, section_text)
                        for match in matches:
                            self.log_extraction(f"발견된 {pattern_name}: {match.group()}")
                
            doc.close()
            return self.section_pages
            
        except Exception as e:
            error_msg = f"Error finding section pages: {e}"
            logger.error(error_msg)
            self.log_extraction(error_msg)
            return {}

class TableExtractor:
    def __init__(self):
        self.font_size_threshold = 10
        self.title_max_length = 50
        self.log_file = f"table_extraction_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    def log_extraction(self, message: str):
        """로그 메시지를 파일에 저장"""
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(f"{datetime.datetime.now()}: {message}\n")

    def extract_tables_from_section(self, pdf_path: str, start_page: int, end_page: int) -> List[Tuple[str, pd.DataFrame]]:
        """섹션 범위 내의 표 추출"""
        try:
            results = []
            
            for page_num in range(start_page, end_page):
                # 페이지 텍스트 분석
                doc = fitz.open(pdf_path)
                page = doc[page_num]
                text = page.get_text()
                
                self.log_extraction(f"\n{'='*50}")
                self.log_extraction(f"페이지 {page_num + 1} 분석:")
                self.log_extraction(f"페이지 텍스트 샘플:\n{text[:200]}...")
                
                # 상해관련 또는 질병관련 특약 확인
                if re.search(r'(상해관련|질병관련)\s*특약', text):
                    self.log_extraction("특약 관련 내용 발견")
                    
                    # 표 위의 제목 찾기
                    title = self.extract_table_title(page)
                    self.log_extraction(f"추출된 표 제목: {title}")
                    
                    # 표 추출
                    tables = self.extract_with_camelot(pdf_path, page_num + 1)
                    
                    for idx, table in enumerate(tables):
                        df = self.clean_table(table.df)
                        if not df.empty:
                            results.append((title, df, page_num + 1))
                            self.log_extraction(f"표 {idx+1} 추출 완료 (크기: {df.shape})")
                            
                doc.close()
                
            return results
            
        except Exception as e:
            error_msg = f"Error extracting tables from section: {e}"
            logger.error(error_msg)
            self.log_extraction(error_msg)
            return []

    def extract_table_title(self, page) -> str:
        """표 위의 제목 추출"""
        try:
            blocks = page.get_text("dict")["blocks"]
            table_block = None
            title_block = None
            
            self.log_extraction("\n표 제목 추출 시도:")
            
            # 표 블록 찾기
            for block in blocks:
                if "lines" in block:
                    text = " ".join(span["text"] for line in block["lines"] 
                                  for span in line["spans"])
                    self.log_extraction(f"분석 중인 블록: {text[:100]}")
                    
                    if "특약" in text and len(block["lines"]) > 1:
                        table_block = block
                        self.log_extraction(f"표 블록 발견: {text[:100]}...")
                        break
            
            if table_block:
                table_top = table_block["bbox"][1]
                # 표 위의 가장 가까운 텍스트 블록 찾기
                potential_titles = [
                    b for b in blocks if "lines" in b 
                    and b["bbox"][3] < table_top
                    and len(" ".join(span["text"] for line in b["lines"] 
                                   for span in line["spans"])) < self.title_max_length
                ]
                
                for b in potential_titles:
                    title_text = " ".join(span["text"] for line in b["lines"] for span in line["spans"])
                    self.log_extraction(f"제목 후보: {title_text}")
                
                if potential_titles:
                    title_block = max(potential_titles, 
                                    key=lambda x: x["bbox"][3])
                    title = " ".join(span["text"] for line in title_block["lines"] 
                                   for span in line["spans"])
                    self.log_extraction(f"선택된 제목: {title}")
                    return title
            
            self.log_extraction("제목을 찾지 못함")
            return "Untitled Table"
            
        except Exception as e:
            error_msg = f"Error extracting table title: {e}"
            logger.error(error_msg)
            self.log_extraction(error_msg)
            return "Untitled Table"

    def extract_with_camelot(self, pdf_path: str, page_num: int) -> List:
        """Camelot을 사용한 표 추출"""
        try:
            self.log_extraction(f"\nCamelot을 사용한 표 추출 시도 (페이지 {page_num}):")
            
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num),
                flavor='lattice'
            )
            if not tables:
                self.log_extraction("lattice 방식 실패, stream 방식 시도")
                tables = camelot.read_pdf(
                    pdf_path,
                    pages=str(page_num),
                    flavor='stream'
                )
            
            self.log_extraction(f"추출된 표 수: {len(tables)}")
            return tables
            
        except Exception as e:
            error_msg = f"Camelot extraction failed: {str(e)}"
            logger.error(error_msg)
            self.log_extraction(error_msg)
            return []

    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """표 데이터 정제"""
        try:
            original_shape = df.shape
            df = df.dropna(how='all')
            df = df[~df.iloc[:, 0].str.contains("※|주)", regex=False, na=False)]
            
            self.log_extraction(f"\n표 정제 결과:")
            self.log_extraction(f"원본 크기: {original_shape}")
            self.log_extraction(f"정제 후 크기: {df.shape}")
            
            return df
            
        except Exception as e:
            error_msg = f"Error cleaning table: {e}"
            logger.error(error_msg)
            self.log_extraction(error_msg)
            return pd.DataFrame()

class ExcelWriter:
    @staticmethod
    def save_to_excel(sections_data: Dict[str, List[Tuple[str, pd.DataFrame, int]]], output_path: str):
        """섹션별 데이터를 Excel 파일로 저장"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for section, tables in sections_data.items():
                    if not tables:
                        continue
                        
                    sheet_name = section.replace("[", "").replace("]", "")
                    current_row = 0
                    
                    for title, df, page_num in tables:
                        # 제목 쓰기
                        title_df = pd.DataFrame([[f"{title} (페이지: {page_num})"]], columns=[''])
                        title_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row,
                            index=False,
                            header=False
                        )
                        
                        # 표 데이터 쓰기
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            startrow=current_row + 2,
                            index=False
                        )
                        
                        # 스타일 적용
                        worksheet = writer.sheets[sheet_name]
                        
                        # 제목 스타일링
                        title_cell = worksheet.cell(row=current_row + 1, column=1)
                        title_cell.font = Font(bold=True, size=12)
                        title_cell.fill = PatternFill(start_color='E6E6E6',
                                                    end_color='E6E6E6',
                                                    fill_type='solid')
                        
                        current_row += len(df) + 5

            logger.info(f"Successfully saved tables to {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")

def main():
    try:
        # 파일 경로 설정
        pdf_path = "/workspaces/automation/uploads/KB 9회주는 암보험Plus(무배당)(24.05)_요약서_10.1판매_v1.0_앞단.pdf"
        output_path = "보험특약표.xlsx"
        
        if not os.path.exists(pdf_path):
            logger.error("PDF file not found")
            return

        # 문서 분석기 초기화
        document_analyzer = InsuranceDocumentAnalyzer()
        section_pages = document_analyzer.find_section_pages(pdf_path)
        
        if not section_pages:
            logger.error("No sections found in the document")
            return

        # 표 추출기 초기화
        table_extractor = TableExtractor()
        sections_data = {}
        
        # 각 섹션별 표 추출
        for section, (start_page, end_page) in document_analyzer.section_ranges.items():
            logger.info(f"Processing {section} (pages {start_page + 1} to {end_page})")
            tables = table_extractor.extract_tables_from_section(pdf_path, start_page, end_page)
            sections_data[section] = tables

        # 결과 저장
        if any(sections_data.values()):
            ExcelWriter.save_to_excel(sections_data, output_path)
            logger.info("Processing completed successfully")
            
            # 추가: 결과 요약 로그 생성
            summary_log_file = f"summary_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            with open(summary_log_file, 'w', encoding='utf-8') as f:
                f.write("=== 처리 결과 요약 ===\n\n")
                f.write(f"처리된 PDF 파일: {pdf_path}\n")
                f.write(f"생성된 Excel 파일: {output_path}\n\n")
                
                for section, tables in sections_data.items():
                    f.write(f"\n{section} 분석 결과:\n")
                    f.write(f"추출된 표 수: {len(tables)}\n")
                    for idx, (title, df, page_num) in enumerate(tables, 1):
                        f.write(f"\n  표 {idx}:\n")
                        f.write(f"  - 페이지: {page_num}\n")
                        f.write(f"  - 제목: {title}\n")
                        f.write(f"  - 크기: {df.shape[0]}행 x {df.shape[1]}열\n")
        else:
            error_msg = "No tables extracted from any section"
            logger.error(error_msg)
            with open("error_log.txt", 'a', encoding='utf-8') as f:
                f.write(f"{datetime.datetime.now()}: {error_msg}\n")

    except Exception as e:
        error_msg = f"Processing error: {str(e)}"
        logger.error(error_msg)
        with open("error_log.txt", 'a', encoding='utf-8') as f:
            f.write(f"{datetime.datetime.now()}: {error_msg}\n")

if __name__ == "__main__":
    main()