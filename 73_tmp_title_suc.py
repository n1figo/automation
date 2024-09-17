import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import logging
import os
from typing import List, Tuple, Any
from collections import defaultdict
import re

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 텍스트를 엑셀에 맞게 정리 (줄바꿈 유지)
def clean_text_for_excel(text: str) -> str:
    if isinstance(text, str):
        # 제어 문자 제거 (줄바꿈은 유지)
        text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
        return text  # 줄바꿈을 제거하지 않음
    return text

# 텍스트 색상 및 배경색 체크 함수
def check_text_color_and_background(span) -> bool:
    # 흰색, 검정색, 회색 제외한 텍스트 색상 필터링
    common_colors = {(1, 1, 1), (0, 0, 0), (0.5, 0.5, 0.5)}  # RGB 값
    color = span.get('color', None)
    background_color = span.get('bgcolor', None)  # 배경색 확인 (bgcolor)

    # 텍스트 색상 또는 배경색이 특정 색상(예: 살색 음영)을 감지하도록 추가
    if color and color not in common_colors:
        return True
    if background_color and background_color not in common_colors:
        return True
    return False

# 텍스트 및 색상 추출 함수
def extract_text_and_colors(page: fitz.Page) -> List[Tuple[str, str]]:
    blocks = page.get_text("dict")["blocks"]
    extracted_text = []

    for block in blocks:
        if 'lines' in block:
            for line in block['lines']:
                for span in line['spans']:
                    text = clean_text_for_excel(span['text'])
                    if check_text_color_and_background(span):  # 배경색도 체크
                        extracted_text.append((text, "추가"))
                    else:
                        extracted_text.append((text, "유지"))
    
    return extracted_text

# PDF 테이블 추출 클래스
class PDFTableExtractor:
    def __init__(self, pdf_path: str, tessdata_dir: str = None):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        if tessdata_dir:
            os.environ['TESSDATA_PREFIX'] = tessdata_dir
        
    def extract_tables_with_titles(self) -> List[Tuple[str, pd.DataFrame]]:
        all_tables = []
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            tables = self.extract_tables_from_page(page)
            titled_tables = self._assign_titles_to_tables(page, tables)
            all_tables.extend(titled_tables)
        return self._merge_tables_with_same_title(all_tables)
    
    def extract_tables_from_page(self, page: fitz.Page) -> List[Any]:
        tables = page.find_tables()
        return tables
    
    def _assign_titles_to_tables(self, page: fitz.Page, tables: List[Any]) -> List[Tuple[str, pd.DataFrame]]:
        titled_tables = []
        for table in tables:
            title = self._find_table_title(page, table)
            df = self._table_to_dataframe(table)
            titled_tables.append((title, df))
        return titled_tables
    
    def _find_table_title(self, page: fitz.Page, table: Any) -> str:
        blocks = page.get_text("dict")["blocks"]
        table_top = table.bbox[1]  # y0 좌표
        potential_titles = []
        for b in blocks:
            if 'lines' in b:
                for l in b['lines']:
                    for s in l['spans']:
                        if s['bbox'][3] < table_top and s['bbox'][3] > table_top - 50:
                            potential_titles.append(s['text'])
        
        if potential_titles:
            return " ".join(potential_titles).strip()
        return "Untitled Table"
    
    def _table_to_dataframe(self, table: Any) -> pd.DataFrame:
        df = pd.DataFrame(table.extract())
        df = df.applymap(clean_text_for_excel)
        return df
    
    def _merge_tables_with_same_title(self, tables: List[Tuple[str, pd.DataFrame]]) -> List[Tuple[str, pd.DataFrame]]:
        merged_tables = defaultdict(list)
        for title, df in tables:
            merged_tables[title].append(df)
        
        return [(title, pd.concat(dfs, ignore_index=True)) for title, dfs in merged_tables.items()]

# 엑셀 작성 및 변경사항 적용 클래스
class ExcelWriterWithChanges:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Extracted Tables"
        
    def write_tables_with_changes(self, tables: List[Tuple[str, pd.DataFrame]]):
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row = 1
        
        for title, df in tables:
            self.sheet.cell(row=row, column=1, value=clean_text_for_excel(title))
            row += 1
            
            # "변경사항" 컬럼 추가
            df["변경사항"] = df.apply(lambda row: "추가" if any(cell == "추가" for cell in row.values) else "유지", axis=1)
            
            for r in dataframe_to_rows(df, index=False, header=True):
                self.sheet.append([clean_text_for_excel(cell) for cell in r])
                
                # 엑셀에서 줄바꿈 적용
                for col_num, cell_value in enumerate(r, start=1):
                    cell = self.sheet.cell(row=row, column=col_num)
                    cell.alignment = Alignment(wrap_text=True)  # 줄바꿈 적용
                    if "추가" in r:
                        cell.fill = yellow_fill
                row += 1
            row += 2
        
        self.workbook.save(self.output_path)
        logger.info(f"Tables saved with '변경사항' to {self.output_path}")

# 메인 함수
def main(pdf_path: str, output_excel_path: str, tessdata_dir: str = None):
    try:
        extractor = PDFTableExtractor(pdf_path, tessdata_dir)
        
        # 테이블 추출 및 색상 감지
        tables = extractor.extract_tables_with_titles()
        
        # Excel 작성 및 변경사항 적용
        writer = ExcelWriterWithChanges(output_excel_path)
        writer.write_tables_with_changes(tables)
        
        logger.info("Table extraction, writing, and text extraction with changes completed successfully.")
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}", exc_info=True)

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables.xlsx"
    tessdata_dir = "/usr/share/tesseract-ocr/4.00/tessdata"  # Tesseract OCR 언어 데이터 파일이 있는 디렉토리
    main(pdf_path, output_excel_path, tessdata_dir)
