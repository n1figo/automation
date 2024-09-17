import fitz  # PyMuPDF
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# 디버깅 모드 설정 (True로 설정하면 색상 정보를 출력합니다)
DEBUG_MODE = False

# 타겟 헤더 정의
TARGET_HEADERS = ["보장명", "지급사유", "지급금액"]

# 허용되지 않는 문자를 제거하는 함수
def remove_illegal_characters(text):
    ILLEGAL_CHARACTERS_RE = re.compile(
        '['
        '\x00-\x08'
        '\x0B-\x0C'
        '\x0E-\x1F'
        ']'
    )
    return ILLEGAL_CHARACTERS_RE.sub('', text)

# 텍스트를 엑셀에 맞게 정리 (줄바꿈 유지)
def clean_text_for_excel(text: str) -> str:
    if isinstance(text, str):
        # 제어 문자 제거 (줄바꿈은 유지)
        text = remove_illegal_characters(text)
        return text  # 줄바꿈을 제거하지 않음
    return text

# 셀의 변경사항 여부를 판단하는 함수 (pdfplumber 사용)
def check_cell_for_changes(cell_rect, fills):
    # 초기값 설정
    change_detected = False
    bg_color = None
    max_overlap_area = 0
    for fill in fills:
        fill_rect = fill['rect']
        overlap_rect = cell_rect & fill_rect  # 교집합 영역 계산
        if overlap_rect.is_empty:
            continue
        overlap_area = overlap_rect.get_area()
        if overlap_area > max_overlap_area:
            max_overlap_area = overlap_area
            bg_color = fill['color']
            change_detected = True
    return change_detected, bg_color

# 페이지에서 타겟 표를 추출하는 함수
def extract_target_tables_from_page(page, fills, page_number):
    print(f"페이지 {page_number + 1} 처리 중...")
    table_data = []
    tables = page.find_tables()
    for table in tables:
        # 테이블 데이터 추출
        table_content = table.extract()
        if not table_content:
            continue
        # 헤더 행 추출 및 정리
        header_row = table_content[0]
        header_texts = [clean_text_for_excel(cell.strip()) if cell else '' for cell in header_row]
        header_texts_normalized = [text.replace(" ", "").replace("\n", "") for text in header_texts]
        # 타겟 헤더가 모두 포함되어 있는지 확인
        if all(any(target_header == header_cell for header_cell in header_texts_normalized) for target_header in TARGET_HEADERS):
            # 셀 딕셔너리 생성 (셀 위치를 기준으로 매핑)
            cell_dict = {}
            for cell in table.cells:
                # cell의 구조를 확인하여 적절히 접근
                if len(cell) == 6:
                    cell_row = cell[0]
                    cell_col = cell[1]
                    cell_bbox = cell[2:6]  # x0, y0, x1, y1
                elif len(cell) >= 3 and isinstance(cell[2], (tuple, list)):
                    cell_row = cell[0]
                    cell_col = cell[1]
                    cell_bbox = cell[2]  # (x0, y0, x1, y1)
                else:
                    continue  # 알 수 없는 셀 구조인 경우 건너뜁니다.

                # cell_bbox가 4개의 값(x0, y0, x1, y1)인지 확인
                if len(cell_bbox) == 4:
                    cell_rect = fitz.Rect(*cell_bbox)
                else:
                    continue  # bbox가 올바르지 않은 경우 건너뜁니다.

                cell_dict[(cell_row, cell_col)] = cell_rect
            # 테이블 데이터 처리
            num_rows = len(table_content)
            num_cols = len(header_texts)
            for row_index in range(1, num_rows):  # 헤더 행 이후부터 처리
                row = table_content[row_index]
                row_data = {}
                change_detected = False
                cell_bg_color = None  # 초기화
                for col_index in range(num_cols):
                    header = header_texts[col_index].replace(" ", "").replace("\n", "")
                    # 셀 값 처리 시 None 체크 추가
                    cell_text = clean_text_for_excel(row[col_index].strip()) if row[col_index] else ''
                    if header in TARGET_HEADERS:
                        # 셀 내용 분할 (줄바꿈 기준)
                        cell_texts = cell_text.split('\n')
                        if header == '보장명':
                            if len(cell_texts) > 1:
                                row_data['보장명1'] = cell_texts[0]
                                row_data['보장명2'] = cell_texts[1]
                            else:
                                row_data['보장명1'] = cell_text
                                row_data['보장명2'] = ''
                        elif header == '지급사유':
                            if len(cell_texts) > 1:
                                row_data['지급사유1'] = cell_texts[0]
                                row_data['지급사유2'] = cell_texts[1]
                            else:
                                row_data['지급사유1'] = cell_text
                                row_data['지급사유2'] = ''
                        else:
                            row_data[header] = cell_text
                        # 셀 객체 가져오기
                        cell_rect = cell_dict.get((row_index, col_index))
                        if cell_rect:
                            # 셀에 변경사항이 있는지 확인
                            detected, bg_color = check_cell_for_changes(cell_rect, fills)
                            if detected:
                                change_detected = True
                            cell_bg_color = bg_color  # 배경색 기록
                if row_data:
                    # 페이지 번호 추가
                    row_data["페이지"] = page_number + 1
                    # 변경사항 설정
                    row_data["변경사항"] = "추가" if change_detected else "유지"
                    # 배경색 추가
                    row_data["배경색"] = str(cell_bg_color) if cell_bg_color else ''
                    table_data.append(row_data)
    return table_data

# 메인 함수
def main(pdf_path, output_excel_path):
    print("PDF에서 개정된 부분을 추출합니다...")
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    all_table_data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_number in range(total_pages):
            page = doc[page_number]
            pdfplumber_page = pdf.pages[page_number]
            page_height = pdfplumber_page.height
            fills = pdfplumber_page.rects
            # 좌표 변환 및 fills 조정
            adjusted_fills = []
            for fill in fills:
                x0 = fill['x0']
                x1 = fill['x1']
                y0 = fill['y0']
                y1 = fill['y1']
                # 좌표 변환 (pdfplumber와 fitz의 좌표계 차이 보정)
                fitz_y0 = page_height - y1
                fitz_y1 = page_height - y0
                fill_rect = fitz.Rect(x0, fitz_y0, x1, fitz_y1)
                # 색상 정보 가져오기
                color = fill.get('non_stroking_color', None)
                adjusted_fill = {
                    'rect': fill_rect,
                    'color': color
                }
                adjusted_fills.append(adjusted_fill)
            # 테이블 데이터 추출
            table_data = extract_target_tables_from_page(page, adjusted_fills, page_number)
            all_table_data.extend(table_data)
    if all_table_data:
        # 데이터프레임 생성
        df = pd.DataFrame(all_table_data)
        # 결측치 처리 (컬럼이 없을 경우 대비)
        for col in ["보장명1", "보장명2", "지급사유1", "지급사유2", "지급금액", "배경색"]:
            if col not in df.columns:
                df[col] = ''
        # 컬럼 순서 지정
        df = df[["페이지", "보장명1", "보장명2", "지급사유1", "지급사유2", "지급금액", "변경사항", "배경색"]]
        # 엑셀로 저장
        print("개정된 부분을 엑셀 파일로 저장합니다...")
        save_revisions_to_excel(df, output_excel_path)
        print("작업이 완료되었습니다.")
    else:
        print("지정된 헤더를 가진 표를 찾을 수 없습니다.")

# 엑셀 파일로 저장하는 함수
def save_revisions_to_excel(df, output_excel_path):
    # 엑셀 파일 생성
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "개정된 부분"
    # 노란색 강조 표시
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # 데이터프레임을 엑셀로 작성
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        sheet.append(row)
        if r_idx == 1:
            continue  # 헤더는 제외
        # 변경사항이 "추가"인 경우 강조 표시
        if df.iloc[r_idx - 2]["변경사항"] == "추가":
            for cell in sheet[r_idx]:
                cell.fill = yellow_fill
                cell.alignment = Alignment(wrap_text=True)
        else:
            for cell in sheet[r_idx]:
                cell.alignment = Alignment(wrap_text=True)
    workbook.save(output_excel_path)
    print(f"개정된 부분이 '{output_excel_path}'에 저장되었습니다.")

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables.xlsx"
    main(pdf_path, output_excel_path)



















