import fitz
import pandas as pd
import numpy as np
import os
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.cluster import KMeans
from collections import Counter
import camelot

DEBUG_MODE = True
IMAGE_OUTPUT_DIR = "/workspaces/automation/output/images"
os.makedirs(IMAGE_OUTPUT_DIR, exist_ok=True)

def is_highlight_color(color, threshold=0.8):
    if isinstance(color, (int, float)):  # 단일 값인 경우 (회색조)
        return color > threshold * 255
    elif len(color) == 3:  # RGB
        r, g, b = color
        return (r > threshold * 255 and g > threshold * 255) or \
               (r > threshold * 255 and b > threshold * 255) or \
               (g > threshold * 255 and b > threshold * 255)
    elif len(color) == 4:  # CMYK
        c, m, y, k = color
        return c < 0.1 and m < 0.1 and y > 0.5 and k < 0.1
    return False

def detect_highlights_pymupdf(page):
    highlight_areas = []
    words = page.get_text("words")
    for word in words:
        if len(word) >= 5:  # 색상 정보가 있는지 확인
            bbox = fitz.Rect(word[:4])
            color = word[4] if isinstance(word[4], (int, float, tuple)) else None
            if color and is_highlight_color(color):
                highlight_areas.append(bbox)
    return highlight_areas

def detect_highlights_pillow(image_path):
    image = Image.open(image_path)
    pixels = list(image.getdata())
    
    # Use K-means clustering to find dominant colors
    kmeans = KMeans(n_clusters=5)
    kmeans.fit(pixels)
    
    # Count occurrences of each color
    color_counts = Counter(kmeans.labels_)
    
    # Find highlight colors
    highlight_colors = [kmeans.cluster_centers_[i] for i, count in color_counts.items() 
                        if is_highlight_color(kmeans.cluster_centers_[i]) and count > len(pixels) * 0.01]
    
    # Create a mask for highlight colors
    highlight_mask = Image.new('1', image.size, 0)
    for y in range(image.height):
        for x in range(image.width):
            pixel = image.getpixel((x, y))
            if any(np.allclose(pixel, color, atol=10) for color in highlight_colors):
                highlight_mask.putpixel((x, y), 1)
    
    # Find contours in the mask
    highlight_areas = []
    for y in range(image.height):
        for x in range(image.width):
            if highlight_mask.getpixel((x, y)) == 1:
                highlight_areas.append((x, y, x+1, y+1))
    
    return highlight_areas

def combine_highlight_areas(areas1, areas2):
    combined = areas1 + areas2
    # Merge overlapping areas
    merged = []
    for area in combined:
        if not merged:
            merged.append(area)
        else:
            overlap = False
            for i, existing in enumerate(merged):
                if check_overlap(area, existing):
                    merged[i] = (min(area[0], existing[0]), min(area[1], existing[1]),
                                 max(area[2], existing[2]), max(area[3], existing[3]))
                    overlap = True
                    break
            if not overlap:
                merged.append(area)
    return merged

def extract_tables_with_camelot(pdf_path, page_number):
    print(f"Extracting tables from page {page_number} using Camelot...")
    tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
    print(f"Found {len(tables)} tables on page {page_number}")
    return tables

def process_tables(tables, highlight_areas):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            row_bbox = table.cells[row_index][0].bbox
            row_highlighted = any(check_overlap(row_bbox, area) for area in highlight_areas)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)
    return pd.DataFrame(processed_data)

def check_overlap(bbox1, bbox2):
    x1 = max(bbox1[0], bbox2[0])
    y1 = max(bbox1[1], bbox2[1])
    x2 = min(bbox1[2], bbox2[2])
    y2 = min(bbox1[3], bbox2[3])
    return x1 < x2 and y1 < y2

def save_to_excel_with_highlight(df, output_path):
    df.to_excel(output_path, index=False)
    
    wb = load_workbook(output_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    change_col_index = df.columns.get_loc('변경사항') + 1

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=change_col_index).value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows")

def main(pdf_path, output_excel_path):
    print("PDF에서 개정된 부분을 추출합니다...")

    doc = fitz.open(pdf_path)
    page_number = 50  # 51페이지 (0-based index)
    page = doc[page_number]

    # PyMuPDF를 사용한 하이라이트 감지
    highlight_areas_pymupdf = detect_highlights_pymupdf(page)

    # 이미지로 변환 후 Pillow를 사용한 하이라이트 감지
    pix = page.get_pixmap()
    img_path = os.path.join(IMAGE_OUTPUT_DIR, f"page_{page_number + 1}.png")
    pix.save(img_path)
    highlight_areas_pillow = detect_highlights_pillow(img_path)

    # 두 방법의 결과 합치기
    highlight_areas = combine_highlight_areas(highlight_areas_pymupdf, highlight_areas_pillow)

    print(f"감지된 강조 영역 수: {len(highlight_areas)}")

    # 디버깅: 강조 영역 시각화
    img = Image.open(img_path)
    overlay = Image.new('RGBA', img.size, (0, 0, 0, 0))
    for area in highlight_areas:
        overlay_draw = Image.new('RGBA', (int(area[2]-area[0]), int(area[3]-area[1])), (255, 0, 0, 128))
        overlay.paste(overlay_draw, (int(area[0]), int(area[1])))
    img = Image.alpha_composite(img.convert('RGBA'), overlay)
    img.save(os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number + 1}_highlighted.png'))

    # Camelot을 사용하여 표 추출
    tables = extract_tables_with_camelot(pdf_path, page_number + 1)

    # 추출된 표 처리
    processed_df = process_tables(tables, highlight_areas)

    print(processed_df)

    save_to_excel_with_highlight(processed_df, output_excel_path)

    print(f"처리된 데이터가 {output_excel_path}에 저장되었습니다.")

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables_camelot.xlsx"
    main(pdf_path, output_excel_path)