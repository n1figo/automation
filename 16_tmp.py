import os
import re
import numpy as np
import PyPDF2
import camelot
import fitz  # PyMuPDF
import pandas as pd
from sentence_transformers import SentenceTransformer
import faiss
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_text_with_positions(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text_chunks = []
        for page_num, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                lines = text.split('\n')
                for line_num, line in enumerate(lines):
                    text_chunks.append({
                        'text': line,
                        'page': page_num + 1,
                        'line': line_num + 1
                    })
    return text_chunks

def preprocess_text(text):
    text = re.sub(r'[\s\W_]+', '', text)
    text = text.lower()
    return text

def detect_table_boundaries(text_chunks):
    start_patterns = [r'표\s*\d+', r'선택\s*특약', r'상해관련\s*특약', r'질병관련\s*특약', r'\[1\s*종\]', r'\[2\s*종\]', r'\[3\s*종\]']
    end_patterns = [r'합\s*계', r'총\s*계', r'주\s*\)', r'※', r'결\s*론']

    tables = []
    table_start = None

    for i, chunk in enumerate(text_chunks):
        text = chunk['text']

        if table_start is None and any(re.search(pattern, text) for pattern in start_patterns):
            table_start = i
            continue

        if table_start is not None and any(re.search(pattern, text) for pattern in end_patterns):
            table_end = i

            start_page = text_chunks[table_start]['page']
            end_page = text_chunks[table_end]['page']
            pages = list(range(start_page, end_page + 1))

            context_start = max(0, table_start - 5)
            context_end = min(len(text_chunks), table_end + 5)
            context = text_chunks[context_start:context_end]

            tables.append({
                'start': text_chunks[table_start],
                'end': text_chunks[table_end],
                'pages': pages,
                'context': context
            })
            table_start = None

    return tables

def extract_text_above_bbox(page, bbox):
    x0, y0, x1, y1 = bbox
    text_blocks = page.get_text("blocks")
    texts_above = []
    for block in text_blocks:
        if len(block) >= 5:
            bx0, by0, bx1, by1, text = block[:5]
            if by1 <= y0:
                texts_above.append((by1, text))
    if texts_above:
        texts_above.sort(reverse=True)
        return texts_above[0][1].strip()
    else:
        return "제목 없음"

def extract_tables_with_camelot(pdf_path, tables_info):
    all_tables = []
    doc = fitz.open(pdf_path)
    for table_info in tables_info:
        pages = table_info['pages']
        pages_str = ','.join(map(str, pages))
        print(f"Extracting table from pages {pages_str} using Camelot...")
        try:
            tables = camelot.read_pdf(pdf_path, pages=pages_str, flavor='lattice')
        except Exception as e:
            print(f"Error extracting tables from pages {pages_str}: {e}")
            continue

        if not tables:
            continue

        for table in tables:
            df = table.df
            bbox = table._bbox
            x0, y0, x1, y1 = bbox
            page_num = table.page - 1
            page = doc.load_page(page_num)
            text_above_table = extract_text_above_bbox(page, bbox)
            title = text_above_table.strip()

            all_tables.append({
                'dataframe': df,
                'title': title,
                'page': table.page
            })
    print(f"Found {len(all_tables)} tables in total")
    return all_tables

def process_tables(all_tables):
    processed_data = []
    for i, table_info in enumerate(all_tables):
        df = table_info['dataframe']
        title = table_info['title']
        page = table_info['page']
        df['Table_Number'] = i + 1
        df['Table_Title'] = title
        df['Page'] = page
        processed_data.append(df)
    if processed_data:
        return pd.concat(processed_data, ignore_index=True)
    else:
        return pd.DataFrame()

def save_tables_to_excel(tables_dict, output_path, document_title=None):
    wb = Workbook()
    wb.remove(wb.active)
    
    for sheet_name, tables in tables_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        current_row = 1

        if document_title:
            ws.cell(row=current_row, column=1, value=document_title)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=1).font = Font(size=16, bold=True)
            current_row += 2

        for idx, table_info in enumerate(tables):
            if 'dataframe' not in table_info:
                print(f"Skipping table in sheet '{sheet_name}' as 'dataframe' key is missing.")
                continue

            df = table_info['dataframe']
            title = table_info['title']
            page = table_info['page']

            ws.cell(row=current_row, column=1, value=f"표 {idx+1}: {title} (페이지 {page})")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=df.shape[1])
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for r in dataframe_to_rows(df, index=False, header=True):
                for c_idx, value in enumerate(r, start=1):
                    ws.cell(row=current_row, column=c_idx, value=value)
                    ws.cell(row=current_row, column=c_idx).alignment = Alignment(wrap_text=True)
                current_row += 1
            current_row += 1

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    wb.save(output_path)
    print(f"Tables have been saved to {output_path}")

def split_text_into_chunks(text, chunk_size=200, overlap=50):
    words = text.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        chunk = " ".join(words[i:i + chunk_size])
        chunks.append(chunk)
    return chunks

def create_index(chunks):
    model = SentenceTransformer('distiluse-base-multilingual-cased')
    embeddings = model.encode(chunks)
    embeddings = embeddings / np.linalg.norm(embeddings, axis=1, keepdims=True)
    dimension = embeddings.shape[1]
    index = faiss.IndexFlatIP(dimension)
    index.add(embeddings.astype('float32'))
    return index, model

def search_rag(query, index, model, chunks, page_numbers=None, k=5, threshold=0.5):
    query_vector = model.encode([query])
    query_vector = query_vector / np.linalg.norm(query_vector, axis=1, keepdims=True)
    distances, indices = index.search(query_vector.astype('float32'), k)
    results = []
    for idx, score in zip(indices[0], distances[0]):
        if score >= threshold:
            result = {'content': chunks[idx], 'score': score}
            if page_numbers:
                result['page'] = page_numbers[idx]
            results.append(result)
    return results

def results_to_dataframe(results, query_type):
    df = pd.DataFrame(results)
    df['type'] = query_type
    return df

def find_pages_with_keyword_in_page(texts_by_page, keyword):
    pages = []
    preprocessed_keyword = preprocess_text(keyword)
    for page_num, text in texts_by_page.items():
        preprocessed_text = preprocess_text(text)
        if preprocessed_keyword in preprocessed_text:
            pages.append(page_num)
    return pages

def extract_text_from_page(page):
    return page.get_text("text")

def main():
    uploads_folder = "uploads"
    output_folder = "output"

    os.makedirs(output_folder, exist_ok=True)

    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_tables.xlsx")
    search_output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_search_results.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        full_text = ""
        page_numbers = []
        texts_by_page = {}
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"
                page_numbers.extend([i+1] * len(page_text.split()))
                texts_by_page[i+1] = page_text

    chunks = split_text_into_chunks(full_text)
    index, model = create_index(chunks)

    select_query = "선택특약"
    select_results = search_rag(select_query, index, model, chunks, page_numbers, k=10, threshold=0.5)
    select_df = results_to_dataframe(select_results, "선택특약")

    select_pages = find_pages_with_keyword_in_page(texts_by_page, "선택특약")
    print("선택특약이 포함된 페이지:", select_pages)

    if not select_pages:
        print("선택특약이 포함된 페이지를 찾지 못했습니다.")
        return

    doc = fitz.open(pdf_path)
    for page_num in select_pages:
        page = doc.load_page(page_num - 1)
        text = extract_text_from_page(page)
        txt_output_path = os.path.join(output_folder, f"page_{page_num}_text.txt")
        with open(txt_output_path, 'w', encoding='utf-8') as f:
            f.write(text)
        print(f"Page {page_num}의 텍스트가 {txt_output_path}에 저장되었습니다.")

    종_sheets = {
        "1종": [],
        "2종": [],
        "3종": []
    }

    종_start_pages = {
        "1종": None,
        "2종": None,
        "3종": None
    }

    for page_num in range(1, len(texts_by_page) + 1):
        page_text = texts_by_page.get(page_num, "")
        for 종 in ["1종", "2종", "3종"]:
            if 종 in page_text and 종_start_pages[종] is None:
                print(f"Page {page_num}: {종}이 검출되었습니다.")
                종_start_pages[종] = page_num
                종_sheets[종].append({
                    'page': page_num,
                    'text': page_text
                })

    injury_query = "상해관련특약"
    injury_results = search_rag(injury_query, index, model, chunks, page_numbers, k=10, threshold=0.5)
    injury_df = results_to_dataframe(injury_results, "상해관련특약")

    disease_query = "질병관련특약"
    disease_results = search_rag(disease_query, index, model, chunks, page_numbers, k=10, threshold=0.5)
    disease_df = results_to_dataframe(disease_results, "질병관련특약")

    final_search_df = pd.concat([select_df, injury_df, disease_df], ignore_index=True)
    final_search_df.to_excel(search_output_excel_path, index=False, engine='openpyxl')
    print(f"검색 결과가 {search_output_excel_path}에 저장되었습니다.")

    stop_keyword = "질병관련 특별약관"
    stop_pages = find_pages_with_keyword_in_page(texts_by_page, stop_keyword)
    if stop_pages:
        print(f"'질병관련 특별약관'이 포함된 페이지 {stop_pages}에서 표 추출을 종료합니다.")
        min_stop_page = min(stop_pages)
    else:
        min_stop_page = float('inf')

    tables_info = detect_table_boundaries(extract_text_with_positions(pdf_path))
    
    tables_sheets = {
        "1종": [],
        "2종": [],
        "3종": [],
        "기타": []  # 1, 2, 3종에 해당하지 않는 테이블을 위한 시트
    }

    for 종, start_page in 종_start_pages.items():
        if start_page is not None:
            end_page = min(종_start_pages.get(next_종(종), float('inf')), min_stop_page) - 1
            종_tables_info = [table for table in tables_info 
                              if any(start_page <= page <= end_page for page in table['pages'])]
            tables = extract_tables_with_camelot(pdf_path, 종_tables_info)
            tables_sheets[종].extend(tables)
        else:
            print(f"{종}의 시작 페이지를 찾지 못했습니다.")

    # 1, 2, 3종에 해당하지 않는 테이블을 '기타' 시트로 분류
    all_종_pages = set()
    for pages in [range(start, 종_start_pages.get(next_종(종), min_stop_page)) 
                  for 종, start in 종_start_pages.items() if start is not None]:
        all_종_pages.update(pages)
    
    other_tables_info = [table for table in tables_info 
                         if not any(page in all_종_pages for page in table['pages'])]
    other_tables = extract_tables_with_camelot(pdf_path, other_tables_info)
    tables_sheets['기타'].extend(other_tables)

    # 첫 번째 페이지에서 문서 제목 추출
    doc = fitz.open(pdf_path)
    first_page = doc.load_page(0)
    first_page_text = extract_text_from_page(first_page)
    document_title = first_page_text.strip().split('\n')[0]

    if any(tables_sheets.values()):
        save_tables_to_excel(tables_sheets, output_excel_path, document_title=document_title)
    else:
        print("추출된 표가 없습니다.")

    print(f"모든 처리된 표가 {output_excel_path}에 저장되었습니다.")

    # 종별 텍스트 정보를 별도의 엑셀 파일로 저장
    text_output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_text_by_type.xlsx")
    with pd.ExcelWriter(text_output_excel_path, engine='openpyxl') as writer:
        for 종, data in 종_sheets.items():
            if data:
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=종, index=False)
    print(f"종별 텍스트 정보가 {text_output_excel_path}에 저장되었습니다.")

def next_종(current_종):
    종_order = ["1종", "2종", "3종"]
    current_index = 종_order.index(current_종)
    if current_index < len(종_order) - 1:
        return 종_order[current_index + 1]
    else:
        return None

if __name__ == "__main__":
    main()