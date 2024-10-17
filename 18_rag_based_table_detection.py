import PyPDF2
import camelot
import pandas as pd
import fitz  # PyMuPDF
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_text_with_positions(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text_chunks = []
        for page_num, page in enumerate(reader.pages):
            text = page.extract_text()
            lines = text.split('\n')
            for line_num, line in enumerate(lines):
                text_chunks.append({
                    'text': line,
                    'page': page_num + 1,
                    'line': line_num + 1
                })
    return text_chunks

def detect_table_boundaries(text_chunks):
    start_patterns = [r'표\s*\d+', r'선택특약\s*내용', r'상해관련\s*특약', r'질병관련\s*특약']
    end_patterns = [r'합\s*계', r'총\s*계', r'주\s*\)', r'※', r'결\s*론']

    tables = []
    table_start = None

    for i, chunk in enumerate(text_chunks):
        text = chunk['text']

        if table_start is None and any(re.search(pattern, text) for pattern in start_patterns):
            table_start = i
            continue

        if table_start is not None and any(re.search(pattern, text) for pattern in end_patterns):
            table_end = i

            start_page = text_chunks[table_start]['page']
            end_page = text_chunks[table_end]['page']
            pages = list(range(start_page, end_page + 1))

            context_start = max(0, table_start - 5)
            context_end = min(len(text_chunks), table_end + 5)
            context = text_chunks[context_start:context_end]

            tables.append({
                'start': text_chunks[table_start],
                'end': text_chunks[table_end],
                'pages': pages,
                'context': context
            })
            table_start = None

    return tables

def extract_tables_with_camelot(pdf_path, tables_info):
    all_tables = []
    doc = fitz.open(pdf_path)
    for table_info in tables_info:
        pages = table_info['pages']
        pages_str = ','.join(map(str, pages))
        print(f"Extracting table from pages {pages_str} using Camelot...")
        tables = camelot.read_pdf(pdf_path, pages=pages_str, flavor='lattice')

        # 여러 페이지에서 추출된 테이블을 하나로 병합
        combined_df = pd.DataFrame()
        for table in tables:
            combined_df = pd.concat([combined_df, table.df], ignore_index=True)

        # 표 바로 위의 텍스트 추출
        first_page_number = pages[0] - 1  # fitz 모듈은 0부터 시작
        page = doc.load_page(first_page_number)
        table_bbox = tables[0]._bbox  # 첫 번째 테이블의 bbox 사용
        text_above_table = extract_text_above_bbox(page, table_bbox)

        all_tables.append({
            'dataframe': combined_df,
            'title': text_above_table.strip(),
            'pages': pages
        })
    print(f"Found {len(all_tables)} tables in total")
    return all_tables

def extract_text_above_bbox(page, bbox):
    x0, y0, x1, y1 = bbox  # bbox: (x0, y0, x1, y1)
    text_blocks = page.get_text("blocks")
    # 테이블 bbox의 y0보다 위에 있는 텍스트 블록 중 가장 아래에 있는 것 선택
    texts_above = []
    for block in text_blocks:
        bx0, by0, bx1, by1, text, block_no = block
        if by1 <= y0:  # 블록의 아래쪽 y좌표가 테이블의 위쪽 y좌표보다 작거나 같으면
            texts_above.append((by1, text))
    if texts_above:
        # y 좌표가 가장 큰 (테이블 바로 위에 있는) 텍스트 선택
        texts_above.sort(reverse=True)
        return texts_above[0][1]
    else:
        return "제목 없음"

def process_tables(all_tables):
    processed_data = []
    for i, table_info in enumerate(all_tables):
        df = table_info['dataframe']
        title = table_info['title']
        pages = table_info['pages']
        df['Table_Number'] = i + 1
        df['Table_Title'] = title
        df['Pages'] = ', '.join(map(str, pages))
        processed_data.append(df)
    return pd.concat(processed_data, ignore_index=True)

def save_to_excel(df_dict, output_path, title=None):
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    for sheet_name, df in df_dict.items():
        ws = wb.create_sheet(title=sheet_name)

        if title:
            ws.cell(row=1, column=1, value=f"{title} - {sheet_name}")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = Font(size=20, bold=True)
            ws.row_dimensions[1].height = 30
            start_row = 2
        else:
            start_row = 1

        current_row = start_row
        # 'Table_Number'로 그룹핑하여 각 테이블을 구분
        grouped = df.groupby('Table_Number')
        for table_number, group in grouped:
            table_title = group['Table_Title'].iloc[0]
            pages = group['Pages'].iloc[0]

            # 테이블 제목 추가
            ws.cell(row=current_row, column=1, value=table_title)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.font = Font(size=14, bold=True)
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            # 헤더 바로 위에 표 위의 텍스트 추가 (폰트 크게, 볼드체)
            header_title = f"{table_title}"
            ws.cell(row=current_row, column=1, value=header_title)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
            header_cell = ws.cell(row=current_row, column=1)
            header_cell.font = Font(size=12, bold=True)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            # 테이블 데이터 작성
            for r_idx, row in enumerate(dataframe_to_rows(group.drop(['Table_Number', 'Table_Title', 'Pages'], axis=1), index=False, header=True), start=current_row):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                # 페이지 번호를 오른편에 추가
                ws.cell(row=r_idx, column=len(row)+1, value=pages)
            current_row = r_idx + 2  # 각 테이블 후에 공백 추가

        # 열 너비 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Data saved to '{output_path}'")

def main():
    uploads_folder = "/workspaces/automation/uploads"
    output_folder = "/workspaces/automation/output"
    
    os.makedirs(output_folder, exist_ok=True)

    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_analysis.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    text_chunks = extract_text_with_positions(pdf_path)
    tables = detect_table_boundaries(text_chunks)

    # 각 타입별 테이블 정보 수집
    types = ["[1종]", "[2종]", "[3종]", "선택특약"]
    type_tables_info = {t: [] for t in types}

    for table in tables:
        table_text = ' '.join([chunk['text'] for chunk in table['context']])
        for t in types:
            if t in table_text:
                type_tables_info[t].append(table)
                break

    # 선택특약 페이지 출력
    if type_tables_info['선택특약']:
        선택특약_pages = [page for table in type_tables_info['선택특약'] for page in table['pages']]
        print(f"선택특약 is on pages: {sorted(set(선택특약_pages))}")

    df_dict = {}
    for insurance_type in ["[1종]", "[2종]", "[3종]"]:
        if type_tables_info[insurance_type]:
            type_tables = type_tables_info[insurance_type]
            camelot_tables = extract_tables_with_camelot(pdf_path, type_tables)
            df = process_tables(camelot_tables)
            df_dict[insurance_type.strip('[]')] = df

    # 선택특약 처리
    if type_tables_info['선택특약']:
        type_tables = type_tables_info['선택특약']
        camelot_tables = extract_tables_with_camelot(pdf_path, type_tables)
        df = process_tables(camelot_tables)
        df_dict['선택특약'] = df

    # 첫 번째 페이지에서 제목 추출
    doc = fitz.open(pdf_path)
    first_page = doc[0]
    page_text = first_page.get_text("text")
    title = page_text.strip().split('\n')[0]
    print(f"Extracted title: {title}")

    # 결과를 엑셀로 저장
    save_to_excel(df_dict, output_excel_path, title=title)

    print(f"All processed data has been saved to {output_excel_path}")

if __name__ == "__main__":
    main()
