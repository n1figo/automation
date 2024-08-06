import os
import base64
import fitz  # PyMuPDF
from openpyxl import Workbook
from PIL import Image, ImageEnhance, ImageFilter
import io
import json
from dotenv import load_dotenv
import pytesseract
from transformers import DetrImageProcessor, DetrForObjectDetection
import torch

# .env 파일에서 환경 변수 로드
load_dotenv()

# 지정된 PDF 파일 경로
PDF_PATH = "/workspaces/automation/uploads/1722922992_5._KB_5.10.10_24.05__0801_v1.0.pdf"
OUTPUT_FOLDER = "highlight_images"

# Hugging Face 모델 및 프로세서 로드
processor = DetrImageProcessor.from_pretrained("facebook/detr-resnet-50")
model = DetrForObjectDetection.from_pretrained("facebook/detr-resnet-50")

def extract_images_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    images = []
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        images.append(img)
    doc.close()
    return images

def detect_highlights(image):
    # 이미지 전처리
    image = image.convert("RGB")
    inputs = processor(images=image, return_tensors="pt")
    
    # 모델 추론
    outputs = model(**inputs)
    
    # 결과 추출
    target_sizes = torch.tensor([image.size[::-1]])
    results = processor.post_process_object_detection(outputs, target_sizes=target_sizes)[0]

    # 하이라이트된 영역 추출 (이 예제에서는 상자 경계로 가정)
    threshold = 0.9
    boxes = []
    for score, label, box in zip(results["scores"], results["labels"], results["boxes"]):
        if score > threshold:
            box = [round(i) for i in box.tolist()]
            boxes.append(box)
    
    return boxes

def ocr_highlight(image, box):
    cropped_image = image.crop((box[0], box[1], box[2], box[3]))
    text = pytesseract.image_to_string(cropped_image, lang='eng')
    return text

def capture_highlight(image, box, output_path):
    cropped_image = image.crop((box[0], box[1], box[2], box[3]))
    cropped_image.save(output_path)

def create_excel_with_highlights(highlights, excel_filename):
    wb = Workbook()
    ws = wb.active
    
    ws.cell(row=1, column=1, value="하이라이트된 텍스트")
    ws.cell(row=1, column=2, value="이미지 파일 경로")
    for i, highlight in enumerate(highlights, start=2):
        ws.cell(row=i, column=1, value=highlight['text'])
        ws.cell(row=i, column=2, value=highlight['image_path'])
    
    wb.save(excel_filename)
    return excel_filename

def process_pdf(pdf_path):
    try:
        print(f"Processing PDF: {pdf_path}")
        
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"The file {pdf_path} does not exist.")
        
        if not os.path.exists(OUTPUT_FOLDER):
            os.makedirs(OUTPUT_FOLDER)
        
        images = extract_images_from_pdf(pdf_path)
        print(f"Extracted {len(images)} pages from PDF")
        
        highlights = []
        for i, image in enumerate(images):
            print(f"Processing page {i+1}")
            detected_boxes = detect_highlights(image)
            for j, box in enumerate(detected_boxes):
                text = ocr_highlight(image, box)
                output_image_path = os.path.join(OUTPUT_FOLDER, f"highlight_page{i+1}_{j+1}.png")
                capture_highlight(image, box, output_image_path)
                highlights.append({
                    'text': text,
                    'image_path': output_image_path
                })
                print(f"Detected highlight on page {i+1}: {text}")
        
        if highlights:
            excel_filename = "output_highlights.xlsx"
            excel_path = create_excel_with_highlights(highlights, excel_filename)
            print(f"Excel file created: {excel_path}")
            print(f"Total highlighted sections: {len(highlights)}")
        else:
            print("No highlights were detected in the PDF.")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    process_pdf(PDF_PATH)
