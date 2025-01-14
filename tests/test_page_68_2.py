import pytest
from pathlib import Path
import json
import pandas as pd
import camelot
import numpy as np
import fitz
from difflib import SequenceMatcher
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import logging
import argparse
import sys
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
from datetime import datetime
import torch
from transformers import AutoModelForObjectDetection, AutoProcessor
from tests.table_analyzer import TableAnalyzer, MergedCell

import sys
import os

# 프로젝트 루트를 Python 경로에 추가
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)

# 이후 src 모듈을 정상적으로 가져올 수 있음
# from src.analyzers.analyzer import AnalyzerClass


@dataclass
class TableQualityMetrics:
    """표 파싱 품질 메트릭"""
    empty_cells_ratio: float
    suspicious_cells_ratio: float
    structure_score: float
    consistency_score: float
    merged_cell_detection_score: float
    text_accuracy_score: float
    overall_score: float
    issues: List[str]
    merged_cell_details: Dict[str, any]
    text_comparison_details: Dict[str, any]

class TableAnalysisEvaluator:
    def __init__(self):
        self.suspicious_patterns = [
            r'^\s*$',              # 빈 셀
            r'^[^\w\s가-힣]+$',    # 특수문자만 있는 셀
            r'^\d{1,2}$',         # 1-2자리 숫자만 있는 셀
            r'^[.]{2,}$'          # 점만 여러 개 있는 셀
        ]

    def _evaluate_text_accuracy(self, extracted_df: pd.DataFrame, pdf_path: str, page_num: int) -> Tuple[float, Dict, List[str]]:
        """텍스트 추출 정확도 평가"""
        issues = []
        text_details = {
            'total_cells': 0,
            'exact_matches': 0,
            'close_matches': 0,
            'mismatches': 0,
            'sample_errors': []
        }

        try:
            # PDF에서 원본 텍스트 추출
            doc = fitz.open(pdf_path)
            page = doc[page_num - 1]
            
            # 표 영역의 텍스트 추출
            page_text = page.get_text("blocks")
            
            # 셀별 비교
            total_similarity = 0
            cell_count = 0
            
            for i, row in extracted_df.iterrows():
                for j, extracted_text in enumerate(row):
                    text_details['total_cells'] += 1
                    
                    try:
                        extracted_text = str(extracted_text).strip()
                        # 가장 가까운 텍스트 블록 찾기
                        closest_text = self._find_closest_text(extracted_text, page_text)
                        
                        if closest_text:
                            similarity = SequenceMatcher(None, closest_text, extracted_text).ratio()
                        else:
                            similarity = 0.0
                            
                        if similarity == 1.0:
                            text_details['exact_matches'] += 1
                        elif similarity >= 0.8:
                            text_details['close_matches'] += 1
                        else:
                            text_details['mismatches'] += 1
                            
                            if len(text_details['sample_errors']) < 5:
                                text_details['sample_errors'].append({
                                    'original': closest_text or "",
                                    'extracted': extracted_text,
                                    'similarity': similarity,
                                    'position': f'({i+1}, {j+1})'
                                })
                        
                        total_similarity += similarity
                        cell_count += 1
                        
                    except Exception as e:
                        issues.append(f"셀 비교 중 오류 ({i+1}, {j+1}): {str(e)}")
                        continue

            doc.close()
            accuracy_score = total_similarity / cell_count if cell_count > 0 else 0.0
            
            if text_details['mismatches'] / text_details['total_cells'] > 0.3:
                issues.append("30% 이상의 셀에서 텍스트 불일치 발견")
            
            return accuracy_score, text_details, issues
            
        except Exception as e:
            issues.append(f"텍스트 정확도 평가 중 오류: {str(e)}")
            return 0.0, text_details, issues

    def _find_closest_text(self, target: str, text_blocks: list) -> str:
        """가장 유사한 텍스트 블록 찾기"""
        best_match = ""
        best_ratio = 0
        
        for block in text_blocks:
            if not isinstance(block, tuple) or len(block) < 4:
                continue
                
            block_text = block[4].strip()
            ratio = SequenceMatcher(None, target, block_text).ratio()
            
            if ratio > best_ratio:
                best_ratio = ratio
                best_match = block_text
                
        return best_match

    def _evaluate_merged_cells(self, df: pd.DataFrame, merged_cells: List) -> Tuple[float, Dict, List[str]]:
        """병합된 셀 검출 품질 평가"""
        issues = []
        merged_cell_details = {
            'total_merged_areas': len(merged_cells),
            'problematic_merges': 0,
            'irregular_patterns': 0
        }

        if not merged_cells:
            return 1.0, merged_cell_details, []

        # 병합된 영역 검증
        for mc in merged_cells:
            if mc.end_row - mc.start_row < 0 or mc.end_col - mc.start_col < 0:
                merged_cell_details['problematic_merges'] += 1
                issues.append(f"잘못된 병합 범위: ({mc.start_row}, {mc.start_col}) -> ({mc.end_row}, {mc.end_col})")

            try:
                merged_area = df.iloc[mc.start_row-1:mc.end_row, mc.start_col-1:mc.end_col]
                non_empty_cells = merged_area.notna().sum().sum()
                if non_empty_cells > 1:
                    merged_cell_details['irregular_patterns'] += 1
                    issues.append(f"병합 영역 내 다중 값 감지: ({mc.start_row}, {mc.start_col})")
            except IndexError:
                merged_cell_details['problematic_merges'] += 1
                issues.append("병합 영역이 표 범위를 벗어남")

        pattern_score = self._analyze_merge_patterns(merged_cells)
        
        problematic_ratio = merged_cell_details['problematic_merges'] / len(merged_cells)
        irregular_ratio = merged_cell_details['irregular_patterns'] / len(merged_cells)
        
        score = (
            (1 - problematic_ratio) * 0.4 +
            (1 - irregular_ratio) * 0.3 +
            pattern_score * 0.3
        )

        return score, merged_cell_details, issues

    def _analyze_merge_patterns(self, merged_cells: List) -> float:
        """병합 패턴의 규칙성 분석"""
        if not merged_cells:
            return 1.0

        pattern_scores = []

        # 수평 병합 패턴 분석
        horizontal_merges = [mc for mc in merged_cells if mc.end_col - mc.start_col > 0]
        if horizontal_merges:
            h_widths = [mc.end_col - mc.start_col + 1 for mc in horizontal_merges]
            h_score = 1 - (np.std(h_widths) / max(np.mean(h_widths), 1))
            pattern_scores.append(h_score)

        # 수직 병합 패턴 분석
        vertical_merges = [mc for mc in merged_cells if mc.end_row - mc.start_row > 0]
        if vertical_merges:
            v_heights = [mc.end_row - mc.start_row + 1 for mc in vertical_merges]
            v_score = 1 - (np.std(v_heights) / max(np.mean(v_heights), 1))
            pattern_scores.append(v_score)

        # 병합 위치의 규칙성 분석
        start_positions = [(mc.start_row, mc.start_col) for mc in merged_cells]
        if len(start_positions) > 1:
            row_diffs = np.diff([pos[0] for pos in start_positions])
            col_diffs = np.diff([pos[1] for pos in start_positions])
            
            position_score = 1 - (
                (np.std(row_diffs) / max(np.mean(abs(row_diffs)), 1) +
                 np.std(col_diffs) / max(np.mean(abs(col_diffs)), 1)
                ) / 2
            )
            pattern_scores.append(position_score)

        return np.mean(pattern_scores) if pattern_scores else 0.0

    def evaluate_table(self, df: pd.DataFrame, pdf_path: str, page_num: int, 
                      merged_cells: List = None) -> TableQualityMetrics:
        """표 품질 평가"""
        empty_ratio = df.isna().sum().sum() / df.size
        suspicious_ratio = self._evaluate_suspicious_cells(df)
        structure_score = self._evaluate_structure(df)
        consistency_score = self._evaluate_consistency(df)
        
        merged_score, merged_details, merged_issues = self._evaluate_merged_cells(
            df, merged_cells or []
        )
        
        text_score, text_details, text_issues = self._evaluate_text_accuracy(
            df, pdf_path, page_num
        )
        
        overall_score = self._calculate_overall_score(
            empty_ratio, suspicious_ratio, structure_score,
            consistency_score, merged_score, text_score
        )
        
        issues = []
        if merged_issues:
            issues.extend(merged_issues)
        if text_issues:
            issues.extend(text_issues)
        
        return TableQualityMetrics(
            empty_cells_ratio=empty_ratio,
            suspicious_cells_ratio=suspicious_ratio,
            structure_score=structure_score,
            consistency_score=consistency_score,
            merged_cell_detection_score=merged_score,
            text_accuracy_score=text_score,
            overall_score=overall_score,
            issues=issues,
            merged_cell_details=merged_details,
            text_comparison_details=text_details
        )

    def _evaluate_suspicious_cells(self, df: pd.DataFrame) -> float:
        """의심스러운 셀 비율 계산"""
        suspicious_count = 0
        total_cells = df.size
        
        for col in df.columns:
            values = df[col].astype(str)
            for pattern in self.suspicious_patterns:
                suspicious_count += values.str.match(pattern).sum()
                
        return suspicious_count / total_cells

    def _evaluate_structure(self, df: pd.DataFrame) -> float:
        """표 구조 평가"""
        scores = []
        
        # 컬럼명 평가
        has_valid_columns = all(isinstance(col, str) and len(col.strip()) > 0 
                              for col in df.columns)
        scores.append(1.0 if has_valid_columns else 0.5)
        
        # 데이터 타입 일관성 평가
        for col in df.columns:
            unique_types = df[col].apply(type).unique()
            scores.append(1.0 if len(unique_types) == 1 else 
                        0.7 if len(unique_types) == 2 else 0.3)
        
        # 행 길이 일관성 평가
        row_lengths = df.apply(lambda x: len(x.dropna()), axis=1)
        length_consistency = 1.0 - (row_lengths.std() / row_lengths.mean() 
                                  if row_lengths.mean() > 0 else 1.0)
        scores.append(length_consistency)
        
        return np.mean(scores)

    def _evaluate_consistency(self, df: pd.DataFrame) -> float:
        """데이터 일관성 평가"""
        scores = []
        
        # 값의 길이 일관성
        for col in df.columns:
            values = df[col].astype(str)
            lengths = values.str.len()
            length_variance = lengths.std() / lengths.mean() if lengths.mean() > 0 else 1.0
            scores.append(1.0 - min(length_variance, 1.0))
        
        # 형식 일관성
        for col in df.columns:
            values = df[col].astype(str)
            has_numbers = values.str.contains(r'\d').sum()
            if has_numbers > 0:
                number_format_consistency = values.str.match(r'^\d+([,.]\d+)?$').sum() / has_numbers
                scores.append(number_format_consistency)
        
        return np.mean(scores) if scores else 0.0

    def _calculate_overall_score(self, empty_ratio: float, suspicious_ratio: float,
                               structure_score: float, consistency_score: float,
                               merged_score: float, text_score: float) -> float:
        """전체 품질 점수 계산"""
        weights = {
            'empty_ratio': 0.15,
            'suspicious_ratio': 0.15,
            'structure_score': 0.15,
            'consistency_score': 0.15,
            'merged_score': 0.2,
            'text_score': 0.2
        }
        
        score = (
            (1 - empty_ratio) * weights['empty_ratio'] +
            (1 - suspicious_ratio) * weights['suspicious_ratio'] +
            structure_score * weights['structure_score'] +
            consistency_score * weights['consistency_score'] +
            merged_score * weights['merged_score'] +
            text_score * weights['text_score']
        )
        
        return round(score * 100, 2)

    def format_evaluation_report(self, metrics: TableQualityMetrics) -> str:
        """평가 보고서 포맷팅"""
        report = [
            "=== 표 품질 평가 보고서 ===",
            f"전체 품질 점수: {metrics.overall_score:.1f}/100점",
            "",
            "세부 메트릭:",
            f"- 빈 셀 비율: {metrics.empty_cells_ratio:.1%}",
            f"- 의심 셀 비율: {metrics.suspicious_cells_ratio:.1%}",
            f"- 구조 점수: {metrics.structure_score:.2f}",
            f"- 일관성 점수: {metrics.consistency_score:.2f}",
            f"- 병합 셀 검출 점수: {metrics.merged_cell_detection_score:.2f}",
            f"- 텍스트 정확도 점수: {metrics.text_accuracy_score:.2f}",
            "",
            "병합 셀 분석:",
            f"- 총 병합 영역 수: {metrics.merged_cell_details['total_merged_areas']}",
            f"- 문제있는 병합 수: {metrics.merged_cell_details['problematic_merges']}",
            f"- 불규칙한 패턴 수: {metrics.merged_cell_details['irregular_patterns']}",
            "",
            "텍스트 정확도 분석:",
            f"- 총 셀 수: {metrics.text_comparison_details['total_cells']}",
            f"- 정확히 일치: {metrics.text_comparison_details['exact_matches']}",
            f"- 유사 일치 (80% 이상): {metrics.text_comparison_details['close_matches']}",
            f"- 불일치: {metrics.text_comparison_details['mismatches']}"
        ]

        if metrics.text_comparison_details['sample_errors']:
            report.extend([
                "",
                "텍스트 오류 예시:"
            ])
            for error in metrics.text_comparison_details['sample_errors']:
                report.extend([
                    f"- 위치 {error['position']}:",
                    f"  원본: '{error['original']}'",
                    f"  추출: '{error['extracted']}'",
                    f"  유사도: {error['similarity']:.2f}"
                ])

        if metrics.issues:
            report.extend([
                "",
                "발견된 문제:",
                *[f"- {issue}" for issue in metrics.issues]
            ])

        return "\n".join(report)

class EnhancedTableAnalyzer:
    def __init__(self):
        self._setup_logging()
        self.evaluator = TableAnalysisEvaluator()
        
        # AI 모델 초기화 추가
        self.model = AutoModelForObjectDetection.from_pretrained("microsoft/table-transformer-detection")
        self.processor = AutoProcessor.from_pretrained("microsoft/table-transformer-detection")
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.model.to(self.device)

    def _setup_logging(self):
        """로깅 설정"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler(
                    f'table_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
                )
            ]
        )
        self.logger = logging.getLogger(__name__)

    def basic_camelot_parse(self, pdf_path: str, page_num: int) -> pd.DataFrame:
        """기본 Camelot 파싱"""
        try:
            tables = camelot.read_pdf(
                pdf_path,
                pages=str(page_num),
                flavor='lattice',
                line_scale=40
            )
            if len(tables) > 0:
                self.logger.info(f"페이지 {page_num}: Camelot 파싱 성공")
                return tables[0].df
            else:
                self.logger.warning(f"페이지 {page_num}: 표를 찾을 수 없음")
                return pd.DataFrame()
        except Exception as e:
            self.logger.error(f"페이지 {page_num} Camelot 파싱 오류: {str(e)}")
            return pd.DataFrame()

    def ai_enhanced_parse(self, pdf_path: str, page_num: int) -> pd.DataFrame:
        """AI 향상된 표 파싱"""
        try:
            # 표 영역 감지
            image = self.extract_page_image(pdf_path, page_num)
            inputs = self.processor(images=image, return_tensors="pt")
            inputs = {k: v.to(self.device) for k, v in inputs.items()}
            
            with torch.no_grad():
                outputs = self.model(**inputs)
            
            # 결과 처리 및 표 추출
            predicted_tables = self.process_model_outputs(outputs, image)
            
            # 데이터 검증 및 보정
            validated_data = self.validate_and_correct_data(predicted_tables)
            
            return validated_data
            
        except Exception as e:
            self.logger.error(f"AI 파싱 중 오류: {str(e)}")
            return pd.DataFrame()

    def analyze_pages(self, pdf_path: str, pages: List[int], output_dir: str) -> str:
        """여러 페이지 분석"""
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            for page_num in pages:
                self.logger.info(f"\n=== 페이지 {page_num} 분석 시작 ===")
                
                # 1. 기본 Camelot 파싱
                basic_df = self.basic_camelot_parse(pdf_path, page_num)
                basic_metrics = self.evaluator.evaluate_table(basic_df, pdf_path, page_num)
                
                self.logger.info(f"\n[기본 Camelot 파싱 평가 - 페이지 {page_num}]")
                self.logger.info(self.evaluator.format_evaluation_report(basic_metrics))
                
                ws_basic = wb.create_sheet(f"P{page_num}_Camelot")
                self._write_dataframe_to_sheet(
                    ws_basic, 
                    basic_df, 
                    f"페이지 {page_num} - Camelot 파싱"
                )
                self._write_metrics_to_sheet(ws_basic, basic_metrics, start_row=len(basic_df) + 5)
                
                # 2. TableAnalyzer 분석
                try:
                    analyzer = TableAnalyzer()
                    result = analyzer.analyze_page(pdf_path, page_num)
                    enhanced_metrics = self.evaluator.evaluate_table(
                        result['data'], 
                        pdf_path, 
                        page_num,
                        result.get('merged_cells', [])
                    )
                    
                    self.logger.info(f"\n[향상된 분석 평가 - 페이지 {page_num}]")
                    self.logger.info(self.evaluator.format_evaluation_report(enhanced_metrics))
                    
                    ws_enhanced = wb.create_sheet(f"P{page_num}_AI")
                    self._write_analysis_result_to_sheet(
                        ws_enhanced, 
                        result, 
                        enhanced_metrics
                    )
                    
                except Exception as e:
                    self.logger.error(f"TableAnalyzer 분석 오류 (페이지 {page_num}): {str(e)}")

                # 3. AI 향상된 분석
                try:
                    ai_df = self.ai_enhanced_parse(pdf_path, page_num)
                    ai_metrics = self.evaluator.evaluate_table(ai_df, pdf_path, page_num)
                    
                    ws_ai = wb.create_sheet(f"P{page_num}_AI")
                    self._write_dataframe_to_sheet(ws_ai, ai_df, f"페이지 {page_num} - AI 분석")
                    self._write_metrics_to_sheet(ws_ai, ai_metrics, start_row=len(ai_df) + 5)
                    
                except Exception as e:
                    self.logger.error(f"AI 분석 오류 (페이지 {page_num}): {str(e)}")

                # 4. 비교 시트 추가 (3가지 방식 비교)
                ws_compare = wb.create_sheet(f"P{page_num}_Compare")
                self._write_comparison_to_sheet(
                    ws_compare,
                    basic_metrics,
                    enhanced_metrics,
                    ai_metrics,
                    page_num
                )

                self.logger.info(f"페이지 {page_num} 분석 완료")

            output_path = Path(output_dir) / f"table_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            self._apply_styles_to_workbook(wb)
            wb.save(str(output_path))
            
            self.logger.info(f"\n분석 결과 저장됨: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"분석 중 오류 발생: {str(e)}")
            raise

    def _write_metrics_to_sheet(self, ws, metrics: TableQualityMetrics, start_row: int):
        """품질 메트릭을 워크시트에 작성"""
        headers = [
            ("전체 품질 점수", f"{metrics.overall_score:.1f}/100점"),
            ("빈 셀 비율", f"{metrics.empty_cells_ratio:.1%}"),
            ("의심 셀 비율", f"{metrics.suspicious_cells_ratio:.1%}"),
            ("구조 점수", f"{metrics.structure_score:.2f}"),
            ("일관성 점수", f"{metrics.consistency_score:.2f}"),
            ("병합 셀 검출 점수", f"{metrics.merged_cell_detection_score:.2f}"),
            ("텍스트 정확도 점수", f"{metrics.text_accuracy_score:.2f}")
        ]
        
        ws.merge_cells(
            start_row=start_row, 
            start_column=1, 
            end_row=start_row, 
            end_column=2
        )
        ws.cell(row=start_row, column=1, value="품질 평가 결과").font = Font(bold=True)
        
        for idx, (key, value) in enumerate(headers, start_row + 1):
            ws.cell(row=idx, column=1, value=key).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)

    def _write_analysis_result_to_sheet(self, ws, result: Dict, metrics: TableQualityMetrics):
        """분석 결과와 메트릭을 워크시트에 작성"""
        if 'data' in result:
            self._write_dataframe_to_sheet(ws, result['data'], "AI 분석 결과")
            
        data_rows = len(result.get('data', [])) + 3
        
        # AI 분석 추가 정보
        additional_info = [
            ("병합된 셀 수", len(result.get('merged_cells', []))),
            ("표 구조", result.get('table_structure', 'N/A')),
            ("AI 신뢰도 점수", f"{result.get('confidence_score', 'N/A')}")
        ]
        
        ws.merge_cells(
            start_row=data_rows, 
            start_column=1, 
            end_row=data_rows, 
            end_column=2
        )
        ws.cell(row=data_rows, column=1, value="AI 분석 정보").font = Font(bold=True)
        
        for idx, (key, value) in enumerate(additional_info, data_rows + 1):
            ws.cell(row=idx, column=1, value=key).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)

        # 품질 메트릭 작성
        self._write_metrics_to_sheet(ws, metrics, data_rows + len(additional_info) + 2)

    def _write_comparison_to_sheet(self, ws, basic_metrics: TableQualityMetrics, 
                                 enhanced_metrics: TableQualityMetrics,
                                 ai_metrics: TableQualityMetrics,
                                 page_num: int):
        """세 가지 분석 방식 비교"""
        ws.title = f"P{page_num}_Compare"
        
        # 제목
        ws.merge_cells('A1:D1')
        ws['A1'] = f"페이지 {page_num} - 분석 방식 비교"
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 헤더
        headers = ['메트릭', 'Camelot', 'Enhanced', 'AI']
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header).font = Font(bold=True)
            
        # 비교 데이터
        comparison_data = [
            ("전체 품질 점수", 
             f"{basic_metrics.overall_score:.1f}", 
             f"{enhanced_metrics.overall_score:.1f}",
             f"{ai_metrics.overall_score:.1f}"),
            ("빈 셀 비율", 
             f"{basic_metrics.empty_cells_ratio:.1%}", 
             f"{enhanced_metrics.empty_cells_ratio:.1%}",
             f"{ai_metrics.empty_cells_ratio:.1%}"),
            # ... 나머지 메트릭들 추가
        ]
        
        for row, (metric, basic, enhanced, ai) in enumerate(comparison_data, 3):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=basic)
            ws.cell(row=row, column=3, value=enhanced)
            ws.cell(row=row, column=4, value=ai)
            
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, title: str):
        """데이터프레임을 워크시트에 작성"""
        ws.title = title[:31]  # 시트 이름 길이 제한
        
        # 제목 작성
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws['A1'] = title
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 열 헤더 작성
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=str(column))
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="CCCCCC")
            
        # 데이터 작성
        for row_idx, row in enumerate(df.values, 3):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value))
                
        # 열 너비 자동 조정
        for column in ws.columns:
            length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(length + 2, 50)

    def _apply_styles_to_workbook(self, wb: Workbook):
        """워크북 전체에 스타일 적용"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

def parse_page_numbers(page_str: str) -> list[int]:
    """
    페이지 범위 문자열을 파싱하여 실제 페이지 번호 리스트 반환
    
    예시:
    "1,3-5,7" -> [1, 3, 4, 5, 7]
    "1-3,5,7-9" -> [1, 2, 3, 5, 7, 8, 9]
    
    Args:
        page_str: 페이지 범위 문자열 (예: "1,3-5,7")
        
    Returns:
        정수 리스트로 변환된 페이지 번호들
        
    Raises:
        ValueError: 잘못된 형식의 입력일 경우
    """
    pages = set()
    
    try:
        # 콤마로 구분된 각 부분 처리
        for part in page_str.strip().split(','):
            if '-' in part:
                # 범위 처리 (예: "3-5")
                start, end = map(int, part.split('-'))
                if start > end:
                    raise ValueError(f"잘못된 범위: {start}-{end}")
                pages.update(range(start, end + 1))
            else:
                # 단일 페이지 처리
                pages.add(int(part))
                
        # 정렬된 리스트로 변환
        result = sorted(list(pages))
        
        # 유효성 검사
        if not result:
            raise ValueError("페이지 번호가 지정되지 않았습니다")
        if any(p < 1 for p in result):
            raise ValueError("페이지 번호는 1 이상이어야 합니다")
            
        return result
        
    except ValueError as e:
        raise ValueError(f"페이지 범위 파싱 오류: {str(e)}")
    except Exception as e:
        raise ValueError(f"잘못된 페이지 범위 형식입니다: {str(e)}")

def main():
    parser = argparse.ArgumentParser(description='PDF 표 분석기')
    parser.add_argument('pdf_path', help='PDF 파일 경로')
    parser.add_argument('pages', help='분석할 페이지 (예: 1,3-5,7)')
    parser.add_argument('--output', '-o', help='출력 디렉토리', default='output')
    
    args = parser.parse_args()
    
    try:
        # 페이지 번호 파싱
        pages = parse_page_numbers(args.pages)
        print(f"\n분석할 페이지: {pages}")
        
        # PDF 파일 존재 확인
        if not Path(args.pdf_path).exists():
            print(f"\n오류: PDF 파일을 찾을 수 없습니다: {args.pdf_path}")
            return 1
            
        # 분석 실행
        analyzer = EnhancedTableAnalyzer()
        output_path = analyzer.analyze_pages(args.pdf_path, pages, args.output)
        print(f"\n분석이 완료되었습니다. 결과 파일: {output_path}")
        return 0
        
    except Exception as e:
        print(f"\n오류 발생: {str(e)}")
        return 1

if __name__ == "__main__":
    sys.exit(main())

# TestEnhancedAnalysis 클래스 추가
class TestEnhancedAnalysis:
    @pytest.fixture
    def test_data_dir(self):
        """테스트 데이터 디렉토리"""
        test_dir = Path("tests/test_data")
        test_dir.mkdir(parents=True, exist_ok=True)
        return test_dir

    @pytest.fixture
    def output_dir(self):
        """결과 파일 저장 디렉토리"""
        output_dir = Path("tests/output")
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir

    @pytest.fixture
    def analyzer(self):
        return EnhancedTableAnalyzer()

    def test_multi_version_analysis(self, analyzer, test_data_dir, output_dir):
        """세 가지 버전의 분석 통합 테스트"""
        pdf_path = test_data_dir / "test.pdf"
        output_path = output_dir / "multi_version_analysis.xlsx"
        
        try:
            result_path = analyzer.analyze_pages(
                str(pdf_path),
                pages=[68],
                output_dir=str(output_dir)
            )
            
            # 결과 검증
            assert Path(result_path).exists()
            wb = load_workbook(result_path)
            
            # 각 시트 존재 확인
            expected_sheets = [
                "P68_Camelot",
                "P68_Enhanced",
                "P68_AI",
                "P68_Compare"
            ]
            for sheet in expected_sheets:
                assert sheet in wb.sheetnames, f"시트 없음: {sheet}"
            
            print(f"\n분석 결과가 저장됨: {result_path}")
            print("포함된 시트:")
            for sheet in wb.sheetnames:
                print(f"- {sheet}")
                
            return result_path
            
        except Exception as e:
            pytest.fail(f"분석 실패: {str(e)}")