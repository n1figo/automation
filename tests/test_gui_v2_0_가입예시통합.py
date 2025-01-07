# 표준 라이브러리
import os
import re
import sys
import threading
from datetime import datetime
from typing import Dict, List, Tuple, Optional

# 데이터 처리 및 분석
import numpy as np
import pandas as pd
from scipy.spatial.distance import cosine

# PDF 처리
import PyPDF2
import fitz
import camelot

# 이미지 처리
import cv2
from PIL import Image

# Excel 처리
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, 
    Border, 
    Side, 
    Alignment, 
    Font
)

# GUI
import tkinter as tk
from tkinter import filedialog, ttk, messagebox

# AI/ML
from sentence_transformers import SentenceTransformer

# 로깅
import logging

# 프로그램 상수
CONSTANTS = {
    # 파일 관련 상수
    'SUPPORTED_EXTENSIONS': ['.pdf'],
    'MAX_FILE_SIZE': 100 * 1024 * 1024,  # 100MB
    
    # 표 추출 관련 상수
    'TABLE_EXTRACTION': {
        'MIN_ROWS': 2,
        'MIN_COLS': 2,
        'MIN_ACCURACY': 50,  # 최소 정확도 (%)
        'MAX_EMPTY_RATIO': 0.5,  # 최대 빈 셀 비율
    },
    
    # 처리 타임아웃 설정
    'TIMEOUTS': {
        'PAGE_PROCESSING': 60,  # 페이지당 최대 처리 시간 (초)
        'TOTAL_PROCESSING': 3600,  # 전체 최대 처리 시간 (초)
    },
    
    # GUI 관련 상수
    'GUI': {
        'WINDOW_SIZE': '800x600',
        'MAX_LOG_LINES': 1000,
        'PROGRESS_UPDATE_INTERVAL': 100,  # ms
    },
    
    # 출력 파일 관련 상수
    'OUTPUT': {
        'MAX_SHEET_NAME_LENGTH': 31,  # Excel 시트 이름 최대 길이
        'MAX_CELL_LENGTH': 32767,  # Excel 셀 최대 길이
    }
}

# 오류 메시지
ERROR_MESSAGES = {
    'FILE_NOT_FOUND': "PDF 파일을 찾을 수 없습니다.",
    'FILE_TOO_LARGE': "파일 크기가 너무 큽니다. (최대 100MB)",
    'INVALID_FILE_TYPE': "지원되지 않는 파일 형식입니다.",
    'EXTRACTION_FAILED': "표 추출에 실패했습니다.",
    'PROCESSING_TIMEOUT': "처리 시간이 초과되었습니다.",
    'MEMORY_ERROR': "메모리가 부족합니다.",
    'UNKNOWN_ERROR': "알 수 없는 오류가 발생했습니다."
}

# 로깅 설정
LOGGING_CONFIG = {
    'version': 1,
    'disable_existing_loggers': False,
    'formatters': {
        'detailed': {
            'format': '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        },
        'simple': {
            'format': '%(levelname)s - %(message)s'
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'level': 'INFO',
            'formatter': 'simple'
        },
        'file': {
            'class': 'logging.FileHandler',
            'filename': 'pdf_analyzer.log',
            'level': 'DEBUG',
            'formatter': 'detailed',
            'encoding': 'utf-8'
        }
    },
    'loggers': {
        '': {
            'handlers': ['console', 'file'],
            'level': 'DEBUG',
            'propagate': True
        }
    }
}

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class HighlightDetector:
    def __init__(self):
        self.saturation_threshold = 30
        self.kernel_size = (5, 5)

    def pdf_to_image(self, page: fitz.Page) -> np.ndarray:
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return np.array(img)

    def detect_highlights(self, image: np.ndarray) -> List[np.ndarray]:
        hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
        s = hsv[:, :, 1]
        v = hsv[:, :, 2]

        saturation_mask = s > self.saturation_threshold
        _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)

        kernel = np.ones(self.kernel_size, np.uint8)
        cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
        cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)

        contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        return contours

    def get_highlight_regions(self, contours: List[np.ndarray], image_height: int) -> List[Tuple[float, float]]:
        regions = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            top = image_height - (y + h)
            bottom = image_height - y
            regions.append((top, bottom))
        return regions

    def check_highlight(self, row_range: Tuple[float, float], highlight_regions: List[Tuple[float, float]]) -> bool:
        row_top, row_bottom = row_range
        for region_top, region_bottom in highlight_regions:
            if (region_top <= row_top <= region_bottom) or \
               (region_top <= row_bottom <= region_bottom) or \
               (row_top <= region_top <= row_bottom) or \
               (row_top <= region_bottom <= row_bottom):
                return True
        return False


class TableExtractor:
    def __init__(self):
        self.matcher = TitleMatcher()
        self.highlight_detector = HighlightDetector()
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]
        self.max_distance = 50

    def extract_tables_from_section(self, pdf_path: str, start_page: int, end_page: int) -> List[Tuple[str, pd.DataFrame, int]]:
        """섹션에서 표 추출"""
        try:
            results = []
            
            for page_num in range(start_page, end_page):
                logger.info(f"Processing page {page_num + 1}")
                doc = fitz.open(pdf_path)
                page = doc[page_num]

                # 이미지 변환 및 하이라이트 감지
                image = self.highlight_detector.pdf_to_image(page)
                contours = self.highlight_detector.detect_highlights(image)
                highlight_regions = self.highlight_detector.get_highlight_regions(contours, image.shape[0])

                # 제목 추출
                titles = self.get_titles_with_positions(page)

                success = False
                tables = None

                # Stream 모드 시도
                try:
                    logger.info(f"Trying stream mode for page {page_num + 1}")
                    tables = camelot.read_pdf(
                        pdf_path,
                        pages=str(page_num + 1),
                        flavor='stream',
                        edge_tol=100,
                        row_tol=15,
                        split_text=True,
                        strip_text=' .\n',
                        columns=['145,280,420,560']
                    )
                    
                    if len(tables) > 0:
                        success = True
                        logger.info(f"Stream mode successful for page {page_num + 1}")

                except Exception as e:
                    logger.warning(f"Stream mode failed for page {page_num + 1}: {str(e)}")

                # Stream 모드 실패시 Lattice 모드 시도
                if not success:
                    try:
                        logger.info(f"Trying lattice mode for page {page_num + 1}")
                        tables = camelot.read_pdf(
                            pdf_path,
                            pages=str(page_num + 1),
                            flavor='lattice',
                            line_scale=40,
                            process_background=True
                        )
                        if len(tables) > 0:
                            success = True
                            logger.info(f"Lattice mode successful for page {page_num + 1}")

                    except Exception as e:
                        logger.error(f"Lattice mode failed for page {page_num + 1}: {str(e)}")
                        continue

                if tables and success:
                    page_height = page.rect.height
                    logger.info(f"Processing {len(tables)} tables from page {page_num + 1}")

                    for table_idx, table in enumerate(tables):
                        df = self.process_table_with_highlights(
                            table, 
                            highlight_regions, 
                            image.shape[0], 
                            table_idx, 
                            page_num + 1
                        )
                        
                        df = self.clean_table(df)
                        
                        if not df.empty and len(df.columns) >= 2:
                            table_position = self.get_table_positions(table, page_height)
                            title, distance = self.match_title_to_table(titles, table_position)
                            
                            if not title:
                                title = f"Table_{table_idx + 1}"
                                
                            results.append((title, df, page_num + 1))
                            logger.info(f"Successfully processed table {table_idx + 1} from page {page_num + 1}")

                doc.close()

            return results

        except Exception as e:
            logger.error(f"Error extracting tables from section: {str(e)}")
            return []

    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """표 데이터 정제"""
        try:
            if df.empty:
                return df

            # 기본 정제
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            
            # 모든 데이터를 문자열로 변환
            df = df.astype(str)
            
            # 문자열 정제
            for col in df.columns:
                df[col] = df[col].str.strip()
                df[col] = df[col].str.replace('\s+', ' ', regex=True)
            
            # 주석 행 제거 (첫 번째 열에 대해서만)
            df = df[~df.iloc[:, 0].str.contains("^※|^주\)", regex=True, na=False)]
            
            # 빈 값 처리
            df = df.replace(r'^\s*$', '', regex=True)
            df = df.replace('nan', '')
            
            # 중복 제거
            df = df.drop_duplicates()
            
            return df
                
        except Exception as e:
            logger.error(f"Table cleaning error: {str(e)}")
            return df

    def get_titles_with_positions(self, page) -> List[Dict]:
        """페이지에서 제목과 위치 추출"""
        titles = []
        blocks = page.get_text("dict")["blocks"]
        
        for block in blocks:
            if block.get("lines"):
                text = " ".join([span["text"] for line in block["lines"] 
                               for span in line["spans"]]).strip()
                
                if text and any(re.search(pattern, text) for pattern in self.title_patterns):
                    titles.append({
                        "text": text,
                        "y_top": block["bbox"][1],
                        "y_bottom": block["bbox"][3],
                        "bbox": block["bbox"],
                        "used": False
                    })

        return sorted(titles, key=lambda x: x["y_top"])

    def get_table_positions(self, table, page_height):
        """표의 위치 정보 추출"""
        try:
            x0, y0, x1, y1 = table._bbox
            return {
                "y_top": page_height - y1,
                "y_bottom": page_height - y0,
                "bbox": (x0, page_height - y1, x1, page_height - y0)
            }
        except:
            return {
                "y_top": 0,
                "y_bottom": page_height,
                "bbox": (0, 0, 0, page_height)
            }

    def match_title_to_table(self, titles, table_position):
        """표와 가장 가까운 제목 매칭"""
        best_title = None
        min_distance = float('inf')

        for title in titles:
            if title["used"]:
                continue

            distance = table_position["y_top"] - title["y_bottom"]
            if 0 < distance < self.max_distance and distance < min_distance:
                best_title = title
                min_distance = distance

        if best_title:
            best_title["used"] = True
            return best_title["text"], min_distance
        return None, None

    def process_table_with_highlights(self, table, highlight_regions, page_height, table_index, page_num):
        """하이라이트된 영역이 있는 표 처리"""
        try:
            df = table.df.copy()
            x1, y1, x2, y2 = table._bbox
            
            if len(df) == 0:
                return df

            table_height = y2 - y1
            row_height = table_height / len(df)
            processed_data = []

            for row_index in range(len(df)):
                row_data = df.iloc[row_index].copy()
                row_top = y2 - (row_index + 1) * row_height
                row_bottom = y2 - row_index * row_height

                row_highlighted = self.highlight_detector.check_highlight(
                    (row_top, row_bottom), 
                    highlight_regions
                )

                row_data['변경사항'] = '추가' if row_highlighted else ''
                row_data['Table_Number'] = table_index + 1
                row_data['페이지'] = page_num
                processed_data.append(row_data)

            return pd.DataFrame(processed_data)
            
        except Exception as e:
            logger.error(f"Table processing error: {str(e)}")
            return pd.DataFrame()
    

class TitleMatcher:
    def __init__(self):
        try:
            # 현재 작업 디렉토리 기준으로 모델 경로 설정
            model_path = "models/distiluse-base-multilingual-cased-v1"
            
            logger.info(f"로컬 모델 로드 시도: {model_path}")
            
            # 로컬 모델 로드
            try:
                self.model = SentenceTransformer(model_path)
                logger.info("로컬 모델 로드 성공")
            except Exception as e:
                error_msg = f"로컬 모델 로드 실패: {str(e)}"
                logger.error(error_msg)
                raise
            
            self.title_patterns = [
                r'기본계약',
                r'의무부가계약',
                r'[\w\s]+관련\s*특약',
                r'[\w\s]+계약',
                r'[\w\s]+보장'
            ]
            self.max_distance = 50
            
        except Exception as e:
            error_msg = f"모델 초기화 실패: {str(e)}"
            logger.error(error_msg)
            raise

    def get_embedding(self, text: str) -> np.ndarray:
        try:
            return self.model.encode(text)
        except Exception as e:
            logger.error(f"임베딩 생성 실패: {str(e)}")
            raise

    def calculate_similarity(self, text1: str, text2: str) -> float:
        try:
            emb1 = self.get_embedding(text1)
            emb2 = self.get_embedding(text2)
            return 1 - cosine(emb1, emb2)
        except Exception as e:
            logger.error(f"유사도 계산 실패: {str(e)}")
            raise

class InsuranceDocumentAnalyzer:
    def __init__(self):
        self.section_patterns = {
            "종류": r'\[(\d)종\]',
        }
        self.section_pages = {"[1종]": None, "[2종]": None, "[3종]": None}
        self.section_ranges = {}
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]

    def find_section_pages(self, pdf_path: str) -> Dict[str, int]:
        try:
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            
            # 각 페이지 검사
            for page_num in range(total_pages):
                text = doc[page_num].get_text()
                
                # 종 구분 찾기
                matches = re.finditer(self.section_patterns["종류"], text)
                for match in matches:
                    종_type = f"[{match.group(1)}종]"
                    if self.section_pages[종_type] is None:
                        self.section_pages[종_type] = page_num
                        logger.info(f"{종_type} 시작 페이지: {page_num + 1}")

            # 섹션 범위 설정
            self._set_section_ranges(total_pages)

            doc.close()
            return self.section_pages

        except Exception as e:
            logger.error(f"섹션 페이지 검색 중 오류: {str(e)}")
            return {}

    def _set_section_ranges(self, total_pages):
        # 유효한 섹션 찾기
        found_sections = [v for v in self.section_pages.values() if v is not None]
        
        if not found_sections:
            logger.warning("종 구분을 찾지 못했습니다. 전체 문서를 처리합니다.")
            self.section_ranges["전체"] = (0, total_pages)
            return

        # 섹션 범위 설정
        sorted_pages = sorted(
            [(k, v) for k, v in self.section_pages.items() if v is not None],
            key=lambda x: x[1]
        )

        for i, (종_type, start_page) in enumerate(sorted_pages):
            if i + 1 < len(sorted_pages):
                end_page = sorted_pages[i + 1][1]
            else:
                end_page = total_pages
                
            self.section_ranges[종_type] = (start_page, end_page)
            logger.info(f"{종_type} 범위: {start_page + 1} ~ {end_page}")

class ExcelWriter:
    @staticmethod
    def save_to_excel(sections_data: Dict[str, List[Tuple[str, pd.DataFrame, int]]], output_path: str):
        try:
            wb = Workbook()
            # 기본 시트 삭제하기 전에 새 시트 생성
            default_sheet = wb.active
            
            # 테두리 스타일 정의
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 데이터가 있는 경우에만 처리
            if sections_data:
                for section, tables in sections_data.items():
                    if not tables:  # 빈 테이블 스킵
                        continue
                        
                    # 시트 이름 생성 (최대 31자로 제한)
                    sheet_name = (section.replace("[", "").replace("]", "") or "전체")[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    current_row = 1

                    for title, df, page_num in tables:
                        # 섹션 제목 추가
                        section_title = "가입담보표" if "전체" in sheet_name else sheet_name
                        section_cell = ws.cell(row=current_row, column=1, value=section_title)
                        section_cell.font = Font(bold=True, size=14)
                        current_row += 2
                        
                        # 표 제목 작성
                        title_cell = ws.cell(row=current_row, column=1, 
                                           value=f"{title} (페이지: {page_num})")
                        title_cell.font = Font(bold=True, size=12)
                        title_cell.fill = PatternFill(start_color='E6E6E6',
                                                    end_color='E6E6E6',
                                                    fill_type='solid')
                        
                        # 제목 셀 병합
                        ws.merge_cells(start_row=current_row, start_column=1, 
                                     end_row=current_row, end_column=len(df.columns))
                        current_row += 2

                        # 열 헤더 작성
                        for col_idx, col_name in enumerate(df.columns, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color='F2F2F2',
                                                  end_color='F2F2F2',
                                                  fill_type='solid')
                            cell.border = border
                            cell.alignment = Alignment(wrap_text=True, 
                                                     horizontal='center',
                                                     vertical='center')
                        current_row += 1

                        # 데이터 작성
                        for _, row in df.iterrows():
                            for col_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=current_row, column=col_idx, value=value)
                                cell.border = border
                                cell.alignment = Alignment(wrap_text=True,
                                                         vertical='center')
                                
                                # '추가' 표시가 있는 행 하이라이트
                                if '변경사항' in df.columns and row['변경사항'] == '추가':
                                    cell.fill = PatternFill(start_color='FFFF00',
                                                          end_color='FFFF00',
                                                          fill_type='solid')
                            current_row += 1
                        
                        current_row += 2

                        # 열 너비 자동 조정
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                                adjusted_width = min(max_length + 2, 50)  # 최대 50
                                ws.column_dimensions[column_letter].width = adjusted_width

                # 기본 시트가 있으면 삭제
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                    
                # 데이터가 없는 경우 기본 시트 유지
                if len(wb.sheetnames) == 0:
                    ws = wb.create_sheet(title="추출 결과")
                    ws.cell(row=1, column=1, value="추출된 데이터가 없습니다.")
                    ws.column_dimensions['A'].width = 30

            else:  # sections_data가 비어있는 경우
                default_sheet.title = "추출 결과"
                default_sheet.cell(row=1, column=1, value="추출된 데이터가 없습니다.")
                default_sheet.column_dimensions['A'].width = 30

            # 첫 번째 시트를 활성화
            if wb.sheetnames:
                wb.active = 0

            wb.save(output_path)
            logger.info(f"Excel 파일 저장 완료: {output_path}")

        except Exception as e:
            logger.error(f"Excel 저장 중 오류 발생: {str(e)}")
            raise

    @staticmethod
    def format_sheet_name(name: str) -> str:
        """Excel 시트 이름 포맷팅"""
        # 유효하지 않은 문자 제거
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        formatted_name = ''.join(c for c in name if c not in invalid_chars)
        
        # 최대 31자로 제한
        formatted_name = formatted_name[:31]
        
        # 빈 문자열인 경우 기본값 설정
        if not formatted_name.strip():
            formatted_name = "Sheet1"
            
        return formatted_name
    

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import logging
from datetime import datetime
import os
from pathlib import Path
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import camelot

class PDFAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("KB손해보험 상품개정 분석기")
        self.root.geometry("800x600")
        self.setup_logging()
        self.setup_gui()

    def setup_logging(self):
        """로깅 설정"""
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        log_filename = log_dir / f'analyzer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def setup_gui(self):
        """GUI 구성 요소 설정"""
        # 메인 프레임
        self.main_frame = ttk.Frame(self.root, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # 제목
        title_label = ttk.Label(
            self.main_frame,
            text="KB손해보험 상품개정 자동화 서비스",
            font=("Helvetica", 16, "bold")
        )
        title_label.pack(pady=10)

        # 보장내용 카테고리 프레임
        coverage_frame = ttk.LabelFrame(self.main_frame, text="보장내용", padding="10")
        coverage_frame.pack(fill=tk.X, pady=5)

        # 요약서 입력란 (grid 사용)
        self.summary_file_path_var = tk.StringVar()
        summary_label = ttk.Label(coverage_frame, text="요약서 파일 선택:")
        summary_label.grid(row=0, column=0, padx=5, pady=5, sticky='w')
        summary_entry = ttk.Entry(coverage_frame, textvariable=self.summary_file_path_var)
        summary_entry.grid(row=0, column=1, padx=5, pady=5, sticky='we')
        summary_browse_button = ttk.Button(coverage_frame, text="찾아보기", command=self.browse_summary_file)
        summary_browse_button.grid(row=0, column=2, padx=5, pady=5)

        # 열 너비 조정
        coverage_frame.columnconfigure(1, weight=1)

        # 가입예시 카테고리 프레임
        example_frame = ttk.LabelFrame(self.main_frame, text="가입예시", padding="10")
        example_frame.pack(fill=tk.X, pady=5)

        # 가입설계서 PDF 파일 선택 (grid 사용)
        self.file_path_var = tk.StringVar()
        example_label = ttk.Label(example_frame, text="가입설계서 파일 선택:")
        example_label.grid(row=0, column=0, padx=5, pady=5, sticky='w')
        example_entry = ttk.Entry(example_frame, textvariable=self.file_path_var)
        example_entry.grid(row=0, column=1, padx=5, pady=5, sticky='we')
        example_browse_button = ttk.Button(example_frame, text="찾아보기", command=self.browse_file)
        example_browse_button.grid(row=0, column=2, padx=5, pady=5)

        # 가입예시 URL 입력란 (아래 줄에 배치)
        url_label = ttk.Label(example_frame, text="URL 입력:")
        url_label.grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.url_var = tk.StringVar()
        url_entry = ttk.Entry(example_frame, textvariable=self.url_var)
        url_entry.grid(row=1, column=1, padx=5, pady=5, sticky='we', columnspan=2)

        # 열 너비 조정
        example_frame.columnconfigure(1, weight=1)

        # 진행 상태 표시
        self.progress_var = tk.StringVar(value="대기 중...")
        progress_label = ttk.Label(self.main_frame, textvariable=self.progress_var)
        progress_label.pack(pady=5)

        self.progress_bar = ttk.Progressbar(
            self.main_frame,
            mode='determinate',
            length=300
        )
        self.progress_bar.pack(pady=5)

        # 처리 시작 버튼
        self.process_button = ttk.Button(
            self.main_frame,
            text="분석 시작",
            command=self.process_start,
            state=tk.DISABLED
        )
        self.process_button.pack(pady=10)

        # 로그 영역
        log_frame = ttk.LabelFrame(self.main_frame, text="처리 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = tk.Text(log_frame, height=15)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)





    def browse_summary_file(self):
        """요약서 파일 선택"""
        file_path = filedialog.askopenfilename(
            title="요약서 파일 선택",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file_path:
            self.summary_file_path_var.set(file_path)
            self.process_button['state'] = tk.NORMAL
            self.log_message(f"요약서 파일 선택됨: {file_path}")


    def browse_file(self):
        """PDF 파일 선택"""
        file_path = filedialog.askopenfilename(
            title="PDF 파일 선택",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.process_button['state'] = tk.NORMAL
            self.log_message(f"파일 선택됨: {file_path}")

    def log_message(self, message, level="INFO"):
        """로그 메시지 출력"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        
        if level == "INFO":
            self.logger.info(message)
        elif level == "ERROR":
            self.logger.error(message)
        elif level == "WARNING":
            self.logger.warning(message)

    def update_progress(self, value, message):
        """진행 상태 업데이트"""
        self.progress_bar['value'] = value
        self.progress_var.set(message)
        self.root.update_idletasks()

    def process_start(self):
        """분석 프로세스 시작"""
        if not self.file_path_var.get():
            messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
            return

        self.process_button['state'] = tk.DISABLED
        self.progress_bar['value'] = 0
        
        thread = threading.Thread(target=self.process_thread)
        thread.start()

    def process_thread(self):
        """별도 스레드에서 실행될 분석 프로세스"""
        try:
            self.log_message("분석 시작")
            pdf_path = self.file_path_var.get()
            web_url = self.url_var.get()

            results = {
                'pdf_data': None,
                'web_data': None,
                'comparison': None
            }

            # PDF 분석을 항상 실행
            self.update_progress(20, "PDF 분석 중...")
            results['pdf_data'] = self.analyze_pdf(pdf_path)

            # 웹 데이터 수집
            if web_url:
                self.update_progress(40, "웹 데이터 수집 중...")
                results['web_data'] = self.collect_web_data(web_url)

            # 데이터 비교
            if results['pdf_data'] is not None and results['web_data'] is not None:
                self.update_progress(60, "데이터 비교 중...")
                results['comparison'] = self.compare_data(results['pdf_data'], results['web_data'])

            # 결과 저장
            if any(results.values()):
                self.update_progress(80, "결과 저장 중...")
                output_path = self.save_results(results)
                self.log_message(f"분석 완료. 결과 저장됨: {output_path}")
                self.update_progress(100, "처리 완료")
                messagebox.showinfo("완료", f"분석이 완료되었습니다.\n저장 위치: {output_path}")
            else:
                self.log_message("처리할 데이터가 없습니다.", "WARNING")
                messagebox.showwarning("경고", "처리할 데이터가 없습니다.")

        except Exception as e:
            error_msg = f"처리 중 오류 발생: {str(e)}"
            self.log_message(error_msg, "ERROR")
            messagebox.showerror("오류", error_msg)

        finally:
            self.progress_var.set("대기 중...")
            self.process_button['state'] = tk.NORMAL


    def analyze_pdf(self, pdf_path):
        """PDF 분석"""
        try:
            self.log_message("PDF 테이블 추출 중...")
            tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
            
            processed_tables = []
            for table in tables:
                df = table.df
                if not df.empty:
                    processed_tables.append(df)
            
            return processed_tables
        except Exception as e:
            self.log_message(f"PDF 분석 중 오류: {str(e)}", "ERROR")
            return None

    def collect_web_data(self, url):
        """웹 데이터 수집"""
        try:
            self.log_message("웹 데이터 수집 중...")
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            tables = []
            for table in soup.find_all('table'):
                df = pd.read_html(str(table))[0]
                tables.append(df)
            
            return tables
        except Exception as e:
            self.log_message(f"웹 데이터 수집 중 오류: {str(e)}", "ERROR")
            return None

    def compare_data(self, pdf_data, web_data):
        """데이터 비교"""
        try:
            self.log_message("데이터 비교 중...")
            comparison_results = []
            
            for pdf_df in pdf_data:
                for web_df in web_data:
                    if pdf_df.shape == web_df.shape:
                        differences = pd.DataFrame()
                        differences['변경사항'] = (pdf_df != web_df).any(axis=1)
                        differences['PDF_데이터'] = pdf_df.apply(lambda x: ' | '.join(map(str, x)), axis=1)
                        differences['웹_데이터'] = web_df.apply(lambda x: ' | '.join(map(str, x)), axis=1)
                        comparison_results.append(differences)
            
            return comparison_results
        except Exception as e:
            self.log_message(f"데이터 비교 중 오류: {str(e)}", "ERROR")
            return None

    def save_results(self, results):
        """결과 저장"""
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"분석결과_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # PDF 데이터 시트
        if results['pdf_data']:
            ws = wb.create_sheet("PDF 데이터")
            self.write_data_to_sheet(ws, results['pdf_data'])
        
        # 웹 데이터 시트
        if results['web_data']:
            ws = wb.create_sheet("웹 데이터")
            self.write_data_to_sheet(ws, results['web_data'])
        
        # 비교 결과 시트
        if results['comparison']:
            ws = wb.create_sheet("비교 결과")
            self.write_comparison_to_sheet(ws, results['comparison'])
        
        # 기본 시트 제거
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        wb.save(output_path)
        return output_path

    def write_data_to_sheet(self, ws, data_list):
        """시트에 데이터 작성"""
        current_row = 1
        
        for df in data_list:
            # 헤더 작성
            for col_idx, header in enumerate(df.columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(header))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            
            current_row += 1
            
            # 데이터 작성
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=str(value))
                    cell.alignment = Alignment(wrap_text=True)
                current_row += 1
            
            current_row += 2

    def write_comparison_to_sheet(self, ws, comparison_results):
        """비교 결과 시트에 데이터 작성"""
        current_row = 1
        
        for df in comparison_results:
            # 헤더 작성
            for col_idx, header in enumerate(df.columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(header))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            
            current_row += 1
            
            # 데이터 작성
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=str(value))
                    
                    # 변경사항이 있는 행은 하이라이트 처리
                    if col_idx == 1 and value:  # '변경사항' 열이고 True인 경우
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    
                    cell.alignment = Alignment(wrap_text=True)
                current_row += 1
            
            current_row += 2

    def format_workbook(self, wb):
        """워크북 전체 서식 설정"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for ws in wb:
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 최대 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # 테두리 및 정렬 설정
            for row in ws.rows:
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

    def extract_table_metadata(self, pdf_path):
        """PDF 테이블 메타데이터 추출"""
        try:
            tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
            metadata = []
            
            for idx, table in enumerate(tables):
                info = {
                    '테이블 번호': idx + 1,
                    '페이지': table.page,
                    '행 수': len(table.df),
                    '열 수': len(table.df.columns),
                    '정확도': table.accuracy,
                    '화이트스페이스': table.whitespace
                }
                metadata.append(info)
            
            return pd.DataFrame(metadata)
            
        except Exception as e:
            self.log_message(f"메타데이터 추출 중 오류: {str(e)}", "ERROR")
            return None

    def extract_tables_with_context(self, pdf_path):
        """컨텍스트 정보와 함께 테이블 추출"""
        try:
            tables_data = []
            tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
            
            for table in tables:
                # 테이블 주변 텍스트 컨텍스트 추출 (미구현)
                context = "테이블 컨텍스트"
                
                table_info = {
                    'dataframe': table.df,
                    'page': table.page,
                    'context': context
                }
                tables_data.append(table_info)
            
            return tables_data
            
        except Exception as e:
            self.log_message(f"테이블 추출 중 오류: {str(e)}", "ERROR")
            return None

    def analyze_table_structure(self, df):
        """테이블 구조 분석"""
        analysis = {
            '열 수': len(df.columns),
            '행 수': len(df),
            '빈 셀 수': df.isna().sum().sum(),
            '데이터 타입': df.dtypes.to_dict(),
            '고유값 수': df.nunique().to_dict()
        }
        return analysis

    def compare_table_structures(self, pdf_df, web_df):
        """테이블 구조 비교"""
        pdf_structure = self.analyze_table_structure(pdf_df)
        web_structure = self.analyze_table_structure(web_df)
        
        differences = {
            key: {'PDF': pdf_structure[key], 'Web': web_structure[key]}
            for key in pdf_structure.keys()
            if pdf_structure[key] != web_structure[key]
        }
        
        return differences

    def run(self):
        """GUI 실행"""
        self.root.mainloop()


def process_pdf(self):
    pdf_path = self.file_path_var.get()

    if not pdf_path:
        messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
        return

    self.process_button['state'] = tk.DISABLED
    self.progress_bar['value'] = 0

    # 별도 스레드에서 처리
    thread = threading.Thread(target=self.process_pdf_thread, args=(pdf_path,))
    thread.daemon = True
    thread.start()


def main():
    try:
        # 로그 디렉토리 생성
        log_dir = "logs"
        os.makedirs(log_dir, exist_ok=True)
        
        # 로그 파일 설정
        log_file = os.path.join(log_dir, f'pdf_analyzer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        
        # 로그 핸들러 설정
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)
        
        logger.info("프로그램 시작")
        
        # GUI 애플리케이션 시작
        app = PDFAnalyzerGUI()
        app.run()
        
    except Exception as e:
        error_msg = f"프로그램 실행 중 오류 발생: {str(e)}"
        logger.error(error_msg)
        if 'app' in locals():
            messagebox.showerror("오류", error_msg)
        raise
    finally:
        logger.info("프로그램 종료")

if __name__ == "__main__":
    # 전체 프로그램 버전 정보
    VERSION = "2.0.0"
    
    # 지원되는 PDF 버전 정보
    SUPPORTED_PDF_VERSION = "1.7"
    
    # 프로그램 시작 시 환경 정보 출력
    print(f"""
    KB손해보험 상품개정 자동화 서비스
    버전: {VERSION}
    Python 버전: {sys.version.split()[0]}
    지원 PDF 버전: {SUPPORTED_PDF_VERSION}
    시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    """)
    
    main()


