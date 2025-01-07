# 2개표 1페이지 출력 성공
# 제목 출력 필요
# 보장명 지급사유 지급금액 행 삭제 완료
# raw 데이터 버전에 불순물 삭제완료
# 제목 위치 수정필요
import camelot
import pandas as pd
from pathlib import Path
import logging
from datetime import datetime
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import PyPDF2

class TableExtractionTester:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.page = "162"
        self.setup_logging()
        
        # 섹션 패턴 정의
        self.section_patterns = {
            '상해': r'[◇◆■□▶]([\s]*)(상해|상해관련|상해 관련)([\s]*)(특약|특별약관)',
            '질병': r'[◇◆■□▶]([\s]*)(질병|질병관련|질병 관련)([\s]*)(특약|특별약관)',
            '상해및질병': r'[◇◆■□▶]([\s]*)(상해\s*및\s*질병|상해와\s*질병)([\s]*)(관련)?([\s]*)(특약|특별약관)'
        }
        
        # enhanced_lattice 설정
        self.test_config = {
            'name': 'enhanced_lattice',
            'params': {
                'flavor': 'lattice',
                'line_scale': 40,
                'split_text': True,
                'process_background': True,
                'line_tol': 3,
                'joint_tol': 3
            }
        }

    def setup_logging(self):
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(
                    log_dir / f'table_test_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log',
                    encoding='utf-8'
                ),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def clean_table(self, df: pd.DataFrame, page_num: int) -> pd.DataFrame:
        """테이블 데이터 정제"""
        try:
            if df.empty:
                return df
                    
            # 모든 값을 문자열로 변환 및 None 값을 빈 문자열로 변환
            df = df.astype(str)
            df = df.fillna('')
                
            # 컬럼 헤더 처리
            if len(df.columns) >= 3:
                df.columns = ['보장명', '지급사유', '지급금액']
            
            # 헤더 행 찾아서 제거 - 행 단위로 체크
            cleaned_rows = []
            for _, row in df.iterrows():
                # "보 장 명" 등의 헤더 행 제외
                if any(x in str(row['보장명']).strip() for x in ['보 장 명', '보장명']) and \
                any(x in str(row['지급사유']).strip() for x in ['지 급 사 유', '지급사유']) and \
                any(x in str(row['지급금액']).strip() for x in ['지 급 금 액', '지급금액']):
                    continue  # 헤더 행이면 건너뛰기

                # 실제 데이터가 있는지 확인
                if row['보장명'].strip() and row['지급사유'].strip() and row['지급금액'].strip():
                    cleaned_rows.append(row)

            # 새로운 데이터프레임 생성
            df = pd.DataFrame(cleaned_rows)
                
            # 페이지 번호와 변경사항 컬럼 추가
            df['페이지'] = page_num
            df['변경사항'] = ''
                
            return df
                    
        except Exception as e:
            self.logger.error(f"테이블 정제 중 오류: {str(e)}")
            return pd.DataFrame(columns=['보장명', '지급사유', '지급금액', '페이지', '변경사항'])

    def extract_tables(self):
        """테이블 추출"""
        try:
            self.logger.info("\n=== 테이블 추출 시작 ===")
            
            # PDF에서 텍스트 추출하여 제목 찾기
            with open(self.pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                page = reader.pages[int(self.page) - 1]
                text = page.extract_text()
                
                # 제목과 위치 정보 추출
                sections = []
                lines = text.split('\n')
                for idx, line in enumerate(lines):
                    for pattern_type, pattern in self.section_patterns.items():
                        if re.search(pattern, line):
                            sections.append({
                                'title': line.strip(),
                                'text': line,
                                'position': idx
                            })
            
            # 테이블 추출
            tables = camelot.read_pdf(
                self.pdf_path,
                pages=self.page,
                **self.test_config['params']
            )
            
            self.logger.info(f"추출된 테이블 수: {len(tables)}")
            
            # 테이블과 제목 매칭
            all_data = []
            current_title = None
            
            for i, table in enumerate(tables):
                df = self.clean_table(table.df, int(self.page))
                if self.validate_table(df):
                    # 첫 번째 테이블은 첫 번째 제목과 매칭
                    if i == 0 and sections:
                        current_title = sections[0]['title']
                    # 두 번째 테이블은 두 번째 제목과 매칭 (있는 경우)
                    elif i == 1 and len(sections) > 1:
                        current_title = sections[1]['title']
                    
                    df['섹션제목'] = current_title
                    all_data.append(df)
                        
            if all_data:
                combined_df = pd.concat(all_data, ignore_index=True)
                return combined_df
            
            return None
                
        except Exception as e:
            self.logger.error(f"표 추출 중 오류: {str(e)}")
            return None

    def validate_table(self, df: pd.DataFrame) -> bool:
        """테이블 유효성 검증"""
        try:
            if df.empty:
                return False
                
            # 필수 컬럼 확인
            required_columns = ['보장명', '지급사유', '지급금액']
            if not all(col in df.columns for col in required_columns):
                return False
                
            # 데이터 존재 확인
            valid_rows = df[required_columns].apply(
                lambda x: x.str.strip().str.len() > 0
            ).any(axis=1)
            
            return valid_rows.any()
            
        except Exception as e:
            self.logger.error(f"테이블 검증 실패: {str(e)}")
            return False

    def save_results(self, df: pd.DataFrame):
        if df is None:
            return
            
        try:
            output_dir = Path("test_results")
            output_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Clean 데이터 생성을 위해 헤더 행 다시 제거
            cleaned_rows = []
            for _, row in df.iterrows():
                # 모든 공백 및 특수문자 제거 후 비교
                row_str = re.sub(r'\s+', '', str(row)).lower()  # 행 전체를 문자열로 변환
                if any(x in row_str for x in ['보장명', '보장내용', '지급사유', '지급금액']):
                    print("삭제할 행 발견:", row)  # 중간 과정 출력
                    continue
                cleaned_rows.append(row)

            # 클린 데이터프레임 생성
            clean_df = pd.DataFrame(cleaned_rows)

            # 클린 데이터 저장
            clean_path = output_dir / f"table_test_{timestamp}_clean.xlsx"
            clean_df.to_excel(clean_path, index=False)
            self.logger.info(f"정제 데이터 저장됨: {clean_path}")

            # Raw 데이터 먼저 저장
            wb_raw = Workbook()
            ws_raw = wb_raw.active
            ws_raw.title = "원본데이터"
            raw_path = output_dir / f"table_test_{timestamp}_raw.xlsx"
            self._write_to_excel(wb_raw, ws_raw, df, "원본")
            wb_raw.save(raw_path)
            self.logger.info(f"원본 데이터 저장됨: {raw_path}")

        except Exception as e:
            self.logger.error(f"결과 저장 중 오류: {str(e)}")

    def _write_to_excel(self, wb, ws, df, version_type):
        """Excel 파일 작성 로직"""
        current_row = 1
        current_section = None
        
        # 데이터를 섹션별로 그룹화하여 저장
        for _, row in df.iterrows():
            # 섹션이 변경된 경우 제목 추가
            if row['섹션제목'] != current_section:
                current_section = row['섹션제목']
                if current_section:
                    # 제목 셀 추가
                    cell = ws.cell(row=current_row, column=1, value=current_section)
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                    current_row += 2
                    
                    # 컬럼 헤더 추가
                    for col_idx, col_name in enumerate(['보장명', '지급사유', '지급금액', '페이지', '변경사항'], 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        ws.column_dimensions[get_column_letter(col_idx)].width = 15
                    current_row += 1
            
            # 데이터 행 추가
            cols_to_write = ['보장명', '지급사유', '지급금액', '페이지', '변경사항']
            for col_idx, col_name in enumerate(cols_to_write, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=row[col_name])
                cell.alignment = Alignment(wrap_text=True, vertical='center')
            current_row += 1
        
        # 열 너비 자동 조정 (최소 15, 최대 50)
        for col_idx in range(1, ws.max_column + 1):
            max_length = 15
            for row in ws.rows:
                cell = row[col_idx - 1]
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) + 2)
                except:
                    continue
            adjusted_width = min(max_length, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    def merge_cell_content(self, existing_content: str, new_content: str) -> str:
        """셀 내용 병합"""
        if not existing_content.strip():
            return new_content
        if not new_content.strip():
            return existing_content
        if new_content in existing_content:
            return existing_content
        return f"{existing_content}\n{new_content}"

def main():
    pdf_path = "D:\github\pdf_local\data\input\KB 금쪽같은 자녀보험Plus(무배당)(24.05)_11월11일판매_요약서_v1.1.pdf"
    
    tester = TableExtractionTester(pdf_path)
    results_df = tester.extract_tables()
    
    if results_df is not None:
        tester.save_results(results_df)
        print("\n테이블 추출 완료")
    else:
        print("\n테이블 추출 실패")

if __name__ == "__main__":
    main()