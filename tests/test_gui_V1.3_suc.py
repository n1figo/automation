# 표준 라이브러리
import os
import re
import sys
import threading
from datetime import datetime
from typing import Dict, List, Tuple, Optional

# 데이터 처리 및 분석
import numpy as np
import pandas as pd
from scipy.spatial.distance import cosine

# PDF 처리
import PyPDF2
import fitz
import camelot

# 이미지 처리
import cv2
from PIL import Image

# Excel 처리
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, 
    Border, 
    Side, 
    Alignment, 
    Font
)

# GUI
import tkinter as tk
from tkinter import filedialog, ttk, messagebox

# AI/ML
from sentence_transformers import SentenceTransformer

# 로깅
import logging

# 프로그램 상수
CONSTANTS = {
    # 파일 관련 상수
    'SUPPORTED_EXTENSIONS': ['.pdf'],
    'MAX_FILE_SIZE': 100 * 1024 * 1024,  # 100MB
    
    # 표 추출 관련 상수
    'TABLE_EXTRACTION': {
        'MIN_ROWS': 2,
        'MIN_COLS': 2,
        'MIN_ACCURACY': 50,  # 최소 정확도 (%)
        'MAX_EMPTY_RATIO': 0.5,  # 최대 빈 셀 비율
    },
    
    # 처리 타임아웃 설정
    'TIMEOUTS': {
        'PAGE_PROCESSING': 60,  # 페이지당 최대 처리 시간 (초)
        'TOTAL_PROCESSING': 3600,  # 전체 최대 처리 시간 (초)
    },
    
    # GUI 관련 상수
    'GUI': {
        'WINDOW_SIZE': '800x600',
        'MAX_LOG_LINES': 1000,
        'PROGRESS_UPDATE_INTERVAL': 100,  # ms
    },
    
    # 출력 파일 관련 상수
    'OUTPUT': {
        'MAX_SHEET_NAME_LENGTH': 31,  # Excel 시트 이름 최대 길이
        'MAX_CELL_LENGTH': 32767,  # Excel 셀 최대 길이
    }
}

# 오류 메시지
ERROR_MESSAGES = {
    'FILE_NOT_FOUND': "PDF 파일을 찾을 수 없습니다.",
    'FILE_TOO_LARGE': "파일 크기가 너무 큽니다. (최대 100MB)",
    'INVALID_FILE_TYPE': "지원되지 않는 파일 형식입니다.",
    'EXTRACTION_FAILED': "표 추출에 실패했습니다.",
    'PROCESSING_TIMEOUT': "처리 시간이 초과되었습니다.",
    'MEMORY_ERROR': "메모리가 부족합니다.",
    'UNKNOWN_ERROR': "알 수 없는 오류가 발생했습니다."
}

# 로깅 설정
LOGGING_CONFIG = {
    'version': 1,
    'disable_existing_loggers': False,
    'formatters': {
        'detailed': {
            'format': '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        },
        'simple': {
            'format': '%(levelname)s - %(message)s'
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'level': 'INFO',
            'formatter': 'simple'
        },
        'file': {
            'class': 'logging.FileHandler',
            'filename': 'pdf_analyzer.log',
            'level': 'DEBUG',
            'formatter': 'detailed',
            'encoding': 'utf-8'
        }
    },
    'loggers': {
        '': {
            'handlers': ['console', 'file'],
            'level': 'DEBUG',
            'propagate': True
        }
    }
}

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class HighlightDetector:
    def __init__(self):
        self.saturation_threshold = 30
        self.kernel_size = (5, 5)

    def pdf_to_image(self, page: fitz.Page) -> np.ndarray:
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return np.array(img)

    def detect_highlights(self, image: np.ndarray) -> List[np.ndarray]:
        hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
        s = hsv[:, :, 1]
        v = hsv[:, :, 2]

        saturation_mask = s > self.saturation_threshold
        _, binary = cv2.threshold(v, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        combined_mask = cv2.bitwise_and(binary, binary, mask=saturation_mask.astype(np.uint8) * 255)

        kernel = np.ones(self.kernel_size, np.uint8)
        cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
        cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)

        contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        return contours

    def get_highlight_regions(self, contours: List[np.ndarray], image_height: int) -> List[Tuple[float, float]]:
        regions = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            top = image_height - (y + h)
            bottom = image_height - y
            regions.append((top, bottom))
        return regions

    def check_highlight(self, row_range: Tuple[float, float], highlight_regions: List[Tuple[float, float]]) -> bool:
        row_top, row_bottom = row_range
        for region_top, region_bottom in highlight_regions:
            if (region_top <= row_top <= region_bottom) or \
               (region_top <= row_bottom <= region_bottom) or \
               (row_top <= region_top <= row_bottom) or \
               (row_top <= region_bottom <= row_bottom):
                return True
        return False


class TableExtractor:
    def __init__(self):
        self.matcher = TitleMatcher()
        self.highlight_detector = HighlightDetector()
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]
        self.max_distance = 50

    def extract_tables_from_section(self, pdf_path: str, start_page: int, end_page: int) -> List[Tuple[str, pd.DataFrame, int]]:
        """섹션에서 표 추출"""
        try:
            results = []
            
            for page_num in range(start_page, end_page):
                logger.info(f"Processing page {page_num + 1}")
                doc = fitz.open(pdf_path)
                page = doc[page_num]

                # 이미지 변환 및 하이라이트 감지
                image = self.highlight_detector.pdf_to_image(page)
                contours = self.highlight_detector.detect_highlights(image)
                highlight_regions = self.highlight_detector.get_highlight_regions(contours, image.shape[0])

                # 제목 추출
                titles = self.get_titles_with_positions(page)

                success = False
                tables = None

                # Stream 모드 시도
                try:
                    logger.info(f"Trying stream mode for page {page_num + 1}")
                    tables = camelot.read_pdf(
                        pdf_path,
                        pages=str(page_num + 1),
                        flavor='stream',
                        edge_tol=100,
                        row_tol=15,
                        split_text=True,
                        strip_text=' .\n',
                        columns=['145,280,420,560']
                    )
                    
                    if len(tables) > 0:
                        success = True
                        logger.info(f"Stream mode successful for page {page_num + 1}")

                except Exception as e:
                    logger.warning(f"Stream mode failed for page {page_num + 1}: {str(e)}")

                # Stream 모드 실패시 Lattice 모드 시도
                if not success:
                    try:
                        logger.info(f"Trying lattice mode for page {page_num + 1}")
                        tables = camelot.read_pdf(
                            pdf_path,
                            pages=str(page_num + 1),
                            flavor='lattice',
                            line_scale=40,
                            process_background=True
                        )
                        if len(tables) > 0:
                            success = True
                            logger.info(f"Lattice mode successful for page {page_num + 1}")

                    except Exception as e:
                        logger.error(f"Lattice mode failed for page {page_num + 1}: {str(e)}")
                        continue

                if tables and success:
                    page_height = page.rect.height
                    logger.info(f"Processing {len(tables)} tables from page {page_num + 1}")

                    for table_idx, table in enumerate(tables):
                        df = self.process_table_with_highlights(
                            table, 
                            highlight_regions, 
                            image.shape[0], 
                            table_idx, 
                            page_num + 1
                        )
                        
                        df = self.clean_table(df)
                        
                        if not df.empty and len(df.columns) >= 2:
                            table_position = self.get_table_positions(table, page_height)
                            title, distance = self.match_title_to_table(titles, table_position)
                            
                            if not title:
                                title = f"Table_{table_idx + 1}"
                                
                            results.append((title, df, page_num + 1))
                            logger.info(f"Successfully processed table {table_idx + 1} from page {page_num + 1}")

                doc.close()

            return results

        except Exception as e:
            logger.error(f"Error extracting tables from section: {str(e)}")
            return []

    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """표 데이터 정제"""
        try:
            if df.empty:
                return df

            # 기본 정제
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            
            # 모든 데이터를 문자열로 변환
            df = df.astype(str)
            
            # 문자열 정제
            for col in df.columns:
                df[col] = df[col].str.strip()
                df[col] = df[col].str.replace('\s+', ' ', regex=True)
            
            # 주석 행 제거 (첫 번째 열에 대해서만)
            df = df[~df.iloc[:, 0].str.contains("^※|^주\)", regex=True, na=False)]
            
            # 빈 값 처리
            df = df.replace(r'^\s*$', '', regex=True)
            df = df.replace('nan', '')
            
            # 중복 제거
            df = df.drop_duplicates()
            
            return df
                
        except Exception as e:
            logger.error(f"Table cleaning error: {str(e)}")
            return df

    def get_titles_with_positions(self, page) -> List[Dict]:
        """페이지에서 제목과 위치 추출"""
        titles = []
        blocks = page.get_text("dict")["blocks"]
        
        for block in blocks:
            if block.get("lines"):
                text = " ".join([span["text"] for line in block["lines"] 
                               for span in line["spans"]]).strip()
                
                if text and any(re.search(pattern, text) for pattern in self.title_patterns):
                    titles.append({
                        "text": text,
                        "y_top": block["bbox"][1],
                        "y_bottom": block["bbox"][3],
                        "bbox": block["bbox"],
                        "used": False
                    })

        return sorted(titles, key=lambda x: x["y_top"])

    def get_table_positions(self, table, page_height):
        """표의 위치 정보 추출"""
        try:
            x0, y0, x1, y1 = table._bbox
            return {
                "y_top": page_height - y1,
                "y_bottom": page_height - y0,
                "bbox": (x0, page_height - y1, x1, page_height - y0)
            }
        except:
            return {
                "y_top": 0,
                "y_bottom": page_height,
                "bbox": (0, 0, 0, page_height)
            }

    def match_title_to_table(self, titles, table_position):
        """표와 가장 가까운 제목 매칭"""
        best_title = None
        min_distance = float('inf')

        for title in titles:
            if title["used"]:
                continue

            distance = table_position["y_top"] - title["y_bottom"]
            if 0 < distance < self.max_distance and distance < min_distance:
                best_title = title
                min_distance = distance

        if best_title:
            best_title["used"] = True
            return best_title["text"], min_distance
        return None, None

    def process_table_with_highlights(self, table, highlight_regions, page_height, table_index, page_num):
        """하이라이트된 영역이 있는 표 처리"""
        try:
            df = table.df.copy()
            x1, y1, x2, y2 = table._bbox
            
            if len(df) == 0:
                return df

            table_height = y2 - y1
            row_height = table_height / len(df)
            processed_data = []

            for row_index in range(len(df)):
                row_data = df.iloc[row_index].copy()
                row_top = y2 - (row_index + 1) * row_height
                row_bottom = y2 - row_index * row_height

                row_highlighted = self.highlight_detector.check_highlight(
                    (row_top, row_bottom), 
                    highlight_regions
                )

                row_data['변경사항'] = '추가' if row_highlighted else ''
                row_data['Table_Number'] = table_index + 1
                row_data['페이지'] = page_num
                processed_data.append(row_data)

            return pd.DataFrame(processed_data)
            
        except Exception as e:
            logger.error(f"Table processing error: {str(e)}")
            return pd.DataFrame()
    

class TitleMatcher:
    def __init__(self):
        try:
            # 현재 작업 디렉토리 기준으로 모델 경로 설정
            model_path = "models/distiluse-base-multilingual-cased-v1"
            
            logger.info(f"로컬 모델 로드 시도: {model_path}")
            
            # 로컬 모델 로드
            try:
                self.model = SentenceTransformer(model_path)
                logger.info("로컬 모델 로드 성공")
            except Exception as e:
                error_msg = f"로컬 모델 로드 실패: {str(e)}"
                logger.error(error_msg)
                raise
            
            self.title_patterns = [
                r'기본계약',
                r'의무부가계약',
                r'[\w\s]+관련\s*특약',
                r'[\w\s]+계약',
                r'[\w\s]+보장'
            ]
            self.max_distance = 50
            
        except Exception as e:
            error_msg = f"모델 초기화 실패: {str(e)}"
            logger.error(error_msg)
            raise

    def get_embedding(self, text: str) -> np.ndarray:
        try:
            return self.model.encode(text)
        except Exception as e:
            logger.error(f"임베딩 생성 실패: {str(e)}")
            raise

    def calculate_similarity(self, text1: str, text2: str) -> float:
        try:
            emb1 = self.get_embedding(text1)
            emb2 = self.get_embedding(text2)
            return 1 - cosine(emb1, emb2)
        except Exception as e:
            logger.error(f"유사도 계산 실패: {str(e)}")
            raise

class InsuranceDocumentAnalyzer:
    def __init__(self):
        self.section_patterns = {
            "종류": r'\[(\d)종\]',
        }
        self.section_pages = {"[1종]": None, "[2종]": None, "[3종]": None}
        self.section_ranges = {}
        self.title_patterns = [
            r'기본계약',
            r'의무부가계약',
            r'[\w\s]+관련\s*특약',
            r'[\w\s]+계약',
            r'[\w\s]+보장'
        ]

    def find_section_pages(self, pdf_path: str) -> Dict[str, int]:
        try:
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            
            # 각 페이지 검사
            for page_num in range(total_pages):
                text = doc[page_num].get_text()
                
                # 종 구분 찾기
                matches = re.finditer(self.section_patterns["종류"], text)
                for match in matches:
                    종_type = f"[{match.group(1)}종]"
                    if self.section_pages[종_type] is None:
                        self.section_pages[종_type] = page_num
                        logger.info(f"{종_type} 시작 페이지: {page_num + 1}")

            # 섹션 범위 설정
            self._set_section_ranges(total_pages)

            doc.close()
            return self.section_pages

        except Exception as e:
            logger.error(f"섹션 페이지 검색 중 오류: {str(e)}")
            return {}

    def _set_section_ranges(self, total_pages):
        # 유효한 섹션 찾기
        found_sections = [v for v in self.section_pages.values() if v is not None]
        
        if not found_sections:
            logger.warning("종 구분을 찾지 못했습니다. 전체 문서를 처리합니다.")
            self.section_ranges["전체"] = (0, total_pages)
            return

        # 섹션 범위 설정
        sorted_pages = sorted(
            [(k, v) for k, v in self.section_pages.items() if v is not None],
            key=lambda x: x[1]
        )

        for i, (종_type, start_page) in enumerate(sorted_pages):
            if i + 1 < len(sorted_pages):
                end_page = sorted_pages[i + 1][1]
            else:
                end_page = total_pages
                
            self.section_ranges[종_type] = (start_page, end_page)
            logger.info(f"{종_type} 범위: {start_page + 1} ~ {end_page}")

class ExcelWriter:
    @staticmethod
    def save_to_excel(sections_data: Dict[str, List[Tuple[str, pd.DataFrame, int]]], output_path: str):
        try:
            wb = Workbook()
            # 기본 시트 삭제하기 전에 새 시트 생성
            default_sheet = wb.active
            
            # 테두리 스타일 정의
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 데이터가 있는 경우에만 처리
            if sections_data:
                for section, tables in sections_data.items():
                    if not tables:  # 빈 테이블 스킵
                        continue
                        
                    # 시트 이름 생성 (최대 31자로 제한)
                    sheet_name = (section.replace("[", "").replace("]", "") or "전체")[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    current_row = 1

                    for title, df, page_num in tables:
                        # 섹션 제목 추가
                        section_title = "가입담보표" if "전체" in sheet_name else sheet_name
                        section_cell = ws.cell(row=current_row, column=1, value=section_title)
                        section_cell.font = Font(bold=True, size=14)
                        current_row += 2
                        
                        # 표 제목 작성
                        title_cell = ws.cell(row=current_row, column=1, 
                                           value=f"{title} (페이지: {page_num})")
                        title_cell.font = Font(bold=True, size=12)
                        title_cell.fill = PatternFill(start_color='E6E6E6',
                                                    end_color='E6E6E6',
                                                    fill_type='solid')
                        
                        # 제목 셀 병합
                        ws.merge_cells(start_row=current_row, start_column=1, 
                                     end_row=current_row, end_column=len(df.columns))
                        current_row += 2

                        # 열 헤더 작성
                        for col_idx, col_name in enumerate(df.columns, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color='F2F2F2',
                                                  end_color='F2F2F2',
                                                  fill_type='solid')
                            cell.border = border
                            cell.alignment = Alignment(wrap_text=True, 
                                                     horizontal='center',
                                                     vertical='center')
                        current_row += 1

                        # 데이터 작성
                        for _, row in df.iterrows():
                            for col_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=current_row, column=col_idx, value=value)
                                cell.border = border
                                cell.alignment = Alignment(wrap_text=True,
                                                         vertical='center')
                                
                                # '추가' 표시가 있는 행 하이라이트
                                if '변경사항' in df.columns and row['변경사항'] == '추가':
                                    cell.fill = PatternFill(start_color='FFFF00',
                                                          end_color='FFFF00',
                                                          fill_type='solid')
                            current_row += 1
                        
                        current_row += 2

                        # 열 너비 자동 조정
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                                adjusted_width = min(max_length + 2, 50)  # 최대 50
                                ws.column_dimensions[column_letter].width = adjusted_width

                # 기본 시트가 있으면 삭제
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                    
                # 데이터가 없는 경우 기본 시트 유지
                if len(wb.sheetnames) == 0:
                    ws = wb.create_sheet(title="추출 결과")
                    ws.cell(row=1, column=1, value="추출된 데이터가 없습니다.")
                    ws.column_dimensions['A'].width = 30

            else:  # sections_data가 비어있는 경우
                default_sheet.title = "추출 결과"
                default_sheet.cell(row=1, column=1, value="추출된 데이터가 없습니다.")
                default_sheet.column_dimensions['A'].width = 30

            # 첫 번째 시트를 활성화
            if wb.sheetnames:
                wb.active = 0

            wb.save(output_path)
            logger.info(f"Excel 파일 저장 완료: {output_path}")

        except Exception as e:
            logger.error(f"Excel 저장 중 오류 발생: {str(e)}")
            raise

    @staticmethod
    def format_sheet_name(name: str) -> str:
        """Excel 시트 이름 포맷팅"""
        # 유효하지 않은 문자 제거
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        formatted_name = ''.join(c for c in name if c not in invalid_chars)
        
        # 최대 31자로 제한
        formatted_name = formatted_name[:31]
        
        # 빈 문자열인 경우 기본값 설정
        if not formatted_name.strip():
            formatted_name = "Sheet1"
            
        return formatted_name
    

class PDFAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("KB손해보험 상품개정 자동화 서비스")
        self.root.geometry("800x600")
        self.setup_logging()
        self.setup_gui()

    def setup_logging(self):
        log_filename = f'pdf_analyzer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
        file_handler = logging.FileHandler(log_filename, encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)

    def setup_gui(self):
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 제목
        title_label = ttk.Label(
            main_frame,
            text="KB손해보험 상품개정 자동화 서비스",
            font=("Helvetica", 16, "bold")
        )
        title_label.pack(pady=10)

        # 파일 선택 프레임
        file_frame = ttk.LabelFrame(main_frame, text="PDF 파일 선택", padding="10")
        file_frame.pack(fill=tk.X, pady=10)

        self.file_path_var = tk.StringVar()
        self.file_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=60)
        self.file_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        browse_button = ttk.Button(
            file_frame,
            text="찾아보기",
            command=self.browse_file
        )
        browse_button.pack(side=tk.LEFT, padx=5)

        # 진행 상태 표시
        self.progress_var = tk.StringVar(value="대기 중...")
        progress_label = ttk.Label(main_frame, textvariable=self.progress_var)
        progress_label.pack(pady=5)

        self.progress_bar = ttk.Progressbar(main_frame, mode='determinate', length=300)
        self.progress_bar.pack(pady=5)

        # 처리 시작 버튼
        self.process_button = ttk.Button(
            main_frame,
            text="분석 시작",
            command=self.process_pdf_start,
            state=tk.DISABLED
        )
        self.process_button.pack(pady=10)

        # 로그 영역
        log_frame = ttk.LabelFrame(main_frame, text="처리 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = tk.Text(log_frame, height=15)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="PDF 파일 선택",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.process_button['state'] = tk.NORMAL
            self.log_message(f"파일 선택됨: {file_path}")

    def process_pdf_start(self):
        pdf_path = self.file_path_var.get()
        if not pdf_path:
            messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
            return

        self.process_button['state'] = tk.DISABLED
        self.progress_bar['value'] = 0

        # 별도 스레드에서 처리
        thread = threading.Thread(target=self.process_pdf_thread, args=(pdf_path,))
        thread.daemon = True
        thread.start()

    def process_pdf_thread(self, pdf_path):
        try:
            # 파일 유효성 검사
            if not os.path.exists(pdf_path):
                raise FileNotFoundError(ERROR_MESSAGES['FILE_NOT_FOUND'])
                
            if os.path.getsize(pdf_path) > CONSTANTS['MAX_FILE_SIZE']:
                raise ValueError(ERROR_MESSAGES['FILE_TOO_LARGE'])
                
            if not any(pdf_path.lower().endswith(ext) for ext in CONSTANTS['SUPPORTED_EXTENSIONS']):
                raise ValueError(ERROR_MESSAGES['INVALID_FILE_TYPE'])

            # 출력 폴더 설정 및 생성
            output_folder = os.path.join(os.path.dirname(pdf_path), "output")
            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(
                output_folder, 
                f"보험특약표_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            # 처리 시작
            self.log_message("문서 분석 시작...")
            self.progress_bar['value'] = 10

            # 문서 분석기 초기화 및 섹션 찾기
            document_analyzer = InsuranceDocumentAnalyzer()
            document_analyzer.find_section_pages(pdf_path)

            # 섹션이 없는 경우 처리
            if not document_analyzer.section_ranges:
                self.log_message("섹션을 찾지 못했습니다. 문서 전체를 처리합니다.", "WARNING")
                doc = fitz.open(pdf_path)
                total_pages = len(doc)
                document_analyzer.section_ranges["전체"] = (0, total_pages)
                doc.close()

            self.log_message("표 추출 시작...")
            self.progress_bar['value'] = 30

            # 표 추출기 초기화
            table_extractor = TableExtractor()
            sections_data = {}
            has_extracted_tables = False  # 표 추출 성공 여부 추적

            # 각 섹션별 처리
            total_sections = len(document_analyzer.section_ranges)
            processing_start_time = datetime.now()

            for idx, (section, (start_page, end_page)) in enumerate(document_analyzer.section_ranges.items(), 1):
                if (datetime.now() - processing_start_time).total_seconds() > CONSTANTS['TIMEOUTS']['TOTAL_PROCESSING']:
                    raise TimeoutError(ERROR_MESSAGES['PROCESSING_TIMEOUT'])

                # 1종, 2종, 3종만 처리
                if section not in ["[1종]", "[2종]", "[3종]", "전체"]:
                    continue

                progress = 30 + (idx / total_sections) * 50
                self.progress_bar['value'] = progress
                self.progress_var.set(f"{section} 처리 중...")
                self.log_message(f"처리 중: {section} (페이지 {start_page + 1} ~ {end_page})")

                # Stream 모드로 첫 시도
                tables = table_extractor.extract_tables_from_section(pdf_path, start_page, end_page)
                
                # Stream 모드로 표를 찾지 못한 경우 Lattice 모드로 재시도
                if not tables:
                    self.log_message(f"{section} stream 모드 실패, lattice 모드로 재시도...", "INFO")
                    try:
                        for page_num in range(start_page, end_page):
                            lattice_tables = camelot.read_pdf(
                                pdf_path,
                                pages=str(page_num + 1),
                                flavor='lattice',
                                line_scale=40,
                                process_background=True
                            )
                            if len(lattice_tables) > 0:
                                processed_tables = []
                                for table in lattice_tables:
                                    df = table.df
                                    if not df.empty and len(df.columns) >= 2:
                                        processed_tables.append((f"Table_{len(processed_tables)+1}", df, page_num + 1))
                                if processed_tables:
                                    if section not in sections_data:
                                        sections_data[section] = []
                                    sections_data[section].extend(processed_tables)
                                    has_extracted_tables = True
                                    self.log_message(f"Lattice 모드로 {len(processed_tables)}개의 표 추출 성공 (페이지 {page_num + 1})")
                    except Exception as e:
                        self.log_message(f"Lattice 모드 실패: {str(e)}", "WARNING")
                else:
                    if tables:
                        sections_data[section] = tables
                        has_extracted_tables = True

            # 결과 저장
            if has_extracted_tables:
                self.progress_bar['value'] = 90
                self.progress_var.set("결과 저장 중...")
                
                try:
                    ExcelWriter.save_to_excel(sections_data, output_path)
                    self.progress_bar['value'] = 100

                    success_message = f"처리가 완료되었습니다.\n저장 위치: {output_path}"
                    self.log_message(success_message)
                    messagebox.showinfo("완료", success_message)
                except Exception as e:
                    raise Exception(f"Excel 파일 저장 중 오류 발생: {str(e)}")
            else:
                self.log_message("추출된 표가 없습니다.", "WARNING")
                messagebox.showwarning("경고", "추출된 표가 없습니다.")

        except Exception as e:
            self.handle_error(e)
        finally:
            self.cleanup()

    def handle_error(self, e):
        """오류 처리"""
        if isinstance(e, FileNotFoundError):
            self.log_message(str(e), "ERROR")
            messagebox.showerror("파일 오류", str(e))
        elif isinstance(e, ValueError):
            self.log_message(str(e), "ERROR")
            messagebox.showerror("입력 오류", str(e))
        elif isinstance(e, TimeoutError):
            self.log_message(str(e), "ERROR")
            messagebox.showerror("시간 초과", str(e))
        elif isinstance(e, MemoryError):
            self.log_message(ERROR_MESSAGES['MEMORY_ERROR'], "ERROR")
            messagebox.showerror("메모리 오류", ERROR_MESSAGES['MEMORY_ERROR'])
        else:
            error_message = f"처리 중 오류가 발생했습니다: {str(e)}"
            self.log_message(error_message, "ERROR")
            messagebox.showerror("오류", error_message)

    def cleanup(self):
        """정리 작업"""
        self.progress_var.set("대기 중...")
        self.process_button['state'] = tk.NORMAL
        self.progress_bar['value'] = 0

    def log_message(self, message, level="INFO"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)

        if level == "INFO":
            logger.info(message)
        elif level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)

    def run(self):
        self.root.mainloop()

def process_pdf(self):
    pdf_path = self.file_path_var.get()
    if not pdf_path:
        messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
        return

    self.process_button['state'] = tk.DISABLED
    self.progress_bar['value'] = 0

    # 별도 스레드에서 처리
    thread = threading.Thread(target=self.process_pdf_thread, args=(pdf_path,))
    thread.daemon = True
    thread.start()


def main():
    try:
        # 로그 디렉토리 생성
        log_dir = "logs"
        os.makedirs(log_dir, exist_ok=True)
        
        # 로그 파일 설정
        log_file = os.path.join(log_dir, f'pdf_analyzer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        
        # 로그 핸들러 설정
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)
        
        logger.info("프로그램 시작")
        
        # GUI 애플리케이션 시작
        app = PDFAnalyzerGUI()
        app.run()
        
    except Exception as e:
        error_msg = f"프로그램 실행 중 오류 발생: {str(e)}"
        logger.error(error_msg)
        if 'app' in locals():
            messagebox.showerror("오류", error_msg)
        raise
    finally:
        logger.info("프로그램 종료")

if __name__ == "__main__":
    # 전체 프로그램 버전 정보
    VERSION = "2.0.0"
    
    # 지원되는 PDF 버전 정보
    SUPPORTED_PDF_VERSION = "1.7"
    
    # 프로그램 시작 시 환경 정보 출력
    print(f"""
    KB손해보험 상품개정 자동화 서비스
    버전: {VERSION}
    Python 버전: {sys.version.split()[0]}
    지원 PDF 버전: {SUPPORTED_PDF_VERSION}
    시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    """)
    
    main()


