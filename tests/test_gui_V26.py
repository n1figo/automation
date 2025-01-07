import sys
import os
from pathlib import Path
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import pandas as pd
import camelot
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import cv2
from PIL import Image
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime
from sentence_transformers import SentenceTransformer
from scipy.spatial.distance import cosine
import threading
import json
import time
from dataclasses import dataclass, field

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'pdf_analyzer_{datetime.now().strftime("%Y%m%d")}.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class SectionInfo:
    title: str
    start_page: int
    end_page: int
    category: str
    line_number: int
    end_line_number: Optional[int] = None
    type_number: str = None  # [1종], [2종] 등을 저장

class PerformanceLogger:
    def __init__(self, base_dir: str = "logs"):
        """
        성능 로깅을 위한 클래스 초기화
        Args:
            base_dir (str): 로그 저장 기본 경로
        """
        self.base_dir = base_dir
        self.test_dir = None
        self.start_time = None
        self.test_name = None
        self.config = None
        self.results = {}
        self.notes = []
        self.warnings = []
        self.sections_found = []
        self.processing_times = {}
        
        # 로그 디렉토리 생성
        os.makedirs(base_dir, exist_ok=True)
        
        # 로거 설정
        self.logger = logging.getLogger(__name__)
        self.setup_logging()
    
    def setup_logging(self):
        """로깅 설정"""
        log_format = '%(asctime)s - %(levelname)s - %(message)s'
        date_format = '%Y-%m-%d %H:%M:%S'
        
        formatter = logging.Formatter(log_format, date_format)
        
        # 파일 핸들러 추가
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = os.path.join(self.base_dir, f'performance_{timestamp}.log')
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setFormatter(formatter)
        
        # 스트림 핸들러 추가
        stream_handler = logging.StreamHandler()
        stream_handler.setFormatter(formatter)
        
        # 로거에 핸들러 추가
        self.logger.addHandler(file_handler)
        self.logger.addHandler(stream_handler)
        self.logger.setLevel(logging.INFO)
    
    def start_test(self, test_name: str, config: Dict[str, Any] = None):
        """새로운 테스트 시작"""
        self.test_name = test_name
        self.start_time = time.time()
        self.config = config or {}
        
        # 테스트별 디렉토리 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.test_dir = os.path.join(self.base_dir, f"{test_name}_{timestamp}")
        os.makedirs(self.test_dir, exist_ok=True)
        
        # 설정 저장
        config_path = os.path.join(self.test_dir, "config.json")
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)
            
        self.logger.info(f"테스트 시작: {test_name}")
        
    def log_result(self, key: str, value: Any):
        """결과 기록"""
        self.results[key] = value
        self.logger.info(f"결과 기록: {key} = {value}")
        
    def add_note(self, note: str):
        """노트 추가"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.notes.append(f"[{timestamp}] {note}")
        self.logger.info(note)
        
    def add_warning(self, warning: str):
        """경고 추가"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.warnings.append(f"[{timestamp}] {warning}")
        self.logger.warning(warning)
    
    def log_section_found(self, section_name: str, page_num: int):
        """섹션 발견 기록"""
        self.sections_found.append({
            'section': section_name,
            'page': page_num,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        self.logger.info(f"섹션 발견: {section_name} (페이지 {page_num})")
    
    def start_operation(self, operation_name: str):
        """작업 시작 시간 기록"""
        self.processing_times[operation_name] = {
            'start': time.time(),
            'end': None,
            'duration': None
        }
        
    def end_operation(self, operation_name: str):
        """작업 종료 및 소요 시간 계산"""
        if operation_name in self.processing_times:
            self.processing_times[operation_name]['end'] = time.time()
            duration = (self.processing_times[operation_name]['end'] - 
                       self.processing_times[operation_name]['start'])
            self.processing_times[operation_name]['duration'] = duration
            
            self.logger.info(f"{operation_name} 완료 (소요시간: {duration:.2f}초)")
            
    def end_test(self, save_results: bool = True) -> Dict[str, Any]:
        """테스트 종료 및 결과 반환"""
        if not self.start_time:
            raise RuntimeError("테스트가 시작되지 않았습니다")
            
        end_time = time.time()
        processing_time = end_time - self.start_time
        
        summary = {
            'test_name': self.test_name,
            'start_time': datetime.fromtimestamp(self.start_time).strftime("%Y-%m-%d %H:%M:%S"),
            'end_time': datetime.fromtimestamp(end_time).strftime("%Y-%m-%d %H:%M:%S"),
            'processing_time': processing_time,
            'results': self.results,
            'notes': self.notes,
            'warnings': self.warnings,
            'sections_found': self.sections_found,
            'operation_times': self.processing_times
        }
        
        if save_results:
            results_path = os.path.join(self.test_dir, "results.json")
            with open(results_path, 'w', encoding='utf-8') as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)
                
        self.logger.info(f"테스트 종료: {self.test_name} (처리 시간: {processing_time:.2f}초)")
        return summary
        
    def save_error(self, error: Exception):
        """오류 정보 저장"""
        if self.test_dir:
            error_path = os.path.join(self.test_dir, "error.txt")
            with open(error_path, 'w', encoding='utf-8') as f:
                f.write(f"Error type: {type(error).__name__}\n")
                f.write(f"Error message: {str(error)}\n")
                f.write("\nTraceback:\n")
                import traceback
                f.write(traceback.format_exc())
                
            self.logger.error(f"오류 저장됨: {error_path}")

    def get_performance_metrics(self) -> Dict[str, float]:
        """성능 지표 계산"""
        metrics = {
            'total_processing_time': 0,
            'avg_operation_time': 0,
            'sections_found_count': len(self.sections_found),
            'warning_count': len(self.warnings)
        }
        
        if self.start_time:
            metrics['total_processing_time'] = time.time() - self.start_time
            
        operation_times = [op['duration'] for op in self.processing_times.values() 
                         if op['duration'] is not None]
        if operation_times:
            metrics['avg_operation_time'] = sum(operation_times) / len(operation_times)
            
        return metrics


class PDFDocument:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        self.similarity_threshold = 0.7  # 유사도 임계값 설정

        # 섹션 제목 패턴 정의 (정규식)
        self.section_patterns = {
            '상해관련 특별약관': r'[◇◆■□▶]([\s]*)(?P<title>상해|상해관련|상해 관련)([\s]*)(특약|특별약관)',
            '질병관련 특별약관': r'[◇◆■□▶]([\s]*)(?P<title>질병|질병관련|질병 관련)([\s]*)(특약|특별약관)',
            '상해및질병관련 특별약관': r'[◇◆■□▶]([\s]*)(?P<title>상해\s*및\s*질병|상해와\s*질병)([\s]*)(관련)?([\s]*)(특약|특별약관)?'
        }

        # 약관 참고 문구 패턴
        self.reference_note_pattern = r'자세한\s*사항은\s*반드시\s*약관을\s*참고하시기\s*바랍니다'
        
        # 검색 시작점 패턴
        self.payment_section_pattern = r'나\.\s*보험금\s*지급.*'
        self.type_pattern = r'\[(\d)종\]'

        # 로컬 모델 로드 및 임베딩 준비
        try:
            model_path = "models/distiluse-base-multilingual-cased-v1"
            logging.getLogger('sentence_transformers').setLevel(logging.WARNING)
            self.model = SentenceTransformer(model_path, device='cpu')
            self.model.max_seq_length = 512
            logger.info("로컬 모델 로드 성공")

            # 섹션 텍스트 기준 임베딩 계산
            self.section_reference_texts = {
                '상해관련 특별약관': ["상해관련 특별약관", "상해관련 특약", "상해 특별약관"],
                '질병관련 특별약관': ["질병관련 특별약관", "질병관련 특약", "질병 특별약관"],
                '상해및질병관련 특별약관': [
                    "상해 및 질병 관련 특별약관",
                    "상해 및 질병 관련 특약",
                    "상해 및 질병 특별약관"
                ]
            }

            # 각 섹션별 임베딩 계산
            self.section_embeddings = {}
            for section_name, reference_texts in self.section_reference_texts.items():
                self.section_embeddings[section_name] = self.model.encode(reference_texts)
            logger.info("임베딩 계산 완료")

        except Exception as e:
            logger.error(f"모델 로드 실패: {str(e)}")
            raise

        # 섹션 정보 초기화
        self.sections = {
            "상해관련 특별약관": {"start": None, "end": None},
            "질병관련 특별약관": {"start": None, "end": None},
            "상해및질병관련 특별약관": {"start": None, "end": None}
        }

    def find_section_ranges(self) -> Dict[str, Dict[str, int]]:
        try:
            # 1. 보험금 지급 섹션과 보종 찾기
            payment_page, types = self.find_payment_and_types()
            if payment_page is None:
                logger.warning("보험금 지급 섹션을 찾을 수 없습니다.")
                payment_page = 0  # 전체 문서 검색을 원하면 0으로 설정

            # '상해및질병관련 특별약관' 시작 페이지 찾기
            injury_disease_start_page = self.find_injury_disease_start_page(payment_page)
            if injury_disease_start_page is not None:
                parsing_end_page = injury_disease_start_page - 1
                print(f"\n=== 파싱 범위: {payment_page + 1}페이지부터 {parsing_end_page + 1}페이지까지 ===")
            else:
                parsing_end_page = len(self.doc) - 1
                print(f"\n=== 파싱 범위: {payment_page + 1}페이지부터 문서 끝까지 ===")

            # 섹션 검색
            self.find_sections_in_range(payment_page, parsing_end_page)

            return self.sections

        except Exception as e:
            logger.error(f"섹션 범위 찾기 실패: {str(e)}")
            # 오류 발생시에도 전체 범위 검색 시도
            logger.info("오류 발생으로 인해 전체 범위에서 섹션을 검색합니다.")
            self.find_sections_in_range(0, len(self.doc) - 1)
            return self.sections


    def find_payment_and_types(self) -> Tuple[Optional[int], List[Dict[str, any]]]:
        """보험금 지급 섹션과 보종 찾기"""
        payment_page = None
        types = []

        # 보험금 지급 섹션 검색 패턴
        payment_pattern = r'나\.\s*보험금\s*지급.*'

        try:
            # 1. 보험금 지급 섹션 찾기
            for page_num in range(len(self.doc)):
                text = self.doc[page_num].get_text()
                if payment_page is None:
                    if re.search(payment_pattern, text, re.IGNORECASE):
                        payment_page = page_num
                        print(f"[찾음] 보험금 지급 섹션: {page_num + 1}페이지")
                        logger.info(f"보험금 지급 섹션 발견: {page_num + 1}페이지")
                        break  # 보험금 지급 섹션을 찾으면 루프 종료

            if payment_page is None:
                logger.warning("보험금 지급 섹션을 찾지 못했습니다.")
                payment_page = 0  # 전체 문서 검색을 원하면 0으로 설정

            # 2. 보종 타입 찾기 (payment_page 찾은 후)
            for page_num in range(payment_page, len(self.doc)):
                text = self.doc[page_num].get_text()
                lines = text.split('\n')
                for line_num, line in enumerate(lines, 1):
                    matches = list(re.finditer(self.type_pattern, line))

                    # 연속된 보종 체크 (예: [1종][2종][3종])
                    if len(matches) > 1:
                        consecutive = all(
                            matches[i + 1].start() - matches[i].end() <= 2
                            for i in range(len(matches) - 1)
                        )
                        if consecutive:
                            logger.info(f"연속된 보종 발견: {line.strip()}")
                            return payment_page, []

                    # 개별 보종 처리
                    for match in matches:
                        type_num = match.group(1)
                        type_info = {
                            'type': f"[{type_num}종]",
                            'page': page_num,
                            'line': line_num
                        }
                        types.append(type_info)
                        logger.info(f"보종 발견: {type_info['type']} ({page_num + 1}페이지)")

            return payment_page, sorted(types, key=lambda x: (x['page'], x.get('line', 0)))

        except Exception as e:
            logger.error(f"섹션 검색 중 오류 발생: {str(e)}")
            return 0, []

    def find_sections_in_range(self, start_page: int, end_page: int):
        try:
            print(f"\n=== 섹션 검색 범위: {start_page + 1} ~ {end_page + 1}페이지 ===")
            current_section = None

            for page_num in range(start_page, end_page + 1):
                text = self.doc[page_num].get_text()
                lines = text.split('\n')

                for line_num, line in enumerate(lines, 1):
                    for section_type, pattern in self.section_patterns.items():
                        if re.search(pattern, line):
                            similarity, has_reference = self.calculate_section_similarity(line, section_type)

                            if similarity >= self.similarity_threshold:
                                logger.info(f"섹션 발견: {section_type} (페이지 {page_num + 1})")
                                print(f"[발견] {section_type}: {page_num + 1}페이지")

                                # 이전 섹션 종료 처리
                                if current_section and current_section != section_type:
                                    if self.sections[current_section]["end"] is None:
                                        self.sections[current_section]["end"] = page_num - 1
                                        logger.info(f"섹션 종료: {current_section} ({page_num}페이지)")

                                # 새 섹션 시작
                                if self.sections[section_type]["start"] is None:
                                    self.sections[section_type]["start"] = page_num
                                    current_section = section_type

            # 마지막 섹션 종료 처리
            if current_section and self.sections[current_section]["end"] is None:
                self.sections[current_section]["end"] = end_page
                logger.info(f"마지막 섹션 종료: {current_section} ({end_page + 1}페이지)")

        except Exception as e:
            logger.error(f"섹션 검색 중 오류: {str(e)}", exc_info=True)


    def calculate_section_similarity(self, text: str, section_type: str) -> Tuple[float, bool]:
        try:
            pattern = self.section_patterns[section_type]
            match = re.search(pattern, text)
            if not match:
                return 0.0, False

            title_text = match.group('title')
            has_reference = bool(re.search(self.reference_note_pattern, text))

            text_embedding = self.model.encode(title_text)
            similarities = [
                1 - cosine(text_embedding, ref_embedding)
                for ref_embedding in self.section_embeddings[section_type]
            ]
            similarity = max(similarities)

            if has_reference:
                similarity = min(similarity * 1.2, 1.0)

            return similarity, has_reference

        except Exception as e:
            logger.error(f"유사도 계산 중 오류: {str(e)}")
            return 0.0, False
    
    def find_injury_disease_start_page(self, start_page: int) -> Optional[int]:
        """'상해및질병관련 특별약관' 시작 페이지 찾기"""
        try:
            for page_num in range(start_page, len(self.doc)):
                text = self.doc[page_num].get_text()
                lines = text.split('\n')
                for line in lines:
                    if re.search(self.section_patterns['상해및질병관련 특별약관'], line):
                        print(f"[찾음] 상해및질병관련 특별약관 시작 페이지: {page_num + 1}페이지")
                        return page_num
            return None
        except Exception as e:
            logger.error(f"'상해및질병관련 특별약관' 시작 페이지 찾기 실패: {str(e)}")
            return None



    def close(self):
        """리소스 정리"""
        if self.doc:
            self.doc.close()
            logger.info(f"PDF 문서 닫힘: {self.pdf_path}")



class TableExtractor:
    def __init__(self):
        self.standard_columns = ['담보명', '지급사유', '지급금액', '변경사항', '페이지']
        
        # 무시할 패턴 정의 (필요에 따라 수정)
        self.ignore_patterns = [
            r'이륜자동차',
            r'Page'
            # r'^\s*$',  # 빈 행은 제거하지 않도록 주석 처리
            # r'특별약관$',
            # r'약관$',
            # r'^표$',
            # r'^그림$',
            # r'^주요', 
            # r'^※', 
            # r'^상기', 
            # r'^위',
            # r'보\s*장\s*명',
            # r'지\s*급\s*사\s*유',
            # r'지\s*급\s*금\s*액'
        ]
        
        # 표 추출 설정 (lattice 모드 우선)
        self.extraction_settings = {
            'lattice': {
                'flavor': 'lattice',
                'line_scale': 40,
                'split_text': True,
                'process_background': True,
                'line_tol': 3,
                'joint_tol': 3
            },
            'stream': {
                'flavor': 'stream',
                'row_tol': 3,
                'col_tol': 3,
                'edge_tol': 50,
                'split_text': True
            }
        }

    def extract_section_tables(self, doc, start_page: int, end_page: int, section_type: str) -> Dict[str, pd.DataFrame]:
        """섹션별 테이블 추출"""
        try:
            if isinstance(doc, fitz.Document):
                pdf_path = doc.name
            elif isinstance(doc, str):
                pdf_path = doc
            else:
                raise ValueError("Invalid document type")

            logger.info(f"\n{section_type} 섹션 테이블 추출 시작 (페이지: {start_page + 1} ~ {end_page + 1})")
            tables_data = []

            # 페이지별 테이블 처리
            for page_num in range(start_page, end_page + 1):
                try:
                    logger.info(f"Processing page-{page_num + 1}")
                    
                    # lattice 모드로 시도
                    tables = camelot.read_pdf(
                        pdf_path,
                        pages=str(page_num + 1),
                        **self.extraction_settings['lattice']
                    )
                    
                    # lattice 모드에서 테이블이 추출되지 않으면 stream 모드로 시도
                    if len(tables) == 0:
                        logger.info("lattice 모드로 테이블을 찾지 못했습니다. stream 모드로 재시도합니다.")
                        tables = camelot.read_pdf(
                            pdf_path,
                            pages=str(page_num + 1),
                            **self.extraction_settings['stream']
                        )

                    # 추출된 테이블 처리
                    for table in tables:
                        if table.df.empty:
                            continue
                            
                        df = table.df.copy()
                        df = df.fillna('')
                        df = self.merge_multiline_cells(df)
                        df = self.standardize_columns(df, page_num + 1)
                        df = self.clean_table_content(df)
                        
                        if self.validate_table(df) and not df.empty:
                            tables_data.append(df)

                except Exception as e:
                    logger.error(f"페이지 {page_num + 1} 처리 중 오류: {str(e)}")
                    continue

            # 테이블 병합
            if tables_data:
                merged_df = self.merge_tables(tables_data, section_type)
                if not merged_df.empty:
                    return {section_type: merged_df}

            return {}

        except Exception as e:
            logger.error(f"섹션 테이블 추출 중 오류: {str(e)}")
            return {}

    def merge_multiline_cells(self, df: pd.DataFrame) -> pd.DataFrame:
        """다중 행으로 분리된 셀 병합"""
        try:
            for col in df.columns:
                # 빈 문자열로 초기화
                merged_values = [''] * len(df)
                for idx, value in enumerate(df[col]):
                    value = str(value).strip()
                    if value:
                        merged_values[idx] += value + ' '
                # 각 셀의 양쪽 공백 제거
                df[col] = [val.strip() for val in merged_values]
            
            return df
            
        except Exception as e:
            logger.error(f"셀 병합 실패: {str(e)}")
            return df

    def clean_table_content(self, df: pd.DataFrame) -> pd.DataFrame:
        """테이블 내용 정제"""
        try:
            if df.empty:
                logger.info("데이터프레임이 비어 있습니다.")
                return df

            # None 값을 빈 문자열로 변환
            df = df.fillna('')
            initial_rows = len(df)
            logger.info(f"정제 전 행 수: {initial_rows}")

            # 헤더 패턴 정의
            header_patterns = [
                r'^보\s*장\s*명$',
                r'^지\s*급\s*사\s*유$',
                r'^지\s*급\s*금\s*액$'
            ]

            # 헤더 행 제거
            df = df[~df.apply(lambda row: all(
                re.match(pattern, str(value).strip()) for pattern, value in zip(header_patterns, row)
            ), axis=1)]

            # 데이터 정제를 위한 임시 리스트
            cleaned_rows = []

            for index, row in df.iterrows():
                # 모든 값을 문자열로 변환
                row = row.astype(str)

                # 무시할 패턴 확인
                should_ignore = False
                for pattern in self.ignore_patterns:
                    for value in row:
                        if re.search(pattern, value.strip(), re.IGNORECASE):
                            logger.debug(f"패턴 '{pattern}'이 값 '{value}'와 일치하여 행 {index} 제거")
                            should_ignore = True
                            break
                    if should_ignore:
                        break

                if should_ignore:
                    continue

                # 각 셀의 내용 정제
                cleaned_row = {}
                for col in df.columns:
                    value = str(row[col]).strip()

                    # 줄바꿈을 스페이스로 변환하고 연속된 공백 제거
                    value = re.sub(r'\n+', ' ', value)
                    value = re.sub(r'\s+', ' ', value)

                    # 특수문자 정리 (필요 시 주석 처리)
                    # value = re.sub(r'[^\w\s가-힣()%,.-]', '', value)

                    cleaned_row[col] = value.strip()

                # 실제 데이터가 있는 행만 추가
                if any(cleaned_row.values()):
                    cleaned_rows.append(cleaned_row)

            # 새로운 데이터프레임 생성
            df = pd.DataFrame(cleaned_rows)

            final_rows = len(df)
            logger.info(f"정제 후 행 수: {final_rows}")

            if initial_rows - final_rows > 0:
                logger.info(f"불필요한 행 {initial_rows - final_rows}개 제거됨")

            # 데이터프레임 내용 샘플 출력
            if not df.empty:
                logger.info(f"데이터프레임 샘플:\n{df.head()}")

            return df

        except Exception as e:
            logger.error(f"테이블 내용 정제 실패: {str(e)}")
            return df

    def standardize_columns(self, df: pd.DataFrame, page_num: int) -> pd.DataFrame:
        """컬럼 표준화 및 페이지 정보 추가"""
        try:
            logger.info(f"컬럼 표준화 시작 (페이지 {page_num})")
            
            # 현재 컬럼명 출력
            logger.info(f"현재 컬럼명: {df.columns.tolist()}")
            
            # 컬럼 매핑 정의
            column_mapping = {
                '보장명': '담보명',
                '급부명': '담보명',
                '보험금지급사유': '지급사유',
                '지급조건': '지급사유',
                '보험금액': '지급금액',
                '지급액': '지급금액',
                '비고': '변경사항'
            }
            
            # 컬럼이 모두 숫자인 경우 처리
            if all(str(col).isdigit() for col in df.columns):
                logger.info("숫자로 된 컬럼명 발견, 기본 컬럼명 적용")
                if len(df.columns) >= 3:
                    df.columns = ['담보명', '지급사유', '지급금액'] + [''] * (len(df.columns) - 3)
                    logger.info(f"새로운 컬럼명: {df.columns.tolist()}")
                else:
                    df.columns = ['담보명', '지급사유'] + [''] * (len(df.columns) - 2)
                    logger.info(f"새로운 컬럼명: {df.columns.tolist()}")
            else:
                # 컬럼명 매핑 적용
                df = df.rename(columns=column_mapping)
                logger.info(f"컬럼명 매핑 적용 후: {df.columns.tolist()}")
            
            # 필수 컬럼 추가
            for col in self.standard_columns:
                if col not in df.columns:
                    df[col] = ''
                    logger.info(f"필수 컬럼 '{col}'이 추가되었습니다.")
            
            # 페이지 번호 설정
            df['페이지'] = page_num
            
            # 컬럼 순서 조정
            df = df[self.standard_columns]
            
            logger.info(f"최종 컬럼명: {df.columns.tolist()}")
            
            return df

        except Exception as e:
            logger.error(f"컬럼 표준화 실패: {str(e)}")
            return pd.DataFrame(columns=self.standard_columns)

    def merge_tables(self, tables: List[pd.DataFrame], section_type: str) -> pd.DataFrame:
        """동일 섹션의 테이블 병합"""
        try:
            logger.info(f"{section_type} 섹션 테이블 병합 시작")
            
            if not tables:
                logger.warning("병합할 테이블 없음")
                return pd.DataFrame(columns=self.standard_columns)

            # 테이블 병합
            merged_df = pd.concat(tables, ignore_index=True)
            
            # 중복 제거
            merged_df = merged_df.drop_duplicates()
            
            # 정렬
            merged_df = merged_df.sort_values(['페이지', '담보명'])
            
            return merged_df

        except Exception as e:
            logger.error(f"테이블 병합 실패: {str(e)}")
            return pd.DataFrame(columns=self.standard_columns)

    def validate_table(self, df: pd.DataFrame) -> bool:
        """테이블 유효성 검사"""
        try:
            # 빈 데이터프레임 체크
            if df.empty:
                logger.warning("빈 데이터프레임입니다.")
                return False

            # 필수 컬럼 체크
            required_columns = ['담보명', '지급사유']
            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                logger.warning(f"필수 컬럼이 없습니다: {missing_cols}")
                return False

            # 데이터가 있는지 확인
            if df[required_columns].isnull().all().all():
                logger.warning("데이터프레임에 데이터가 없습니다.")
                return False

            # 유효한 데이터가 있는지 확인
            valid_rows = df[required_columns].dropna(how='all')
            if valid_rows.empty:
                logger.warning("유효한 데이터가 없습니다.")
                return False

            logger.info("테이블 유효성 검사 통과")
            return True

        except Exception as e:
            logger.error(f"테이블 유효성 검사 실패: {str(e)}")
            return False

    def merge_cell_content(self, existing_content: str, new_content: str) -> str:
        """셀 내용 병합"""
        try:
            if not existing_content.strip():
                return new_content.strip()
            if not new_content.strip():
                return existing_content.strip()
            if new_content.strip() in existing_content:
                return existing_content.strip()
                
            # 줄바꿈 처리
            return f"{existing_content.strip()} {new_content.strip()}"
            
        except Exception as e:
            logger.error(f"셀 내용 병합 실패: {str(e)}")
            return existing_content

        

class ExcelWriter:
    def __init__(self):
        self.styles = {
            'header': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'section_header': {
                'font': Font(bold=True, size=14),
                'fill': PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'cell': {
                'alignment': Alignment(vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'highlight': {
                'fill': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            }
        }

    def apply_cell_style(self, cell, style_name: str):
        """스타일 적용"""
        style = self.styles[style_name]
        for attr, value in style.items():
            setattr(cell, attr, value)

    def write_table(self, worksheet, df: pd.DataFrame, start_row: int, title: str) -> int:
        """테이블 작성"""
        try:
            # 섹션별로 데이터 그룹화
            grouped_data = df.groupby('섹션명')
            current_row = start_row

            for section_name, section_df in grouped_data:
                # 섹션 제목 작성
                section_title = f"{section_name} 특별약관" if section_name else title
                title_cell = worksheet.cell(row=current_row, column=1, value=section_title)
                self.apply_cell_style(title_cell, 'section_header')
                worksheet.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=len(df.columns)
                )
                current_row += 2

                # 컬럼 헤더 작성
                columns_to_display = ['담보명', '지급사유', '지급금액', '변경사항', '페이지']
                for col_idx, column in enumerate(columns_to_display, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx, value=column)
                    self.apply_cell_style(cell, 'header')
                current_row += 1

                # 데이터 작성
                for _, row in section_df.iterrows():
                    for col_idx, column in enumerate(columns_to_display, 1):
                        cell = worksheet.cell(row=current_row, column=col_idx, value=row[column])
                        self.apply_cell_style(cell, 'cell')
                        
                        # 변경사항이 '추가'인 행 하이라이트
                        if column == '변경사항' and str(row[column]).strip() == '추가':
                            for highlight_col in range(1, len(columns_to_display) + 1):
                                highlight_cell = worksheet.cell(row=current_row, column=highlight_col)
                                self.apply_cell_style(highlight_cell, 'highlight')
                    
                    current_row += 1
                
                current_row += 2  # 섹션 간 간격
            
            return current_row

        except Exception as e:
            logger.error(f"테이블 작성 실패: {str(e)}")
            return start_row + 1

    def adjust_column_widths(self, worksheet):
        """열 너비 자동 조정"""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 최대 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"열 너비 조정 실패: {str(e)}")

    def save_to_excel(self, tables_data: Dict[str, pd.DataFrame], output_path: str):
        """엑셀 파일로 저장"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "보장내용"
            current_row = 1

            for section_type, df in tables_data.items():
                if not isinstance(df, pd.DataFrame) or df.empty:
                    logger.warning(f"{section_type}: 유효하지 않은 데이터")
                    continue

                # 섹션 제목 가공
                section_title = section_type.replace('_', ' ').title()
                
                # 테이블 작성
                current_row = self.write_table(ws, df, current_row, section_title)
                current_row += 2  # 섹션 간 간격

            # 열 너비 조정
            self.adjust_column_widths(ws)

            # 파일 저장
            wb.save(output_path)
            logger.info(f"Excel 파일 저장 완료: {output_path}")

        except Exception as e:
            logger.error(f"Excel 저장 실패: {str(e)}")
            raise

class PDFAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("KB손해보험 상품분석기")
        self.root.geometry("800x600")
        self.file_path_var = tk.StringVar()
        self.progress_var = tk.StringVar(value="대기 중...")
        self.setup_gui()

    def setup_gui(self):
        """GUI 구성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 제목
        title_label = ttk.Label(
            main_frame,
            text="보험약관 테이블 분석기",
            font=("Helvetica", 16, "bold")
        )
        title_label.pack(pady=10)

        # 파일 선택 프레임
        file_frame = ttk.LabelFrame(main_frame, text="PDF 파일 선택", padding="10")
        file_frame.pack(fill=tk.X, pady=10)

        file_entry = ttk.Entry(
            file_frame,
            textvariable=self.file_path_var,
            width=60
        )
        file_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        browse_button = ttk.Button(
            file_frame,
            text="찾아보기",
            command=self.browse_file
        )
        browse_button.pack(side=tk.LEFT, padx=5)

        # 진행 상태 표시
        progress_label = ttk.Label(
            main_frame,
            textvariable=self.progress_var
        )
        progress_label.pack(pady=5)

        self.progress_bar = ttk.Progressbar(
            main_frame,
            mode='determinate',
            length=300
        )
        self.progress_bar.pack(pady=5)

        # 분석 버튼
        self.analyze_button = ttk.Button(
            main_frame,
            text="분석 시작",
            command=self.start_analysis,
            state=tk.DISABLED
        )
        self.analyze_button.pack(pady=10)

        # 로그 영역
        log_frame = ttk.LabelFrame(main_frame, text="처리 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = tk.Text(log_frame, height=15, width=70)
        scrollbar = ttk.Scrollbar(
            log_frame,
            orient=tk.VERTICAL,
            command=self.log_text.yview
        )
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def browse_file(self):
        """파일 선택"""
        file_path = filedialog.askopenfilename(
            title="PDF 파일 선택",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.analyze_button['state'] = tk.NORMAL
            self.log_message(f"파일 선택됨: {file_path}")

    def log_message(self, message: str, level: str = "INFO"):
        """로그 메시지 추가"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {level}: {message}\n"
        
        self.log_text.insert(tk.END, log_entry)
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_progress(self, value: int, message: str = None):
        """진행 상태 업데이트"""
        self.progress_bar['value'] = value
        if message:
            self.progress_var.set(message)
        self.root.update_idletasks()

    def analyze_pdf(self, pdf_path: str):
        """PDF 분석 실행"""
        try:
            # 출력 폴더 생성
            output_dir = os.path.join(os.path.dirname(pdf_path), "output")
            os.makedirs(output_dir, exist_ok=True)

            self.log_message("PDF 분석 시작")
            self.update_progress(10, "PDF 문서 로드 중...")

            # PDF 문서 로드 및 섹션 분석
            pdf_doc = PDFDocument(pdf_path)
            sections = pdf_doc.find_section_ranges()

            if not any(sections[section]["start"] is not None for section in sections):
                self.log_message("섹션을 찾을 수 없습니다.", "WARNING")
                messagebox.showwarning("경고", "문서에서 섹션을 찾을 수 없습니다.")
                return

            # 테이블 추출기 초기화
            extractor = TableExtractor()
            all_tables = {}

            # 섹션별 테이블 추출
            section_count = len([s for s in sections if sections[s]["start"] is not None])
            current_section = 0

            for section_type, ranges in sections.items():
                if ranges["start"] is not None:
                    current_section += 1
                    progress = 10 + (current_section / section_count) * 60
                    self.update_progress(
                        progress,
                        f"{section_type} 섹션 테이블 추출 중..."
                    )
                    
                    tables = extractor.extract_section_tables(
                        pdf_path,  # 수정된 부분: pdf_path 전달
                        ranges["start"],
                        ranges["end"],
                        section_type
                    )
                    
                    all_tables.update(tables)

            # PDF 문서 닫기
            pdf_doc.close()

            if any(not df.empty for df in all_tables.values()):
                self.update_progress(70, "Excel 파일 생성 중...")
                
                # 결과 저장
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = os.path.join(
                    output_dir,
                    f"analyzed_tables_{timestamp}.xlsx"
                )
                
                excel_writer = ExcelWriter()
                excel_writer.save_to_excel(all_tables, output_path)
                
                self.update_progress(100, "완료")
                self.log_message(f"분석 완료. 결과 파일: {output_path}")
                
                # 결과 파일 경로를 상대 경로로 변환
                try:
                    rel_path = os.path.relpath(output_path, os.path.dirname(pdf_path))
                    display_path = rel_path
                except:
                    display_path = output_path
                    
                messagebox.showinfo(
                    "완료",
                    f"분석이 완료되었습니다.\n저장 위치: {display_path}"
                )
            else:
                self.log_message("추출된 테이블이 없습니다.", "WARNING")
                messagebox.showwarning("경고", "추출된 테이블이 없습니다.")

        except Exception as e:
            self.log_message(f"오류 발생: {str(e)}", "ERROR")
            messagebox.showerror(
                "오류",
                f"처리 중 오류가 발생했습니다:\n{str(e)}"
            )
        finally:
            self.analyze_button['state'] = tk.NORMAL
            self.update_progress(0, "대기 중...")


    def start_analysis(self):
        """분석 시작"""
        pdf_path = self.file_path_var.get()
        if not pdf_path:
            messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
            return

        self.analyze_button['state'] = tk.DISABLED
        self.progress_bar['value'] = 0
        
        # 별도 스레드에서 처리
        threading.Thread(
            target=self.analyze_pdf,
            args=(pdf_path,),
            daemon=True
        ).start()

    def run(self):
        """애플리케이션 실행"""
        self.root.mainloop()


def main():
    try:
        app = PDFAnalyzerGUI()
        app.run()
    except Exception as e:
        logger.error(f"프로그램 실행 중 오류 발생: {str(e)}")
        raise

if __name__ == "__main__":
    main()