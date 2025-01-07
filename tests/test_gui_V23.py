# 컬럼 클린작업

import sys
import os
from pathlib import Path
import re
import logging
import fitz
import numpy as np
from typing import Dict, List, Tuple, Optional
import pandas as pd
import camelot
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import cv2
from PIL import Image
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from datetime import datetime
from sentence_transformers import SentenceTransformer
from scipy.spatial.distance import cosine
import threading
import os
import json
import time
from datetime import datetime
import logging
from typing import Dict, Any, Optional




# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'pdf_analyzer_{datetime.now().strftime("%Y%m%d")}.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class PerformanceLogger:
    def __init__(self, base_dir: str = "logs"):
        self.base_dir = base_dir
        self.test_dir = None
        self.start_time = None
        self.test_name = None
        self.config = None
        self.results = {}
        self.notes = []
        
        # 로그 디렉토리 생성
        os.makedirs(base_dir, exist_ok=True)
        
        # 로거 설정
        self.logger = logging.getLogger(__name__)
        
    def start_test(self, test_name: str, config: Dict[str, Any] = None):
        """새로운 테스트 시작"""
        self.test_name = test_name
        self.start_time = time.time()
        self.config = config or {}
        
        # 테스트별 디렉토리 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.test_dir = os.path.join(self.base_dir, f"{test_name}_{timestamp}")
        os.makedirs(self.test_dir, exist_ok=True)
        
        # 설정 저장
        config_path = os.path.join(self.test_dir, "config.json")
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)
            
        self.logger.info(f"테스트 시작: {test_name}")
        
    def log_result(self, key: str, value: Any):
        """결과 기록"""
        self.results[key] = value
        self.logger.info(f"결과 기록: {key} = {value}")
        
    def add_note(self, note: str):
        """노트 추가"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.notes.append(f"[{timestamp}] {note}")
        self.logger.info(note)
        
    def end_test(self, save_results: bool = True) -> Dict[str, Any]:
        """테스트 종료 및 결과 반환"""
        if not self.start_time:
            raise RuntimeError("테스트가 시작되지 않았습니다")
            
        end_time = time.time()
        processing_time = end_time - self.start_time
        
        summary = {
            'test_name': self.test_name,
            'start_time': datetime.fromtimestamp(self.start_time).strftime("%Y-%m-%d %H:%M:%S"),
            'end_time': datetime.fromtimestamp(end_time).strftime("%Y-%m-%d %H:%M:%S"),
            'processing_time': processing_time,
            'results': self.results,
            'notes': self.notes
        }
        
        if save_results:
            results_path = os.path.join(self.test_dir, "results.json")
            with open(results_path, 'w', encoding='utf-8') as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)
                
        self.logger.info(f"테스트 종료: {self.test_name} (처리 시간: {processing_time:.2f}초)")
        
        return summary
        
    def save_error(self, error: Exception):
        """오류 정보 저장"""
        if self.test_dir:
            error_path = os.path.join(self.test_dir, "error.txt")
            with open(error_path, 'w', encoding='utf-8') as f:
                f.write(f"Error type: {type(error).__name__}\n")
                f.write(f"Error message: {str(error)}\n")
            self.logger.error(f"오류 저장됨: {error_path}")


class PDFDocument:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        self.similarity_threshold = 0.7
        
        # 섹션 제목 패턴 정의
        self.section_patterns = {
            '상해관련 특별약관': r'[◇◆■□▶]([\s]*)(?P<title>상해|상해관련|상해 관련)([\s]*)(특약|특별약관)',
            '질병관련 특별약관': r'[◇◆■□▶]([\s]*)(?P<title>질병|질병관련|질병 관련)([\s]*)(특약|특별약관)',
            '상해및질병관련 특별약관': r'[◇◆■□▶]([\s]*)(?P<title>상해\s*및\s*질병|상해와\s*질병)([\s]*)(관련)?([\s]*)(특약|특별약관)?'
        }
        
        # 약관 참고 문구 패턴
        self.reference_note_pattern = r'자세한\s*사항은\s*반드시\s*약관을\s*참고하시기\s*바랍니다'
        
        # 검색 시작점 패턴
        self.payment_section_pattern = r'나\.\s*보험금\s*지급.*'
        self.type_pattern = r'\[(\d)종\]'
        
        # 로컬 모델 로드 및 임베딩 준비
        try:
            model_path = "models/distiluse-base-multilingual-cased-v1"
            logging.getLogger('sentence_transformers').setLevel(logging.WARNING)
            self.model = SentenceTransformer(model_path, device='cpu')
            self.model.max_seq_length = 512
            logger.info("로컬 모델 로드 성공")
            
            # 섹션 텍스트 기준 임베딩 계산
            self.section_reference_texts = {
                '상해관련 특별약관': ["상해관련 특별약관", "상해관련 특약", "상해 특별약관"],
                '질병관련 특별약관': ["질병관련 특별약관", "질병관련 특약", "질병 특별약관"],
                '상해및질병관련 특별약관': [
                    "상해 및 질병 관련 특별약관",
                    "상해 및 질병 관련 특약",
                    "상해와 질병 관련 특별약관",
                    "상해와 질병 관련 특약"
                ]
            }
            
            # 각 섹션별 임베딩 계산
            self.section_embeddings = {}
            for section_name, reference_texts in self.section_reference_texts.items():
                self.section_embeddings[section_name] = self.model.encode(reference_texts)
            logger.info("임베딩 계산 완료")
            
        except Exception as e:
            logger.error(f"모델 로드 실패: {str(e)}")
            raise
        
        # 섹션 정보 초기화
        self.sections = {
            "상해관련 특별약관": {"start": None, "end": None},
            "질병관련 특별약관": {"start": None, "end": None},
            "상해및질병관련 특별약관": {"start": None, "end": None}  # 시작 페이지만 파악용
        }

    def calculate_section_similarity(self, text: str, section_name: str) -> Tuple[float, bool]:
        try:
            pattern = self.section_patterns[section_name]
            match = re.search(pattern, text)
            if not match:
                return 0.0, False
            
            title_text = match.group('title')
            has_reference = bool(re.search(self.reference_note_pattern, text))
            
            text_embedding = self.model.encode(title_text)
            similarities = [
                1 - cosine(text_embedding, ref_embedding)
                for ref_embedding in self.section_embeddings[section_name]
            ]
            similarity = max(similarities)
            
            if has_reference:
                similarity = min(similarity * 1.2, 1.0)
            
            return similarity, has_reference
            
        except Exception as e:
            logger.error(f"유사도 계산 중 오류: {str(e)}")
            return 0.0, False

    def find_sections_in_range(self, start_page: int, end_page: int):
        try:
            print(f"\n=== 섹션 검색 범위: {start_page + 1} ~ {end_page + 1}페이지 ===")
            current_section = None
            
            for page_num in range(start_page, end_page + 1):
                text = self.doc[page_num].get_text()
                lines = text.split('\n')
                
                for line_num, line in enumerate(lines, 1):
                    for section_name, pattern in self.section_patterns.items():
                        if re.search(pattern, line):
                            similarity, has_reference = self.calculate_section_similarity(line, section_name)
                            
                            if similarity > self.similarity_threshold:
                                # 상해및질병 섹션 발견시 질병 섹션 종료 처리
                                if section_name == "상해및질병관련 특별약관":
                                    if "질병관련 특별약관" in self.sections:
                                        self.sections["질병관련 특별약관"]["end"] = page_num  # 같은 페이지로 설정
                                        print(f"[종료] 질병관련 특별약관: {page_num + 1}페이지 (상해및질병 섹션 시작)")
                                    # 상해및질병 섹션 시작 페이지 기록
                                    self.sections[section_name]["start"] = page_num
                                    print(f"[시작] {section_name}: {page_num + 1}페이지")
                                    current_section = section_name
                                    continue

                                # 이전 섹션 종료 (상해및질병 제외)
                                if current_section and current_section != "상해및질병관련 특별약관":
                                    if self.sections[current_section]["end"] is None:
                                        self.sections[current_section]["end"] = page_num - 1
                                        print(f"[종료] {current_section}: {page_num}페이지")
                                
                                # 새 섹션 시작
                                if self.sections[section_name]["start"] is None:
                                    self.sections[section_name]["start"] = page_num
                                    print(f"[시작] {section_name}: {page_num + 1}페이지")
                                    current_section = section_name

        except Exception as e:
            logger.error(f"섹션 검색 중 오류: {str(e)}")

    def find_section_ranges(self) -> Dict[str, Dict[str, int]]:
        try:
            # 보험금 지급 섹션과 보종 찾기
            payment_page, types = self.find_payment_and_types()
            if payment_page is None:
                print("[오류] 보험금 지급 섹션을 찾을 수 없습니다")
                return self.sections

            print(f"\n=== 섹션 검색 시작 (시작 페이지: {payment_page + 1}) ===")
            
            # 보종이 없거나 연속된 보종인 경우
            if not types:
                self.find_sections_in_range(payment_page, len(self.doc) - 1)
            else:
                # 보종별로 처리
                for i, type_info in enumerate(types):
                    type_start = type_info['page']
                    type_end = types[i + 1]['page'] - 1 if i < len(types) - 1 else len(self.doc) - 1
                    self.find_sections_in_range(max(payment_page, type_start), type_end)
            
            return self.sections

        except Exception as e:
            logger.error(f"섹션 범위 찾기 실패: {str(e)}")
            return self.sections

    def find_payment_and_types(self) -> Tuple[Optional[int], List[Dict[str, any]]]:
        payment_page = None
        types = []
        
        for page_num in range(len(self.doc)):
            text = self.doc[page_num].get_text()
            
            # 보험금 지급 섹션 찾기
            if payment_page is None and re.search(self.payment_section_pattern, text):
                payment_page = page_num
                print(f"[찾음] 보험금 지급 섹션: {page_num + 1}페이지")
            
            # 보종 타입 찾기
            if payment_page is not None:
                lines = text.split('\n')
                for line in lines:
                    matches = list(re.finditer(self.type_pattern, line))
                    if len(matches) > 1:
                        consecutive = all(
                            matches[i+1].start() - matches[i].end() <= 2 
                            for i in range(len(matches)-1)
                        )
                        if consecutive:
                            return payment_page, []
                    
                    for match in matches:
                        type_num = match.group(1)
                        types.append({
                            'type': f"[{type_num}종]",
                            'page': page_num
                        })

        return payment_page, sorted(types, key=lambda x: x['page'])

    def close(self):
        if self.doc:
            self.doc.close()
            logger.info(f"PDF 문서 닫힘: {self.pdf_path}")


class HighlightDetector:
    def __init__(self):
        self.yellow_range = {
            'lower': np.array([20, 100, 100]),
            'upper': np.array([30, 255, 255])
        }
        self.kernel_size = (5, 5)

    def pdf_to_image(self, page: fitz.Page) -> np.ndarray:
        """PDF 페이지를 이미지로 변환"""
        try:
            pix = page.get_pixmap(alpha=False)
            img_array = np.frombuffer(pix.samples, dtype=np.uint8)
            img_array = img_array.reshape(pix.height, pix.width, 3)
            return cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
        except Exception as e:
            logger.error(f"PDF 페이지 이미지 변환 실패: {str(e)}")
            return None

    def detect_highlights(self, image: np.ndarray) -> List[np.ndarray]:
        """하이라이트 영역 감지"""
        try:
            if image is None:
                return []

            # BGR -> HSV 변환
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            
            # 노란색 마스크 생성
            mask = cv2.inRange(hsv, self.yellow_range['lower'], self.yellow_range['upper'])
            
            # 노이즈 제
            kernel = np.ones(self.kernel_size, np.uint8)
            mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
            mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
            
            # 윤곽선 찾기
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            return [cnt for cnt in contours if cv2.contourArea(cnt) > 100]
            
        except Exception as e:
            logger.error(f"하이라이트 감지 실패: {str(e)}")
            return []

    def get_highlight_regions(self, contours: List[np.ndarray], image_height: int) -> List[Tuple[int, int]]:
        """하이라이트된 영역의 위치 정보 추출"""
        try:
            regions = []
            for contour in contours:
                x, y, w, h = cv2.boundingRect(contour)
                top = image_height - (y + h)
                bottom = image_height - y
                regions.append((top, bottom))
            return regions
        except Exception as e:
            logger.error(f"하이라이트 영역 추출 실패: {str(e)}")
            return []

    def check_highlight(self, row_range: Tuple[float, float], highlight_regions: List[Tuple[float, float]]) -> bool:
        """특정 행이 하이라이트된 영역과 겹치는지 확인"""
        try:
            row_top, row_bottom = row_range
            for region_top, region_bottom in highlight_regions:
                overlap = min(row_bottom, region_bottom) - max(row_top, region_top)
                if overlap > 0:
                    return True
            return False
        except Exception as e:
            logger.error(f"하이라이트 체크 실패: {str(e)}")
            return False
        
class TableExtractor:
    def __init__(self):
        self.standard_columns = ['담보명', '지급사유', '지급금액', '변경사항', '페이지']
        self.ignore_patterns = [
            r'이륜자동차',
            r'Page',
            r'^\s*$',
            r'특별약관$',
            r'약관$',
            r'^표$',
            r'^그림$',
            r'^주요',
            r'^※',
            r'^상기',
            r'^위'
        ]

        # Section mapping for processing
        self.process_sections = {
            "상해관련 특별약관": "injury_section",
            "질병관련 특별약관": "disease_section",
            "상해및질병관련 특별약관": "both_section"
        }

        # Initialize the SentenceTransformer model
        self.model = SentenceTransformer('models/distiluse-base-multilingual-cased-v1')
        self.model.max_seq_length = 512

        # Reference texts for title classification
        self.title_references = {
            'exclusion': [
                "보상하지 않는 사항",
                "면책사항",
                "보상하지 않는 손해",
                "보상하지 아니하는 사항"
            ],
            'coverage': [
                "보장내용",
                "보험금 지급사유",
                "보상하는 사항",
                "기본계약"
            ]
        }

        # Precompute embeddings for reference titles
        self.reference_embeddings = {
            category: [self.model.encode(text) for text in texts]
            for category, texts in self.title_references.items()
        }

        # Table extraction settings
        self.extraction_settings = {
            'lattice': {
                'line_scale': 40,
                'process_background': True,
                'copy_text': ['v'],
                'line_tol': 2,
                'joint_tol': 2
            },
            'stream': {
                'row_tol': 3,
                'col_tol': 3,
                'edge_tol': 50
            }
        }

    def identify_title_type(self, text: str) -> Tuple[Optional[str], float]:
        """Identify the type of the title and return its similarity score."""
        try:
            text_embedding = self.model.encode(text)
            best_category = None
            best_score = -1

            for category, ref_embeddings in self.reference_embeddings.items():
                for ref_embedding in ref_embeddings:
                    similarity = 1 - cosine(text_embedding, ref_embedding)
                    if similarity > best_score:
                        best_score = similarity
                        best_category = category

            return best_category, best_score
        except Exception as e:
            logger.error(f"Error in identify_title_type: {str(e)}")
            return None, 0.0

    def extract_table_contexts(self, page: fitz.Page) -> List[Dict]:
        """Extract contexts related to tables from the page."""
        try:
            blocks = page.get_text("dict")["blocks"]
            contexts = []
            current_context = {
                'text_before': [],
                'table_area': None,
                'type': None
            }

            for block in blocks:
                if block.get("type") == 0:  # Text block
                    text = block.get("text", '').strip()
                    if text:
                        # Identify title type
                        title_type, score = self.identify_title_type(text)
                        if score > 0.7:  # Threshold
                            if current_context['table_area']:
                                contexts.append(current_context)
                            current_context = {
                                'text_before': [text],
                                'table_area': None,
                                'type': title_type
                            }
                        else:
                            current_context['text_before'].append(text)
                elif block.get("type") == 1:  # Image or table block
                    # Check if 'bbox' key exists
                    if 'bbox' in block:
                        current_context['table_area'] = block['bbox']

            if current_context['table_area']:
                contexts.append(current_context)

            return contexts
        except Exception as e:
            logger.error(f"Error in extract_table_contexts: {str(e)}")
            return []

    def extract_tables_with_context(self, pdf_path: str, page_num: int) -> List[Dict]:
        """Extract tables along with their context from a PDF page."""
        try:
            doc = fitz.open(pdf_path)
            page = doc[page_num]

            # Extract contexts
            contexts = self.extract_table_contexts(page)
            tables = []

            for context in contexts:
                if context['table_area']:
                    # Extract table using Camelot
                    table_areas = [
                        f"{context['table_area'][0]},{context['table_area'][1]},"
                        f"{context['table_area'][2]},{context['table_area'][3]}"
                    ]
                    extracted_tables = camelot.read_pdf(
                        pdf_path,
                        pages=str(page_num + 1),
                        flavor='lattice',
                        table_areas=table_areas
                    )

                    for table in extracted_tables:
                        df = self.process_table(table.df, page_num + 1)
                        if not df.empty:
                            tables.append({
                                'table': df,
                                'type': context['type'],
                                'metadata': {
                                    'text_before': context['text_before'],
                                    'position': context['table_area']
                                }
                            })

            doc.close()
            return tables
        except Exception as e:
            logger.error(f"Error in extract_tables_with_context: {str(e)}")
            return []

    def process_table(self, df: pd.DataFrame, page_num: int) -> pd.DataFrame:
        """표 데이터 처리"""
        df = df.fillna('')
        df = self.clean_table_content(df)
        df = self.standardize_columns(df, page_num)
        
        if self.validate_table(df):
            return df
        return pd.DataFrame()

    def merge_tables(self, tables: List[Dict], section_type: str) -> Dict[str, pd.DataFrame]:
        """섹션별 테이블 병합"""
        merged_tables = {
            'coverage': pd.DataFrame(),
            'exclusion': pd.DataFrame()
        }
        
        for table_info in tables:
            table_type = table_info['type']
            if table_type in merged_tables:
                if merged_tables[table_type].empty:
                    merged_tables[table_type] = table_info['table']
                else:
                    merged_tables[table_type] = pd.concat(
                        [merged_tables[table_type], table_info['table']],
                        ignore_index=True
                    )
        
        # 중복 제거 및 정렬
        for table_type in merged_tables:
            if not merged_tables[table_type].empty:
                merged_tables[table_type] = merged_tables[table_type].drop_duplicates()
                merged_tables[table_type] = merged_tables[table_type].sort_values(['페이지', '담보명'])
        
        return merged_tables
    
    def extract_section_tables(self, doc, start_page: int, end_page: int, section_type: str) -> Dict[str, pd.DataFrame]:
        """섹션별 테이블 추출"""
        try:
            # 상해및질병 섹션 처리 제외
            if "상해및질병" in section_type:
                return {}

            # 처리 대상 섹션인지 확인
            if section_type not in self.process_sections:
                logger.info(f"처리 대상이 아닌 섹션: {section_type}")
                return {}

            # PDF 경로 설정
            if isinstance(doc, fitz.Document):
                self.pdf_path = doc.name
            elif isinstance(doc, str):
                self.pdf_path = doc
            else:
                raise ValueError("Invalid document type")

            if not self.pdf_path:
                raise ValueError("PDF path not set")

            logger.info(f"{section_type} 섹션 테이블 추출 시작 (페이지: {start_page + 1} ~ {end_page + 1})")
            tables = []

            # 페이지별 테이블 추출
            for page_num in range(start_page, end_page + 1):
                try:
                    logger.info(f"Processing page-{page_num + 1}")
                    
                    # 페이지 컨텍스트 추출
                    page_tables = self.extract_tables_with_context(self.pdf_path, page_num)
                    
                    for table_info in page_tables:
                        df = table_info['table']
                        df = self.merge_multiline_cells(df)  # 다중행 셀 병합 적용
                        df = self.clean_table_content(df)
                        
                        if self.validate_table(df) and not df.empty:
                            tables.append(df)

                except Exception as e:
                    logger.error(f"페이지 {page_num + 1} 처리 중 오류: {str(e)}")
                    continue

            # 테이블 병합 및 반환
            if tables:
                merged_df = self.merge_tables(tables, section_type)
                if not all(df.empty for df in merged_df.values()):
                    section_key = self.process_sections[section_type]
                    return {section_key: merged_df}

            logger.warning(f"{section_type} 섹션에서 추출된 테이블 없음")
            return {}

        except Exception as e:
            logger.error(f"섹션 테이블 추출 중 오류: {str(e)}")
            return {}

    def merge_multiline_cells(self, df: pd.DataFrame) -> pd.DataFrame:
        """다중 행으로 분리된 셀 병합"""
        try:
            for col in df.columns:
                current_value = []
                merged_values = []
                
                for value in df[col]:
                    value = str(value).strip()
                    if value:
                        current_value.append(value)
                    else:
                        if current_value:
                            merged_values.append(' '.join(current_value))
                            current_value = []
                        merged_values.append('')
                
                if current_value:
                    merged_values.append(' '.join(current_value))
                    
                if len(merged_values) == len(df):
                    df[col] = merged_values
            
            return df
            
        except Exception as e:
            logger.error(f"셀 병합 실패: {str(e)}")
            return df

    def clean_table_content(self, df: pd.DataFrame) -> pd.DataFrame:
        """테이블 내용 정제"""
        try:
            if df.empty:
                return df

            # None 값을 빈 문자열로 변환
            df = df.fillna('')
            
            # 모든 컬럼의 데이터를 문자열로 변환
            for col in df.columns:
                df[col] = df[col].astype(str)
                df[col] = df[col].apply(lambda x: x.strip() if x else '')
            
            initial_rows = len(df)
            
            # 패턴 매칭을 위한 데이터 정제
            rows_to_keep = []
            for idx, row in df.iterrows():
                cell_value = str(row.get('담보명', '')).strip()
                
                # 무시할 패턴 확인
                should_ignore = any(
                    re.search(pattern, cell_value, re.IGNORECASE) 
                    for pattern in self.ignore_patterns
                )
                
                if not should_ignore and cell_value:  # 빈 값이 아닐 경우만 유지
                    rows_to_keep.append(idx)
            
            # 선택된 행만 유지
            df = df.loc[rows_to_keep]
            
            # 줄바꿈 및 연속된 공백 처리
            for col in df.columns:
                if col != '페이지':
                    df[col] = df[col].str.replace('\n', ' ')
                    df[col] = df[col].str.replace(r'\s+', ' ', regex=True)
                    df[col] = df[col].str.strip()
            
            final_rows = len(df)
            removed_rows = initial_rows - final_rows
            if removed_rows > 0:
                logger.info(f"불필요한 행 {removed_rows}개 제거됨")
            
            return df

        except Exception as e:
            logger.error(f"테이블 내용 정제 실패: {str(e)}")
            return df

    def standardize_columns(self, df: pd.DataFrame, page_num: int) -> pd.DataFrame:
        """컬럼 표준화 및 페이지 정보 추가"""
        try:
            logger.info(f"컬럼 표준화 시작 (페이지 {page_num})")
            
            # None 값을 빈 문자열로 변환
            df = df.fillna('')
            df = df.copy()
            
            # 컬럼 매핑 적용
            column_mapping = {
                '보장명': '담보명',
                '급부명': '담보명',
                '보험금지급사유': '지급사유',
                '지급조건': '지급사유',
                '보험금액': '지급금액',
                '지급액': '지급금액',
                '비고': '변경사항'
            }
            
            # 컬럼이 모두 숫자인 경우 처리
            if all(str(col).isdigit() for col in df.columns):
                logger.info("숫자로 된 컬럼명 발견, 기본 컬럼명 적용")
                if len(df.columns) >= 3:
                    df.columns = ['담보명', '지급사유', '지급금액'] + [''] * (len(df.columns) - 3)
            else:
                # 컬럼명 매핑 적용
                for old_col, new_col in column_mapping.items():
                    if old_col in df.columns:
                        df = df.rename(columns={old_col: new_col})
            
            # 필수 컬럼 추가
            for col in self.standard_columns:
                if col not in df.columns:
                    df[col] = ''
            
            # 페이지 번호 설정
            df['페이지'] = page_num
            
            # 컬럼 순서 조정
            df = df[self.standard_columns]
            
            return df

        except Exception as e:
            logger.error(f"컬럼 표준화 실패: {str(e)}")
            return pd.DataFrame(columns=self.standard_columns)

    def merge_tables(self, tables: List[pd.DataFrame], section_type: str) -> pd.DataFrame:
        """동일 섹션의 테이블 병합"""
        try:
            logger.info(f"{section_type} 섹션 테이블 병합 시작")
            
            if not tables:
                logger.warning("병합할 테이블 없음")
                return pd.DataFrame(columns=self.standard_columns)

            # 테이블 병합
            merged_df = pd.concat(tables, ignore_index=True)
            
            # 중복 제거
            merged_df = merged_df.drop_duplicates()
            
            # 정렬
            merged_df = merged_df.sort_values(['페이지', '담보명'])
            
            return merged_df

        except Exception as e:
            logger.error(f"테이블 병합 실패: {str(e)}")
            return pd.DataFrame(columns=self.standard_columns)

    def validate_table(self, df: pd.DataFrame) -> bool:
        """테이블 유효성 검사"""
        try:
            # 빈 데이터프레임 체크
            if df.empty:
                logger.warning("빈 데이터프레임")
                return False

            # 필수 컬럼 존재 여부 체크
            missing_cols = [col for col in self.standard_columns if col not in df.columns]
            if missing_cols:
                logger.warning(f"필수 컬럼 없음: {missing_cols}")
                return False

            # 데이터 존재 여부 체크
            if df['담보명'].isna().all() or df['담보명'].str.strip().eq('').all():
                logger.warning("담보명 데이터 없음")
                return False

            # 유효한 데이터가 있는지 확인
            valid_data = df.apply(lambda x: x.astype(str).str.strip().ne('').any())
            if not valid_data.any():
                logger.warning("유효한 데이터가 없음")
                return False

            logger.info("테이블 유효성 검사 통과")
            return True

        except Exception as e:
            logger.error(f"테이블 유효성 검사 실패: {str(e)}")
            return False
        

class ExcelWriter:
    def __init__(self):
        self.styles = {
            'header': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'cell': {
                'alignment': Alignment(vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'highlight': {
                'fill': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            }
        }

    def apply_cell_style(self, cell, style_name: str):
        """스타일 적용"""
        for attr, value in self.styles[style_name].items():
            setattr(cell, attr, value)

    def write_table(self, worksheet, df: pd.DataFrame, start_row: int, 
                    title: str) -> int:
        """테이블 작성"""
        try:
            # 제목 작성
            title_cell = worksheet.cell(row=start_row, column=1, value=title)
            self.apply_cell_style(title_cell, 'header')
            worksheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=start_row,
                end_column=len(df.columns)
            )
            
            # 열 헤더 작성
            for col_idx, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=start_row + 2, column=col_idx, value=column)
                self.apply_cell_style(cell, 'header')
            
            # 데이터 작성
            for row_idx, row in enumerate(df.itertuples(index=False), start_row + 3):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    self.apply_cell_style(cell, 'cell')
                    
                    # 변경사항이 '추가'인 행 하이라이트
                    if col_idx == df.columns.get_loc('변경사항') + 1 and value == '추가':
                        for highlight_col in range(1, len(df.columns) + 1):
                            self.apply_cell_style(
                                worksheet.cell(row=row_idx, column=highlight_col),
                                'highlight'
                            )
            
            return start_row + len(df) + 4

        except Exception as e:
            logger.error(f"테이블 작성 실패: {str(e)}")
            return start_row + 1

    def adjust_column_widths(self, worksheet):
        """열 너비 자동 조정"""
        try:
            for column in worksheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except:
                        pass
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = \
                    min(max_length + 2, 50)
        except Exception as e:
            logger.error(f"열 너비 조정 실패: {str(e)}")

    def save_to_excel(self, tables_data: Dict[str, pd.DataFrame], output_path: str):
        """엑셀 파일로 저장"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "보장내용"
            current_row = 1

            for section_type, df in tables_data.items():
                # DataFrame 검사를 명시적으로 수정
                if not isinstance(df, pd.DataFrame):
                    logger.warning(f"{section_type}: 유효하지 않은 데이터 형식")
                    continue
                
                if df.empty:  # df.empty를 사용하여 명시적으로 빈 DataFrame 체크
                    logger.warning(f"{section_type}: 데이터 없음")
                    continue

                # 섹션 제목 작성
                title_cell = ws.cell(row=current_row, column=1, value=section_type)
                title_cell.font = Font(bold=True, size=14)
                current_row += 2

                # 컬럼 헤더 작성
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                current_row += 1

                # 데이터 작성
                for _, row in df.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                    current_row += 1
                current_row += 2

            # 열 너비 자동 조정
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

            wb.save(output_path)
            logger.info(f"Excel 파일 저장 완료: {output_path}")

        except Exception as e:
            logger.error(f"Excel 저장 실패: {str(e)}")
            raise


class PDFAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("KB손해보험 상품 분석기")
        self.root.geometry("800x600")
        self.file_path_var = tk.StringVar()
        self.progress_var = tk.StringVar(value="대기 중...")
        self.setup_gui()

    def setup_gui(self):
        """GUI 구성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 제목
        title_label = ttk.Label(
            main_frame,
            text="보험약관 테이블 분석기",
            font=("Helvetica", 16, "bold")
        )
        title_label.pack(pady=10)

        # 파일 선택 프레임
        file_frame = ttk.LabelFrame(main_frame, text="PDF 파일 선택", padding="10")
        file_frame.pack(fill=tk.X, pady=10)

        file_entry = ttk.Entry(
            file_frame,
            textvariable=self.file_path_var,
            width=60
        )
        file_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        browse_button = ttk.Button(
            file_frame,
            text="찾아보기",
            command=self.browse_file
        )
        browse_button.pack(side=tk.LEFT, padx=5)

        # 진행 상태 표시
        progress_label = ttk.Label(
            main_frame,
            textvariable=self.progress_var
        )
        progress_label.pack(pady=5)

        self.progress_bar = ttk.Progressbar(
            main_frame,
            mode='determinate',
            length=300
        )
        self.progress_bar.pack(pady=5)

        # 분석 버튼
        self.analyze_button = ttk.Button(
            main_frame,
            text="분석 시작",
            command=self.start_analysis,
            state=tk.DISABLED
        )
        self.analyze_button.pack(pady=10)

        # 로그 영역
        log_frame = ttk.LabelFrame(main_frame, text="처리 로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = tk.Text(log_frame, height=15, width=70)
        scrollbar = ttk.Scrollbar(
            log_frame,
            orient=tk.VERTICAL,
            command=self.log_text.yview
        )
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def browse_file(self):
        """파일 선택"""
        file_path = filedialog.askopenfilename(
            title="PDF 파일 선택",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.analyze_button['state'] = tk.NORMAL
            self.log_message(f"파일 선택됨: {file_path}")

    def log_message(self, message: str, level: str = "INFO"):
        """로그 메시지 추가"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {level}: {message}\n"
        
        self.log_text.insert(tk.END, log_entry)
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_progress(self, value: int, message: str = None):
        """진행 상태 업데이트"""
        self.progress_bar['value'] = value
        if message:
            self.progress_var.set(message)
        self.root.update_idletasks()

    def analyze_pdf(self, pdf_path: str):
        """PDF 분석 실행"""
        try:
            # 출력 폴더 생성
            output_dir = os.path.join(os.path.dirname(pdf_path), "output")
            os.makedirs(output_dir, exist_ok=True)

            self.log_message("PDF 분석 시작")
            self.update_progress(10, "PDF 문서 로드 중...")

            # PDF 문서 로드 및 섹션 분석
            pdf_doc = PDFDocument(pdf_path)
            sections = pdf_doc.find_section_ranges()

            if not any(sections[section]["start"] is not None for section in sections):
                self.log_message("섹션을 찾을 수 없습니다.", "WARNING")
                messagebox.showwarning("경고", "문서에서 섹션을 찾을 수 없습니다.")
                return

            # 테이블 추출기 초기화
            extractor = TableExtractor()
            all_tables = {}

            # 섹션별 테이블 추출
            section_count = len([s for s in sections if sections[s]["start"] is not None])
            current_section = 0

            for section_type, ranges in sections.items():
                if ranges["start"] is not None:
                    current_section += 1
                    progress = 10 + (current_section / section_count) * 60
                    self.update_progress(
                        progress,
                        f"{section_type} 섹션 테이블 추출 중..."
                    )
                    
                    tables = extractor.extract_section_tables(
                        pdf_doc.doc,
                        ranges["start"],
                        ranges["end"],
                        section_type
                    )
                    
                    all_tables.update(tables)

            # PDF 문서 닫기
            pdf_doc.close()

            if any(not df.empty for df in all_tables.values()):
                self.update_progress(70, "Excel 파일 생성 중...")
                
                # 결과 저장
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = os.path.join(
                    output_dir,
                    f"analyzed_tables_{timestamp}.xlsx"
                )
                
                excel_writer = ExcelWriter()
                excel_writer.save_to_excel(all_tables, output_path)
                
                # Excel 정제 시작
                self.update_progress(90, "Excel 파일 정제 중...")
                self.log_message("Excel 파일 정제 작업 시작")
                
                try:
                    cleaner = ExcelCleaner()
                    final_path = cleaner.clean_excel(output_path)
                    
                    if final_path != output_path:
                        self.log_message("헤더 행이 제거된 새 파일이 생성되었습니다")
                    else:
                        self.log_message("제거할 헤더 행이 발견되지 않았습니다")
                        
                except Exception as e:
                    self.log_message(f"Excel 정제 중 오류 발생: {str(e)}", "ERROR")
                    final_path = output_path  # 오류 발생 시 원본 파일 사용
                
                self.update_progress(100, "완료")
                self.log_message(f"분석 완료. 결과 파일: {final_path}")
                
                # 결과 파일 경로를 상대 경로로 변환
                try:
                    rel_path = os.path.relpath(final_path, os.path.dirname(pdf_path))
                    display_path = rel_path
                except:
                    display_path = final_path
                    
                messagebox.showinfo(
                    "완료",
                    f"분석이 완료되었습니다.\n저장 위치: {display_path}"
                )
            else:
                self.log_message("추출된 테이블이 없습니다.", "WARNING")
                messagebox.showwarning("경고", "추출된 테이블이 없습니다.")

        except Exception as e:
            self.log_message(f"오류 발생: {str(e)}", "ERROR")
            messagebox.showerror(
                "오류",
                f"처리 중 오류가 발생했습니다:\n{str(e)}"
            )
        finally:
            # 버튼 상태 복구
            self.analyze_button['state'] = tk.NORMAL
            self.update_progress(0, "대기 중...")

    def start_analysis(self):
        """분석 시작"""
        pdf_path = self.file_path_var.get()
        if not pdf_path:
            messagebox.showerror("오류", "PDF 파일을 선택해주세요.")
            return

        self.analyze_button['state'] = tk.DISABLED
        self.progress_bar['value'] = 0
        self.progress_var.set("처리 중...")

        # 별도 스레드에서 처리
        threading.Thread(
            target=self.analyze_pdf,
            args=(pdf_path,),
            daemon=True
        ).start()

    def run(self):
        """애플리케이션 실행"""
        self.root.mainloop()


class ExcelCleaner:
    def __init__(self):
        self.header_pattern = re.compile(r'^보\s+장\s+명\s+지\s+급\s+사\s+유\s+지\s+급\s+금\s+액$')
        self.logger = logging.getLogger(__name__)

    def clean_excel(self, input_path: str) -> str:
        """엑셀 파일을 읽어서 헤더 행을 제거하고 새 파일로 저장"""
        try:
            # 입력 파일 경로 확인
            if not os.path.exists(input_path):
                raise FileNotFoundError(f"파일을 찾을 수 없습니다: {input_path}")

            # 새 파일명 생성 (원본파일명_cleaned.xlsx)
            output_path = input_path.replace('.xlsx', '_cleaned.xlsx')
            
            # 워크북 로드
            self.logger.info(f"파일 로드 중: {input_path}")
            wb = load_workbook(input_path)
            ws = wb.active
            
            # 데이터 처리
            rows_to_delete = []
            processed = False
            
            # 모든 행 검사
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
                # 첫 번째 열(담보명)의 값 확인
                if row[0]:  # None이 아닌 경우에만 처리
                    cell_value = str(row[0]).strip()
                    # 공백 정규화
                    normalized_value = ' '.join(cell_value.split())
                    
                    # 패턴 매칭 전 로깅
                    self.logger.debug(f"검사 중인 값: '{normalized_value}'")
                    
                    if self.header_pattern.match(normalized_value):
                        rows_to_delete.append(row_idx)
                        self.logger.info(f"헤더 행 발견: 행 {row_idx}")
                        processed = True

            # 행 삭제 (뒤에서부터 삭제하여 인덱스 변화 방지)
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)
                
            if processed:
                # 새 파일로 저장
                wb.save(output_path)
                self.logger.info(f"처리 완료. 저장된 파일: {output_path}")
                self.logger.info(f"제거된 행 수: {len(rows_to_delete)}")
                return output_path
            else:
                self.logger.info("제거할 헤더 행이 없습니다.")
                return input_path

        except Exception as e:
            self.logger.error(f"엑셀 정제 중 오류 발생: {str(e)}")
            raise

    def process_directory(self, directory: str):
        """디렉토리 내의 모든 엑셀 파일 처리"""
        try:
            self.logger.info(f"디렉토리 처리 시작: {directory}")
            
            # 엑셀 파일 찾기
            excel_files = glob.glob(os.path.join(directory, "*.xlsx"))
            
            if not excel_files:
                self.logger.warning(f"엑셀 파일을 찾을 수 없습니다: {directory}")
                return
            
            for excel_file in excel_files:
                try:
                    self.logger.info(f"파일 처리 중: {excel_file}")
                    output_path = self.clean_excel(excel_file)
                    if output_path != excel_file:
                        self.logger.info(f"처리 완료: {output_path}")
                except Exception as e:
                    self.logger.error(f"파일 처리 실패 ({excel_file}): {str(e)}")
                    continue

        except Exception as e:
            self.logger.error(f"디렉토리 처리 중 오류 발생: {str(e)}")
            raise


def main():
    try:
        app = PDFAnalyzerGUI()
        app.run()
    except Exception as e:
        logger.error(f"프로그램 실행 중 오류 발생: {str(e)}")
        raise

if __name__ == "__main__":
    main()