import camelot
import cv2
import numpy as np
from pathlib import Path
import pandas as pd
import json
from typing import List, Dict, Any, Optional
import fitz
from dataclasses import dataclass
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

@dataclass
class MergedCell:
    start_row: int
    end_row: int
    start_col: int
    end_col: int

class TableAnalyzer:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.INFO)

    def analyze_page_68(self, pdf_path: str) -> dict:
        """
        PDF의 68페이지 표를 분석하고 병합 셀 정보를 추출
        """
        try:
            # 1. Camelot으로 기본 표 구조 추출
            tables = camelot.read_pdf(
                pdf_path,
                pages='68',
                flavor='lattice',  # 테두리가 있는 표에 적합
                line_scale=40      # 선 감지 민감도 조정
            )
            
            if not tables:
                raise ValueError("No tables found on page 68")

            # 기본 표 구조
            base_df = tables[0].df

            # 2. PDF를 이미지로 변환하여 병합 셀 분석
            doc = fitz.open(pdf_path)
            page = doc[67]  # 0-based index
            pix = page.get_pixmap()
            img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                pix.height, pix.width, pix.n
            )
            
            # 3. 병합 셀 감지
            merged_cells = self._detect_merged_cells(img)
            
            # 4. 결과 생성
            result = {
                'data': base_df,
                'merged_cells': merged_cells,
                'num_rows': len(base_df),
                'num_cols': len(base_df.columns)
            }

            doc.close()
            return result

        except Exception as e:
            self.logger.error(f"Error analyzing table on page 68: {str(e)}")
            raise

    def _detect_merged_cells(self, img: np.ndarray) -> List[MergedCell]:
        """이미지 처리를 통한 병합 셀 감지"""
        # 이미지 전처리
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
        _, binary = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)

        # 수평/수직 선 감지
        h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
        v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))

        horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, h_kernel)
        vertical = cv2.morphologyEx(binary, cv2.MORPH_OPEN, v_kernel)

        # 교차점 찾기
        joints = cv2.bitwise_and(horizontal, vertical)
        
        # 연결된 컴포넌트 분석
        num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(
            cv2.bitwise_not(joints), connectivity=8
        )

        merged_cells = []
        for i in range(1, num_labels):  # 0은 배경
            x, y, w, h, area = stats[i]
            if w > 50 and h > 50:  # 최소 크기 조건
                merged_cells.append(MergedCell(
                    start_row=y // 50,  # 대략적인 행 인덱스
                    end_row=(y + h) // 50,
                    start_col=x // 100,  # 대략적인 열 인덱스
                    end_col=(x + w) // 100
                ))

        return merged_cells

    def save_to_excel(self, result: dict, output_path: str):
        """결과를 Excel 파일로 저장"""
        wb = Workbook()
        ws = wb.active
        
        # 스타일 설정
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 열 폭 자동 조정을 위한 최대 길이 계산
        column_widths = []
        df = result['data']
        for i in range(len(df.columns)):
            column_cells = df.iloc[:, i].astype(str)
            max_length = max(
                max(len(str(cell)) for cell in column_cells),
                len(str(df.columns[i]))
            )
            column_widths.append(max_length)

        # 데이터 입력
        for i, row in enumerate(df.values, 1):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                # 열 폭 설정
                ws.column_dimensions[chr(64 + j)].width = column_widths[j-1] + 2

        # 병합 셀 처리
        for cell in result['merged_cells']:
            try:
                ws.merge_cells(
                    start_row=cell.start_row + 1,
                    start_column=cell.start_col + 1,
                    end_row=cell.end_row + 1,
                    end_column=cell.end_col + 1
                )
            except ValueError as e:
                self.logger.warning(f"Merging cells failed: {str(e)}")
                continue

        self.logger.info(f"Saving Excel file to {output_path}")
        wb.save(output_path)

    def generate_html(self, result: dict) -> str:
        """HTML 테이블 코드 생성"""
        df = result['data']
        merged_cells = result['merged_cells']
        
        html = ['<table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">']
        
        # 헤더 스타일
        header_style = 'style="background-color: #f2f2f2; font-weight: bold;"'
        
        for i, row in enumerate(df.values):
            html.append('<tr>')
            for j, value in enumerate(row):
                # 병합된 셀 확인
                merged = next(
                    (cell for cell in merged_cells 
                     if cell.start_row == i and cell.start_col == j),
                    None
                )
                
                if merged:
                    rowspan = merged.end_row - merged.start_row + 1
                    colspan = merged.end_col - merged.start_col + 1
                    style = header_style if i == 0 else ''
                    html.append(
                        f'<td {style} rowspan="{rowspan}" colspan="{colspan}">{value}</td>'
                    )
                else:
                    # 일반 셀 처리
                    if not any(
                        i >= cell.start_row and i <= cell.end_row and
                        j >= cell.start_col and j <= cell.end_col
                        for cell in merged_cells
                    ):
                        style = header_style if i == 0 else ''
                        html.append(f'<td {style}>{value}</td>')
            
            html.append('</tr>')
        
        html.append('</table>')
        return '\n'.join(html)

    def save_as_json(self, result: dict, output_path: str):
        """결과를 JSON 형식으로 저장"""
        json_data = {
            'data': result['data'].to_dict('records'),
            'merged_cells': [
                {
                    'start_row': cell.start_row,
                    'end_row': cell.end_row,
                    'start_col': cell.start_col,
                    'end_col': cell.end_col
                }
                for cell in result['merged_cells']
            ],
            'num_rows': result['num_rows'],
            'num_cols': result['num_cols']
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"Saving JSON file to {output_path}")