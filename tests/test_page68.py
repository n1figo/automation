import pytest
from pathlib import Path
import json
import pandas as pd
from openpyxl import load_workbook
import logging
from table_analyzer import TableAnalyzer, MergedCell

class TestPage68Analysis:
    @pytest.fixture
    def test_data_dir(self):
        """테스트 데이터 디렉토리"""
        test_dir = Path("tests/test_data")
        test_dir.mkdir(parents=True, exist_ok=True)
        return test_dir

    @pytest.fixture
    def output_dir(self):
        """결과 파일 저장 디렉토리"""
        output_dir = Path("tests/output")
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir

    @pytest.fixture
    def analyzer(self):
        return TableAnalyzer()

    @pytest.fixture
    def pdf_path(self, test_data_dir):
        """테스트할 PDF 파일 경로"""
        return test_data_dir / "test_document.pdf"

    def test_page_68_table_structure(self, analyzer, pdf_path):
        """68페이지 표 구조 분석 테스트"""
        try:
            result = analyzer.analyze_page_68(str(pdf_path))
            
            # 기본 구조 검증
            assert isinstance(result, dict)
            assert 'data' in result
            assert 'merged_cells' in result
            assert 'num_rows' in result
            assert 'num_cols' in result
            
            # DataFrame 검증
            assert isinstance(result['data'], pd.DataFrame)
            assert not result['data'].empty
            
            print("\n=== 68페이지 분석 결과 ===")
            print(f"행 수: {result['num_rows']}")
            print(f"열 수: {result['num_cols']}")
            print(f"병합된 셀 수: {len(result['merged_cells'])}")
            print("\n=== 데이터 미리보기 ===")
            print(result['data'].head())
            
            return result  # 다른 테스트에서 재사용할 수 있도록 반환
            
        except Exception as e:
            pytest.fail(f"68페이지 분석 실패: {str(e)}")

    def test_excel_output(self, analyzer, pdf_path, output_dir):
        """Excel 출력 테스트"""
        result = analyzer.analyze_page_68(str(pdf_path))
        excel_path = output_dir / "page_68_table.xlsx"
        
        analyzer.save_to_excel(result, str(excel_path))
        
        # Excel 파일 검증
        assert excel_path.exists()
        wb = load_workbook(str(excel_path))
        ws = wb.active
        
        # 데이터 존재 확인
        assert ws['A1'].value is not None
        
        # 병합된 셀 확인
        merged_ranges = ws.merged_cells.ranges
        assert len(merged_ranges) > 0
        
        print(f"\nExcel 파일 저장됨: {excel_path}")
        print(f"병합된 셀 수: {len(merged_ranges)}")
        
        return excel_path

    def test_json_output(self, analyzer, pdf_path, output_dir):
        """JSON 출력 테스트"""
        result = analyzer.analyze_page_68(str(pdf_path))
        json_path = output_dir / "page_68_table.json"
        
        analyzer.save_as_json(result, str(json_path))
        
        # JSON 파일 검증
        assert json_path.exists()
        
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        assert 'data' in data
        assert 'merged_cells' in data
        assert len(data['data']) > 0
        
        print(f"\nJSON 파일 저장됨: {json_path}")
        print(f"데이터 행 수: {len(data['data'])}")
        print(f"병합된 셀 정보: {len(data['merged_cells'])}개")
        
        return json_path

    def test_html_generation(self, analyzer, pdf_path, output_dir):
        """HTML 생성 테스트"""
        result = analyzer.analyze_page_68(str(pdf_path))
        html_code = analyzer.generate_html(result)
        
        # HTML 파일로 저장
        html_path = output_dir / "page_68_table.html"
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write('''
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <style>
                        table { 
                            border-collapse: collapse; 
                            width: 100%;
                            margin: 20px 0;
                        }
                        th, td { 
                            padding: 8px; 
                            border: 1px solid #ddd;
                            text-align: left;
                            vertical-align: top;
                        }
                        th { 
                            background-color: #f2f2f2;
                            font-weight: bold;
                        }
                        tr:nth-child(even) { 
                            background-color: #f9f9f9; 
                        }
                    </style>
                </head>
                <body>
                    <h2>68페이지 표 분석 결과</h2>
            ''')
            f.write(html_code)
            f.write('</body></html>')
        
        print(f"\nHTML 파일 저장됨: {html_path}")
        return html_path

    def test_all_outputs(self, analyzer, pdf_path, output_dir):
        """모든 출력 형식 통합 테스트"""
        # 1. 표 분석
        result = self.test_page_68_table_structure(analyzer, pdf_path)
        
        # 2. 각 형식으로 저장
        excel_path = self.test_excel_output(analyzer, pdf_path, output_dir)
        json_path = self.test_json_output(analyzer, pdf_path, output_dir)
        html_path = self.test_html_generation(analyzer, pdf_path, output_dir)
        
        # 3. 모든 출력 파일 존재 확인
        assert excel_path.exists()
        assert json_path.exists()
        assert html_path.exists()
        
        print("\n=== 모든 출력 파일 생성 완료 ===")
        print(f"Excel: {excel_path}")
        print(f"JSON: {json_path}")
        print(f"HTML: {html_path}")

if __name__ == "__main__":
    pytest.main(["-v", __file__])