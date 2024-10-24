from langchain_community.document_loaders import PyPDFLoader
import camelot
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import re

def extract_titles_and_tables(pdf_path, page_num):
    """특정 페이지의 표 제목과 표 내용을 추출"""
    # 1. PyPDFLoader로 페이지 텍스트 추출
    loader = PyPDFLoader(pdf_path)
    pages = loader.load()
    page_content = pages[page_num-1].page_content
    print("\n=== 페이지 텍스트 ===")
    print(page_content[:200])  # 처음 200자만 출력

    # 2. 제목 찾기 (패턴: 줄바꿈으로 구분된 텍스트 중 "관련" 또는 "특약"이 포함된 줄)
    lines = page_content.split('\n')
    titles = []
    for i, line in enumerate(lines):
        if ("관련" in line or "특약" in line or "기본계약" in line or "의무부가계약" in line) and len(line.strip()) > 5:
            titles.append(line.strip())
            print(f"\n발견된 제목: {line.strip()}")

    # 3. Camelot으로 표 추출
    print("\n=== 표 추출 시도 ===")
    tables = camelot.read_pdf(
        pdf_path,
        pages=str(page_num),
        flavor='lattice'
    )
    if not tables:
        tables = camelot.read_pdf(
            pdf_path,
            pages=str(page_num),
            flavor='stream'
        )
    print(f"추출된 표 수: {len(tables)}")

    # 4. 결과 저장을 위한 데이터 구조
    results = []
    for i, table in enumerate(tables):
        title = titles[i] if i < len(titles) else f"Table {i+1}"
        results.append({
            'title': title,
            'table': table.df,
            'page': page_num
        })

    return results

def save_to_excel(results, output_path):
    """추출된 표와 제목을 Excel 파일로 저장"""
    wb = Workbook()
    ws = wb.active
    current_row = 1

    for item in results:
        # 제목 셀 작성
        title_cell = ws.cell(row=current_row, column=1, 
                           value=f"{item['title']} (Page: {item['page']})")
        title_cell.font = Font(bold=True, size=12)
        title_cell.fill = PatternFill(start_color='E6E6E6', 
                                    end_color='E6E6E6', 
                                    fill_type='solid')
        
        current_row += 2

        # 표 데이터 작성
        df = item['table']
        for r_idx, row in enumerate(df.values):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=current_row + r_idx, 
                             column=c_idx + 1, 
                             value=value)
                cell.alignment = Alignment(wrap_text=True)

        current_row += len(df) + 3

    wb.save(output_path)
    print(f"\n결과가 {output_path}에 저장되었습니다.")

def main():
    # 설정
    pdf_path = "/workspaces/automation/uploads/KB 9회주는 암보험Plus(무배당)(24.05)_요약서_10.1판매_v1.0_앞단.pdf"  # 실제 PDF 경로로 변경
    output_path = "test_page59_tables.xlsx"
    test_page = 59

    print(f"\n=== {test_page}페이지 처리 시작 ===")
    
    # 제목과 표 추출
    results = extract_titles_and_tables(pdf_path, test_page)
    
    # Excel 파일로 저장
    save_to_excel(results, output_path)

    print("\n처리 완료!")
    print(f"- 발견된 표 수: {len(results)}")
    print(f"- 저장된 파일: {output_path}")

if __name__ == "__main__":
    main()