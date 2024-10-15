import camelot
import pandas as pd
import numpy as np
import cv2
import os
import fitz
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def pdf_to_image(page):
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return np.array(img)

def detect_highlights(image):
    hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
    
    # 여러 색상 범위 정의 (HSV)
    color_ranges = [
        ((20, 100, 100), (40, 255, 255)),  # 노란색
        ((100, 100, 100), (140, 255, 255)),  # 파란색
        ((125, 100, 100), (155, 255, 255))  # 보라색
    ]
    
    masks = []
    for lower, upper in color_ranges:
        mask = cv2.inRange(hsv, np.array(lower), np.array(upper))
        masks.append(mask)
    
    # 모든 마스크 결합
    combined_mask = np.zeros_like(masks[0])
    for mask in masks:
        combined_mask = cv2.bitwise_or(combined_mask, mask)
    
    kernel = np.ones((5,5), np.uint8)
    cleaned_mask = cv2.morphologyEx(combined_mask, cv2.MORPH_CLOSE, kernel)
    cleaned_mask = cv2.morphologyEx(cleaned_mask, cv2.MORPH_OPEN, kernel)
    
    contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    return contours

def get_highlight_regions(contours, image_height):
    regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        top = image_height - (y + h)
        bottom = image_height - y
        regions.append((top, bottom))
    return regions

def extract_tables_with_camelot(pdf_path, page_number):
    print(f"Extracting tables from page {page_number} using Camelot...")
    tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='lattice')
    print(f"Found {len(tables)} tables on page {page_number}")
    return tables

def process_tables(tables, highlight_regions, page_height):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df
        x1, y1, x2, y2 = table._bbox

        table_height = y2 - y1
        row_height = table_height / len(df)

        for row_index in range(len(df)):
            row_data = df.iloc[row_index].copy()
            
            row_top = y2 - (row_index + 1) * row_height
            row_bottom = y2 - row_index * row_height
            
            row_highlighted = check_highlight((row_top, row_bottom), highlight_regions)
            row_data["변경사항"] = "추가" if row_highlighted else ""
            row_data["Table_Number"] = i + 1
            processed_data.append(row_data)

    return pd.DataFrame(processed_data)

def check_highlight(row_range, highlight_regions):
    row_top, row_bottom = row_range
    for region_top, region_bottom in highlight_regions:
        if (region_top <= row_top <= region_bottom) or (region_top <= row_bottom <= region_bottom) or \
           (row_top <= region_top <= row_bottom) or (row_top <= region_bottom <= row_bottom):
            return True
    return False

def save_to_excel_with_highlight(df, output_path):
    df.to_excel(output_path, index=False)
    
    wb = load_workbook(output_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    if '변경사항' in df.columns:
        change_col_index = df.columns.get_loc('변경사항') + 1
    else:
        raise ValueError("DataFrame에 '변경사항' 컬럼이 없습니다.")

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=change_col_index).value
        if cell_value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 행 높이 자동 조정
    for row in ws.rows:
        max_height = 0
        for cell in row:
            if cell.value:
                text_lines = str(cell.value).count('\n') + 1
                text_height = text_lines * 15  # 대략적인 텍스트 높이
                if text_height > max_height:
                    max_height = text_height
        ws.row_dimensions[row[0].row].height = max_height

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows and auto-adjusted cell sizes")

def main():
    uploads_folder = "/workspaces/automation/uploads"
    output_folder = "/workspaces/automation/output"
    
    # Ensure output folder exists
    os.makedirs(output_folder, exist_ok=True)

    # Get the first PDF file from the uploads folder
    pdf_files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
    if not pdf_files:
        print("No PDF files found in the uploads folder.")
        return

    pdf_file = pdf_files[0]
    pdf_path = os.path.join(uploads_folder, pdf_file)
    output_excel_path = os.path.join(output_folder, f"{os.path.splitext(pdf_file)[0]}_extracted_tables.xlsx")

    print(f"Processing PDF file: {pdf_file}")

    doc = fitz.open(pdf_path)
    all_processed_data = []

    for page_number in range(len(doc)):
        print(f"Processing page: {page_number + 1}/{len(doc)}")

        page = doc[page_number]
        image = pdf_to_image(page)

        contours = detect_highlights(image)
        highlight_regions = get_highlight_regions(contours, image.shape[0])

        print(f"Page {page_number + 1}: Detected {len(highlight_regions)} highlighted regions")

        tables = extract_tables_with_camelot(pdf_path, page_number + 1)

        if not tables:
            print(f"Page {page_number + 1}: No tables extracted")
            continue

        processed_df = process_tables(tables, highlight_regions, image.shape[0])
        processed_df['Page_Number'] = page_number + 1
        all_processed_data.append(processed_df)

    if not all_processed_data:
        print("No processed data available.")
        return

    final_df = pd.concat(all_processed_data, ignore_index=True)

    save_to_excel_with_highlight(final_df, output_excel_path)

    print(f"All processed data has been saved to {output_excel_path}")

if __name__ == "__main__":
    main()