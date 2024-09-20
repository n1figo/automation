import fitz
import camelot
import pandas as pd
import numpy as np
import cv2
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DEBUG_MODE = True
IMAGE_OUTPUT_DIR = "/workspaces/automation/output/images"
os.makedirs(IMAGE_OUTPUT_DIR, exist_ok=True)

def int_to_rgb(color_int):
    b = color_int & 255
    g = (color_int >> 8) & 255
    r = (color_int >> 16) & 255
    return r, g, b

def is_highlighted(color, threshold=0.8):
    if isinstance(color, int):
        r, g, b = int_to_rgb(color)
    else:
        r, g, b = color
    return r > threshold * 255 and g > threshold * 255 and b < 0.5 * 255

def get_highlighted_areas(page):
    highlighted_areas = []
    for block in page.get_text("dict")["blocks"]:
        if "lines" in block:
            for line in block["lines"]:
                for span in line["spans"]:
                    if DEBUG_MODE:
                        print(f"Span color: {span['color']}")
                    if is_highlighted(span["color"]):
                        highlighted_areas.append(span["bbox"])
    return highlighted_areas

def extract_highlighted_rows_with_camelot(pdf_path, page_number, highlighted_areas):
    doc = fitz.open(pdf_path)
    page = doc[page_number]
    height = page.rect.height
    
    highlighted_rows = []
    for area in highlighted_areas:
        x1, y1, x2, y2 = area
        y1, y2 = height - y2, height - y1  # Camelot uses bottom-left as origin
        
        tables = camelot.read_pdf(
            pdf_path,
            pages=str(page_number + 1),
            flavor='stream',
            table_areas=[f"{0},{y1},{page.rect.width},{y2}"],
            line_scale=40,
            process_background=True,
            strip_text=' .\n'
        )
        
        if tables:
            row_text = ' '.join(tables[0].df.iloc[0])
            highlighted_rows.append(row_text)
            print(f"Highlighted row: {row_text}")
    
    return highlighted_rows

def extract_tables_with_camelot(pdf_path, page_number):
    print(f"Extracting tables from page {page_number} using Camelot...")
    tables = camelot.read_pdf(
        pdf_path,
        pages=str(page_number),
        flavor='lattice',
        line_scale=40,
        process_background=True,
        strip_text=' .\n'
    )
    print(f"Found {len(tables)} tables on page {page_number}")
    return tables

def process_tables(tables, highlighted_rows):
    processed_data = []
    for i, table in enumerate(tables):
        df = table.df.copy()
        df['변경사항'] = ''
        df['Table_Number'] = i + 1
        
        for idx, row in df.iterrows():
            row_text = ' '.join(str(cell) for cell in row)  # 모든 셀을 문자열로 변환
            if any(highlighted_row.lower() in row_text.lower() for highlighted_row in highlighted_rows):
                df.at[idx, '변경사항'] = '추가'
        
        processed_data.append(df)
    
    return pd.concat(processed_data, ignore_index=True)

def save_to_excel_with_highlight(df, output_path):
    df.to_excel(output_path, index=False)
    
    wb = load_workbook(output_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    change_col_index = df.columns.get_loc('변경사항') + 1

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=change_col_index).value == '추가':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)
    print(f"Data saved to '{output_path}' with highlighted rows")

def visualize_results(page, highlighted_areas, tables, output_path):
    img = page.get_pixmap()
    img_np = np.frombuffer(img.samples, dtype=np.uint8).reshape(img.height, img.width, 3)
    
    img_np_copy = img_np.copy()
    
    for area in highlighted_areas:
        cv2.rectangle(img_np_copy, (int(area[0]), int(area[1])), (int(area[2]), int(area[3])), (0, 255, 0), 2)
    
    for table in tables:
        x1, y1, x2, y2 = table._bbox
        cv2.rectangle(img_np_copy, (int(x1), int(y1)), (int(x2), int(y2)), (255, 0, 0), 2)
    
    cv2.imwrite(output_path, cv2.cvtColor(img_np_copy, cv2.COLOR_RGB2BGR))
    print(f"Visualization saved to {output_path}")

def main(pdf_path, output_excel_path):
    try:
        print("Extracting structure and tables from PDF...")

        page_number = 50  # 51페이지 (0-based index)

        doc = fitz.open(pdf_path)
        page = doc[page_number]
        highlighted_areas = get_highlighted_areas(page)

        print(f"Found {len(highlighted_areas)} highlighted areas")

        highlighted_rows = extract_highlighted_rows_with_camelot(pdf_path, page_number, highlighted_areas)

        tables = extract_tables_with_camelot(pdf_path, page_number + 1)

        processed_df = process_tables(tables, highlighted_rows)

        print(processed_df)

        viz_output_path = os.path.join(IMAGE_OUTPUT_DIR, f'page_{page_number + 1}_visualization.png')
        visualize_results(page, highlighted_areas, tables, viz_output_path)

        save_to_excel_with_highlight(processed_df, output_excel_path)

        print(f"처리된 데이터가 {output_excel_path}에 저장되었습니다.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables_pymupdf_camelot.xlsx"
    main(pdf_path, output_excel_path)