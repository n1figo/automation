import fitz  # PyMuPDF (여전히 사용 중, PDF 정보에 접근)
import pytesseract
import cv2
from pytesseract import Output
from pdf2image import convert_from_path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import os
from typing import List, Tuple, Any  # List, Tuple, Any를 import

# PDF를 이미지로 변환
def convert_pdf_to_images(pdf_path: str):
    pages = convert_from_path(pdf_path)
    image_paths = []
    for i, page in enumerate(pages):
        image_path = f'page_{i+1}.png'
        page.save(image_path, 'PNG')
        image_paths.append(image_path)
    return image_paths

# 텍스트 및 배경색을 감지
def extract_text_and_background_from_image(image_path: str):
    image = cv2.imread(image_path)
    d = pytesseract.image_to_data(image, output_type=Output.DICT)

    extracted_data = []
    n_boxes = len(d['text'])
    for i in range(n_boxes):
        if int(d['conf'][i]) > 60:  # 신뢰도 60 이상만 처리
            (x, y, w, h) = (d['left'][i], d['top'][i], d['width'][i], d['height'][i])
            text = d['text'][i]

            # 텍스트 영역의 배경 색상 확인
            background_color = image[y:y+h, x:x+w].mean(axis=(0, 1))  # 평균 색상 추출 (BGR 값)
            
            # 배경색이 특정 조건을 만족하는 경우 "추가"로 판단
            if not is_common_background_color(background_color):
                extracted_data.append((text, "추가", background_color))
            else:
                extracted_data.append((text, "유지", background_color))

    return extracted_data

# 일반적인 배경색 (흰색, 회색 등)을 제외하는 함수
def is_common_background_color(color) -> bool:
    white = (255, 255, 255)
    gray = (127, 127, 127)
    color_tolerance = 30  # 허용 오차

    def is_similar(c1, c2):
        return all(abs(c1[i] - c2[i]) <= color_tolerance for i in range(3))

    # 흰색이나 회색 배경은 일반적이므로 배경색으로 제외
    if is_similar(color, white) or is_similar(color, gray):
        return True
    return False

# PDF 테이블 추출 클래스
class PDFTableExtractor:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        
    def extract_tables_with_titles(self) -> List[Tuple[str, pd.DataFrame]]:
        all_tables = []
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            tables = self.extract_tables_from_page(page)
            titled_tables = self._assign_titles_to_tables(page, tables)
            all_tables.extend(titled_tables)
        return self._merge_tables_with_same_title(all_tables)
    
    def extract_tables_from_page(self, page: fitz.Page) -> List[Any]:
        tables = page.find_tables()
        return tables
    
    def _assign_titles_to_tables(self, page: fitz.Page, tables: List[Any]) -> List[Tuple[str, pd.DataFrame]]:
        titled_tables = []
        for table in tables:
            title = self._find_table_title(page, table)
            df = self._table_to_dataframe(table)
            titled_tables.append((title, df))
        return titled_tables
    
    def _find_table_title(self, page: fitz.Page, table: Any) -> str:
        blocks = page.get_text("dict")["blocks"]
        table_top = table.bbox[1]  # y0 좌표
        potential_titles = []
        for b in blocks:
            if 'lines' in b:
                for l in b['lines']:
                    for s in l['spans']:
                        if s['bbox'][3] < table_top and s['bbox'][3] > table_top - 50:
                            potential_titles.append(s['text'])
        
        if potential_titles:
            return " ".join(potential_titles).strip()
        return "Untitled Table"
    
    def _table_to_dataframe(self, table: Any) -> pd.DataFrame:
        df = pd.DataFrame(table.extract())
        df = df.applymap(clean_text_for_excel)
        return df
    
    def _merge_tables_with_same_title(self, tables: List[Tuple[str, pd.DataFrame]]) -> List[Tuple[str, pd.DataFrame]]:
        merged_tables = defaultdict(list)
        for title, df in tables:
            merged_tables[title].append(df)
        
        return [(title, pd.concat(dfs, ignore_index=True)) for title, dfs in merged_tables.items()]

# 엑셀 작성 및 변경사항 적용 클래스
class ExcelWriterWithChanges:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Extracted Tables"
        
    def write_tables_with_changes(self, extracted_data: List[Tuple[str, str, Tuple[int, int, int]]]):
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row = 1
        
        for text, status, bg_color in extracted_data:
            self.sheet.cell(row=row, column=1, value=text)
            self.sheet.cell(row=row, column=2, value=status)
            self.sheet.cell(row=row, column=3, value=str(bg_color))  # 배경색을 기록
            
            if status == "추가":
                for col_num in range(1, 4):
                    self.sheet.cell(row=row, column=col_num).fill = yellow_fill
                
            self.sheet.row_dimensions[row].height = 20  # 줄바꿈 허용
            self.sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            
            row += 1

        self.workbook.save(self.output_path)
        print(f"Tables saved with changes to {self.output_path}")

# PDF에서 텍스트와 배경색을 추출하고 엑셀로 출력
def main(pdf_path: str, output_excel_path: str):
    # PDF를 이미지로 변환
    image_paths = convert_pdf_to_images(pdf_path)
    
    extracted_data = []
    
    # 각 이미지에서 텍스트와 배경색을 추출
    for image_path in image_paths:
        extracted_data.extend(extract_text_and_background_from_image(image_path))
    
    # 엑셀 파일로 작성
    writer = ExcelWriterWithChanges(output_excel_path)
    writer.write_tables_with_changes(extracted_data)

if __name__ == "__main__":
    pdf_path = "/workspaces/automation/uploads/5. KB 5.10.10 플러스 건강보험(무배당)(24.05)_요약서_0801_v1.0.pdf"
    output_excel_path = "/workspaces/automation/output/extracted_tables.xlsx"
    main(pdf_path, output_excel_path)
