import os
from flask import Flask, render_template, request, send_file, abort
from werkzeug.utils import secure_filename
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
import torch
from transformers import AutoProcessor, AutoModel
from PIL import Image, ImageDraw
import fitz  # PyMuPDF
import io
import requests
from bs4 import BeautifulSoup
import difflib
import re

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

ALLOWED_EXTENSIONS = {'pdf'}

# Create uploads directory if it doesn't exist
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# CLIP model and processor loading
processor = AutoProcessor.from_pretrained("openai/clip-vit-base-patch32")
model = AutoModel.from_pretrained("openai/clip-vit-base-patch32")

def setup_browser():
    playwright = sync_playwright().start()
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()
    return playwright, browser, page

def capture_full_page(url):
    playwright, browser, page = setup_browser()
    try:
        page.goto(url)
        page.wait_for_timeout(5000)  # Wait for the page to load
        screenshot_path = os.path.join(app.config['UPLOAD_FOLDER'], "full_page.png")
        page.screenshot(path=screenshot_path, full_page=True)
        return screenshot_path
    finally:
        browser.close()
        playwright.stop()

def pdf_to_images(pdf_path):
    doc = fitz.open(pdf_path)
    images = []
    for page in doc:
        pix = page.get_pixmap()
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        images.append(img)
    doc.close()
    return images

def detect_highlights(image):
    width, height = image.size
    sections = []
    for i in range(0, height, 100):
        for j in range(0, width, 100):
            section = image.crop((j, i, j+100, i+100))
            sections.append((section, (j, i, j+100, i+100)))
    
    section_features = []
    for section, _ in sections:
        inputs = processor(images=section, return_tensors="pt")
        with torch.no_grad():
            features = model.get_image_features(**inputs)
        section_features.append(features)
    
    text_queries = ["노란색 하이라이트", "강조된 텍스트"]
    text_features = []
    for query in text_queries:
        text_inputs = processor(text=query, return_tensors="pt", padding=True)
        with torch.no_grad():
            features = model.get_text_features(**text_inputs)
        text_features.append(features)
    
    highlighted_sections = []
    for i, section_feature in enumerate(section_features):
        for text_feature in text_features:
            similarity = torch.nn.functional.cosine_similarity(section_feature, text_feature)
            if similarity > 0.5:  # Threshold
                highlighted_sections.append(sections[i][1])
                break
    
    return highlighted_sections

def create_excel_with_highlights(image_path, excel_filename, data, highlighted_sections):
    wb = Workbook()
    ws = wb.active
    
    img = XLImage(image_path)
    img.width = 800
    img.height = 600
    
    ws.add_image(img, 'A1')
    
    # Add scraped data
    info_data = [
        ['상품명', data.get('상품명', '')],
        ['보장내용', data.get('보장내용', '')],
        ['보험기간', data.get('보험기간', '')]
    ]
    for row, (key, value) in enumerate(info_data, start=ws.max_row + 2):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
    
    # Add table data
    if data.get('테이블 데이터'):
        table_data = data['테이블 데이터']
        headers = list(table_data[0].keys())
        ws.append(headers)
        for row in table_data:
            ws.append([row.get(header, '') for header in headers])
    
    # Add highlighted sections
    ws.cell(row=ws.max_row + 2, column=1, value="하이라이트된 섹션")
    for i, section in enumerate(highlighted_sections):
        img = Image.open(image_path)
        highlight = img.crop(section)
        
        # Save highlight as image
        with io.BytesIO() as output:
            highlight.save(output, format="PNG")
            highlight_image = XLImage(output)
        
        # Add highlight image to Excel
        ws.add_image(highlight_image, f'A{ws.max_row + 1}')
        ws.row_dimensions[ws.max_row].height = 75  # Adjust row height
    
    output_dir = 'output/excel'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    excel_path = os.path.join(output_dir, excel_filename)
    wb.save(excel_path)
    return excel_path

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# New functions for web-PDF comparison
def extract_pdf_text(pdf_path):
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def extract_web_content(url):
    response = requests.get(url)
    return response.text

def parse_html(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    return soup

def extract_text_with_positions(soup):
    text_positions = []
    for element in soup.find_all(text=True):
        if element.parent.name not in ['script', 'style']:
            text = element.strip()
            if text:
                parent = element.parent
                style = parent.get('style', '')
                position = re.search(r'position:\s*absolute;\s*left:\s*(\d+)px;\s*top:\s*(\d+)px', style)
                if position:
                    left, top = map(int, position.groups())
                    text_positions.append((text, left, top))
    return text_positions

def compare_texts(text1, text2):
    differ = difflib.Differ()
    diff = list(differ.compare(text1.splitlines(), text2.splitlines()))
    return diff

def find_changes(diff):
    changes = []
    for line in diff:
        if line.startswith('+ ') or line.startswith('- '):
            changes.append(line[2:])
    return changes

def highlight_changes_on_html_capture(html_capture_path, changes, text_positions):
    with Image.open(html_capture_path) as img:
        draw = ImageDraw.Draw(img)
        
        for change in changes:
            matching_positions = [pos for text, left, top in text_positions if change in text]
            for left, top in matching_positions:
                # 변경된 부분에 빨간색 테두리의 노란색 네모 표시
                draw.rectangle([left-5, top-5, left+100, top+20], outline="red", fill="yellow")
                draw.text((left, top), change[:20], fill="black")

        highlighted_capture_path = os.path.join(app.config['UPLOAD_FOLDER'], "highlighted_html_capture.png")
        img.save(highlighted_capture_path)
        return highlighted_capture_path

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        try:
            url = request.form.get('url')
            if not url:
                return "URL is required", 400

            image_path = capture_full_page(url)
            
            # Here you should implement the web scraping function to get the data
            # For now, I'll use a dummy data structure
            data = {
                '상품명': 'Sample Product',
                '보장내용': 'Sample Coverage',
                '보험기간': 'Sample Period',
                '테이블 데이터': [{'Column1': 'Value1', 'Column2': 'Value2'}]
            }
            
            highlighted_sections = []
            
            if 'file' not in request.files:
                return "No file part in the request", 400
            file = request.files['file']
            if file.filename == '':
                return "No file selected", 400
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                
                file.save(filepath)
                
                # Process PDF and detect highlights
                pdf_images = pdf_to_images(filepath)
                for img in pdf_images:
                    highlighted_sections.extend(detect_highlights(img))
                
                # Compare web content with PDF
                pdf_text = extract_pdf_text(filepath)
                html_content = extract_web_content(url)
                soup = parse_html(html_content)
                web_text = soup.get_text()
                text_positions = extract_text_with_positions(soup)

                diff = compare_texts(web_text, pdf_text)
                changes = find_changes(diff)
                highlighted_capture_path = highlight_changes_on_html_capture(image_path, changes, text_positions)
                
                # Clean up temporary files
                os.remove(filepath)
            else:
                return "Invalid file type. Please upload only PDF files.", 400
            
            excel_filename = "output.xlsx"
            excel_path = create_excel_with_highlights(highlighted_capture_path, excel_filename, data, highlighted_sections)
            return send_file(excel_path, as_attachment=True)
        except Exception as e:
            app.logger.error(f"An error occurred: {str(e)}")
            return f"An error occurred: {str(e)}", 500

    return render_template('index.html')

if __name__ == "__main__":
    app.run(debug=True)